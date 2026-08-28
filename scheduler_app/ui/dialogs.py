"""All dialog windows: Setup, AddClass, PlaceClass, SelectClass, Warnings, OpenSlots, PostAdd, BulkAdd."""

import html
import os

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QSpinBox, QCheckBox,
    QTextEdit, QTabWidget, QTreeWidget, QTreeWidgetItem, QGroupBox,
    QMessageBox, QScrollArea, QWidget, QHeaderView, QFrame, QSizePolicy,
    QTableWidget, QTableWidgetItem, QAbstractItemView, QApplication,
    QProgressDialog, QSlider, QFileDialog,
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QColor, QCursor, QShortcut, QKeySequence

from scheduler_app.translations import TRANSLATIONS, tr
from scheduler_app.models import (
    new_class, room_fits_class, new_lecturer_availability,
    needs_physical_room, LOCATION_TYPES,
    LOCATION_FACE_TO_FACE, LOCATION_ONLINE, LOCATION_LECTURER_OFFICE,
    normalize_class_data, get_effective_room_resource_for_class,
    get_location_label, get_protection_label, parse_location_type_label,
    validate_class_fields,
)
from scheduler_app import storage
from scheduler_app.logic import (
    find_valid_options, get_placed_classes,
    occupied_slots_of, classroom_of,
    parse_slot_lines, slot_meaning_changes,
    SLOT_ERROR_DUPLICATE,
)
from scheduler_app.widgets import MultiSelectButton
from scheduler_app.i18n.day_keys import DAY_KEYS, day_label, normalize_day_list, normalize_day_value


_DIALOG_STYLESHEET_TEMPLATE = """
QDialog {
    background: #F8FAFC;
    font-family: "Segoe UI", sans-serif;
}
QLabel {
    font-family: "Segoe UI", sans-serif;
}
QGroupBox {
    font-weight: bold;
    font-size: 9pt;
    border: 1px solid #CBD5E1;
    border-radius: 8px;
    margin-top: 12px;
    padding-top: 16px;
    background: white;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
    color: #1E40AF;
}
QLineEdit, QTextEdit {
    border: 1px solid #CBD5E1;
    border-radius: 4px;
    padding: 4px 8px;
    background: white;
    font-size: 9pt;
}
QLineEdit:focus, QTextEdit:focus {
    border-color: #3B82F6;
}
QCheckBox {
    font-size: 9pt;
    spacing: 6px;
}
QCheckBox::indicator {
    width: 16px;
    height: 16px;
    border-radius: 4px;
    border: 1px solid #CBD5E1;
}
QCheckBox::indicator:checked {
    background: #3B82F6;
    border-color: #3B82F6;
}
QTabBar::scroller {
    width: 56px;
}
QTabBar QToolButton {
    background: #F1F5F9;
    border: 1px solid #CBD5E1;
    border-radius: 4px;
    padding: 2px;
    margin: 2px 1px;
    width: 22px;
    height: 22px;
    min-width: 22px;
    max-width: 22px;
    min-height: 22px;
    max-height: 22px;
}
QTabBar QToolButton:hover {
    background: #DBEAFE;
    border-color: #93C5FD;
}
QTabBar QToolButton::left-arrow {
    image: url(ARROW_LEFT);
    width: 12px;
    height: 12px;
}
QTabBar QToolButton::right-arrow {
    image: url(ARROW_RIGHT);
    width: 12px;
    height: 12px;
}
"""

_dialog_stylesheet_cache = None


def _get_dialog_stylesheet():
    global _dialog_stylesheet_cache
    if _dialog_stylesheet_cache is None:
        import os
        from scheduler_app.icons import get_arrow_dir
        arrows = get_arrow_dir()
        left = os.path.join(arrows, "left.png").replace("\\", "/")
        right = os.path.join(arrows, "right.png").replace("\\", "/")
        _dialog_stylesheet_cache = (_DIALOG_STYLESHEET_TEMPLATE
                                     .replace("ARROW_LEFT", left)
                                     .replace("ARROW_RIGHT", right))
    return _dialog_stylesheet_cache


def DIALOG_STYLESHEET():
    return _get_dialog_stylesheet()


_CLASS_IO_FIELDS = [
    ("class_code", "labels.class_code", "Class Code"),
    ("name", "labels.class_name", "Class Name"),
    ("lecturer", "labels.lecturer", "Lecturer"),
    ("duration", "labels.duration", "Duration"),
    ("participants", "labels.participants", "Participants"),
    ("location_type", "labels.location_type", "Location Type"),
    ("targets", "labels.targets", "Targets"),
    ("joint", "labels.joint", "Joint"),
    ("pinned", "labels.pinned", "Pinned"),
    ("pin_day", "labels.pin_day", "Pin Day"),
    ("pin_time", "labels.pin_time", "Pin Time"),
    ("pin_room", "labels.pin_room", "Pin Room"),
    ("allowed_days", "constraints.allowed_days", "Allowed Days"),
    ("excluded_days", "constraints.excluded_days", "Excluded Days"),
    ("allowed_times", "constraints.allowed_start_times", "Allowed Start Times"),
    ("excluded_times", "constraints.excluded_start_times", "Excluded Start Times"),
    ("required_classrooms", "constraints.required_classrooms", "Required Classrooms"),
    ("excluded_classrooms", "constraints.excluded_classrooms", "Excluded Classrooms"),
]


def _trim_label(text):
    return str(text or "").strip().rstrip(":").strip()


def _excel_file_filter(include_legacy=False):
    pattern = "*.xlsx *.xls" if include_legacy else "*.xlsx"
    return f"{tr('labels.excel_files')} ({pattern})"


def _safe_export_stem(text, fallback="export"):
    stem = _trim_label(text) or fallback
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if ch in invalid else ch for ch in stem)
    return cleaned.rstrip(". ") or fallback


def _translated_bool(value):
    return tr("common.yes") if value else tr("common.no")


def _tr_or(label_key, fallback):
    """``tr(label_key)``, or *fallback* when the catalogue has no such key.

    ST-UI-011. The idiom this replaces — ``tr(label_key) or fallback`` — is dead
    code: ``tr`` returns the *key itself* when it cannot resolve one, and a key
    is a non-empty string, so ``or`` never selects the fallback. That is exactly
    why ``labels.targets`` reached the user as the literal text
    ``labels.targets`` instead of the ``"Targets"`` sitting right beside it in
    ``_CLASS_IO_FIELDS``. ``tier_enforcement._display_name`` already had the
    working form; this is it.
    """
    translated = tr(label_key)
    return fallback if translated == label_key else translated


def _class_io_headers():
    return [_trim_label(_tr_or(label_key, fallback))
            for _, label_key, fallback in _CLASS_IO_FIELDS]


def _class_io_header_map():
    return {
        field: _trim_label(_tr_or(label_key, fallback))
        for field, label_key, fallback in _CLASS_IO_FIELDS
    }


def _class_io_reverse_header_map():
    reverse = {}
    for field, label_key, fallback in _CLASS_IO_FIELDS:
        # ST-UI-011, backward compatibility. The RAW KEY is a permanent import
        # alias, and dropping it would be a silent data loss the fix itself
        # caused. While `labels.targets` was missing from the catalogue,
        # `tr()` returned the key, so every workbook DERSİS exported wrote
        # `labels.targets` as its Targets header — and re-imported it, because
        # this same expression registered that string. Now that the key
        # resolves, the literal would vanish from the map and every previously
        # exported file would import its classes with no target groups at all.
        for label in (_trim_label(fallback), _trim_label(label_key),
                      _trim_label(_tr_or(label_key, fallback))):
            if label:
                reverse[label.casefold()] = field
        for lang_dict in TRANSLATIONS.values():
            label = _trim_label(lang_dict.get(label_key, fallback))
            if label:
                reverse[label.casefold()] = field
    return reverse


def _canonicalize_class_io_row(row):
    canonical = {}
    reverse = _class_io_reverse_header_map()
    for key, value in dict(row).items():
        field = reverse.get(_trim_label(key).casefold())
        if field:
            canonical[field] = value
    return canonical


def _install_select_all_shortcut(owner, widget, callback=None):
    """Bind Ctrl+A to select-all behavior for widgets with multi-selection."""
    sc = QShortcut(QKeySequence.StandardKey.SelectAll, widget)
    sc.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
    sc.activated.connect(callback or widget.selectAll)
    if not hasattr(owner, "_select_all_shortcuts"):
        owner._select_all_shortcuts = []
    owner._select_all_shortcuts.append(sc)


def _install_table_keyboard_shortcuts(owner, table, delete_callback,
                                       copy_callback=None, paste_callback=None):
    """Install Delete, Ctrl+C, and Ctrl+V shortcuts on a QTableWidget.

    Args:
        owner: dialog that owns the shortcuts (prevents garbage collection)
        table: QTableWidget to add shortcuts to
        delete_callback: called when Delete is pressed (no args)
        copy_callback: called when Ctrl+C is pressed (no args); if None,
                       a default tab-separated copy is used
        paste_callback: called when Ctrl+V is pressed (no args); if None,
                        paste is not installed
    """
    if not hasattr(owner, "_table_kbd_shortcuts"):
        owner._table_kbd_shortcuts = []
    ctx = Qt.ShortcutContext.WidgetWithChildrenShortcut

    # Delete
    sc_del = QShortcut(QKeySequence(QKeySequence.StandardKey.Delete), table)
    sc_del.setContext(ctx)
    sc_del.activated.connect(delete_callback)
    owner._table_kbd_shortcuts.append(sc_del)

    # Ctrl+C
    if copy_callback is None:
        copy_callback = lambda: _default_copy_table_rows(table)
    sc_copy = QShortcut(QKeySequence(QKeySequence.StandardKey.Copy), table)
    sc_copy.setContext(ctx)
    sc_copy.activated.connect(copy_callback)
    owner._table_kbd_shortcuts.append(sc_copy)

    # Ctrl+V
    if paste_callback is not None:
        sc_paste = QShortcut(QKeySequence(QKeySequence.StandardKey.Paste), table)
        sc_paste.setContext(ctx)
        sc_paste.activated.connect(paste_callback)
        owner._table_kbd_shortcuts.append(sc_paste)


def _default_copy_table_rows(table):
    """Copy selected rows from a QTableWidget to clipboard as tab-separated text."""
    rows = sorted(set(idx.row() for idx in table.selectedIndexes()))
    if not rows:
        return
    lines = []
    col_count = table.columnCount()
    for row in rows:
        cells = []
        for col in range(col_count):
            widget = table.cellWidget(row, col)
            item = table.item(row, col)
            if widget:
                # Handle common widget types
                if hasattr(widget, 'value'):  # QSpinBox
                    cells.append(str(widget.value()))
                elif hasattr(widget, 'currentText'):  # QComboBox
                    cells.append(widget.currentText())
                elif hasattr(widget, 'isChecked'):  # QCheckBox
                    cells.append(_translated_bool(widget.isChecked()))
                elif hasattr(widget, 'text'):  # QPushButton / MultiSelectButton
                    cells.append(widget.text())
                else:
                    cells.append("")
            elif item:
                cells.append(item.text())
            else:
                cells.append("")
        lines.append("\t".join(cells))
    text = "\n".join(lines)
    QApplication.clipboard().setText(text)


def _ensure_excel_deps(parent):
    """Return True if pandas+openpyxl are available, auto-installing if needed."""
    try:
        import pandas  # noqa: F401
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        pass
    reply = QMessageBox.question(
        parent, tr("dialogs.import.excel_title"),
        tr("errors.pandas_openpyxl_required")
        + "\n\n" + tr("dialogs.install.prompt"),
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
    )
    if reply != QMessageBox.StandardButton.Yes:
        return False
    import subprocess
    import sys
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "pandas", "openpyxl"],
            timeout=120,
        )
        return True
    except Exception as exc:
        QMessageBox.warning(parent, tr("status.import_failed"), str(exc))
        return False


def _offer_template_or_nothing(parent, sheet_name, headers):
    """When export has no data, offer to generate a template instead."""
    reply = QMessageBox.question(
        parent, tr("dialogs.export.excel_title"),
        tr("dialogs.export.no_data_template"),
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
    )
    if reply != QMessageBox.StandardButton.Yes:
        QMessageBox.information(parent, tr("dialogs.export.excel_title"), tr("warnings.nothing_to_export"))
        return
    if not _ensure_excel_deps(parent):
        return
    import pandas as pd
    path, _ = QFileDialog.getSaveFileName(
        parent, tr("dialogs.export.excel_title"),
        os.path.join(
            storage.sub_dir(storage.EXPORTS_DIR),
            f"{_safe_export_stem(sheet_name)}_template.xlsx",
        ),
        _excel_file_filter())
    if not path:
        return
    try:
        df = pd.DataFrame(columns=headers)
        df.to_excel(path, index=False, sheet_name=sheet_name)
        QMessageBox.information(
            parent, tr("status.template_generated"),
            f"{tr('status.excel_template_saved')} {os.path.basename(path)}")
    except Exception as exc:
        QMessageBox.warning(parent, tr("dialogs.export.excel_title"), str(exc))


def _safe_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _split_csv(value):
    return [x.strip() for x in _safe_text(value).split(",") if x.strip()]


def _parse_bool(value, default=False):
    text = _safe_text(value).lower()
    if not text:
        return default
    true_aliases = {"1", "true", "yes", "y", "on", "checked", "x"}
    false_aliases = {"0", "false", "no", "n", "off", "unchecked"}
    for lang_dict in TRANSLATIONS.values():
        yes_label = _trim_label(lang_dict.get("common.yes", "")).casefold()
        no_label = _trim_label(lang_dict.get("common.no", "")).casefold()
        if yes_label:
            true_aliases.add(yes_label)
        if no_label:
            false_aliases.add(no_label)
    if text in true_aliases:
        return True
    if text in false_aliases:
        return False
    return default


def _to_int(value, default=0, minimum=0):
    try:
        iv = int(float(value))
    except Exception:
        iv = default
    return max(minimum, iv)


def _serialize_targets(targets):
    return ", ".join(f"{t['year']}/{t['branch']}" for t in targets)


def _parse_targets(raw, state):
    lookup = {}
    for yr in state.get("years", {}):
        for br in state["years"][yr]:
            lookup[(yr.lower(), br.lower())] = {"year": yr, "branch": br}

    out = []
    for token in _split_csv(raw):
        if "/" not in token:
            continue
        yr_raw, br_raw = token.split("/", 1)
        key = (yr_raw.strip().lower(), br_raw.strip().lower())
        if key in lookup and lookup[key] not in out:
            out.append(lookup[key])
    return out


def _class_to_io_row(cls):
    headers = _class_io_header_map()
    return {
        headers["class_code"]: cls.get("class_code", ""),
        headers["name"]: cls.get("name", ""),
        headers["lecturer"]: cls.get("lecturer", ""),
        headers["duration"]: int(cls.get("duration", 1) or 1),
        headers["participants"]: int(cls.get("participants", 0) or 0),
        headers["location_type"]: get_location_label(
            cls.get("location_type", LOCATION_FACE_TO_FACE)),
        headers["targets"]: _serialize_targets(cls.get("targets", [])),
        headers["joint"]: _translated_bool(cls.get("joint_session", False)),
        headers["pinned"]: _translated_bool(cls.get("pinned", False)),
        headers["pin_day"]: day_label(cls["pinned_day"]) if cls.get("pinned_day") else "",
        headers["pin_time"]: cls.get("pinned_time", ""),
        headers["pin_room"]: cls.get("pinned_classroom", "") if needs_physical_room(cls) else "",
        headers["allowed_days"]: ", ".join(day_label(d) for d in cls.get("allowed_days", [])),
        headers["excluded_days"]: ", ".join(day_label(d) for d in cls.get("excluded_days", [])),
        headers["allowed_times"]: ", ".join(cls.get("allowed_times", [])),
        headers["excluded_times"]: ", ".join(cls.get("excluded_times", [])),
        headers["required_classrooms"]: ", ".join(cls.get("required_classrooms", [])) if needs_physical_room(cls) else "",
        headers["excluded_classrooms"]: ", ".join(cls.get("excluded_classrooms", [])) if needs_physical_room(cls) else "",
    }


def _class_from_io_row(row, state):
    row = _canonicalize_class_io_row(row)
    cls = new_class()

    cls["class_code"] = _safe_text(row.get("class_code"))
    cls["name"] = _safe_text(row.get("name"))
    cls["lecturer"] = _safe_text(row.get("lecturer"))
    cls["duration"] = _to_int(row.get("duration"), default=1, minimum=1)
    cls["participants"] = _to_int(row.get("participants"), default=0, minimum=0)
    cls["location_type"] = parse_location_type_label(row.get("location_type"))

    cls["targets"] = _parse_targets(row.get("targets"), state)
    cls["joint_session"] = _parse_bool(row.get("joint"), default=False)

    days = set(state.get("days", []))
    slots = set(state.get("slots", []))
    rooms = set(state.get("classrooms", []))

    def _normalize_days(raw_csv):
        """Normalize imported day values case-insensitively to matching day keys."""
        result = []
        for d in _split_csv(raw_csv):
            nk = normalize_day_value(d)
            if nk and nk in days:
                result.append(nk)
        return result

    cls["allowed_days"] = _normalize_days(row.get("allowed_days"))
    cls["excluded_days"] = _normalize_days(row.get("excluded_days"))
    cls["allowed_times"] = [t for t in _split_csv(row.get("allowed_times")) if t in slots]
    cls["excluded_times"] = [t for t in _split_csv(row.get("excluded_times")) if t in slots]
    cls["required_classrooms"] = [r for r in _split_csv(row.get("required_classrooms")) if r in rooms]
    cls["excluded_classrooms"] = [r for r in _split_csv(row.get("excluded_classrooms")) if r in rooms]

    pinned = _parse_bool(row.get("pinned"), default=False)
    pin_day_raw = _safe_text(row.get("pin_day"))
    pin_day = normalize_day_value(pin_day_raw) if pin_day_raw else None
    pin_time = _safe_text(row.get("pin_time"))
    pin_room = _safe_text(row.get("pin_room"))
    if pinned and pin_day and pin_day in days and pin_time in slots:
        if needs_physical_room(cls):
            if pin_room in rooms:
                cls["pinned"] = True
                cls["pinned_day"] = pin_day
                cls["pinned_time"] = pin_time
                cls["pinned_classroom"] = pin_room
        else:
            cls["pinned"] = True
            cls["pinned_day"] = pin_day
            cls["pinned_time"] = pin_time
            cls["pinned_classroom"] = None

    return normalize_class_data(cls)


def _import_class_rows(parent, state):
    if not _ensure_excel_deps(parent):
        return []
    import pandas as pd

    path, _ = QFileDialog.getOpenFileName(
        parent, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
        _excel_file_filter(include_legacy=True))
    if not path:
        return []
    try:
        df = pd.read_excel(path, sheet_name=0, header=0)
        if df.empty:
            QMessageBox.information(parent, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
            return []
        classes = []
        for _, row in df.iterrows():
            canonical = _canonicalize_class_io_row(row)
            row_name = _safe_text(canonical.get("name"))
            row_lecturer = _safe_text(canonical.get("lecturer"))
            if not row_name and not row_lecturer:
                continue
            classes.append(_class_from_io_row(row, state))
        if not classes:
            QMessageBox.information(parent, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
            return []
        QMessageBox.information(
            parent, tr("status.import_successful"),
            f"{len(classes)} {tr('status.classes')}")
        return classes
    except Exception as exc:
        QMessageBox.warning(parent, tr("status.import_failed"), str(exc))
        return []


def _export_class_rows(parent, rows, sheet_name=None):
    sheet_name = sheet_name or (_trim_label(tr("labels.classes")) or "Classes")
    headers = _class_io_headers()
    if not rows:
        _offer_template_or_nothing(parent, sheet_name, headers)
        return
    if not _ensure_excel_deps(parent):
        return
    import pandas as pd
    path, _ = QFileDialog.getSaveFileName(
        parent, tr("dialogs.export.excel_title"),
        os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{_safe_export_stem(sheet_name)}.xlsx"),
        _excel_file_filter())
    if not path:
        return
    try:
        df = pd.DataFrame(rows)
        for col in headers:
            if col not in df.columns:
                df[col] = ""
        df = df[headers]
        df.to_excel(path, index=False, sheet_name=sheet_name)
        QMessageBox.information(
            parent, tr("dialogs.export.excel_title"),
            f"{tr('status.exported_to')} {os.path.basename(path)}")
    except Exception as exc:
        QMessageBox.warning(parent, tr("dialogs.export.excel_title"), str(exc))


def _generate_class_template(parent):
    """Generate a blank Excel template with headers, guidance, and an example row."""
    if not _ensure_excel_deps(parent):
        return
    import pandas as pd
    sheet_name = _trim_label(tr("labels.classes")) or "Classes"
    path, _ = QFileDialog.getSaveFileName(
        parent, tr("dialogs.template.title"),
        os.path.join(
            storage.sub_dir(storage.EXPORTS_DIR),
            f"{_safe_export_stem(sheet_name)}_template.xlsx",
        ),
        _excel_file_filter())
    if not path:
        return
    try:
        headers = _class_io_header_map()
        ordered_headers = _class_io_headers()
        # Example row showing expected value formats
        example = {
            headers["class_code"]: "CS101",
            headers["name"]: tr("template.class_example.name"),
            headers["lecturer"]: tr("template.class_example.lecturer"),
            headers["duration"]: 2,
            headers["participants"]: 30,
            headers["location_type"]: get_location_label(LOCATION_FACE_TO_FACE),
            headers["targets"]: tr("template.class_example.targets"),
            headers["joint"]: _translated_bool(True),
            headers["pinned"]: _translated_bool(False),
            headers["pin_day"]: "",
            headers["pin_time"]: "",
            headers["pin_room"]: "",
            headers["allowed_days"]: "",
            headers["excluded_days"]: "",
            headers["allowed_times"]: "",
            headers["excluded_times"]: "",
            headers["required_classrooms"]: "",
            headers["excluded_classrooms"]: "",
        }
        df = pd.DataFrame([example], columns=ordered_headers)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            # Add guidance comment to header row
            import openpyxl
            from openpyxl.comments import Comment
            guidance = {
                "A1": tr("template.class_guide.class_code"),
                "B1": tr("template.class_guide.name"),
                "C1": tr("template.class_guide.lecturer"),
                "D1": tr("template.class_guide.duration"),
                "E1": tr("template.class_guide.participants"),
                "F1": tr("template.class_guide.location_type"),
                "G1": tr("template.class_guide.targets"),
                "H1": tr("template.class_guide.joint"),
                "I1": tr("template.class_guide.pinned"),
                "J1": tr("template.class_guide.pin_day"),
                "K1": tr("template.class_guide.pin_time"),
                "L1": tr("template.class_guide.pin_room"),
                "M1": tr("template.class_guide.allowed_days"),
                "N1": tr("template.class_guide.excluded_days"),
                "O1": tr("template.class_guide.allowed_times"),
                "P1": tr("template.class_guide.excluded_times"),
                "Q1": tr("template.class_guide.required_classrooms"),
                "R1": tr("template.class_guide.excluded_classrooms"),
            }
            for cell_ref, text in guidance.items():
                cell = ws[cell_ref]
                cell.comment = Comment(text, tr("template.class_guide.author"))
            # Mark example row
            from openpyxl.styles import Font as XlFont, PatternFill
            example_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            example_font = XlFont(italic=True, color="856404")
            for col_idx in range(1, len(ordered_headers) + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.fill = example_fill
                cell.font = example_font
            # Auto-fit column widths
            for col_idx, header in enumerate(ordered_headers, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) + 4, 15)
        QMessageBox.information(
            parent, tr("dialogs.template.title"),
            tr("dialogs.template.saved").format(path=os.path.basename(path)))
    except Exception as exc:
        QMessageBox.warning(parent, tr("dialogs.template.title"), str(exc))


# ── Simple Text Input Dialog ────────────────────────────────────────────────

class TextInputDialog(QDialog):
    def __init__(self, parent, title, prompt, default=""):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle(title)
        self.setMinimumWidth(350)
        self.result = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(prompt))
        self.entry = QLineEdit(default)
        self.entry.selectAll()
        layout.addWidget(self.entry)

        btn_row = QHBoxLayout()
        ok_btn = QPushButton(tr("buttons.ok"))
        ok_btn.clicked.connect(self._ok)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        btn_row.addStretch()
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        self.entry.returnPressed.connect(self._ok)
        self.entry.setFocus()

    def _ok(self):
        self.result = self.entry.text()
        self.accept()


# ── Setup Dialog ─────────────────────────────────────────────────────────────

class SetupDialog(QDialog):
    def __init__(self, parent, state, focus_lecturer=None):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u2699  " + tr("dialogs.setup.title"))
        self.state = state
        self.result = False
        self.setMinimumSize(520, 480)
        self._focus_lecturer_name = (focus_lecturer or "").strip()
        self._select_all_shortcuts = []

        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        tabs = self.tabs
        layout.addWidget(tabs)

        # ── Days & Time tab ──
        dt_w = QWidget()
        dt = QVBoxLayout(dt_w)
        dt.setSpacing(10)
        dt.setContentsMargins(8, 8, 8, 8)

        day_box = QGroupBox(tr("setup.days"))
        day_layout = QVBoxLayout(day_box)
        day_layout.setSpacing(8)
        day_layout.addWidget(QLabel(tr("setup.select_weekdays")))
        chip_row = QHBoxLayout()
        chip_row.setSpacing(8)
        self.day_buttons = {}
        selected_days = normalize_day_list(state.get("days", []))
        if not selected_days:
            selected_days = DAY_KEYS[:5]
        for key in DAY_KEYS:
            btn = QPushButton(day_label(key))
            btn.setCheckable(True)
            btn.setChecked(key in selected_days)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(
                """
                QPushButton {
                    border: 1px solid #CBD5E1;
                    border-radius: 14px;
                    padding: 5px 12px;
                    background: #F8FAFC;
                    color: #334155;
                    font-weight: 600;
                }
                QPushButton:checked {
                    background: #DBEAFE;
                    border-color: #3B82F6;
                    color: #1D4ED8;
                }
                """
            )
            chip_row.addWidget(btn)
            self.day_buttons[key] = btn
        chip_row.addStretch()
        day_layout.addLayout(chip_row)

        day_quick = QHBoxLayout()
        day_quick.setSpacing(6)
        quick_actions = [
            (tr("buttons.select_all"), self._select_all_days),
            (tr("setup.select_weekdays_only"), self._select_weekdays_only),
            (tr("setup.select_weekends"), self._select_weekends),
            (tr("buttons.clear"), self._clear_days),
            (tr("setup.invert_day_selection"), self._invert_days),
        ]
        _quick_btn_style = ("font-size: 8pt; padding: 3px 8px; "
                            "border: 1px solid #CBD5E1; border-radius: 4px; "
                            "background: #F1F5F9; color: #475569;")
        for label, handler in quick_actions:
            qb = QPushButton(label)
            qb.setStyleSheet(_quick_btn_style)
            qb.clicked.connect(handler)
            day_quick.addWidget(qb)
        day_quick.addStretch()
        day_layout.addLayout(day_quick)

        days_io = QHBoxLayout()
        days_io.addStretch()
        _b = QPushButton("⬇ " + tr("dialogs.import.excel_title"))
        _b.clicked.connect(self._import_days_from_excel)
        days_io.addWidget(_b)
        _b = QPushButton("⬆ " + tr("dialogs.export.excel_title"))
        _b.clicked.connect(self._export_days_to_excel)
        days_io.addWidget(_b)
        day_layout.addLayout(days_io)
        dt.addWidget(day_box)

        slot_box = QGroupBox(tr("setup.time_slots"))
        slot_layout = QVBoxLayout(slot_box)
        slot_layout.setSpacing(8)
        slot_layout.addWidget(QLabel(tr("setup.enter_time_slots")))
        self.slots_text = QTextEdit()
        self.slots_text.setMaximumHeight(180)
        default_slots = "\n".join(state["slots"]) if state["slots"] else "08:00\n09:00\n10:00\n11:00\n12:00\n13:00\n14:00\n15:00"
        self.slots_text.setPlainText(default_slots)
        slot_layout.addWidget(self.slots_text)

        # ST-UI-021. A live status strip rather than a structured widget.
        # A QTimeEdit-per-row would hard-code HH:MM -- the one rule the grid
        # must NOT have, since nothing parses a slot as a time and "1. Ders"
        # is a legitimate label -- and a list with move-up/move-down would
        # turn reordering, which silently corrupts a committed timetable,
        # into a one-click affordance. Validate in place instead; both Excel
        # buttons below keep working untouched because the widget is the same.
        self._slots_status = QLabel("")
        self._slots_status.setWordWrap(True)
        self._slots_status.setStyleSheet("font-size: 8pt; padding: 2px;")
        slot_layout.addWidget(self._slots_status)
        self.slots_text.textChanged.connect(self._refresh_slot_status)
        # `setPlainText` above ran BEFORE that connection, so nothing evaluated
        # the text the dialog opens with. Without this call the strip is blank
        # until the user types — and the case it exists for is precisely a
        # saved file that ALREADY contains a duplicate, where the user has no
        # reason to touch the box at all and would meet the problem as a modal
        # after making ten other edits.
        self._refresh_slot_status()

        slots_io = QHBoxLayout()
        slots_io.addStretch()
        _b = QPushButton("⬇ " + tr("dialogs.import.excel_title"))
        _b.clicked.connect(lambda: self._import_text_from_excel(self.slots_text, tr("setup.time_slots")))
        slots_io.addWidget(_b)
        _b = QPushButton("⬆ " + tr("dialogs.export.excel_title"))
        _b.clicked.connect(lambda: self._export_text_to_excel(self.slots_text, tr("setup.time_slots"), [tr("setup.time_slot")]))
        slots_io.addWidget(_b)
        slot_layout.addLayout(slots_io)
        dt.addWidget(slot_box, 1)

        tabs.addTab(dt_w, "📅⏰  " + tr("setup.days_time"))

        # ── Classrooms tab ──
        rooms_w = QWidget()
        rl = QVBoxLayout(rooms_w)
        rl.setSpacing(8)
        rl.setContentsMargins(8, 8, 8, 8)
        rl.addWidget(QLabel(tr("setup.enter_classrooms_capacity")))
        # Search bar for classrooms
        self._rooms_search = QLineEdit()
        self._rooms_search.setPlaceholderText(tr("setup.classrooms_search_placeholder"))
        self._rooms_search.setClearButtonEnabled(True)
        self._rooms_search.textChanged.connect(self._filter_rooms_table)
        rl.addWidget(self._rooms_search)
        self.rooms_table = QTableWidget()
        self.rooms_table.setColumnCount(2)
        self.rooms_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.rooms_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.rooms_table.setAlternatingRowColors(True)
        self.rooms_table.setStyleSheet(
            "QTableWidget { gridline-color: #E2E8F0; font-size: 9pt; }"
            "QTableWidget::item:selected { background: #DBEAFE; color: #1E293B; }")
        self.rooms_table.setHorizontalHeaderLabels([tr("setup.classroom_name"), tr("setup.capacity")])
        self.rooms_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.rooms_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.ResizeToContents)
        capacities = state.get("classroom_capacities", {})
        if state["classrooms"]:
            self.rooms_table.setRowCount(len(state["classrooms"]))
            for i, rm in enumerate(state["classrooms"]):
                self.rooms_table.setItem(i, 0, QTableWidgetItem(rm))
                cap_spin = QSpinBox()
                cap_spin.setRange(0, 9999)
                cap_spin.setValue(capacities.get(rm, 0))
                self.rooms_table.setCellWidget(i, 1, cap_spin)
        rl.addWidget(self.rooms_table)
        rooms_btn_row = QHBoxLayout()
        add_room_btn = QPushButton(tr("buttons.add_classroom"))
        add_room_btn.clicked.connect(self._add_room_row)
        rooms_btn_row.addWidget(add_room_btn)
        del_room_btn = QPushButton(tr("buttons.remove_classroom"))
        del_room_btn.clicked.connect(self._remove_room_row)
        rooms_btn_row.addWidget(del_room_btn)
        rooms_btn_row.addStretch()
        _b = QPushButton("\u2B07 " + tr("dialogs.import.excel_title"))
        _b.clicked.connect(self._import_rooms_from_excel)
        rooms_btn_row.addWidget(_b)
        _b = QPushButton("\u2B06 " + tr("dialogs.export.excel_title"))
        _b.clicked.connect(self._export_rooms_to_excel)
        rooms_btn_row.addWidget(_b)
        rl.addLayout(rooms_btn_row)
        tabs.addTab(rooms_w, "\U0001F3E0  " + tr("setup.classrooms"))
        _install_select_all_shortcut(self, self.rooms_table)
        _install_table_keyboard_shortcuts(
            self, self.rooms_table,
            delete_callback=self._remove_room_row,
            paste_callback=self._paste_room_rows,
        )

        # ── Lecturers tab (master-detail) ──
        self._lec_tab_index = 2  # index of Lecturers tab
        self._lec_availability = {}  # {lecturer_name: availability_dict}
        lec_w = QWidget()
        ll = QVBoxLayout(lec_w)
        ll.setSpacing(8)
        ll.setContentsMargins(8, 8, 8, 8)

        # Search and filter bar
        lec_toolbar = QHBoxLayout()
        self._lec_search = QLineEdit()
        self._lec_search.setPlaceholderText(tr("setup.lec_search_placeholder"))
        self._lec_search.setClearButtonEnabled(True)
        self._lec_search.textChanged.connect(self._filter_lec_table)
        lec_toolbar.addWidget(self._lec_search)
        self._lec_filter_combo = QComboBox()
        self._lec_filter_combo.addItems([
            tr("setup.lec_filter_all"),
            tr("setup.lec_filter_with"),
            tr("setup.lec_filter_without"),
            tr("setup.lec_filter_conflicts"),
        ])
        self._lec_filter_combo.currentIndexChanged.connect(self._filter_lec_table)
        lec_toolbar.addWidget(self._lec_filter_combo)
        ll.addLayout(lec_toolbar)

        # Overview label
        self._lec_overview_label = QLabel("")
        self._lec_overview_label.setStyleSheet(
            "font-size: 8pt; color: #64748B; padding: 2px 0;")
        ll.addWidget(self._lec_overview_label)

        # Master table: Name, Status, Summary
        self.lec_table = QTableWidget()
        self.lec_table.setColumnCount(3)
        self.lec_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.lec_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.lec_table.setAlternatingRowColors(True)
        self.lec_table.setStyleSheet(
            "QTableWidget { gridline-color: #E2E8F0; font-size: 9pt; }"
            "QTableWidget::item:selected { background: #DBEAFE; color: #1E293B; }")
        self.lec_table.setHorizontalHeaderLabels([
            tr("setup.lecturer_name"),
            tr("setup.lec_constraint_status"),
            tr("setup.lec_constraint_summary"),
        ])
        self.lec_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.lec_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Interactive)
        self.lec_table.setColumnWidth(1, 80)
        self.lec_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch)
        self.lec_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed)
        self.lec_table.doubleClicked.connect(self._on_lec_double_click)

        avail_data = state.get("lecturer_availability", {})
        existing_lecs = state.get("lecturers", [])
        if existing_lecs:
            self.lec_table.setRowCount(len(existing_lecs))
            for i, name in enumerate(existing_lecs):
                self.lec_table.setItem(i, 0, QTableWidgetItem(name))
                av = avail_data.get(name, new_lecturer_availability())
                self._lec_availability[name] = dict(av)
                self._update_lec_row_display(i, av)
        ll.addWidget(self.lec_table)

        lec_btn_row = QHBoxLayout()
        add_lec_btn = QPushButton(tr("buttons.add_lecturer"))
        add_lec_btn.clicked.connect(self._add_lec_row)
        lec_btn_row.addWidget(add_lec_btn)
        del_lec_btn = QPushButton(tr("buttons.remove_lecturer"))
        del_lec_btn.clicked.connect(self._remove_lec_row)
        lec_btn_row.addWidget(del_lec_btn)
        edit_lec_btn = QPushButton(tr("setup.lec_edit_constraints"))
        edit_lec_btn.clicked.connect(self._edit_lec_constraints)
        lec_btn_row.addWidget(edit_lec_btn)
        clear_lec_btn = QPushButton(tr("setup.lec_clear_constraints"))
        clear_lec_btn.setStyleSheet(
            "QPushButton { color: #DC2626; } QPushButton:hover { background: #FEE2E2; }")
        clear_lec_btn.clicked.connect(self._clear_all_lec_constraints)
        lec_btn_row.addWidget(clear_lec_btn)
        lec_btn_row.addStretch()
        _b = QPushButton("\u2B07 " + tr("dialogs.import.excel_title"))
        _b.clicked.connect(self._import_lecturers_from_excel)
        lec_btn_row.addWidget(_b)
        _b = QPushButton("\u2B06 " + tr("dialogs.export.excel_title"))
        _b.clicked.connect(self._export_lecturers_to_excel)
        lec_btn_row.addWidget(_b)
        ll.addLayout(lec_btn_row)
        tabs.addTab(lec_w, "\U0001F468\u200D\U0001F3EB  " + tr("setup.lecturers"))
        _install_select_all_shortcut(self, self.lec_table)
        _install_table_keyboard_shortcuts(
            self, self.lec_table,
            delete_callback=self._remove_lec_row,
            paste_callback=self._paste_lec_rows,
        )
        self._update_lec_overview()

        # ── Years & Branches tab ──
        years_w = QWidget()
        yl = QVBoxLayout(years_w)
        yl.setSpacing(8)
        yl.setContentsMargins(8, 8, 8, 8)
        self.years_tree = QTreeWidget()
        self.years_tree.setHeaderLabels([
            tr("setup.year_name_col"),
            tr("setup.year_groups_col"),
            tr("setup.year_count_col"),
        ])
        self.years_tree.setColumnCount(3)
        self.years_tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.years_tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.years_tree.header().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.years_tree.setColumnWidth(0, 140)
        self.years_tree.setAlternatingRowColors(True)
        self.years_tree.setStyleSheet(
            "QTreeWidget { font-size: 9pt; }"
            "QTreeWidget::item { padding: 3px 0; }"
            "QTreeWidget::item:selected { background: #DBEAFE; color: #1E293B; }")
        yl.addWidget(self.years_tree)

        self._years_data = dict(state["years"])
        self._refresh_years_tree()

        btn_frame = QHBoxLayout()
        for label, fn in [(tr("setup.add_year"), self._add_year),
                          (tr("setup.edit_year"), self._edit_year),
                          (tr("setup.remove_year"), self._remove_year)]:
            b = QPushButton(label)
            b.clicked.connect(fn)
            btn_frame.addWidget(b)
        btn_frame.addStretch()
        _b = QPushButton("\u2B07 " + tr("dialogs.import.excel_title"))
        _b.clicked.connect(self._import_years_from_excel)
        btn_frame.addWidget(_b)
        _b = QPushButton("\u2B06 " + tr("dialogs.export.excel_title"))
        _b.clicked.connect(self._export_years_to_excel)
        btn_frame.addWidget(_b)
        yl.addLayout(btn_frame)
        tabs.addTab(years_w, "\U0001F393  " + tr("setup.years_branches"))
        # Keyboard shortcuts for Years tree
        _install_select_all_shortcut(self, self.years_tree)
        if not hasattr(self, "_table_kbd_shortcuts"):
            self._table_kbd_shortcuts = []
        _sc_del_yr = QShortcut(QKeySequence(QKeySequence.StandardKey.Delete), self.years_tree)
        _sc_del_yr.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _sc_del_yr.activated.connect(self._remove_year)
        self._table_kbd_shortcuts.append(_sc_del_yr)
        _sc_copy_yr = QShortcut(QKeySequence(QKeySequence.StandardKey.Copy), self.years_tree)
        _sc_copy_yr.setContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        _sc_copy_yr.activated.connect(self._copy_years)
        self._table_kbd_shortcuts.append(_sc_copy_yr)

        # Connect tab switch to sync dropdowns
        tabs.currentChanged.connect(self._on_tab_changed)

        # ── Bottom buttons ──
        bottom = QHBoxLayout()
        bottom.setContentsMargins(0, 8, 0, 0)
        bottom.addStretch()
        ok_btn = QPushButton(tr("buttons.ok"))
        ok_btn.setStyleSheet(
            "QPushButton { background: #3B82F6; color: white; border: none;"
            "  border-radius: 6px; font-weight: bold; padding: 8px 24px;"
            "  font-size: 10pt; }"
            "QPushButton:hover { background: #2563EB; }")
        ok_btn.clicked.connect(self._ok)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.setStyleSheet(
            "QPushButton { padding: 8px 20px; font-size: 10pt;"
            "  border: 1px solid #CBD5E1; border-radius: 6px; }"
            "QPushButton:hover { background: #F1F5F9; }")
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(ok_btn)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

        if self._focus_lecturer_name:
            QTimer.singleShot(0, self._focus_lecturer_row)

    def _add_room_row(self):
        from scheduler_app.ui.tier_enforcement import TierEnforcement
        from scheduler_app.plans import ENTITY_CLASSROOMS
        enforcer = TierEnforcement.instance()
        # Count non-empty room names as the current count
        current = sum(
            1 for r in range(self.rooms_table.rowCount())
            if (self.rooms_table.item(r, 0) or QTableWidgetItem("")).text().strip()
        )
        if not enforcer.require_entity_limit(ENTITY_CLASSROOMS, current, self):
            return
        row = self.rooms_table.rowCount()
        self.rooms_table.insertRow(row)
        self.rooms_table.setItem(row, 0, QTableWidgetItem(""))
        cap_spin = QSpinBox()
        cap_spin.setRange(0, 9999)
        cap_spin.setValue(0)
        self.rooms_table.setCellWidget(row, 1, cap_spin)

    def _remove_room_row(self):
        rows = sorted(set(idx.row() for idx in self.rooms_table.selectedIndexes()), reverse=True)
        for row in rows:
            self.rooms_table.removeRow(row)

    def _filter_rooms_table(self):
        """Filter classrooms table by search text."""
        search = self._rooms_search.text().strip().lower()
        for row in range(self.rooms_table.rowCount()):
            item = self.rooms_table.item(row, 0)
            name = (item.text().strip() if item else "").lower()
            self.rooms_table.setRowHidden(row, bool(search and search not in name))

    def _add_lec_row(self):
        from scheduler_app.ui.tier_enforcement import TierEnforcement
        from scheduler_app.plans import ENTITY_LECTURERS
        enforcer = TierEnforcement.instance()
        # Count non-empty lecturer names as the current count
        current = sum(
            1 for r in range(self.lec_table.rowCount())
            if (self.lec_table.item(r, 0) or QTableWidgetItem("")).text().strip()
        )
        if not enforcer.require_entity_limit(ENTITY_LECTURERS, current, self):
            return
        row = self.lec_table.rowCount()
        self.lec_table.insertRow(row)
        self.lec_table.setItem(row, 0, QTableWidgetItem(""))
        av = new_lecturer_availability()
        self._update_lec_row_display(row, av)
        self._update_lec_overview()

    def _remove_lec_row(self):
        rows = sorted(set(idx.row() for idx in self.lec_table.selectedIndexes()), reverse=True)
        for row in rows:
            item = self.lec_table.item(row, 0)
            name = item.text().strip() if item else ""
            if name and name in self._lec_availability:
                del self._lec_availability[name]
            self.lec_table.removeRow(row)
        self._update_lec_overview()

    def _paste_room_rows(self):
        """Paste tab-separated clipboard data into the classrooms table."""
        from scheduler_app.ui.tier_enforcement import TierEnforcement
        from scheduler_app.plans import ENTITY_CLASSROOMS, get_limit, UNLIMITED
        text = QApplication.clipboard().text()
        if not text or not text.strip():
            return
        enforcer = TierEnforcement.instance()
        limit = get_limit(enforcer.tier_slug, ENTITY_CLASSROOMS)
        lines = [l for l in text.strip().split("\n") if l.strip()]
        for line in lines:
            cols = line.split("\t")
            name = cols[0].strip() if cols else ""
            cap = 0
            if len(cols) > 1:
                try:
                    cap = max(0, int(float(cols[1].strip())))
                except (ValueError, TypeError):
                    cap = 0
            if not name:
                continue
            # Check limit before adding each row
            current = sum(
                1 for r in range(self.rooms_table.rowCount())
                if (self.rooms_table.item(r, 0) or QTableWidgetItem("")).text().strip()
            )
            if limit != UNLIMITED and current >= limit:
                enforcer.require_entity_limit(ENTITY_CLASSROOMS, current, self)
                return
            row = self.rooms_table.rowCount()
            self.rooms_table.insertRow(row)
            self.rooms_table.setItem(row, 0, QTableWidgetItem(name))
            cap_spin = QSpinBox()
            cap_spin.setRange(0, 9999)
            cap_spin.setValue(cap)
            self.rooms_table.setCellWidget(row, 1, cap_spin)

    def _paste_lec_rows(self):
        """Paste tab-separated clipboard data into the lecturers table."""
        from scheduler_app.ui.tier_enforcement import TierEnforcement
        from scheduler_app.plans import ENTITY_LECTURERS, get_limit, UNLIMITED
        text = QApplication.clipboard().text()
        if not text or not text.strip():
            return
        enforcer = TierEnforcement.instance()
        limit = get_limit(enforcer.tier_slug, ENTITY_LECTURERS)
        lines = [l for l in text.strip().split("\n") if l.strip()]
        for line in lines:
            cols = line.split("\t")
            name = cols[0].strip() if cols else ""
            if not name:
                continue
            # Check limit before adding each row
            current = sum(
                1 for r in range(self.lec_table.rowCount())
                if (self.lec_table.item(r, 0) or QTableWidgetItem("")).text().strip()
            )
            if limit != UNLIMITED and current >= limit:
                enforcer.require_entity_limit(ENTITY_LECTURERS, current, self)
                return
            row = self.lec_table.rowCount()
            self.lec_table.insertRow(row)
            self.lec_table.setItem(row, 0, QTableWidgetItem(name))
            # Parse availability from remaining columns if present
            av = new_lecturer_availability()
            days = self._get_current_days()
            slots = self._get_current_slots()
            col_keys = ["allowed_days", "allowed_hours", "excluded_days", "excluded_hours"]
            col_is_day = [True, False, True, False]
            col_valid = [days, slots, days, slots]
            for i, key in enumerate(col_keys):
                if i + 1 < len(cols) and cols[i + 1].strip():
                    vals = [v.strip() for v in cols[i + 1].split(",") if v.strip()]
                    if col_is_day[i]:
                        av[key] = [nk for v in vals
                                   for nk in [normalize_day_value(v)]
                                   if nk and nk in col_valid[i]]
                    else:
                        av[key] = [v for v in vals if v in col_valid[i]]
            if any(av[k] for k in av):
                self._lec_availability[name] = av
            self._update_lec_row_display(row, av)
        self._update_lec_overview()

    def _copy_years(self):
        """Copy selected years to clipboard as tab-separated text."""
        items = self.years_tree.selectedItems()
        if not items:
            return
        lines = []
        for item in items:
            yr = item.data(0, Qt.ItemDataRole.UserRole)
            branches = self._years_data.get(yr, [])
            lines.append(f"{yr}\t{', '.join(branches)}")
        QApplication.clipboard().setText("\n".join(lines))

    def _get_current_days(self):
        """Read currently selected day keys from day chips."""
        return [key for key in DAY_KEYS if self.day_buttons[key].isChecked()]

    def _set_selected_days(self, keys):
        selected = set(keys)
        for key, btn in self.day_buttons.items():
            btn.setChecked(key in selected)

    def _select_all_days(self):
        self._set_selected_days(DAY_KEYS)

    def _select_weekdays_only(self):
        self._set_selected_days(DAY_KEYS[:5])

    def _select_weekends(self):
        self._set_selected_days(DAY_KEYS[5:])

    def _clear_days(self):
        self._set_selected_days([])

    def _invert_days(self):
        current = set(self._get_current_days())
        self._set_selected_days([k for k in DAY_KEYS if k not in current])

    def _get_current_slots(self):
        """Read current time slots from the Slots text widget.

        De-duplicated (ST-UI-021). Callers reading this mid-edit — notably
        ``LecturerConstraintsDialog``, which builds one checkbox per slot into a
        ``{slot: checkbox}`` map — must never receive a repeated label, or the
        map loses a box to the collision and that hour's availability silently
        cannot be edited. The user is separately refused at OK; this keeps every
        reader safe in the meantime.
        """
        slots, _problems = parse_slot_lines(self.slots_text.toPlainText())
        return slots

    def _slot_problems(self):
        """``(slots, problems)`` for the current text (ST-UI-021)."""
        return parse_slot_lines(self.slots_text.toPlainText())

    def _refresh_slot_status(self):
        """Show slot-list problems as the user types, before they press OK.

        Diagnosed on the way in rather than only on the way out: a user whose
        saved file already carries a duplicate sees it the moment the dialog
        opens, instead of being blocked by a modal after they have made ten
        other edits.
        """
        if not hasattr(self, "_slots_status"):
            return
        _slots, problems = self._slot_problems()
        dupes = [p for p in problems if p[1] == SLOT_ERROR_DUPLICATE]
        if dupes:
            listed = ", ".join(
                tr("setup.slots_dup_item").format(line=ln, value=val)
                for ln, _kind, val in dupes[:4])
            self._slots_status.setText(
                f"⚠ {tr('setup.slots_duplicate')} {listed}")
            self._slots_status.setStyleSheet(
                "font-size: 8pt; padding: 3px; color: #991B1B; "
                "background: #FEE2E2; border: 1px solid #DC2626; "
                "border-radius: 3px;")
        else:
            self._slots_status.setText("")
            self._slots_status.setStyleSheet("font-size: 8pt; padding: 2px;")

    def _update_lec_row_display(self, row, av):
        """Update the Status and Summary columns for a lecturer row."""
        status = self._lec_compute_status(av)
        summary = self._lec_compute_summary(av)
        # Status badge
        status_item = QTableWidgetItem(status)
        status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        status_colors = {
            tr("setup.lec_status_none"): "#6B7280",
            tr("setup.lec_status_custom"): "#2563EB",
            tr("setup.lec_status_conflict"): "#DC2626",
        }
        color = status_colors.get(status, "#6B7280")
        status_item.setForeground(QColor(color))
        self.lec_table.setItem(row, 1, status_item)
        # Summary
        summary_item = QTableWidgetItem(summary)
        summary_item.setFlags(summary_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        summary_item.setForeground(QColor("#475569"))
        self.lec_table.setItem(row, 2, summary_item)

    def _lec_compute_status(self, av):
        """Compute constraint status label for a lecturer."""
        has_allowed = bool(av.get("allowed_days") or av.get("allowed_hours"))
        has_excluded = bool(av.get("excluded_days") or av.get("excluded_hours"))
        if not has_allowed and not has_excluded:
            return tr("setup.lec_status_none")
        # Check for conflicts: all allowed days are excluded
        if has_allowed and has_excluded:
            allowed_d = set(av.get("allowed_days", []))
            excluded_d = set(av.get("excluded_days", []))
            if allowed_d and allowed_d.issubset(excluded_d):
                return tr("setup.lec_status_conflict")
            allowed_h = set(av.get("allowed_hours", []))
            excluded_h = set(av.get("excluded_hours", []))
            if allowed_h and allowed_h.issubset(excluded_h):
                return tr("setup.lec_status_conflict")
        return tr("setup.lec_status_custom")

    def _lec_compute_summary(self, av):
        """Compute a concise summary string for lecturer constraints."""
        parts = []
        if av.get("allowed_days"):
            days_str = ", ".join(day_label(d) for d in av["allowed_days"])
            parts.append(f"{tr('setup.allowed_days')}: {days_str}")
        if av.get("excluded_days"):
            days_str = ", ".join(day_label(d) for d in av["excluded_days"])
            parts.append(f"{tr('setup.excluded_days')}: {days_str}")
        if av.get("allowed_hours"):
            parts.append(f"{tr('setup.allowed_hours')}: {', '.join(av['allowed_hours'])}")
        if av.get("excluded_hours"):
            parts.append(f"{tr('setup.excluded_hours')}: {', '.join(av['excluded_hours'])}")
        return " | ".join(parts) if parts else tr("setup.lec_no_constraints")

    def _on_lec_double_click(self, index):
        """Open constraints dialog on double-click (except on name column for editing)."""
        if index.column() != 0:
            self._edit_lec_constraints_for_row(index.row())

    def _edit_lec_constraints(self):
        """Open constraints dialog for the currently selected lecturer."""
        rows = sorted(set(idx.row() for idx in self.lec_table.selectedIndexes()))
        if not rows:
            return
        self._edit_lec_constraints_for_row(rows[0])

    def _edit_lec_constraints_for_row(self, row):
        """Open the LecturerConstraintsDialog for a specific row."""
        item = self.lec_table.item(row, 0)
        name = item.text().strip() if item else ""
        av = self._lec_availability.get(name, new_lecturer_availability())
        days = self._get_current_days()
        slots = self._get_current_slots()
        dlg = LecturerConstraintsDialog(self, name, av, days, slots)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            new_av = dlg.get_availability()
            if any(new_av[k] for k in new_av):
                self._lec_availability[name] = new_av
            elif name in self._lec_availability:
                del self._lec_availability[name]
            self._update_lec_row_display(row, new_av)
            self._update_lec_overview()

    def _filter_lec_table(self):
        """Filter the lecturer table by search text and filter combo."""
        search = self._lec_search.text().strip().lower()
        filter_idx = self._lec_filter_combo.currentIndex()
        for row in range(self.lec_table.rowCount()):
            item = self.lec_table.item(row, 0)
            name = item.text().strip() if item else ""
            # Search filter
            if search and search not in name.lower():
                self.lec_table.setRowHidden(row, True)
                continue
            # Status filter
            if filter_idx == 0:  # All
                self.lec_table.setRowHidden(row, False)
            elif filter_idx == 1:  # With constraints
                av = self._lec_availability.get(name, new_lecturer_availability())
                has = any(av[k] for k in av)
                self.lec_table.setRowHidden(row, not has)
            elif filter_idx == 2:  # Without constraints
                av = self._lec_availability.get(name, new_lecturer_availability())
                has = any(av[k] for k in av)
                self.lec_table.setRowHidden(row, has)
            elif filter_idx == 3:  # Conflicts
                av = self._lec_availability.get(name, new_lecturer_availability())
                status = self._lec_compute_status(av)
                self.lec_table.setRowHidden(row, status != tr("setup.lec_status_conflict"))

    def _update_lec_overview(self):
        """Update the overview label with counts."""
        total = self.lec_table.rowCount()
        custom = 0
        for row in range(total):
            item = self.lec_table.item(row, 0)
            name = item.text().strip() if item else ""
            av = self._lec_availability.get(name, new_lecturer_availability())
            if any(av[k] for k in av):
                custom += 1
        self._lec_overview_label.setText(
            tr("setup.lec_overview").format(total=total, custom=custom))

    def _clear_all_lec_constraints(self):
        """Clear all lecturer constraints after confirmation."""
        if not self._lec_availability:
            return
        if QMessageBox.question(
                self, tr("setup.lec_clear_constraints"),
                tr("setup.lec_clear_constraints_confirm"),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        self._lec_availability.clear()
        for row in range(self.lec_table.rowCount()):
            self._update_lec_row_display(row, new_lecturer_availability())
        self._update_lec_overview()

    def _on_tab_changed(self, index):
        """When switching to Lecturers tab, refresh display."""
        if index == self._lec_tab_index:
            self._update_lec_overview()

    def _refresh_years_tree(self):
        self.years_tree.clear()
        for yr, branches in self._years_data.items():
            item = QTreeWidgetItem([yr, ", ".join(branches), str(len(branches))])
            item.setData(0, Qt.ItemDataRole.UserRole, yr)
            self.years_tree.addTopLevelItem(item)

    def _focus_lecturer_row(self):
        """Open Lecturers tab and highlight the requested lecturer, if present."""
        if not self._focus_lecturer_name:
            return
        self.tabs.setCurrentIndex(self._lec_tab_index)
        target = self._focus_lecturer_name.casefold()
        for row in range(self.lec_table.rowCount()):
            item = self.lec_table.item(row, 0)
            if item and item.text().strip().casefold() == target:
                self.lec_table.clearSelection()
                self.lec_table.selectRow(row)
                self.lec_table.setCurrentCell(row, 0)
                self.lec_table.scrollToItem(
                    item,
                    QAbstractItemView.ScrollHint.PositionAtCenter,
                )
                return

    def _add_year(self):
        d = TextInputDialog(self, tr("setup.add_year"), tr("setup.year_name_prompt"))
        if d.exec() != QDialog.DialogCode.Accepted or not d.result or not d.result.strip():
            return
        yr_name = d.result.strip()
        d2 = TextInputDialog(
            self,
            tr("setup.branches"),
            tr("setup.branches_for_year").format(year=yr_name),
            "A, B",
        )
        if d2.exec() != QDialog.DialogCode.Accepted or d2.result is None:
            return
        branches = [b.strip() for b in d2.result.split(",") if b.strip()]
        self._years_data[yr_name] = branches
        self._refresh_years_tree()

    def _edit_year(self):
        sel = self.years_tree.currentItem()
        if not sel:
            QMessageBox.information(self, tr("setup.edit_year"), tr("setup.select_year_first"))
            return
        yr = sel.data(0, Qt.ItemDataRole.UserRole)
        branches = self._years_data.get(yr, [])
        d = TextInputDialog(
            self,
            tr("setup.edit_year"),
            tr("setup.branches_for_year").format(year=yr),
            ", ".join(branches),
        )
        if d.exec() != QDialog.DialogCode.Accepted or d.result is None:
            return
        self._years_data[yr] = [b.strip() for b in d.result.split(",") if b.strip()]
        self._refresh_years_tree()

    def _remove_year(self):
        sel = self.years_tree.currentItem()
        if not sel:
            QMessageBox.information(self, tr("setup.remove_year"), tr("setup.select_year_first"))
            return
        yr = sel.data(0, Qt.ItemDataRole.UserRole)
        if QMessageBox.question(
                self, tr("setup.remove_year"),
                f"{tr('setup.remove_year')}: '{yr}'?") == QMessageBox.StandardButton.Yes:
            del self._years_data[yr]
            self._refresh_years_tree()

    # ── Excel dependency helper ───────────────────────────────────────────
    def _ensure_excel_deps(self):
        """Return True if pandas+openpyxl are available, auto-installing if needed."""
        return _ensure_excel_deps(self)

    # ── Template-offer for empty exports ─────────────────────────────────
    def _offer_template_or_nothing(self, sheet_name, headers):
        """When export has no data, offer to generate a template instead."""
        _offer_template_or_nothing(self, sheet_name, headers)

    # ── Text-based import/export (Days, Slots) ────────────────────────────
    def _import_days_from_excel(self):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
            _excel_file_filter(include_legacy=True))
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name=0, header=0)
            if df.empty:
                QMessageBox.information(self, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
                return
            col = df.columns[0]
            imported = normalize_day_list(df[col].tolist())
            for key, btn in self.day_buttons.items():
                btn.setChecked(key in imported)
            QMessageBox.information(self, tr("status.import_successful"), f"{len(imported)} {tr('setup.days')}")
        except Exception as exc:
            QMessageBox.warning(self, tr("status.import_failed"), str(exc))

    def _export_days_to_excel(self):
        day_keys = self._get_current_days()
        if not day_keys:
            return self._offer_template_or_nothing(tr("setup.days"), [tr("labels.day")])
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getSaveFileName(
            self, tr("dialogs.export.excel_title"),
            os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{tr('setup.days')}.xlsx"),
            _excel_file_filter())
        if not path:
            return
        try:
            df = pd.DataFrame([day_label(k) for k in day_keys], columns=[tr("labels.day")])
            df.to_excel(path, index=False, sheet_name=tr("setup.days"))
            QMessageBox.information(self, tr("dialogs.export.excel_title"),
                                    f"{tr('status.exported_to')} {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.warning(self, tr("dialogs.export.excel_title"), str(exc))

    def _import_text_from_excel(self, text_widget, sheet_name):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
            _excel_file_filter(include_legacy=True))
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name=0, header=0)
            if df.empty:
                QMessageBox.information(self, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
                return
            col = df.columns[0]
            values = [str(v).strip() for v in df[col] if str(v).strip()]
            text_widget.setPlainText("\n".join(values))
            QMessageBox.information(self, tr("status.import_successful"),
                                    f"{len(values)} {sheet_name}")
        except Exception as exc:
            QMessageBox.warning(self, tr("status.import_failed"), str(exc))

    def _export_text_to_excel(self, text_widget, sheet_name, headers):
        lines = [l.strip() for l in text_widget.toPlainText().strip().split("\n") if l.strip()]
        if not lines:
            return self._offer_template_or_nothing(sheet_name, headers)
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getSaveFileName(
            self, tr("dialogs.export.excel_title"),
            os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{sheet_name}.xlsx"),
            _excel_file_filter())
        if not path:
            return
        try:
            df = pd.DataFrame(lines, columns=headers)
            df.to_excel(path, index=False, sheet_name=sheet_name)
            QMessageBox.information(
                self, tr("dialogs.export.excel_title"),
                f"{tr('status.exported_to')} {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.warning(self, tr("dialogs.export.excel_title"), str(exc))

    # ── Rooms import/export ────────────────────────────────────────────────
    def _import_rooms_from_excel(self):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
            _excel_file_filter(include_legacy=True))
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name=0, header=0)
            if df.empty:
                QMessageBox.information(self, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
                return
            name_col = df.columns[0]
            cap_col = df.columns[1] if len(df.columns) > 1 else None
            count = 0
            for _, row in df.iterrows():
                name = str(row[name_col]).strip()
                if not name:
                    continue
                cap = 0
                if cap_col is not None:
                    try:
                        cap = int(row[cap_col])
                    except (ValueError, TypeError):
                        cap = 0
                r = self.rooms_table.rowCount()
                self.rooms_table.insertRow(r)
                self.rooms_table.setItem(r, 0, QTableWidgetItem(name))
                cap_spin = QSpinBox()
                cap_spin.setRange(0, 9999)
                cap_spin.setValue(max(0, cap))
                self.rooms_table.setCellWidget(r, 1, cap_spin)
                count += 1
            QMessageBox.information(self, tr("status.import_successful"),
                                    f"{count} {tr('labels.classrooms')}")
        except Exception as exc:
            QMessageBox.warning(self, tr("status.import_failed"), str(exc))

    def _export_rooms_to_excel(self):
        rows = []
        for r in range(self.rooms_table.rowCount()):
            item = self.rooms_table.item(r, 0)
            name = item.text().strip() if item else ""
            if not name:
                continue
            cap_w = self.rooms_table.cellWidget(r, 1)
            cap = cap_w.value() if cap_w else 0
            rows.append({tr("labels.classroom"): name, tr("setup.capacity"): cap})
        sheet_name = _trim_label(tr("tabs.rooms")) or "Rooms"
        if not rows:
            return self._offer_template_or_nothing(
                sheet_name, [tr("labels.classroom"), tr("setup.capacity")])
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getSaveFileName(
            self, tr("dialogs.export.excel_title"),
            os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{_safe_export_stem(sheet_name)}.xlsx"),
            _excel_file_filter())
        if not path:
            return
        try:
            pd.DataFrame(rows).to_excel(path, index=False, sheet_name=sheet_name)
            QMessageBox.information(
                self, tr("dialogs.export.excel_title"),
                f"{tr('status.exported_to')} {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.warning(self, tr("dialogs.export.excel_title"), str(exc))

    # ── Lecturers import/export ────────────────────────────────────────────
    def _import_lecturers_from_excel(self):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
            _excel_file_filter(include_legacy=True))
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name=0, header=0)
            if df.empty:
                QMessageBox.information(self, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
                return
            cols = list(df.columns)
            count = 0
            for _, row in df.iterrows():
                name = str(row[cols[0]]).strip() if len(cols) > 0 else ""
                if not name:
                    continue
                # Parse availability from comma-separated strings
                av = new_lecturer_availability()
                keys = ["allowed_days", "allowed_hours", "excluded_days", "excluded_hours"]
                for ci, key in enumerate(keys, 1):
                    val = str(row[cols[ci]]).strip() if ci < len(cols) else ""
                    if val.lower() == "nan":
                        val = ""
                    av[key] = [v.strip() for v in val.split(",") if v.strip()]
                r = self.lec_table.rowCount()
                self.lec_table.insertRow(r)
                self.lec_table.setItem(r, 0, QTableWidgetItem(name))
                if any(av[k] for k in av):
                    self._lec_availability[name] = av
                self._update_lec_row_display(r, av)
                count += 1
            QMessageBox.information(self, tr("status.import_successful"),
                                    f"{count} {tr('setup.lecturers')}")
        except Exception as exc:
            QMessageBox.warning(self, tr("status.import_failed"), str(exc))

    def _export_lecturers_to_excel(self):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        rows = []
        for r in range(self.lec_table.rowCount()):
            name_item = self.lec_table.item(r, 0)
            name = name_item.text().strip() if name_item else ""
            if not name:
                continue
            av = self._lec_availability.get(name, new_lecturer_availability())
            row_data = {tr("labels.lecturer"): name}
            for key, label in [
                ("allowed_days", tr("setup.allowed_days")),
                ("allowed_hours", tr("setup.allowed_hours")),
                ("excluded_days", tr("setup.excluded_days")),
                ("excluded_hours", tr("setup.excluded_hours")),
            ]:
                row_data[label] = ", ".join(av.get(key, []))
            rows.append(row_data)
        sheet_name = _trim_label(tr("setup.lecturers")) or "Lecturers"
        if not rows:
            return self._offer_template_or_nothing(
                sheet_name,
                [
                    tr("labels.lecturer"),
                    tr("setup.allowed_days"),
                    tr("setup.allowed_hours"),
                    tr("setup.excluded_days"),
                    tr("setup.excluded_hours"),
                ],
            )
        path, _ = QFileDialog.getSaveFileName(
            self, tr("dialogs.export.excel_title"),
            os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{_safe_export_stem(sheet_name)}.xlsx"),
            _excel_file_filter())
        if not path:
            return
        try:
            pd.DataFrame(rows).to_excel(path, index=False, sheet_name=sheet_name)
            QMessageBox.information(
                self, tr("dialogs.export.excel_title"),
                f"{tr('status.exported_to')} {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.warning(self, tr("dialogs.export.excel_title"), str(exc))

    # ── Years import/export ────────────────────────────────────────────────
    def _import_years_from_excel(self):
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        path, _ = QFileDialog.getOpenFileName(
            self, tr("dialogs.import.excel_title"), storage.sub_dir(storage.EXPORTS_DIR),
            _excel_file_filter(include_legacy=True))
        if not path:
            return
        try:
            df = pd.read_excel(path, sheet_name=0, header=0)
            if df.empty:
                QMessageBox.information(self, tr("dialogs.import.excel_title"), tr("warnings.no_data_found"))
                return
            cols = list(df.columns)
            count = 0
            for _, row in df.iterrows():
                year = str(row[cols[0]]).strip() if len(cols) > 0 else ""
                if not year:
                    continue
                branches_str = str(row[cols[1]]).strip() if len(cols) > 1 else ""
                if branches_str.lower() == "nan":
                    branches_str = ""
                branches = [b.strip() for b in branches_str.split(",") if b.strip()]
                self._years_data[year] = branches
                count += 1
            self._refresh_years_tree()
            QMessageBox.information(self, tr("status.import_successful"),
                                    f"{count} {tr('labels.years_branches')}")
        except Exception as exc:
            QMessageBox.warning(self, tr("status.import_failed"), str(exc))

    def _export_years_to_excel(self):
        sheet_name = _trim_label(tr("labels.years_branches")) or "Years"
        headers = [tr("labels.year"), tr("setup.branches")]
        if not self._years_data:
            return self._offer_template_or_nothing(sheet_name, headers)
        if not self._ensure_excel_deps():
            return
        import pandas as pd
        rows = []
        for yr, branches in self._years_data.items():
            rows.append({
                tr("labels.year"): yr,
                tr("setup.branches"): ", ".join(branches),
            })
        path, _ = QFileDialog.getSaveFileName(
            self, tr("dialogs.export.excel_title"),
            os.path.join(storage.sub_dir(storage.EXPORTS_DIR), f"{_safe_export_stem(sheet_name)}.xlsx"),
            _excel_file_filter())
        if not path:
            return
        try:
            pd.DataFrame(rows).to_excel(path, index=False, sheet_name=sheet_name)
            QMessageBox.information(
                self, tr("dialogs.export.excel_title"),
                f"{tr('status.exported_to')} {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.warning(self, tr("dialogs.export.excel_title"), str(exc))

    def _ok(self):
        from scheduler_app.ui.tier_enforcement import TierEnforcement
        from scheduler_app.plans import ENTITY_CLASSROOMS, ENTITY_LECTURERS

        days = self._get_current_days()

        # ST-UI-021. Uniqueness is the ONE hard rule the grid has, because every
        # lookup is `list.index()` and the first match wins: a repeated label
        # makes every later row with that name permanently unreachable.
        # Measured on a four-hour day with 09:00 typed twice — 4 rows drawn,
        # 3 addressable, 3 of 4 classes placed — and reconcile_placements sees
        # nothing, because it is a membership test and every label is a member.
        #
        # REFUSED, not silently de-duplicated: dropping the second 09:00
        # shortens the grid by a row, which re-points every multi-row lesson
        # below it and can push one off the end entirely. That would be a silent
        # repair of a silent corruption.
        slots, slot_problems = self._slot_problems()
        dupes = [p for p in slot_problems if p[1] == SLOT_ERROR_DUPLICATE]
        if dupes:
            listed = "\n".join(
                "  " + tr("setup.slots_dup_item").format(line=ln, value=val)
                for ln, _kind, val in dupes)
            QMessageBox.warning(
                self, tr("setup.time_slots"),
                tr("setup.slots_duplicate_body").format(lines=listed))
            return

        # A slot label is a by-name reference into an ORDERED list, so editing
        # that list can change which hours an existing lesson occupies without
        # touching the lesson at all. Measured: sorting a clean schedule gave 6
        # hard violations — a double-booked room and a lecturer in two places —
        # with reconcile_placements reporting []. Compare covered CELLS rather
        # than trying to recognise the dangerous edit shapes: a reorder, a
        # mid-list substitution and a removal are one defect seen three ways,
        # and an edit-shaped test catches only the ones someone thought of.
        #
        # Reported and confirmed, never repaired: pinned lessons are included,
        # and a pin is the user's instruction (ST-SCHED-002).
        moved = slot_meaning_changes(self.state, slots)
        if moved:
            names = "\n".join(
                "  • " + (c.get("name") or c.get("class_code") or "?")
                for c, _b, _a in moved[:10])
            if len(moved) > 10:
                names += "\n  …"
            if QMessageBox.question(
                    self, tr("setup.time_slots"),
                    tr("setup.slots_moved_body").format(
                        n=len(moved), names=names)
            ) != QMessageBox.StandardButton.Yes:
                return

        rooms = []
        capacities = {}
        for row in range(self.rooms_table.rowCount()):
            item = self.rooms_table.item(row, 0)
            name = item.text().strip() if item else ""
            if not name:
                continue
            cap_widget = self.rooms_table.cellWidget(row, 1)
            cap = cap_widget.value() if cap_widget else 0
            rooms.append(name)
            if cap > 0:
                capacities[name] = cap

        missing = []
        if not days: missing.append(tr("labels.days"))
        if not slots: missing.append(tr("labels.time_slots"))
        if not rooms: missing.append(tr("labels.classrooms"))
        if not self._years_data: missing.append(tr("labels.years_branches"))

        if missing:
            if QMessageBox.question(
                    self, tr("dialogs.missing_data.title"),
                    tr("dialogs.setup.still_missing") + f" {', '.join(missing)}.\n" + tr("dialogs.setup.continue_anyway")) != QMessageBox.StandardButton.Yes:
                return

        lecturers = []
        lecturer_availability = {}
        for row in range(self.lec_table.rowCount()):
            name_item = self.lec_table.item(row, 0)
            name = name_item.text().strip() if name_item else ""
            if not name:
                continue
            lecturers.append(name)
            av = self._lec_availability.get(name, new_lecturer_availability())
            if any(av[k] for k in av):
                lecturer_availability[name] = av

        # --- Tier enforcement: check classroom and lecturer limits ---
        # require_entity_limit(entity, current_count) checks whether adding
        # one more is allowed, i.e. current_count < limit.  We want to check
        # whether the *final* count is within limits, so we pass count - 1
        # (effectively asking "if I had count-1, could I add one more?").
        # When count is 0, skip the check (nothing to enforce).
        enforcer = TierEnforcement.instance()
        if rooms and not enforcer.require_entity_limit(
                ENTITY_CLASSROOMS, len(rooms) - 1, self):
            return
        if lecturers and not enforcer.require_entity_limit(
                ENTITY_LECTURERS, len(lecturers) - 1, self):
            return

        self.state["days"] = days
        self.state["slots"] = slots
        self.state["classrooms"] = rooms
        self.state["classroom_capacities"] = capacities
        self.state["lecturers"] = lecturers
        self.state["lecturer_availability"] = lecturer_availability
        self.state["years"] = dict(self._years_data)
        self.result = True
        self.accept()


# ── Lecturer Constraints Dialog ──────────────────────────────────────────────

class LecturerConstraintsDialog(QDialog):
    """Dialog for editing a single lecturer's availability constraints."""

    def __init__(self, parent, lecturer_name, availability, days, slots):
        super().__init__(parent)
        self.setWindowTitle(tr("setup.lec_dialog_title"))
        self.setMinimumWidth(520)
        self.setMinimumHeight(500)
        self.setStyleSheet(_DIALOG_STYLESHEET_TEMPLATE)

        self._days = list(days)
        self._slots = list(slots)
        self._av = {
            "allowed_days": list(availability.get("allowed_days", [])),
            "allowed_hours": list(availability.get("allowed_hours", [])),
            "excluded_days": list(availability.get("excluded_days", [])),
            "excluded_hours": list(availability.get("excluded_hours", [])),
        }

        layout = QVBoxLayout(self)

        # Header
        header_lbl = QLabel(f"\U0001F468\u200D\U0001F3EB  {lecturer_name}" if lecturer_name
                            else tr("setup.lecturer_name"))
        header_lbl.setStyleSheet("font-size: 14pt; font-weight: bold; color: #1E293B;")
        layout.addWidget(header_lbl)
        helper_lbl = QLabel(tr("setup.lec_dialog_helper"))
        helper_lbl.setWordWrap(True)
        helper_lbl.setStyleSheet("font-size: 8pt; color: #64748B; margin-bottom: 8px;")
        layout.addWidget(helper_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        # ── Availability Section ──
        avail_group = QGroupBox(tr("setup.lec_availability"))
        avail_layout = QVBoxLayout(avail_group)

        # Allowed Days
        avail_layout.addWidget(QLabel(tr("setup.allowed_days")))
        self._day_chips = {}
        day_chips_layout = QHBoxLayout()
        for d in self._days:
            btn = QPushButton(day_label(d))
            btn.setCheckable(True)
            btn.setChecked(d in self._av["allowed_days"])
            btn.setStyleSheet(
                "QPushButton { padding: 4px 10px; border: 1px solid #CBD5E1; "
                "border-radius: 12px; font-size: 8pt; background: white; }"
                "QPushButton:checked { background: #3B82F6; color: white; "
                "border-color: #2563EB; }")
            btn.clicked.connect(self._on_change)
            self._day_chips[d] = btn
            day_chips_layout.addWidget(btn)
        day_chips_layout.addStretch()
        avail_layout.addLayout(day_chips_layout)

        # Quick actions for days
        day_actions = QHBoxLayout()
        for label, fn in [
            (tr("buttons.select_all"), self._select_all_allowed_days),
            (tr("buttons.weekdays"), self._select_weekdays_allowed),
            (tr("buttons.weekends"), self._select_weekends_allowed),
            (tr("buttons.clear_all"), self._clear_allowed_days),
        ]:
            btn = QPushButton(label)
            btn.setStyleSheet("font-size: 7pt; padding: 2px 6px;")
            btn.clicked.connect(fn)
            day_actions.addWidget(btn)
        day_actions.addStretch()
        avail_layout.addLayout(day_actions)

        # Allowed Hours
        avail_layout.addWidget(QLabel(tr("setup.allowed_hours")))
        self._hour_checks = {}
        hours_layout = QGridLayout()
        for i, slot in enumerate(self._slots):
            cb = QCheckBox(slot)
            cb.setChecked(slot in self._av["allowed_hours"])
            cb.stateChanged.connect(self._on_change)
            self._hour_checks[slot] = cb
            hours_layout.addWidget(cb, i // 4, i % 4)
        avail_layout.addLayout(hours_layout)

        # Quick actions for hours
        hour_actions = QHBoxLayout()
        sa_btn = QPushButton(tr("buttons.select_all"))
        sa_btn.setStyleSheet("font-size: 7pt; padding: 2px 6px;")
        sa_btn.clicked.connect(self._select_all_allowed_hours)
        hour_actions.addWidget(sa_btn)
        clr_btn = QPushButton(tr("buttons.clear_all"))
        clr_btn.setStyleSheet("font-size: 7pt; padding: 2px 6px;")
        clr_btn.clicked.connect(self._clear_allowed_hours)
        hour_actions.addWidget(clr_btn)
        hour_actions.addStretch()
        avail_layout.addLayout(hour_actions)

        scroll_layout.addWidget(avail_group)

        # ── Excluded (Advanced) Section ──
        excl_group = QGroupBox(tr("setup.lec_excluded_advanced"))
        excl_layout = QVBoxLayout(excl_group)

        # Excluded Days
        excl_layout.addWidget(QLabel(tr("setup.excluded_days")))
        self._excl_day_chips = {}
        excl_day_layout = QHBoxLayout()
        for d in self._days:
            btn = QPushButton(day_label(d))
            btn.setCheckable(True)
            btn.setChecked(d in self._av["excluded_days"])
            btn.setStyleSheet(
                "QPushButton { padding: 4px 10px; border: 1px solid #CBD5E1; "
                "border-radius: 12px; font-size: 8pt; background: white; }"
                "QPushButton:checked { background: #EF4444; color: white; "
                "border-color: #DC2626; }")
            btn.clicked.connect(self._on_change)
            self._excl_day_chips[d] = btn
            excl_day_layout.addWidget(btn)
        excl_day_layout.addStretch()
        excl_layout.addLayout(excl_day_layout)

        # Excluded Hours
        excl_layout.addWidget(QLabel(tr("setup.excluded_hours")))
        self._excl_hour_checks = {}
        excl_hours_layout = QGridLayout()
        for i, slot in enumerate(self._slots):
            cb = QCheckBox(slot)
            cb.setChecked(slot in self._av["excluded_hours"])
            cb.stateChanged.connect(self._on_change)
            self._excl_hour_checks[slot] = cb
            excl_hours_layout.addWidget(cb, i // 4, i % 4)
        excl_layout.addLayout(excl_hours_layout)

        scroll_layout.addWidget(excl_group)

        # ── Live Summary ──
        summary_group = QGroupBox(tr("setup.lec_live_summary"))
        summary_layout = QVBoxLayout(summary_group)
        self._summary_label = QLabel("")
        self._summary_label.setWordWrap(True)
        self._summary_label.setStyleSheet("font-size: 9pt; color: #334155;")
        summary_layout.addWidget(self._summary_label)
        scroll_layout.addWidget(summary_group)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton(tr("buttons.ok"))
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

        self._update_summary()

    def _on_change(self):
        self._update_summary()

    def _collect(self):
        """Read current widget state into self._av."""
        self._av["allowed_days"] = [d for d in self._days
                                     if self._day_chips[d].isChecked()]
        self._av["allowed_hours"] = [s for s in self._slots
                                      if self._hour_checks[s].isChecked()]
        self._av["excluded_days"] = [d for d in self._days
                                      if self._excl_day_chips[d].isChecked()]
        self._av["excluded_hours"] = [s for s in self._slots
                                       if self._excl_hour_checks[s].isChecked()]

    def _update_summary(self):
        self._collect()
        parts = []
        if self._av["allowed_days"]:
            parts.append(f"{tr('setup.allowed_days')}: "
                         + ", ".join(day_label(d) for d in self._av["allowed_days"]))
        if self._av["excluded_days"]:
            parts.append(f"{tr('setup.excluded_days')}: "
                         + ", ".join(day_label(d) for d in self._av["excluded_days"]))
        if self._av["allowed_hours"]:
            parts.append(f"{tr('setup.allowed_hours')}: "
                         + ", ".join(self._av["allowed_hours"]))
        if self._av["excluded_hours"]:
            parts.append(f"{tr('setup.excluded_hours')}: "
                         + ", ".join(self._av["excluded_hours"]))
        if not parts:
            self._summary_label.setText(tr("setup.lec_no_constraints"))
        else:
            self._summary_label.setText("\n".join(parts))

    def get_availability(self):
        self._collect()
        return dict(self._av)

    # Quick actions
    def _select_all_allowed_days(self):
        for btn in self._day_chips.values():
            btn.setChecked(True)
        self._on_change()

    def _select_weekdays_allowed(self):
        for d, btn in self._day_chips.items():
            btn.setChecked(d in DAY_KEYS[:5])
        self._on_change()

    def _select_weekends_allowed(self):
        for d, btn in self._day_chips.items():
            btn.setChecked(d in DAY_KEYS[5:])
        self._on_change()

    def _clear_allowed_days(self):
        for btn in self._day_chips.values():
            btn.setChecked(False)
        self._on_change()

    def _select_all_allowed_hours(self):
        for cb in self._hour_checks.values():
            cb.setChecked(True)
        self._on_change()

    def _clear_allowed_hours(self):
        for cb in self._hour_checks.values():
            cb.setChecked(False)
        self._on_change()


# ── Add / Edit Class Dialog ──────────────────────────────────────────────────

class AddClassDialog(QDialog):
    def __init__(self, parent, state, edit_cls=None):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        title = ("\u270E  " + tr("dialogs.edit_class.title")) if edit_cls else ("\u2795  " + tr("toolbar.add_single"))
        self.setWindowTitle(title)
        self.state = state
        self.result = None
        self._select_all_shortcuts = []
        self.setMinimumSize(540, 620)

        # Scrollable form
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        form_widget = QWidget()
        f = QVBoxLayout(form_widget)
        f.setContentsMargins(15, 15, 15, 15)

        # ── Class Code ──
        code_row = QHBoxLayout()
        code_lbl = QLabel(tr("dialogs.add_class.code_optional") + ":")
        code_lbl.setStyleSheet("color: #475569;")
        code_row.addWidget(code_lbl)
        self.code_edit = QLineEdit(edit_cls.get("class_code", "") if edit_cls else "")
        self.code_edit.setPlaceholderText(tr("dialogs.add_class.code_optional"))
        code_row.addWidget(self.code_edit)
        f.addLayout(code_row)

        # ── Name ──
        row = QHBoxLayout()
        row.addWidget(QLabel(tr("dialogs.add_class.name_label")))
        self.name_edit = QLineEdit(edit_cls["name"] if edit_cls else "")
        row.addWidget(self.name_edit)
        f.addLayout(row)

        # ── Lecturer ──
        row2 = QHBoxLayout()
        row2.addWidget(QLabel(tr("dialogs.add_class.lecturer_label")))
        self.lecturer_combo = QComboBox()
        self.lecturer_combo.setEditable(True)
        lecturers = state.get("lecturers", [])
        if lecturers:
            self.lecturer_combo.addItems(lecturers)
        if edit_cls and edit_cls["lecturer"]:
            idx = self.lecturer_combo.findText(edit_cls["lecturer"])
            if idx >= 0:
                self.lecturer_combo.setCurrentIndex(idx)
            else:
                self.lecturer_combo.setCurrentText(edit_cls["lecturer"])
        else:
            self.lecturer_combo.setCurrentText("")
        row2.addWidget(self.lecturer_combo)
        f.addLayout(row2)

        # ── Duration ──
        row3 = QHBoxLayout()
        row3.addWidget(QLabel(tr("dialogs.add_class.duration_label")))
        self.duration_spin = QSpinBox()
        self.duration_spin.setRange(1, max(len(state["slots"]), 1))
        self.duration_spin.setValue(edit_cls["duration"] if edit_cls else 1)
        row3.addWidget(self.duration_spin)
        row3.addStretch()
        f.addLayout(row3)

        # ── Participants ──
        row_part = QHBoxLayout()
        row_part.addWidget(QLabel(tr("labels.participants_colon")))
        self.participants_spin = QSpinBox()
        self.participants_spin.setRange(0, 9999)
        self.participants_spin.setSpecialValueText(tr("labels.no_limit"))
        self.participants_spin.setValue(edit_cls.get("participants", 0) if edit_cls else 0)
        row_part.addWidget(self.participants_spin)
        row_part.addStretch()
        f.addLayout(row_part)

        # ── Location Type ──
        loc_row = QHBoxLayout()
        loc_row.addWidget(QLabel(tr("labels.location_type") + ":"))
        self.location_type_combo = QComboBox()
        for lt in LOCATION_TYPES:
            self.location_type_combo.addItem(get_location_label(lt), lt)
        if edit_cls:
            lt_val = edit_cls.get("location_type", LOCATION_FACE_TO_FACE)
            idx = self.location_type_combo.findData(lt_val)
            if idx >= 0:
                self.location_type_combo.setCurrentIndex(idx)
        self.location_type_combo.currentIndexChanged.connect(
            self._on_location_type_changed)
        loc_row.addWidget(self.location_type_combo)
        loc_row.addStretch()
        f.addLayout(loc_row)

        # ── Import / Generate Template ──
        io_row = QHBoxLayout()
        io_row.addStretch()
        import_btn = QPushButton("\u2B07 " + tr("dialogs.import.excel_title"))
        import_btn.clicked.connect(self._import_from_excel)
        io_row.addWidget(import_btn)
        template_btn = QPushButton("\U0001F4C4 " + tr("buttons.generate_template"))
        template_btn.clicked.connect(self._generate_template)
        io_row.addWidget(template_btn)
        f.addLayout(io_row)

        # ── Targets ──
        f.addWidget(self._separator())
        tgt_header = QHBoxLayout()
        tgt_header.addWidget(QLabel(f"<b>{tr('dialogs.add_class.target_groups')}</b>"))
        tgt_header.addStretch()
        sel_all_btn = QPushButton(tr("buttons.select_all"))
        sel_all_btn.setStyleSheet("font-size: 8pt; padding: 2px 8px;")
        sel_all_btn.clicked.connect(lambda: self._set_all_targets(True))
        desel_all_btn = QPushButton(tr("buttons.deselect_all"))
        desel_all_btn.setStyleSheet("font-size: 8pt; padding: 2px 8px;")
        desel_all_btn.clicked.connect(lambda: self._set_all_targets(False))
        tgt_header.addWidget(sel_all_btn)
        tgt_header.addWidget(desel_all_btn)
        f.addLayout(tgt_header)
        _install_select_all_shortcut(self, self, callback=lambda: self._set_all_targets(True))

        existing_targets = set()
        if edit_cls:
            existing_targets = {(t["year"], t["branch"]) for t in edit_cls["targets"]}

        self.target_vars = {}
        tgt_grid = QGridLayout()
        tgt_grid.setSpacing(4)
        idx = 0
        for yr in sorted(state["years"].keys()):
            for br in state["years"][yr]:
                cb = QCheckBox(f"{yr} / {br}")
                cb.setChecked((yr, br) in existing_targets)
                cb.stateChanged.connect(self._update_joint_visibility)
                self.target_vars[(yr, br)] = cb
                tgt_grid.addWidget(cb, idx // 3, idx % 3)
                idx += 1

        year_quick_row = QHBoxLayout()
        year_quick_row.addWidget(QLabel(tr("labels.year") + ":"))
        self._year_quick_combo = QComboBox()
        self._year_quick_combo.addItems(sorted(state["years"].keys()))
        year_quick_row.addWidget(self._year_quick_combo)
        sel_year_btn = QPushButton(tr("buttons.select_all"))
        sel_year_btn.setStyleSheet("font-size: 8pt; padding: 2px 8px;")
        sel_year_btn.clicked.connect(lambda: self._set_year_targets(True))
        year_quick_row.addWidget(sel_year_btn)
        desel_year_btn = QPushButton(tr("buttons.deselect_all"))
        desel_year_btn.setStyleSheet("font-size: 8pt; padding: 2px 8px;")
        desel_year_btn.clicked.connect(lambda: self._set_year_targets(False))
        year_quick_row.addWidget(desel_year_btn)
        year_quick_row.addStretch()
        f.addLayout(year_quick_row)
        f.addLayout(tgt_grid)

        # ── Joint Session ──
        self.joint_group = QWidget()
        jl = QVBoxLayout(self.joint_group)
        jl.setContentsMargins(20, 0, 0, 0)
        self.joint_cb = QCheckBox(tr("dialogs.add_class.joint_session"))
        self.joint_cb.setChecked(edit_cls.get("joint_session", False) if edit_cls else False)
        jl.addWidget(self.joint_cb)
        hint = QLabel(tr("dialogs.add_class.non_joint_info"))
        hint.setStyleSheet("color: #64748B; font-size: 8pt;")
        jl.addWidget(hint)
        f.addWidget(self.joint_group)
        self._update_joint_visibility()

        # ── Pinned ──
        f.addWidget(self._separator())
        self.pinned_cb = QCheckBox(tr("dialogs.add_class.pinned"))
        self.pinned_cb.setChecked(edit_cls["pinned"] if edit_cls else False)
        self.pinned_cb.stateChanged.connect(self._toggle_pinned)
        f.addWidget(self.pinned_cb)

        self.pin_widget = QWidget()
        pg = QFormLayout(self.pin_widget)
        pg.setContentsMargins(20, 0, 0, 0)
        self.pin_day = QComboBox()
        for d in state["days"]:
            self.pin_day.addItem(day_label(d), d)
        if edit_cls and edit_cls.get("pinned_day"):
            idx = self.pin_day.findData(edit_cls["pinned_day"])
            if idx >= 0: self.pin_day.setCurrentIndex(idx)
        pg.addRow(tr("labels.day_colon"), self.pin_day)

        self.pin_time = QComboBox()
        self.pin_time.addItems(state["slots"])
        if edit_cls and edit_cls.get("pinned_time"):
            idx = self.pin_time.findText(edit_cls["pinned_time"])
            if idx >= 0: self.pin_time.setCurrentIndex(idx)
        pg.addRow(tr("labels.time_colon"), self.pin_time)

        self.pin_room = QComboBox()
        self.pin_room.addItems(state["classrooms"])
        if edit_cls and edit_cls.get("pinned_classroom"):
            idx = self.pin_room.findText(edit_cls["pinned_classroom"])
            if idx >= 0: self.pin_room.setCurrentIndex(idx)
        pg.addRow(tr("labels.room_colon"), self.pin_room)

        f.addWidget(self.pin_widget)
        self.pin_widget.setVisible(self.pinned_cb.isChecked())

        # ── Protection Level ──
        from scheduler_app.models import PROTECTION_LEVELS
        prot_row = QHBoxLayout()
        prot_row.addWidget(QLabel(tr("protection.label")))
        self.protection_combo = QComboBox()
        for level in PROTECTION_LEVELS:
            self.protection_combo.addItem(get_protection_label(level), level)
        current_prot = edit_cls.get("protection", "none") if edit_cls else "none"
        idx = PROTECTION_LEVELS.index(current_prot) if current_prot in PROTECTION_LEVELS else 0
        self.protection_combo.setCurrentIndex(idx)
        self.protection_combo.setToolTip(
            tr("tooltips.protection_info"))
        prot_row.addWidget(self.protection_combo)
        prot_row.addStretch()
        f.addLayout(prot_row)

        # ── Constraints (collapsible) ──
        f.addWidget(self._separator())
        self._constraints_toggle = QPushButton(f"\u25B6 {tr('dialogs.add_class.constraints')}")
        self._constraints_toggle.setStyleSheet(
            "text-align: left; font-weight: bold; border: none; padding: 4px 0;")
        self._constraints_toggle.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._constraints_toggle.clicked.connect(self._toggle_constraints)
        f.addWidget(self._constraints_toggle)

        self._constraints_widget = QWidget()
        cw_layout = QVBoxLayout(self._constraints_widget)
        cw_layout.setContentsMargins(0, 0, 0, 0)

        # Allowed days
        cw_layout.addWidget(QLabel(tr("constraints.allowed_days")))
        ad_row = QHBoxLayout()
        existing_ad = set(edit_cls["allowed_days"]) if edit_cls else set()
        self.allowed_day_cbs = {}
        for d in state["days"]:
            cb = QCheckBox(day_label(d))
            cb.setChecked(d in existing_ad)
            self.allowed_day_cbs[d] = cb
            ad_row.addWidget(cb)
        ad_row.addStretch()
        cw_layout.addLayout(ad_row)

        # Excluded days
        cw_layout.addWidget(QLabel(tr("constraints.excluded_days")))
        ed_row = QHBoxLayout()
        existing_ed = set(edit_cls.get("excluded_days", [])) if edit_cls else set()
        self.excluded_day_cbs = {}
        for d in state["days"]:
            cb = QCheckBox(day_label(d))
            cb.setChecked(d in existing_ed)
            self.excluded_day_cbs[d] = cb
            ed_row.addWidget(cb)
        ed_row.addStretch()
        cw_layout.addLayout(ed_row)

        # Allowed times
        cw_layout.addWidget(QLabel(tr("constraints.allowed_start_times")))
        at_grid = QGridLayout()
        existing_at = set(edit_cls["allowed_times"]) if edit_cls else set()
        self.allowed_time_cbs = {}
        for i, s in enumerate(state["slots"]):
            cb = QCheckBox(s)
            cb.setChecked(s in existing_at)
            self.allowed_time_cbs[s] = cb
            at_grid.addWidget(cb, i // 4, i % 4)
        cw_layout.addLayout(at_grid)

        # Excluded times
        cw_layout.addWidget(QLabel(tr("constraints.excluded_start_times")))
        et_grid = QGridLayout()
        existing_et = set(edit_cls.get("excluded_times", [])) if edit_cls else set()
        self.excluded_time_cbs = {}
        for i, s in enumerate(state["slots"]):
            cb = QCheckBox(s)
            cb.setChecked(s in existing_et)
            self.excluded_time_cbs[s] = cb
            et_grid.addWidget(cb, i // 4, i % 4)
        cw_layout.addLayout(et_grid)

        # Required classrooms (wrapped for location_type toggling)
        self._classroom_constraints_widget = QWidget()
        cc_layout = QVBoxLayout(self._classroom_constraints_widget)
        cc_layout.setContentsMargins(0, 0, 0, 0)
        cc_layout.addWidget(QLabel(tr("constraints.required_classrooms")))
        rr_grid = QGridLayout()
        existing_rr = set(edit_cls["required_classrooms"]) if edit_cls else set()
        self.req_room_cbs = {}
        for i, rm in enumerate(state["classrooms"]):
            cb = QCheckBox(rm)
            cb.setChecked(rm in existing_rr)
            self.req_room_cbs[rm] = cb
            rr_grid.addWidget(cb, i // 3, i % 3)
        cc_layout.addLayout(rr_grid)

        # Excluded classrooms
        cc_layout.addWidget(QLabel(tr("constraints.excluded_classrooms")))
        er_grid = QGridLayout()
        existing_er = set(edit_cls["excluded_classrooms"]) if edit_cls else set()
        self.excl_room_cbs = {}
        for i, rm in enumerate(state["classrooms"]):
            cb = QCheckBox(rm)
            cb.setChecked(rm in existing_er)
            self.excl_room_cbs[rm] = cb
            er_grid.addWidget(cb, i // 3, i % 3)
        cc_layout.addLayout(er_grid)
        cw_layout.addWidget(self._classroom_constraints_widget)

        # If editing and any constraint is set, expand constraints
        has_constraints = edit_cls and any([
            edit_cls.get("allowed_days"), edit_cls.get("excluded_days", []),
            edit_cls.get("allowed_times"), edit_cls.get("excluded_times", []),
            edit_cls.get("required_classrooms"), edit_cls.get("excluded_classrooms"),
        ])
        self._constraints_widget.setVisible(bool(has_constraints))
        if has_constraints:
            self._constraints_toggle.setText(f"\u25BC {tr('dialogs.add_class.constraints')}")
        f.addWidget(self._constraints_widget)

        # Apply initial location type visibility
        self._on_location_type_changed()

        # ── Bottom buttons ──
        f.addWidget(self._separator())
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn = QPushButton(tr("buttons.ok"))
        ok_btn.clicked.connect(self._ok)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        f.addLayout(btn_row)

        scroll.setWidget(form_widget)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

    @staticmethod
    def _separator():
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        return line

    def _set_all_targets(self, checked):
        for cb in self.target_vars.values():
            cb.setChecked(checked)

    def _set_year_targets(self, checked):
        year = self._year_quick_combo.currentText().strip()
        if not year:
            return
        for (yr, _), cb in self.target_vars.items():
            if yr == year:
                cb.setChecked(checked)

    def _on_location_type_changed(self):
        """Show/hide classroom-related widgets based on selected location type."""
        lt = self.location_type_combo.currentData() or LOCATION_FACE_TO_FACE
        is_physical = (lt == LOCATION_FACE_TO_FACE)
        # Toggle classroom constraints
        if hasattr(self, "_classroom_constraints_widget"):
            self._classroom_constraints_widget.setVisible(is_physical)
        # Toggle pin-room selector
        if hasattr(self, "pin_room"):
            self.pin_room.setVisible(is_physical)
            # Also hide label if present
            if hasattr(self, "_pin_room_label"):
                self._pin_room_label.setVisible(is_physical)

    def _import_from_excel(self):
        classes = _import_class_rows(self, self.state)
        if not classes:
            return
        cls = classes[0]
        self.code_edit.setText(cls.get("class_code", ""))
        self.name_edit.setText(cls.get("name", ""))
        self.lecturer_combo.setCurrentText(cls.get("lecturer", ""))
        self.duration_spin.setValue(max(1, cls.get("duration", 1)))
        self.participants_spin.setValue(max(0, cls.get("participants", 0)))
        # Set location type from import
        lt = cls.get("location_type", LOCATION_FACE_TO_FACE)
        idx = self.location_type_combo.findData(lt)
        if idx >= 0:
            self.location_type_combo.setCurrentIndex(idx)
        self._set_all_targets(False)
        for t in cls.get("targets", []):
            cb = self.target_vars.get((t.get("year"), t.get("branch")))
            if cb:
                cb.setChecked(True)
        self.joint_cb.setChecked(bool(cls.get("joint_session", False)))

        self.pinned_cb.setChecked(bool(cls.get("pinned", False)))
        if cls.get("pinned_day"):
            i = self.pin_day.findData(cls["pinned_day"])
            if i >= 0:
                self.pin_day.setCurrentIndex(i)
        if cls.get("pinned_time"):
            i = self.pin_time.findText(cls["pinned_time"])
            if i >= 0:
                self.pin_time.setCurrentIndex(i)
        if cls.get("pinned_classroom"):
            i = self.pin_room.findText(cls["pinned_classroom"])
            if i >= 0:
                self.pin_room.setCurrentIndex(i)

        for d, cb in self.allowed_day_cbs.items():
            cb.setChecked(d in cls.get("allowed_days", []))
        for d, cb in self.excluded_day_cbs.items():
            cb.setChecked(d in cls.get("excluded_days", []))
        for t, cb in self.allowed_time_cbs.items():
            cb.setChecked(t in cls.get("allowed_times", []))
        for t, cb in self.excluded_time_cbs.items():
            cb.setChecked(t in cls.get("excluded_times", []))
        for r, cb in self.req_room_cbs.items():
            cb.setChecked(r in cls.get("required_classrooms", []))
        for r, cb in self.excl_room_cbs.items():
            cb.setChecked(r in cls.get("excluded_classrooms", []))
        self._update_joint_visibility()
        self._toggle_pinned()

    def _generate_template(self):
        _generate_class_template(self)

    def _update_joint_visibility(self):
        checked = sum(1 for cb in self.target_vars.values() if cb.isChecked())
        self.joint_group.setVisible(checked > 1)

    def _toggle_pinned(self):
        self.pin_widget.setVisible(self.pinned_cb.isChecked())

    def _toggle_constraints(self):
        visible = not self._constraints_widget.isVisible()
        self._constraints_widget.setVisible(visible)
        arrow = "\u25BC" if visible else "\u25B6"
        self._constraints_toggle.setText(f"{arrow} {tr('dialogs.add_class.constraints')}")

    def _ok(self):
        targets = [{"year": yr, "branch": br}
                    for (yr, br), cb in self.target_vars.items() if cb.isChecked()]

        cls = new_class()
        cls["class_code"] = self.code_edit.text().strip()
        cls["name"] = self.name_edit.text().strip()
        cls["lecturer"] = self.lecturer_combo.currentText().strip()
        cls["targets"] = targets
        cls["duration"] = self.duration_spin.value()
        cls["participants"] = self.participants_spin.value()
        cls["location_type"] = self.location_type_combo.currentData() or LOCATION_FACE_TO_FACE
        cls["joint_session"] = self.joint_cb.isChecked() if len(targets) > 1 else True
        cls["protection"] = self.protection_combo.currentData() or "none"

        is_physical = needs_physical_room(cls)

        if self.pinned_cb.isChecked():
            cls["pinned"] = True
            cls["pinned_day"] = self.pin_day.currentData() or ""
            cls["pinned_time"] = self.pin_time.currentText()
            cls["pinned_classroom"] = self.pin_room.currentText() if is_physical else None

        # Validate using canonical validator
        errors = validate_class_fields(cls)
        if errors:
            QMessageBox.critical(self, tr("dialogs.error.title"), errors[0])
            return

        cls["allowed_days"] = [d for d, cb in self.allowed_day_cbs.items() if cb.isChecked()]
        cls["excluded_days"] = [d for d, cb in self.excluded_day_cbs.items() if cb.isChecked()]
        cls["allowed_times"] = [t for t, cb in self.allowed_time_cbs.items() if cb.isChecked()]
        cls["excluded_times"] = [t for t, cb in self.excluded_time_cbs.items() if cb.isChecked()]
        if is_physical:
            cls["required_classrooms"] = [r for r, cb in self.req_room_cbs.items() if cb.isChecked()]
            cls["excluded_classrooms"] = [r for r, cb in self.excl_room_cbs.items() if cb.isChecked()]
        else:
            cls["required_classrooms"] = []
            cls["excluded_classrooms"] = []

        self.result = normalize_class_data(cls)
        self.accept()


# ── Place Class Dialog ───────────────────────────────────────────────────────

class PlaceClassDialog(QDialog):
    def __init__(self, parent, state):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u25BC  " + tr("dialogs.place.title"))
        self.state = state
        self.result = None
        self._select_all_shortcuts = []
        self.setMinimumSize(620, 480)

        self.unplaced = [(i, c) for i, c in enumerate(state["classes"])
                         if not c["pinned"] and not c["placed"]]

        if not self.unplaced:
            QMessageBox.information(parent, tr("dialogs.place.title"),
                                   tr("warnings.no_unplaced_classes"))
            return

        layout = QVBoxLayout(self)

        # Class selector
        top = QHBoxLayout()
        top.addWidget(QLabel(tr("dialogs.place.select_class")))
        self.class_combo = QComboBox()
        self.class_combo.setMinimumWidth(350)
        labels = []
        for _, c in self.unplaced:
            code = c.get("class_code", "")
            name = f"[{code}] {c['name']}" if code else c["name"]
            labels.append(f"{name}  ({c['lecturer']}, dur={c['duration']})")

        self.class_combo.addItems(labels)
        self.class_combo.currentIndexChanged.connect(self._on_class_selected)
        top.addWidget(self.class_combo)
        top.addStretch()
        layout.addLayout(top)

        # Options table
        layout.addWidget(QLabel(tr("dialogs.place.valid_placements")))
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels([tr("labels.day"), tr("labels.start_time"), tr("labels.classroom")])
        self.tree.setRootIsDecorated(False)
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tree.itemDoubleClicked.connect(lambda: self._place())
        layout.addWidget(self.tree)

        sel_row = QHBoxLayout()
        sel_all_btn = QPushButton(tr("buttons.select_all"))
        sel_all_btn.clicked.connect(self.tree.selectAll)
        sel_row.addWidget(sel_all_btn)
        clear_sel_btn = QPushButton(tr("buttons.deselect_all"))
        clear_sel_btn.clicked.connect(self.tree.clearSelection)
        sel_row.addWidget(clear_sel_btn)
        sel_row.addStretch()
        layout.addLayout(sel_row)
        _install_select_all_shortcut(self, self.tree)

        # ST-UI-015. The dialog used to dead-end on its most important case:
        # "0 gecerli yerlestirme bulundu" over an empty table, with the
        # "Yerlestir" button still ENABLED -- and pressing it said "select a
        # placement option first", instructing the user to pick a row from a
        # table with no rows. This panel says why, and what to change.
        self._explain_label = QLabel("")
        self._explain_label.setWordWrap(True)
        self._explain_label.setTextFormat(Qt.TextFormat.RichText)
        self._explain_label.setStyleSheet(
            "background: #FEF3C7; color: #92400E; padding: 8px; "
            "border: 1px solid #F59E0B; border-radius: 4px; font-size: 9pt;")
        self._explain_label.setVisible(False)
        layout.addWidget(self._explain_label)

        # Bottom
        bottom = QHBoxLayout()
        self.info_label = QLabel(tr("dialogs.place.select_class_above"))
        bottom.addWidget(self.info_label)
        bottom.addStretch()
        self._place_btn = QPushButton(tr("buttons.place"))
        self._place_btn.clicked.connect(self._place)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(self._place_btn)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

        self._valid_options = []
        if self.unplaced:
            self.class_combo.setCurrentIndex(0)
            self._on_class_selected(0)

    def _on_class_selected(self, idx):
        if idx < 0 or idx >= len(self.unplaced):
            return
        _, candidate = self.unplaced[idx]
        options = find_valid_options(self.state, candidate)
        self._valid_options = options

        self.tree.clear()
        for day, slot, room in options:
            display_room = get_effective_room_resource_for_class(
                candidate, room_override=room)
            item = QTreeWidgetItem(self.tree, [day_label(day), slot, display_room])
            item.setData(0, Qt.ItemDataRole.UserRole, day)
            item.setData(2, Qt.ItemDataRole.UserRole, room)

        self.info_label.setText(f"{len(options)} {tr('dialogs.place.valid_found')}")

        # A button that cannot succeed must not invite a click -- but disabling
        # it ALONE would just convert a loud dead end into a quiet one, so the
        # reason goes on screen at the same time.
        self._place_btn.setEnabled(bool(options))
        if options:
            self._explain_label.setVisible(False)
            self._explain_label.clear()
        else:
            self._explain_label.setText(self._explain_unplaceable(candidate))
            self._explain_label.setVisible(True)

    def _explain_unplaceable(self, candidate):
        """Why this class has no valid placement, and what would change it.

        Uses the per-class negotiator entry point, not the full report:
        `negotiate_class` costs 0.3 ms (small) to 11 ms (large) and does not
        build the conflict graph, while `full_diagnostic` is 18 ms to 2.3 s --
        the ST-PERF-007 number. This runs on every combo change, so the
        difference is the whole reason the panel is affordable.
        """
        from scheduler_app.constraint_negotiator import ConstraintNegotiator

        try:
            report = ConstraintNegotiator(self.state).negotiate_class(candidate)
        except Exception:
            # Never let the explanation be the thing that breaks the dialog.
            return f"<b>{tr('dialogs.place.no_options_title')}</b>"

        parts = [f"<b>{tr('dialogs.place.no_options_title')}</b>"]
        for reason in report.get("blocking_reasons", [])[:3]:
            parts.append(f"&nbsp;&nbsp;• {html.escape(str(reason))}")
        suggestions = report.get("suggestions", [])
        if suggestions:
            parts.append(f"<b>{tr('dialogs.place.what_to_change')}</b>")
            for sug in suggestions[:3]:
                # The impact is already inside `description` for the
                # move_conflicting type ("...frees 3 slots"), so appending
                # impact_label there reads as a stutter.
                if sug.get("type") == "move_conflicting":
                    parts.append(
                        f"&nbsp;&nbsp;→ {html.escape(str(sug['description']))}")
                else:
                    parts.append(
                        f"&nbsp;&nbsp;→ {html.escape(str(sug['description']))} "
                        f"<i>({html.escape(str(sug['impact_label']))})</i>")
        n_off_grid = report.get("off_grid_blockers", 0)
        if n_off_grid:
            # ST-DATA-003: these lessons are skipped when counting blockers
            # because they occupy no cell -- and they are mentioned nowhere
            # else in the UI, so saying nothing here would hide them a third
            # time.
            parts.append(
                f"&nbsp;&nbsp;⚠ "
                + tr("dialogs.place.off_grid_note").format(n=n_off_grid))
        return "<br>".join(parts)

    def _place(self):
        selected = self.tree.selectedItems()
        sel = self.tree.currentItem() or (selected[0] if selected else None)
        if not sel:
            QMessageBox.information(self, tr("buttons.place"),
                                   tr("dialogs.place.select_option_first"))
            return
        day = sel.data(0, Qt.ItemDataRole.UserRole)
        slot = sel.text(1)
        room = sel.data(2, Qt.ItemDataRole.UserRole)
        idx = self.class_combo.currentIndex()
        orig_idx, _ = self.unplaced[idx]
        self.result = (orig_idx, day, slot, room)
        self.accept()


# ── Select Class Dialog ──────────────────────────────────────────────────────

class SelectClassDialog(QDialog):
    def __init__(self, parent, title, classes_with_indices, show_placement=False):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle(title)
        self.result = None
        self.setMinimumSize(580, 380)
        self._classes = classes_with_indices

        layout = QVBoxLayout(self)

        cols = [tr("labels.class_name"), tr("labels.lecturer")]
        if show_placement:
            cols.append(tr("labels.placement"))
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(cols)
        self.tree.setRootIsDecorated(False)
        self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.tree)

        for idx, c in classes_with_indices:
            if c["pinned"]:
                status = (
                    f"{tr('labels.pinned')}: {day_label(c['pinned_day'])} "
                    f"{c['pinned_time']} {classroom_of(c)}"
                )
            elif c["placed"]:
                status = f"{day_label(c['placed_day'])} {c['placed_time']} {classroom_of(c)}"
            else:
                status = tr("labels.unplaced")
            code = c.get("class_code", "")
            display_name = f"[{code}] {c['name']}" if code else c["name"]
            vals = [display_name, c["lecturer"]]
            if show_placement:
                vals.append(status)
            QTreeWidgetItem(self.tree, vals)

        self.tree.itemDoubleClicked.connect(lambda: self._ok())

        bottom = QHBoxLayout()
        bottom.addStretch()
        ok_btn = QPushButton(tr("buttons.ok"))
        ok_btn.clicked.connect(self._ok)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(ok_btn)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

    def _ok(self):
        sel = self.tree.currentItem()
        if not sel:
            return
        tree_idx = self.tree.indexOfTopLevelItem(sel)
        self.result = self._classes[tree_idx][0]
        self.accept()


class MultiSelectClassDialog(QDialog):
    """Dialog that allows selecting multiple classes for batch operations."""

    def __init__(self, parent, title, classes_with_indices,
                 show_placement=False, action_label=None):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle(title)
        self.result = []  # List of indices
        self.setMinimumSize(620, 420)
        self._classes = classes_with_indices
        self._select_all_shortcuts = []

        layout = QVBoxLayout(self)

        cols = [tr("labels.class_name"), tr("labels.lecturer")]
        if show_placement:
            cols.append(tr("labels.placement"))
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(cols)
        self.tree.setRootIsDecorated(False)
        self.tree.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.tree)
        _install_select_all_shortcut(self, self.tree)

        for idx, c in classes_with_indices:
            if c["pinned"]:
                status = (
                    f"{tr('labels.pinned')}: {day_label(c['pinned_day'])} "
                    f"{c['pinned_time']} {classroom_of(c)}"
                )
            elif c["placed"]:
                status = f"{day_label(c['placed_day'])} {c['placed_time']} {classroom_of(c)}"
            else:
                status = tr("labels.unplaced")
            code = c.get("class_code", "")
            display_name = f"[{code}] {c['name']}" if code else c["name"]
            vals = [display_name, c["lecturer"]]
            if show_placement:
                vals.append(status)
            QTreeWidgetItem(self.tree, vals)

        # Select All / Deselect All
        sel_row = QHBoxLayout()
        sel_all = QPushButton(tr("buttons.select_all"))
        sel_all.clicked.connect(self.tree.selectAll)
        sel_row.addWidget(sel_all)
        desel = QPushButton(tr("buttons.deselect_all"))
        desel.clicked.connect(self.tree.clearSelection)
        sel_row.addWidget(desel)
        sel_row.addStretch()
        count_lbl_text = tr("labels.selected_colon")
        self._sel_count = QLabel(f"{count_lbl_text} 0")
        sel_row.addWidget(self._sel_count)
        layout.addLayout(sel_row)

        self.tree.itemSelectionChanged.connect(self._update_count)

        bottom = QHBoxLayout()
        bottom.addStretch()
        ok_btn = QPushButton(action_label or tr("buttons.ok"))
        ok_btn.setStyleSheet(
            "background: #1D4ED8; color: white; font-weight: bold; padding: 6px 16px;")
        ok_btn.clicked.connect(self._ok)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(ok_btn)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

    def _update_count(self):
        n = len(self.tree.selectedItems())
        self._sel_count.setText(f"{tr('labels.selected_colon')} {n}")

    def _ok(self):
        selected = self.tree.selectedItems()
        if not selected:
            return
        self.result = []
        for item in selected:
            tree_idx = self.tree.indexOfTopLevelItem(item)
            self.result.append(self._classes[tree_idx][0])
        self.accept()


# ── Warnings Dialog ──────────────────────────────────────────────────────────

class WarningsDialog(QDialog):
    def __init__(self, parent, state):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u26A0  " + tr("panels.workload_warnings"))
        self.setMinimumSize(520, 420)

        layout = QVBoxLayout(self)
        self.text = QTextEdit()
        self.text.setReadOnly(True)
        layout.addWidget(self.text)

        placed = get_placed_classes(state)
        if not placed:
            self.text.append(tr("warnings.no_classes_placed"))
        else:
            for yr in sorted(state["years"].keys()):
                for br in state["years"][yr]:
                    day_counts = {}
                    for day in state["days"]:
                        count = 0
                        for c in placed:
                            if not any(t["year"] == yr and t["branch"] == br
                                       for t in c["targets"]):
                                continue
                            occ = occupied_slots_of(state, c)
                            count += sum(1 for d, s in occ if d == day)
                        day_counts[day] = count

                    total = sum(day_counts.values())
                    self.text.append(f"\n<b>{yr} / {br}:</b>")
                    if total == 0:
                        self.text.append(f"  {tr('warnings.no_classes_scheduled')}")
                        continue

                    loads = ", ".join(f"{day_label(d)}={n}" for d, n in day_counts.items())
                    self.text.append(f"  {tr('warnings.slots_per_day')} {loads}")

                    heavy = [d for d, n in day_counts.items()
                             if n >= len(state["slots"]) * 0.75]
                    light = [d for d, n in day_counts.items() if n == 0]
                    if heavy:
                        self.text.append(
                            f"  <span style='color:#DC2626'>{tr('warnings.heavy_days')} {', '.join(day_label(d) for d in heavy)}</span>")
                    if light:
                        self.text.append(
                            f"  <span style='color:#DC2626'>{tr('warnings.empty_days')} {', '.join(day_label(d) for d in light)}</span>")
                    if not heavy and not light:
                        self.text.append(
                            f"  <span style='color:#16A34A'>{tr('warnings.balanced')}</span>")

        close_btn = QPushButton(tr("buttons.close"))
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)


# ── Open Slots Dialog ────────────────────────────────────────────────────────

class OpenSlotsDialog(QDialog):
    def __init__(self, parent, state):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u2B50  " + tr("panels.open_slots"))
        self.setMinimumSize(520, 420)

        layout = QVBoxLayout(self)
        text = QTextEdit()
        text.setReadOnly(True)
        layout.addWidget(text)

        placed = get_placed_classes(state)
        for room in state["classrooms"]:
            free = []
            for day in state["days"]:
                for slot in state["slots"]:
                    occupied = False
                    for c in placed:
                        if classroom_of(c) != room:
                            continue
                        if (day, slot) in occupied_slots_of(state, c):
                            occupied = True
                            break
                    if not occupied:
                        free.append(f"{day_label(day)} {slot}")
            text.append(f"\n<b>{room}: {len(free)} {tr('labels.free_slots')}</b>")
            for f_slot in free:
                text.append(f"  {f_slot}")

        close_btn = QPushButton(tr("buttons.close"))
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)


# ── Post-Add Placement Dialog ──────────────────────────────────────────────

class PostAddDialog(QDialog):
    """After adding a class, let user choose: place now or skip."""

    def __init__(self, parent, state, cls):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u25BC  " + tr("menus.edit_place_class"))
        self.result = None
        self.setMinimumSize(620, 480)

        layout = QVBoxLayout(self)

        options = find_valid_options(state, cls)
        self._valid_options = options

        if not options:
            lbl = QLabel(tr("warnings.no_valid_placements"))
            lbl.setStyleSheet("color: #DC2626; font-weight: bold; font-size: 10pt;")
            layout.addWidget(lbl)
        else:
            lbl = QLabel(f"{len(options)} {tr('dialogs.place.valid_found')}")
            lbl.setStyleSheet("font-weight: bold; font-size: 10pt;")
            layout.addWidget(lbl)

        layout.addWidget(QLabel(tr("dialogs.post_add.prompt")))

        # Options table
        if options:
            layout.addWidget(QLabel(tr("dialogs.place.valid_placements")))
            self.tree = QTreeWidget()
            self.tree.setHeaderLabels([tr("labels.day"), tr("labels.start_time"), tr("labels.classroom")])
            self.tree.setRootIsDecorated(False)
            self.tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            self.tree.itemDoubleClicked.connect(lambda: self._place_selected())
            layout.addWidget(self.tree)

            for day, slot, room in options:
                display_room = get_effective_room_resource_for_class(
                    cls, room_override=room)
                item = QTreeWidgetItem(self.tree, [day_label(day), slot, display_room])
                item.setData(0, Qt.ItemDataRole.UserRole, day)
                item.setData(2, Qt.ItemDataRole.UserRole, room)

        # Bottom buttons
        bottom = QHBoxLayout()
        if options:
            place_btn = QPushButton(tr("dialogs.post_add.place_now"))
            place_btn.clicked.connect(self._place_selected)
            bottom.addWidget(place_btn)
        skip_btn = QPushButton(tr("dialogs.post_add.add_without_placing"))
        skip_btn.clicked.connect(self._skip)
        bottom.addWidget(skip_btn)
        bottom.addStretch()
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self._cancel)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

    def _place_selected(self):
        if not self._valid_options:
            return
        sel = self.tree.currentItem()
        if not sel:
            QMessageBox.information(self, tr("buttons.place"),
                                   tr("dialogs.place.select_option_first"))
            return
        day = sel.data(0, Qt.ItemDataRole.UserRole)
        slot = sel.text(1)
        room = sel.data(2, Qt.ItemDataRole.UserRole)
        self.result = (day, slot, room)
        self.accept()

    def _skip(self):
        self.result = "skip"
        self.accept()

    def _cancel(self):
        self.result = None
        self.reject()


# ── Bulk Add Classes Dialog ─────────────────────────────────────────────────

class BulkAddDialog(QDialog):
    """Dialog for entering multiple classes at once with full constraint support."""

    def __init__(self, parent, state):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u2795  " + tr("dialogs.bulk_add.title"))
        self.state = state
        self.result = None  # Will be list of class dicts on accept
        self._select_all_shortcuts = []
        self.setMinimumSize(950, 650)

        layout = QVBoxLayout(self)
        layout.setSpacing(6)
        layout.setContentsMargins(10, 10, 10, 10)

        # Instructions
        info = QLabel(tr("dialogs.bulk_add.instructions"))
        info.setWordWrap(True)
        info.setStyleSheet("font-size: 9pt; color: #475569; padding: 2px;")
        layout.addWidget(info)

        # Build target group labels
        self._target_labels = []
        for yr in sorted(state["years"].keys()):
            for br in state["years"][yr]:
                self._target_labels.append((yr, br, f"{yr}/{br}"))

        # Search and filter bar
        search_row = QHBoxLayout()
        self._bulk_search = QLineEdit()
        self._bulk_search.setPlaceholderText(tr("dialogs.bulk_add.search_placeholder"))
        self._bulk_search.setClearButtonEnabled(True)
        self._bulk_search.textChanged.connect(self._apply_bulk_filter)
        search_row.addWidget(self._bulk_search)
        self._bulk_filter_combo = QComboBox()
        self._bulk_filter_combo.addItems([
            tr("dialogs.bulk_add.filter_all"),
            tr("dialogs.bulk_add.filter_filled"),
            tr("dialogs.bulk_add.filter_empty"),
            tr("dialogs.bulk_add.filter_pinned"),
            tr("dialogs.bulk_add.filter_joint"),
            tr("dialogs.bulk_add.filter_with_constraints"),
        ])
        self._bulk_filter_combo.currentIndexChanged.connect(self._apply_bulk_filter)
        search_row.addWidget(self._bulk_filter_combo)
        layout.addLayout(search_row)

        # Table
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            "QTableWidget { gridline-color: #E2E8F0; font-size: 9pt; }"
            "QTableWidget::item:selected { background: #DBEAFE; color: #1E293B; }")
        self.table.verticalHeader().setDefaultSectionSize(30)
        self._setup_columns()
        layout.addWidget(self.table)
        _install_select_all_shortcut(self, self.table)

        # Row buttons
        btn_row = QHBoxLayout()
        add_btn = QPushButton(tr("buttons.add_row"))
        add_btn.clicked.connect(self._add_row_with_undo)
        btn_row.addWidget(add_btn)
        add5_btn = QPushButton(tr("buttons.add_5_rows"))
        add5_btn.clicked.connect(self._add_5_rows_with_undo)
        btn_row.addWidget(add5_btn)
        del_btn = QPushButton(tr("buttons.remove_selected_row"))
        del_btn.clicked.connect(self._remove_selected)
        btn_row.addWidget(del_btn)
        import_btn = QPushButton("\u2B07 " + tr("dialogs.import.excel_title"))
        import_btn.clicked.connect(self._import_from_excel_with_undo)
        btn_row.addWidget(import_btn)
        template_btn = QPushButton("\U0001F4C4 " + tr("buttons.generate_template"))
        template_btn.clicked.connect(self._generate_template)
        btn_row.addWidget(template_btn)
        btn_row.addStretch()

        count_label_text = tr("labels.rows")
        self.row_count_label = QLabel(f"{count_label_text} 0")
        btn_row.addWidget(self.row_count_label)
        layout.addLayout(btn_row)

        year_row = QHBoxLayout()
        year_row.addWidget(QLabel(tr("labels.year") + ":"))
        self._year_group_combo = QComboBox()
        self._year_group_combo.addItems(sorted(self.state["years"].keys()))
        year_row.addWidget(self._year_group_combo)
        sel_year_btn = QPushButton(tr("buttons.select_all"))
        sel_year_btn.clicked.connect(lambda: self._apply_year_targets_to_rows(True))
        year_row.addWidget(sel_year_btn)
        desel_year_btn = QPushButton(tr("buttons.deselect_all"))
        desel_year_btn.clicked.connect(lambda: self._apply_year_targets_to_rows(False))
        year_row.addWidget(desel_year_btn)
        year_row.addStretch()
        layout.addLayout(year_row)

        # Bottom buttons
        bottom = QHBoxLayout()
        bottom.setContentsMargins(0, 6, 0, 0)
        bottom.addStretch()
        schedule_btn = QPushButton(tr("buttons.schedule_all"))
        schedule_btn.setStyleSheet(
            "QPushButton { background: #3B82F6; color: white; border: none;"
            "  border-radius: 6px; font-weight: bold; padding: 8px 24px;"
            "  font-size: 10pt; }"
            "QPushButton:hover { background: #2563EB; }")
        schedule_btn.clicked.connect(self._ok)
        bottom.addWidget(schedule_btn)
        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.setStyleSheet(
            "QPushButton { padding: 8px 20px; font-size: 10pt;"
            "  border: 1px solid #CBD5E1; border-radius: 6px; }"
            "QPushButton:hover { background: #F1F5F9; }")
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

        # Keyboard shortcuts: Delete, Ctrl+C, Ctrl+V, Ctrl+Z, Ctrl+Y
        _install_table_keyboard_shortcuts(
            self, self.table,
            delete_callback=self._remove_selected,
            copy_callback=self._copy_rows,
            paste_callback=self._paste_rows,
        )
        # Undo / Redo (dialog-local, operates on table row snapshots)
        self._undo_stack_local = []  # list of (label, [class_dicts])
        self._redo_stack_local = []
        self._max_undo_local = 30
        if not hasattr(self, "_table_kbd_shortcuts"):
            self._table_kbd_shortcuts = []
        _ctx = Qt.ShortcutContext.WidgetWithChildrenShortcut
        _sc_undo = QShortcut(QKeySequence(QKeySequence.StandardKey.Undo), self.table)
        _sc_undo.setContext(_ctx)
        _sc_undo.activated.connect(self._local_undo)
        self._table_kbd_shortcuts.append(_sc_undo)
        _sc_redo = QShortcut(QKeySequence(QKeySequence.StandardKey.Redo), self.table)
        _sc_redo.setContext(_ctx)
        _sc_redo.activated.connect(self._local_redo)
        self._table_kbd_shortcuts.append(_sc_redo)

        # Start with 5 rows
        for _ in range(5):
            self._add_row()

        # Initially hide pin columns
        self._update_pin_columns_visibility()

    def _on_bulk_location_changed(self, row):
        """Show/hide room-related widgets for a row based on location type."""
        loc_w = self.table.cellWidget(row, self._col_location_type)
        lt = loc_w.currentData() if loc_w else LOCATION_FACE_TO_FACE
        is_physical = (lt == LOCATION_FACE_TO_FACE)
        for col in (self._col_pin_room, self._col_req_rooms, self._col_excl_rooms):
            w = self.table.cellWidget(row, col)
            if w:
                w.setEnabled(is_physical)
                if not is_physical:
                    # Clear stale room values for non-physical classes
                    if hasattr(w, 'set_checked'):
                        w.set_checked([])  # MultiSelectButton
                    elif hasattr(w, 'setCurrentIndex'):
                        w.setCurrentIndex(0)  # QComboBox

    def _update_pin_columns_visibility(self):
        """Show Pin Day/Time/Room columns only if any row has Pinned checked."""
        any_pinned = any(
            self._get_checkbox(r, self._col_pinned)
            for r in range(self.table.rowCount())
        )
        for col in (self._col_pin_day, self._col_pin_time, self._col_pin_room):
            self.table.setColumnHidden(col, not any_pinned)

    def _setup_columns(self):
        # Columns: Code | Name | Lecturer | Duration | Participants | Targets (checkboxes) | Joint | Pinned |
        #          Pin Day | Pin Time | Pin Room |
        #          Allowed Days | Allowed Times | Required Rooms | Excluded Rooms
        headers = [
            tr("dialogs.add_class.code"), tr("labels.class_name"), tr("labels.lecturer"),
            tr("labels.duration"), tr("labels.participants"), tr("labels.location_type"),
        ]
        # Add a column per target group
        for _, _, label in self._target_labels:
            headers.append(label)

        headers += [
            tr("labels.all"),  # Select-all-targets toggle per row
            tr("labels.joint"), tr("labels.pinned"),
            tr("labels.pin_day"), tr("labels.pin_time"), tr("labels.pin_room"),
        ]

        # Constraint columns - use combobox-like text fields
        headers += [
            tr("constraints.allowed_days"), tr("constraints.excluded_days"),
            tr("constraints.allowed_start_times"), tr("constraints.excluded_start_times"),
            tr("constraints.required_classrooms"), tr("constraints.excluded_classrooms"),
        ]

        self._col_code = 0
        self._col_name = 1
        self._col_lecturer = 2
        self._col_duration = 3
        self._col_participants = 4
        self._col_location_type = 5
        self._col_targets_start = 6
        self._col_targets_end = 6 + len(self._target_labels)
        self._col_all_targets = self._col_targets_end
        self._col_joint = self._col_all_targets + 1
        self._col_pinned = self._col_joint + 1
        self._col_pin_day = self._col_pinned + 1
        self._col_pin_time = self._col_pin_day + 1
        self._col_pin_room = self._col_pin_time + 1
        self._col_allowed_days = self._col_pin_room + 1
        self._col_excluded_days = self._col_allowed_days + 1
        self._col_allowed_times = self._col_excluded_days + 1
        self._col_excluded_times = self._col_allowed_times + 1
        self._col_req_rooms = self._col_excluded_times + 1
        self._col_excl_rooms = self._col_req_rooms + 1

        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents)
        # Make name and lecturer columns wider
        self.table.horizontalHeader().setSectionResizeMode(
            self._col_name, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(
            self._col_lecturer, QHeaderView.ResizeMode.Stretch)
        # Give constraint columns readable minimum widths
        for col in (self._col_allowed_days, self._col_excluded_days,
                    self._col_allowed_times, self._col_excluded_times,
                    self._col_req_rooms, self._col_excl_rooms):
            self.table.horizontalHeader().setSectionResizeMode(
                col, QHeaderView.ResizeMode.Interactive)
            self.table.setColumnWidth(col, 120)
        # Code column interactive width
        self.table.horizontalHeader().setSectionResizeMode(
            self._col_code, QHeaderView.ResizeMode.Interactive)
        self.table.setColumnWidth(self._col_code, 80)

    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)

        # Class Code (text)
        self.table.setItem(row, self._col_code, QTableWidgetItem(""))
        # Name (text)
        self.table.setItem(row, self._col_name, QTableWidgetItem(""))
        # Lecturer (dropdown — matches Pin Day/Time/Room pattern)
        lec_combo = QComboBox()
        lec_combo.addItem("")
        lecturers = self.state.get("lecturers", [])
        if lecturers:
            lec_combo.addItems(lecturers)
        lec_combo.setCurrentIndex(0)
        self.table.setCellWidget(row, self._col_lecturer, lec_combo)

        # Duration (spinbox)
        dur_spin = QSpinBox()
        dur_spin.setRange(1, max(len(self.state["slots"]), 1))
        dur_spin.setValue(1)
        self.table.setCellWidget(row, self._col_duration, dur_spin)

        # Participants (spinbox)
        part_spin = QSpinBox()
        part_spin.setRange(0, 9999)
        part_spin.setSpecialValueText(tr("labels.no_limit"))
        part_spin.setValue(0)
        self.table.setCellWidget(row, self._col_participants, part_spin)

        # Location Type (dropdown)
        loc_combo = QComboBox()
        for lt in LOCATION_TYPES:
            loc_combo.addItem(get_location_label(lt), lt)
        loc_combo.currentIndexChanged.connect(
            lambda _, r=row: self._on_bulk_location_changed(r))
        self.table.setCellWidget(row, self._col_location_type, loc_combo)

        # Target checkboxes
        for i in range(len(self._target_labels)):
            cb = QCheckBox()
            w = QWidget()
            lo = QHBoxLayout(w)
            lo.addWidget(cb)
            lo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lo.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, self._col_targets_start + i, w)

        # All-targets toggle checkbox
        all_tgt_cb = QCheckBox()
        all_tgt_cb.setToolTip(tr("buttons.select_all") + " / " + tr("buttons.deselect_all"))
        all_tgt_cb.stateChanged.connect(lambda state, r=row: self._toggle_all_targets(r, state))
        atw = QWidget()
        atl = QHBoxLayout(atw)
        atl.addWidget(all_tgt_cb)
        atl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        atl.setContentsMargins(0, 0, 0, 0)
        self.table.setCellWidget(row, self._col_all_targets, atw)

        # Joint checkbox
        joint_cb = QCheckBox()
        joint_cb.setChecked(False)
        jw = QWidget()
        jl = QHBoxLayout(jw)
        jl.addWidget(joint_cb)
        jl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        jl.setContentsMargins(0, 0, 0, 0)
        self.table.setCellWidget(row, self._col_joint, jw)

        # Pinned checkbox
        pin_cb = QCheckBox()
        pin_cb.stateChanged.connect(self._update_pin_columns_visibility)
        pw = QWidget()
        pl = QHBoxLayout(pw)
        pl.addWidget(pin_cb)
        pl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pl.setContentsMargins(0, 0, 0, 0)
        self.table.setCellWidget(row, self._col_pinned, pw)

        # Pin Day combo
        day_combo = QComboBox()
        day_combo.addItem("", "")
        for d in self.state["days"]:
            day_combo.addItem(day_label(d), d)
        self.table.setCellWidget(row, self._col_pin_day, day_combo)

        # Pin Time combo
        time_combo = QComboBox()
        time_combo.addItem("")
        time_combo.addItems(self.state["slots"])
        self.table.setCellWidget(row, self._col_pin_time, time_combo)

        # Pin Room combo
        room_combo = QComboBox()
        room_combo.addItem("")
        room_combo.addItems(self.state["classrooms"])
        self.table.setCellWidget(row, self._col_pin_room, room_combo)

        # Constraint multi-select buttons
        days = self.state["days"]
        slots = self.state["slots"]
        rooms = self.state["classrooms"]
        day_labels = {d: day_label(d) for d in days}
        for col, items, lbls in [
            (self._col_allowed_days, days, day_labels),
            (self._col_excluded_days, days, day_labels),
            (self._col_allowed_times, slots, None),
            (self._col_excluded_times, slots, None),
            (self._col_req_rooms, rooms, None),
            (self._col_excl_rooms, rooms, None),
        ]:
            btn = MultiSelectButton(items, labels=lbls)
            self.table.setCellWidget(row, col, btn)

        self._update_count()

    def _selected_rows_or_all(self):
        rows = sorted(set(idx.row() for idx in self.table.selectedIndexes()))
        if rows:
            return rows
        return list(range(self.table.rowCount()))

    def _target_col_indices_for_year(self, year):
        cols = []
        for i, (yr, _br, _label) in enumerate(self._target_labels):
            if yr == year:
                cols.append(self._col_targets_start + i)
        return cols

    def _apply_year_targets_to_rows(self, checked):
        year = self._year_group_combo.currentText().strip()
        if not year:
            return
        target_cols = self._target_col_indices_for_year(year)
        if not target_cols:
            return
        state_val = Qt.CheckState.Checked.value if checked else Qt.CheckState.Unchecked.value
        for row in self._selected_rows_or_all():
            for col in target_cols:
                w = self.table.cellWidget(row, col)
                if not w:
                    continue
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(state_val == Qt.CheckState.Checked.value)

    def _fill_row_from_class(self, row, cls):
        code_item = self.table.item(row, self._col_code)
        if code_item:
            code_item.setText(cls.get("class_code", ""))
        self.table.item(row, self._col_name).setText(cls.get("name", ""))
        lec_combo = self.table.cellWidget(row, self._col_lecturer)
        if lec_combo:
            lec_combo.setCurrentText(cls.get("lecturer", ""))
        dur_spin = self.table.cellWidget(row, self._col_duration)
        if dur_spin:
            dur_spin.setValue(max(1, int(cls.get("duration", 1) or 1)))
        part_spin = self.table.cellWidget(row, self._col_participants)
        if part_spin:
            part_spin.setValue(max(0, int(cls.get("participants", 0) or 0)))

        loc_combo = self.table.cellWidget(row, self._col_location_type)
        if loc_combo:
            lt = cls.get("location_type", LOCATION_FACE_TO_FACE)
            idx = loc_combo.findData(lt)
            if idx >= 0:
                loc_combo.setCurrentIndex(idx)

        target_pairs = {(t.get("year"), t.get("branch")) for t in cls.get("targets", [])}
        for i, (yr, br, _label) in enumerate(self._target_labels):
            w = self.table.cellWidget(row, self._col_targets_start + i)
            if not w:
                continue
            cb = w.findChild(QCheckBox)
            if cb:
                cb.setChecked((yr, br) in target_pairs)

        joint_w = self.table.cellWidget(row, self._col_joint)
        if joint_w:
            joint_cb = joint_w.findChild(QCheckBox)
            if joint_cb:
                joint_cb.setChecked(bool(cls.get("joint_session", False)))

        pin_w = self.table.cellWidget(row, self._col_pinned)
        if pin_w:
            pin_cb = pin_w.findChild(QCheckBox)
            if pin_cb:
                pin_cb.setChecked(bool(cls.get("pinned", False)))

        day_combo = self.table.cellWidget(row, self._col_pin_day)
        if day_combo:
            idx = day_combo.findData(cls.get("pinned_day", ""))
            if idx >= 0:
                day_combo.setCurrentIndex(idx)
        time_combo = self.table.cellWidget(row, self._col_pin_time)
        if time_combo:
            time_combo.setCurrentText(cls.get("pinned_time", ""))
        room_combo = self.table.cellWidget(row, self._col_pin_room)
        if room_combo:
            room_combo.setCurrentText(cls.get("pinned_classroom", ""))

        for attr, col in [
            ("allowed_days", self._col_allowed_days),
            ("excluded_days", self._col_excluded_days),
            ("allowed_times", self._col_allowed_times),
            ("excluded_times", self._col_excluded_times),
            ("required_classrooms", self._col_req_rooms),
            ("excluded_classrooms", self._col_excl_rooms),
        ]:
            w = self.table.cellWidget(row, col)
            if isinstance(w, MultiSelectButton):
                w.set_checked(cls.get(attr, []))

    def _import_from_excel(self):
        classes = _import_class_rows(self, self.state)
        if not classes:
            return
        self.table.setRowCount(0)
        for cls in classes:
            if not cls.get("name") and not cls.get("lecturer"):
                continue
            self._add_row()
            self._fill_row_from_class(self.table.rowCount() - 1, cls)
        self._update_pin_columns_visibility()
        self._update_count()

    def _generate_template(self):
        _generate_class_template(self)

    def _toggle_all_targets(self, row, state):
        """Toggle all target checkboxes in the given row."""
        checked = state == Qt.CheckState.Checked.value
        for i in range(len(self._target_labels)):
            w = self.table.cellWidget(row, self._col_targets_start + i)
            if w:
                cb = w.findChild(QCheckBox)
                if cb:
                    cb.setChecked(checked)

    # ── Local undo/redo for table operations ────────────────────────────

    def _read_row_as_class(self, row):
        """Extract all data from a table row into a class dict."""
        cls = new_class()
        code_item = self.table.item(row, self._col_code)
        cls["class_code"] = code_item.text().strip() if code_item else ""
        name_item = self.table.item(row, self._col_name)
        cls["name"] = name_item.text().strip() if name_item else ""
        lec_w = self.table.cellWidget(row, self._col_lecturer)
        cls["lecturer"] = lec_w.currentText().strip() if lec_w else ""
        dur_w = self.table.cellWidget(row, self._col_duration)
        cls["duration"] = dur_w.value() if dur_w else 1
        part_w = self.table.cellWidget(row, self._col_participants)
        cls["participants"] = part_w.value() if part_w else 0
        loc_w = self.table.cellWidget(row, self._col_location_type)
        cls["location_type"] = loc_w.currentData() if loc_w else LOCATION_FACE_TO_FACE
        targets = []
        for i, (yr, br, _) in enumerate(self._target_labels):
            if self._get_checkbox(row, self._col_targets_start + i):
                targets.append({"year": yr, "branch": br})
        cls["targets"] = targets
        cls["joint_session"] = self._get_checkbox(row, self._col_joint)
        cls["pinned"] = self._get_checkbox(row, self._col_pinned)
        day_w = self.table.cellWidget(row, self._col_pin_day)
        cls["pinned_day"] = (day_w.currentData() or "") if day_w else ""
        time_w = self.table.cellWidget(row, self._col_pin_time)
        cls["pinned_time"] = time_w.currentText() if time_w else ""
        room_w = self.table.cellWidget(row, self._col_pin_room)
        cls["pinned_classroom"] = room_w.currentText() if room_w else ""
        for attr, col in [
            ("allowed_days", self._col_allowed_days),
            ("excluded_days", self._col_excluded_days),
            ("allowed_times", self._col_allowed_times),
            ("excluded_times", self._col_excluded_times),
            ("required_classrooms", self._col_req_rooms),
            ("excluded_classrooms", self._col_excl_rooms),
        ]:
            w = self.table.cellWidget(row, col)
            cls[attr] = w.checked_items() if w else []
        return normalize_class_data(cls)

    def _snapshot_table(self):
        """Return a list of class dicts representing the entire table."""
        return [self._read_row_as_class(r) for r in range(self.table.rowCount())]

    def _restore_table(self, classes):
        """Clear and repopulate the table from a list of class dicts."""
        self.table.setRowCount(0)
        for cls in classes:
            self._add_row()
            self._fill_row_from_class(self.table.rowCount() - 1, cls)
        self._update_pin_columns_visibility()
        self._update_count()

    def _push_local_undo(self, label=""):
        """Snapshot current table state before a change."""
        snapshot = self._snapshot_table()
        self._undo_stack_local.append((label, snapshot))
        if len(self._undo_stack_local) > self._max_undo_local:
            self._undo_stack_local.pop(0)
        self._redo_stack_local.clear()

    def _local_undo(self):
        if not self._undo_stack_local:
            return
        label, snapshot = self._undo_stack_local.pop()
        current = self._snapshot_table()
        self._redo_stack_local.append((label, current))
        self._restore_table(snapshot)

    def _local_redo(self):
        if not self._redo_stack_local:
            return
        label, snapshot = self._redo_stack_local.pop()
        current = self._snapshot_table()
        self._undo_stack_local.append((label, current))
        self._restore_table(snapshot)

    def _add_row_with_undo(self):
        self._push_local_undo("add row")
        self._add_row()

    def _add_5_rows_with_undo(self):
        self._push_local_undo("add 5 rows")
        for _ in range(5):
            self._add_row()

    def _import_from_excel_with_undo(self):
        self._push_local_undo("import")
        self._import_from_excel()

    # ── Copy / Paste ─────────────────────────────────────────────────────

    def _copy_rows(self):
        """Copy selected rows to clipboard as tab-separated class data."""
        rows = sorted(set(idx.row() for idx in self.table.selectedIndexes()))
        if not rows:
            return
        lines = []
        for row in rows:
            cls = self._read_row_as_class(row)
            io_row = _class_to_io_row(cls)
            lines.append("\t".join(str(io_row.get(h, "")) for h in _class_io_headers()))
        # Header line + data lines
        header = "\t".join(_class_io_headers())
        QApplication.clipboard().setText(header + "\n" + "\n".join(lines))

    def _paste_rows(self):
        """Paste tab-separated clipboard data as new rows."""
        text = QApplication.clipboard().text()
        if not text or not text.strip():
            return
        self._push_local_undo("paste")
        lines = [l for l in text.strip().split("\n") if l.strip()]
        if not lines:
            return
        # Detect if first line is a header row
        first_cells = lines[0].split("\t")
        expected_headers = {
            _class_io_header_map()["name"].casefold(),
            _class_io_header_map()["class_code"].casefold(),
            _class_io_header_map()["lecturer"].casefold(),
        }
        is_header = any(cell.strip().casefold() in expected_headers for cell in first_cells)
        if is_header and len(lines) > 1:
            col_map = {cell.strip(): i for i, cell in enumerate(first_cells)}
            data_lines = lines[1:]
        else:
            col_map = {h: i for i, h in enumerate(_class_io_headers())}
            data_lines = lines
        for line in data_lines:
            cells = line.split("\t")
            row_dict = {}
            for header, idx in col_map.items():
                if idx < len(cells):
                    row_dict[header] = cells[idx].strip()
            cls = _class_from_io_row(row_dict, self.state)
            if not cls.get("name") and not cls.get("lecturer"):
                continue
            self._add_row()
            self._fill_row_from_class(self.table.rowCount() - 1, cls)
        self._update_pin_columns_visibility()
        self._update_count()

    # ── Row operations ───────────────────────────────────────────────────

    def _remove_selected(self):
        rows = sorted(set(idx.row() for idx in self.table.selectedIndexes()), reverse=True)
        if not rows:
            return
        self._push_local_undo("delete rows")
        for row in rows:
            self.table.removeRow(row)
        self._update_count()

    def _update_count(self):
        self.row_count_label.setText(f"{tr('labels.rows')} {self.table.rowCount()}")

    def _apply_bulk_filter(self):
        """Filter BulkAdd table rows by search text and filter combo."""
        search = self._bulk_search.text().strip().lower()
        filter_idx = self._bulk_filter_combo.currentIndex()
        for row in range(self.table.rowCount()):
            # Gather row text for search
            name_item = self.table.item(row, self._col_name)
            name = (name_item.text().strip() if name_item else "").lower()
            code_item = self.table.item(row, self._col_code)
            code = (code_item.text().strip() if code_item else "").lower()
            lec_w = self.table.cellWidget(row, self._col_lecturer)
            lecturer = (lec_w.currentText().strip() if lec_w else "").lower()

            # Search match
            if search and search not in name and search not in code and search not in lecturer:
                self.table.setRowHidden(row, True)
                continue

            # Filter match
            is_filled = bool(name or lecturer)
            if filter_idx == 0:  # All
                self.table.setRowHidden(row, False)
            elif filter_idx == 1:  # Filled
                self.table.setRowHidden(row, not is_filled)
            elif filter_idx == 2:  # Empty
                self.table.setRowHidden(row, is_filled)
            elif filter_idx == 3:  # Pinned
                self.table.setRowHidden(row, not self._get_checkbox(row, self._col_pinned))
            elif filter_idx == 4:  # Joint
                self.table.setRowHidden(row, not self._get_checkbox(row, self._col_joint))
            elif filter_idx == 5:  # With constraints
                has_constraints = False
                for col in (self._col_allowed_days, self._col_excluded_days,
                            self._col_allowed_times, self._col_excluded_times,
                            self._col_req_rooms, self._col_excl_rooms):
                    w = self.table.cellWidget(row, col)
                    if isinstance(w, MultiSelectButton) and w.checked_items():
                        has_constraints = True
                        break
                self.table.setRowHidden(row, not has_constraints)

    def _get_checkbox(self, row, col):
        w = self.table.cellWidget(row, col)
        if w:
            cb = w.findChild(QCheckBox)
            if cb:
                return cb.isChecked()
        return False

    def _parse_list(self, text, valid_set=None):
        """Parse comma-separated text into a list, optionally filtering by valid set."""
        items = [x.strip() for x in text.split(",") if x.strip()]
        if valid_set is not None:
            items = [x for x in items if x in valid_set]
        return items

    def _ok(self):
        classes = []
        errors = []

        for row in range(self.table.rowCount()):
            code_item = self.table.item(row, self._col_code)
            class_code = code_item.text().strip() if code_item else ""
            name_item = self.table.item(row, self._col_name)
            name = name_item.text().strip() if name_item else ""
            lec_widget = self.table.cellWidget(row, self._col_lecturer)
            lecturer = lec_widget.currentText().strip() if lec_widget else ""

            # Skip completely empty rows
            if not name and not lecturer:
                continue

            dur_widget = self.table.cellWidget(row, self._col_duration)
            duration = dur_widget.value() if dur_widget else 1

            part_widget = self.table.cellWidget(row, self._col_participants)
            participants = part_widget.value() if part_widget else 0

            # Targets
            targets = []
            for i, (yr, br, _) in enumerate(self._target_labels):
                if self._get_checkbox(row, self._col_targets_start + i):
                    targets.append({"year": yr, "branch": br})

            joint = self._get_checkbox(row, self._col_joint)
            pinned = self._get_checkbox(row, self._col_pinned)

            cls = new_class()
            cls["class_code"] = class_code
            cls["name"] = name
            cls["lecturer"] = lecturer
            cls["duration"] = duration
            cls["participants"] = participants

            loc_w = self.table.cellWidget(row, self._col_location_type)
            cls["location_type"] = loc_w.currentData() if loc_w else LOCATION_FACE_TO_FACE

            cls["targets"] = targets
            cls["joint_session"] = joint if len(targets) > 1 else True

            is_physical = needs_physical_room(cls)

            if pinned:
                day_w = self.table.cellWidget(row, self._col_pin_day)
                time_w = self.table.cellWidget(row, self._col_pin_time)
                room_w = self.table.cellWidget(row, self._col_pin_room)
                cls["pinned"] = True
                cls["pinned_day"] = (day_w.currentData() or "") if day_w else ""
                cls["pinned_time"] = time_w.currentText() if time_w else ""
                cls["pinned_classroom"] = (room_w.currentText() if room_w else "") if is_physical else None

            # Validate using canonical validator
            row_errors = validate_class_fields(cls)
            if row_errors:
                label = name or f"{tr('labels.row')} {row + 1}"
                for e in row_errors:
                    errors.append(f"{tr('labels.row')} {row + 1} ({label}): {e}")
                continue

            # Constraints (read from MultiSelectButton widgets)
            for attr, col in [
                ("allowed_days", self._col_allowed_days),
                ("excluded_days", self._col_excluded_days),
                ("allowed_times", self._col_allowed_times),
                ("excluded_times", self._col_excluded_times),
            ]:
                w = self.table.cellWidget(row, col)
                cls[attr] = w.checked_items() if w else []
            # Room constraints only for face-to-face
            if is_physical:
                for attr, col in [
                    ("required_classrooms", self._col_req_rooms),
                    ("excluded_classrooms", self._col_excl_rooms),
                ]:
                    w = self.table.cellWidget(row, col)
                    cls[attr] = w.checked_items() if w else []

            classes.append(normalize_class_data(cls))

        if errors:
            QMessageBox.warning(
                self, tr("dialogs.validation.title"),
                tr("dialogs.validation.fix_errors") + "\n\n" + "\n".join(errors))
            return

        if not classes:
            QMessageBox.information(
                self, tr("dialogs.bulk_add.title"),
                tr("errors.no_classes_entered"))
            return

        self.result = classes
        self.accept()


# ── Bulk Schedule Results Dialog ────────────────────────────────────────────

class BulkResultsDialog(QDialog):
    """Shows results of batch scheduling: what was placed and what wasn't.

    Optionally displays schedule analytics and optimization insights
    when analytics and/or reschedule_explanation are provided.
    """

    def __init__(self, parent, placed, unplaced, rescheduled=False,
                 analytics=None, reschedule_explanation=None,
                 negotiation_result=None, negotiation_source=None,
                 infeasibility=None):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\U0001F4CB  " + tr("dialogs.bulk_results.title"))
        self.setMinimumSize(700, 500)
        self.result = True  # Accept by default
        # `negotiation_source` is a zero-argument callable producing the report
        # (ST-PERF-007); `negotiation_result` stays accepted so existing callers
        # keep working unchanged.
        self._negotiation_source = negotiation_source
        self._negotiation_result = negotiation_result
        self._negotiation_built = False
        self._negotiation_tab_index = None

        layout = QVBoxLayout(self)

        total = len(placed) + len(unplaced)

        # Reschedule notice
        if rescheduled:
            notice = QLabel(
                f"<b>{tr('labels.note')}</b> "
                f"{tr('status.classes_reorganized')}")
            notice.setStyleSheet(
                "background: #FEF3C7; color: #92400E; padding: 8px; "
                "border: 1px solid #F59E0B; font-size: 9pt;")
            notice.setWordWrap(True)
            layout.addWidget(notice)

        # Summary with quality score
        summary_text = ""
        if unplaced:
            summary_text = (
                f"<b>{len(placed)}</b> / {total} "
                f"{tr('status.classes_placed_count')} "
                f"<span style='color:#DC2626'><b>{len(unplaced)}</b> "
                f"{tr('status.could_not_place')}</span>")
        else:
            summary_text = (
                f"<span style='color:#16A34A'><b>{tr('labels.all')} {total} "
                f"{tr('status.classes_placed_count')}</b></span>")
        if analytics:
            grade = analytics.get("grade", "")
            score = analytics.get("global_score", 0)
            grade_color = {"A": "#16A34A", "B": "#2563EB", "C": "#D97706",
                           "D": "#EA580C", "F": "#DC2626"}.get(grade, "#6B7280")
            summary_text += (
                f"<br><span style='color:{grade_color}'>"
                f"{tr('analytics.schedule_quality')}: <b>{score:.0f}/100</b> "
                f"({tr('labels.grade')}: <b>{grade}</b>)</span>")
        summary = QLabel(summary_text)
        summary.setStyleSheet("font-size: 11pt; padding: 8px;")
        summary.setWordWrap(True)
        layout.addWidget(summary)

        # ST-SCHED-014. The global bottleneck, above the per-class list and
        # visually distinct from it, because it answers a different question.
        # A list of unplaced classes says which lessons did not fit; this says
        # the instance CANNOT be built -- 'you are asking for 14 class-hours
        # and the building offers 8 room-hours' -- which is arithmetic rather
        # than search, and is the only kind of problem that no amount of
        # rearranging can fix. Told 'all candidate slots are occupied' forty
        # times instead, a user starts adding rooms at random.
        #
        # diagnose_infeasibility is deliberately one-sided: it reports only
        # what it can PROVE, and passing its checks does NOT mean the instance
        # is satisfiable. The wording must not promise the converse.
        if infeasibility and infeasibility.get("message"):
            bottleneck = QLabel(
                f"<b>{tr('dialogs.bulk_results.impossible_title')}</b><br>"
                f"{html.escape(str(infeasibility['message']))}")
            bottleneck.setWordWrap(True)
            bottleneck.setStyleSheet(
                "background: #FEE2E2; color: #991B1B; padding: 8px; "
                "border: 1px solid #DC2626; border-radius: 4px; "
                "font-size: 9pt;")
            layout.addWidget(bottleneck)
            self._bottleneck_label = bottleneck

        tabs = QTabWidget()
        layout.addWidget(tabs)

        # Placed tab
        if placed:
            placed_tree = QTreeWidget()
            placed_tree.setHeaderLabels([
                tr("labels.class_name"), tr("labels.lecturer"),
                tr("labels.day"), tr("labels.start_time"), tr("labels.classroom")])
            placed_tree.setRootIsDecorated(False)
            placed_tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            for cls, day, slot, room in placed:
                code = cls.get("class_code", "")
                display_name = f"[{code}] {cls['name']}" if code else cls["name"]
                display_room = get_effective_room_resource_for_class(
                    cls, room_override=room)
                QTreeWidgetItem(placed_tree, [
                    display_name, cls["lecturer"], day, slot, display_room])
            tabs.addTab(placed_tree,
                        f"{tr('labels.placed')} ({len(placed)})")

        # Unplaced tab
        if unplaced:
            unplaced_tree = QTreeWidget()
            unplaced_tree.setHeaderLabels([
                tr("labels.class_name"), tr("labels.lecturer"), tr("labels.reason")])
            unplaced_tree.setRootIsDecorated(False)
            unplaced_tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            for cls, reason in unplaced:
                code = cls.get("class_code", "")
                display_name = f"[{code}] {cls['name']}" if code else cls["name"]
                item = QTreeWidgetItem(unplaced_tree, [
                    display_name, cls["lecturer"], reason])
                # ST-UI-015: the reason is ELIDED by Qt, not truncated -- the
                # full text is present but unreachable (measured: 163 chars
                # needing 1956 px in a 224 px column, with no tooltip). The
                # data was always there; nothing surfaced it.
                item.setToolTip(2, reason)
            tabs.addTab(unplaced_tree,
                        f"{tr('labels.unplaced')} ({len(unplaced)})")

        # Optimization Insights tab
        if reschedule_explanation or analytics:
            insights_widget = QWidget()
            insights_layout = QVBoxLayout(insights_widget)
            insights_layout.setContentsMargins(8, 8, 8, 8)

            if reschedule_explanation:
                # Verdict
                verdict = reschedule_explanation.get("verdict", "")
                verdict_label = QLabel(f"<b>{verdict}</b>")
                verdict_label.setStyleSheet("font-size: 11pt; padding: 4px;")
                insights_layout.addWidget(verdict_label)

                # Improvement details
                for imp in reschedule_explanation.get("improvements", []):
                    imp_label = QLabel(
                        f"  + {imp['description']}")
                    imp_label.setStyleSheet("color: #16A34A; font-size: 9pt;")
                    insights_layout.addWidget(imp_label)
                for deg in reschedule_explanation.get("degradations", []):
                    deg_label = QLabel(
                        f"  - {deg['description']}")
                    deg_label.setStyleSheet("color: #DC2626; font-size: 9pt;")
                    insights_layout.addWidget(deg_label)

                # Engine info
                engine = reschedule_explanation.get("engine", "")
                if engine:
                    engine_label = QLabel(
                        f"{tr('labels.engine')}: {engine}")
                    engine_label.setStyleSheet(
                        "color: #6B7280; font-size: 8pt; padding-top: 4px;")
                    insights_layout.addWidget(engine_label)

            if analytics and analytics.get("insights"):
                if reschedule_explanation:
                    sep = QFrame()
                    sep.setFrameShape(QFrame.Shape.HLine)
                    sep.setStyleSheet("color: #E5E7EB;")
                    insights_layout.addWidget(sep)

                insights_header = QLabel(f"<b>{tr('analytics.schedule_insights')}</b>")
                insights_header.setStyleSheet("font-size: 10pt; padding: 4px;")
                insights_layout.addWidget(insights_header)

                for insight in analytics["insights"]:
                    itype = insight.get("type", "info")
                    color = {"success": "#16A34A", "warning": "#D97706",
                             "info": "#2563EB"}.get(itype, "#6B7280")
                    icon = {"success": "+", "warning": "!", "info": "*"}.get(
                        itype, "-")
                    insight_label = QLabel(
                        f"  [{icon}] {insight['message']}")
                    insight_label.setStyleSheet(
                        f"color: {color}; font-size: 9pt;")
                    insight_label.setWordWrap(True)
                    insights_layout.addWidget(insight_label)

            insights_layout.addStretch()

            scroll = QScrollArea()
            scroll.setWidget(insights_widget)
            scroll.setWidgetResizable(True)
            tabs.addTab(scroll, tr("analytics.insights"))

        # Negotiation tab — contents built on first selection, not here
        # (ST-PERF-007). The label needs only len(unplaced), already known.
        if unplaced and (self._negotiation_source is not None
                         or self._negotiation_result):
            self._negotiation_container = QWidget()
            self._negotiation_layout = QVBoxLayout(self._negotiation_container)
            self._negotiation_layout.setContentsMargins(8, 8, 8, 8)
            self._tabs = tabs
            self._negotiation_tab_index = tabs.addTab(
                self._negotiation_container,
                f"{tr('buttons.negotiate')} ({len(unplaced)})")
            tabs.currentChanged.connect(self._on_results_tab_changed)

        # Bottom
        bottom = QHBoxLayout()
        bottom.addStretch()
        if placed:
            ok_btn = QPushButton(tr("buttons.accept_results"))
            ok_btn.setStyleSheet(
                "background: #16A34A; color: white; font-weight: bold; padding: 6px 16px;")
            ok_btn.clicked.connect(self.accept)
            bottom.addWidget(ok_btn)
        cancel_btn = QPushButton(tr("buttons.discard_all"))
        cancel_btn.clicked.connect(self._discard)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

    def _on_results_tab_changed(self, index):
        if index == self._negotiation_tab_index:
            self._build_negotiation_tab()

    def _build_negotiation_tab(self):
        """Populate the negotiation tab. Runs once, on first selection.

        ST-PERF-007: producing this report costs about as much as the solve
        itself (measured at ~727 ms on 250 classes, ~5.8 s on 600), and this
        dialog is constructed after EVERY reschedule. Reading it in
        __init__ would move that cost a few lines inside the same frozen
        stretch, and the user would feel no difference at all.
        """
        if self._negotiation_built:
            return
        self._negotiation_built = True
        if (self._negotiation_result is None
                and self._negotiation_source is not None):
            self._negotiation_result = self._negotiation_source()
        # Negotiation tab (when unplaced classes have suggestions)
        if self._negotiation_result:
            neg_reports = self._negotiation_result.get("class_reports", [])
            neg_diag = self._negotiation_result.get("diagnostic_summary", {})
            if neg_reports:
                neg_widget = QWidget()
                neg_layout = QVBoxLayout(neg_widget)
                neg_layout.setContentsMargins(8, 8, 8, 8)

                # Assessment
                assessment = neg_diag.get("overall_assessment", "")
                if assessment:
                    assess_label = QLabel(f"<b>{assessment}</b>")
                    assess_label.setStyleSheet("font-size: 10pt; padding: 4px;")
                    assess_label.setWordWrap(True)
                    neg_layout.addWidget(assess_label)

                for report in neg_reports:
                    status_color = {"infeasible": "#DC2626",
                                    "constrained": "#D97706",
                                    "ok": "#16A34A"}.get(
                        report["status"], "#6B7280")
                    name_label = QLabel(
                        f"<b style='color:{status_color}'>"
                        f"{report['class_name']}</b> — "
                        f"{report['summary']}")
                    name_label.setWordWrap(True)
                    name_label.setStyleSheet("font-size: 9pt; padding: 2px;")
                    neg_layout.addWidget(name_label)

                    for sug in report.get("suggestions", [])[:3]:
                        sug_label = QLabel(
                            f"  → {sug['description']}  "
                            f"({sug['impact_label']})")
                        sug_label.setStyleSheet(
                            "color: #2563EB; font-size: 8pt;")
                        sug_label.setWordWrap(True)
                        neg_layout.addWidget(sug_label)

                neg_layout.addStretch()
                neg_scroll = QScrollArea()
                neg_scroll.setWidget(neg_widget)
                neg_scroll.setWidgetResizable(True)
                self._negotiation_layout.addWidget(neg_scroll)
                self._tabs.setTabText(
                    self._negotiation_tab_index,
                    f"{tr('buttons.negotiate')} ({len(neg_reports)})")


    def _discard(self):
        self.result = False
        self.reject()


# ── Reschedule Confirmation Dialog ──────────────────────────────────────────

class RescheduleDialog(QDialog):
    """Reschedule confirmation with optional Advanced Optimization Goals panel.

    If the user does not expand or change the panel, result_goals is None
    and the scheduler uses default weights (full backward compatibility).
    """

    def __init__(self, parent, has_ortools=False):
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle(tr("buttons.reschedule"))
        self.setMinimumWidth(520)
        self.result_mode = None   # "standard" or "deep"
        self.result_goals = None  # None = use defaults

        from scheduler_app.optimization_goals import (
            GOAL_KEYS, DEFAULT_GOALS, PRESETS, PRESET_LABEL_KEYS, DEFAULT_PRESET,
        )
        self._GOAL_KEYS = GOAL_KEYS
        self._DEFAULT_GOALS = DEFAULT_GOALS
        self._PRESETS = PRESETS
        self._PRESET_LABEL_KEYS = PRESET_LABEL_KEYS
        self._DEFAULT_PRESET = DEFAULT_PRESET
        self._goals_modified = False

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # ── Confirmation text ──
        msg = QLabel(tr("dialogs.reschedule.confirm"))
        msg.setWordWrap(True)
        msg.setStyleSheet("font-size: 10pt; padding: 6px;")
        layout.addWidget(msg)

        # ── Advanced Optimization Goals (collapsible) ──
        self._goals_toggle = QPushButton(
            "\u25B6  " + tr("optimization.goals_title"))
        self._goals_toggle.setStyleSheet(
            "text-align: left; font-weight: bold; border: 1px solid #CBD5E1; "
            "border-radius: 6px; padding: 8px 12px; background: #F8FAFC;")
        self._goals_toggle.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._goals_toggle.clicked.connect(self._toggle_goals_panel)
        layout.addWidget(self._goals_toggle)

        self._goals_panel = QWidget()
        self._goals_panel.setVisible(False)
        gp_layout = QVBoxLayout(self._goals_panel)
        gp_layout.setContentsMargins(8, 4, 8, 8)

        # Preset row
        preset_row = QHBoxLayout()
        preset_row.addWidget(QLabel(tr("optimization.preset_label")))
        self._preset_combo = QComboBox()
        self._preset_combo.addItem(tr("optimization.custom"))
        for preset_id in PRESETS:
            self._preset_combo.addItem(tr(PRESET_LABEL_KEYS[preset_id]))
        default_preset_index = list(PRESETS.keys()).index(DEFAULT_PRESET) + 1
        self._preset_combo.setCurrentIndex(default_preset_index)
        self._preset_combo.currentIndexChanged.connect(self._on_preset_changed)
        preset_row.addWidget(self._preset_combo, 1)

        reset_btn = QPushButton(tr("buttons.reset_to_default"))
        reset_btn.setStyleSheet("font-size: 8pt; padding: 3px 10px;")
        reset_btn.clicked.connect(self._reset_to_default)
        preset_row.addWidget(reset_btn)
        gp_layout.addLayout(preset_row)

        # Slider grid
        self._sliders = {}
        self._slider_labels = {}

        slider_defs = [
            ("lecturer_compactness", tr("optimization.lecturer_compactness"),
             tr("optimization.lecturer_compactness_desc")),
            ("student_compactness", tr("optimization.student_compactness"),
             tr("optimization.student_compactness_desc")),
            ("room_utilization", tr("optimization.room_utilization"),
             tr("optimization.room_utilization_desc")),
            ("fairness", tr("optimization.fairness"),
             tr("optimization.fairness_desc")),
            ("minimal_disruption", tr("optimization.minimal_disruption"),
             tr("optimization.minimal_disruption_desc")),
            ("early_hour_preference", tr("optimization.early_hour_preference"),
             tr("optimization.early_hour_desc")),
        ]

        grid = QGridLayout()
        grid.setColumnStretch(1, 1)
        for i, (key, label, tooltip) in enumerate(slider_defs):
            name_label = QLabel(label)
            name_label.setToolTip(tooltip)
            name_label.setStyleSheet("font-size: 9pt;")
            grid.addWidget(name_label, i, 0)

            slider = QSlider(Qt.Orientation.Horizontal)
            slider.setRange(0, 100)
            slider.setValue(DEFAULT_GOALS[key])
            slider.setTickPosition(QSlider.TickPosition.TicksBelow)
            slider.setTickInterval(25)
            slider.setToolTip(tooltip)
            slider.valueChanged.connect(
                lambda val, k=key: self._on_slider_changed(k, val))
            grid.addWidget(slider, i, 1)

            val_label = QLabel(str(DEFAULT_GOALS[key]))
            val_label.setFixedWidth(30)
            val_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            val_label.setStyleSheet("font-weight: bold; font-size: 9pt;")
            grid.addWidget(val_label, i, 2)

            self._sliders[key] = slider
            self._slider_labels[key] = val_label

        gp_layout.addLayout(grid)

        # Hint text
        hint = QLabel(tr("optimization.slider_info"))
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #64748B; font-size: 8pt; padding-top: 4px;")
        gp_layout.addWidget(hint)

        layout.addWidget(self._goals_panel)

        # ── Bottom buttons ──
        layout.addWidget(self._separator())
        bottom = QHBoxLayout()
        bottom.addStretch()

        # Task 6. The two modes used to be equally-primary buttons labelled
        # "Standart" and "Derin (CP-SAT)", with tooltips reading "Multi-start
        # LNS optimization" and "Deep optimization with constraint solver".
        # A school administrator cannot choose between two acronyms, and
        # neither label said what the choice costs. They now say what the user
        # GETS, and Quick is marked as the recommended default so a first
        # reschedule does not require an engine decision.
        std_btn = QPushButton(tr("optimization.standard"))
        std_btn.setStyleSheet(
            "background: #1D4ED8; color: white; font-weight: bold; "
            "padding: 8px 20px; font-size: 10pt;")
        # "normally the same timetable", not "the same timetable": the
        # production budget is multi_start_time_limit=120 s, and when that cap
        # fires the search is truncated and summary['deterministic'] goes False
        # — on an 80-class instance it measurably does. The result toast says so
        # when it happens (_reproducibility_note); the tooltip must not promise
        # more than the engine delivers.
        std_btn.setToolTip(
            tr("optimization.lns_tooltip")
            + "\n(" + tr("optimization.recommended") + ")")
        std_btn.setDefault(True)
        std_btn.setAutoDefault(True)
        std_btn.clicked.connect(lambda: self._accept_mode("standard"))
        bottom.addWidget(std_btn)

        if has_ortools:
            deep_btn = QPushButton(tr("optimization.deep_cpsat"))
            deep_btn.setStyleSheet(
                "background: #7C3AED; color: white; font-weight: bold; "
                "padding: 8px 20px; font-size: 10pt;")
            # ST-SCHED-013. summary['deterministic'] is
            # `(not clock_capped) and (not cpsat_used)`, so pressing this
            # button ALWAYS forfeits reproducibility. Phase 1 was careful the
            # engine never claims a reproducibility it cannot deliver; saying
            # nothing here would let the UI make the claim on its behalf.
            deep_btn.setToolTip(
                tr("optimization.cpsat_tooltip")
                + "\n⚠ " + tr("optimization.not_reproducible"))
            deep_btn.setDefault(False)
            deep_btn.clicked.connect(lambda: self._accept_mode("deep"))
            bottom.addWidget(deep_btn)
        else:
            # Without OR-Tools the button simply was not rendered, so a user
            # following a colleague's instructions saw a dialog that did not
            # match the description and was told nothing.
            note = QLabel(tr("optimization.deep_unavailable"))
            note.setWordWrap(True)
            note.setStyleSheet("color: #64748B; font-size: 8pt; padding: 4px;")
            layout.addWidget(note)

        cancel_btn = QPushButton(tr("buttons.cancel"))
        cancel_btn.clicked.connect(self.reject)
        bottom.addWidget(cancel_btn)
        layout.addLayout(bottom)

    @staticmethod
    def _separator():
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        return line

    def _toggle_goals_panel(self):
        visible = not self._goals_panel.isVisible()
        self._goals_panel.setVisible(visible)
        arrow = "\u25BC" if visible else "\u25B6"
        self._goals_toggle.setText(
            f"{arrow}  {tr('optimization.goals_title')}")
        self.adjustSize()

    def _on_preset_changed(self, index):
        if index <= 0:
            return  # "Custom" selected — don't change sliders
        preset_ids = list(self._PRESETS.keys())
        preset_id = preset_ids[index - 1]
        values = self._PRESETS[preset_id]
        self._apply_slider_values(values)

    def _reset_to_default(self):
        self._apply_slider_values(self._DEFAULT_GOALS)
        default_preset_index = list(self._PRESETS.keys()).index(self._DEFAULT_PRESET) + 1
        self._preset_combo.setCurrentIndex(default_preset_index)
        self._goals_modified = False

    def _apply_slider_values(self, values):
        for key, slider in self._sliders.items():
            slider.blockSignals(True)
            slider.setValue(values.get(key, 50))
            self._slider_labels[key].setText(str(slider.value()))
            slider.blockSignals(False)

    def _on_slider_changed(self, key, value):
        self._slider_labels[key].setText(str(value))
        self._goals_modified = True
        # Check if current values match any preset
        current = self._get_current_goals()
        matched = False
        for i, (_preset_id, preset) in enumerate(self._PRESETS.items()):
            if all(current.get(k) == preset.get(k) for k in self._GOAL_KEYS):
                self._preset_combo.blockSignals(True)
                self._preset_combo.setCurrentIndex(i + 1)
                self._preset_combo.blockSignals(False)
                matched = True
                break
        if not matched:
            self._preset_combo.blockSignals(True)
            self._preset_combo.setCurrentIndex(0)  # "Custom"
            self._preset_combo.blockSignals(False)

    def _get_current_goals(self):
        return {key: slider.value() for key, slider in self._sliders.items()}

    def _accept_mode(self, mode):
        self.result_mode = mode
        # Only set goals if user actually opened and modified the panel
        if self._goals_modified or self._goals_panel.isVisible():
            self.result_goals = self._get_current_goals()
        else:
            self.result_goals = None
        self.accept()


# ── Edit Classes Dialog ──────────────────────────────────────────────────────

class EditClassesDialog(QDialog):
    """Dialog for viewing, searching, editing, deleting, and exporting classes."""

    def __init__(self, parent, state, edit_callback=None):
        """
        Args:
            parent: parent widget (SchedulerApp)
            state: the full scheduler state dict
            edit_callback: callable(cls) to open edit form for a single class
        """
        super().__init__(parent)
        self.setStyleSheet(DIALOG_STYLESHEET())
        self.setWindowTitle("\u270E  " + tr("dialogs.edit_classes.title"))
        self.state = state
        self._edit_callback = edit_callback
        self.result = None  # will be set to list of deleted class refs
        self._deleted_classes = []
        self._edited_classes = []
        self.setMinimumSize(900, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # ── Search bar ──
        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("\U0001F50D"))
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText(tr("dialogs.edit_classes.search_placeholder"))
        self._search_edit.textChanged.connect(self._apply_filter)
        search_row.addWidget(self._search_edit)
        self._count_label = QLabel()
        search_row.addWidget(self._count_label)
        layout.addLayout(search_row)

        # ── Table ──
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(
            "QTableWidget { gridline-color: #E2E8F0; font-size: 9pt; }"
            "QTableWidget::item:selected { background: #DBEAFE; color: #1E293B; }")
        self.table.doubleClicked.connect(self._on_double_click)
        _install_select_all_shortcut(self, self.table)
        _install_table_keyboard_shortcuts(
            self, self.table,
            delete_callback=self._delete_selected,
            copy_callback=lambda: _default_copy_table_rows(self.table),
        )
        layout.addWidget(self.table)

        # ── Action buttons ──
        btn_row = QHBoxLayout()
        edit_btn = QPushButton("\u270E " + tr("dialogs.edit_classes.edit_selected"))
        edit_btn.clicked.connect(self._edit_selected)
        btn_row.addWidget(edit_btn)

        del_btn = QPushButton("\u2715 " + tr("dialogs.edit_classes.delete_selected"))
        del_btn.setStyleSheet(
            "QPushButton { color: #DC2626; } QPushButton:hover { background: #FEE2E2; }")
        del_btn.clicked.connect(self._delete_selected)
        btn_row.addWidget(del_btn)

        export_btn = QPushButton("\u21D7 " + tr("dialogs.edit_classes.export_selected"))
        export_btn.clicked.connect(self._export_to_excel)
        btn_row.addWidget(export_btn)

        btn_row.addStretch()

        close_btn = QPushButton(tr("buttons.close"))
        close_btn.setStyleSheet(
            "QPushButton { background: #3B82F6; color: white; border: none;"
            "  border-radius: 6px; font-weight: bold; padding: 8px 20px; }"
            "QPushButton:hover { background: #2563EB; }")
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        # Build table
        self._build_table()

    # ── Column definitions ──

    _COLUMNS = [
        "labels.class_code", "labels.class_name", "labels.lecturer",
        "labels.duration", "labels.participants", "labels.targets",
        "labels.joint", "labels.status", "protection.title",
    ]

    def _build_table(self):
        classes = self.state.get("classes", [])
        self.table.setColumnCount(len(self._COLUMNS))
        self.table.setHorizontalHeaderLabels([tr(key) for key in self._COLUMNS])
        self.table.setRowCount(len(classes))
        self._class_refs = []  # row index -> class dict reference

        for row, cls in enumerate(classes):
            self._class_refs.append(cls)
            self._populate_row(row, cls)

        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self._update_count()

    def _populate_row(self, row, cls):
        self.table.setItem(row, 0, QTableWidgetItem(cls.get("class_code", "")))
        self.table.setItem(row, 1, QTableWidgetItem(cls.get("name", "")))
        self.table.setItem(row, 2, QTableWidgetItem(cls.get("lecturer", "")))
        self.table.setItem(row, 3, QTableWidgetItem(str(cls.get("duration", 1))))
        self.table.setItem(row, 4, QTableWidgetItem(str(cls.get("participants", 0))))
        targets = cls.get("targets", [])
        target_str = ", ".join(f"{t['year']}/{t['branch']}" for t in targets)
        self.table.setItem(row, 5, QTableWidgetItem(target_str))
        self.table.setItem(row, 6, QTableWidgetItem(
            _translated_bool(cls.get("joint_session", True))))
        placed = cls.get("placed", False) or cls.get("pinned", False)
        status = tr("dialogs.edit_classes.placed_badge") if placed else tr("dialogs.edit_classes.unplaced_badge")
        status_item = QTableWidgetItem(status)
        if placed:
            status_item.setForeground(QColor("#16A34A"))
        else:
            status_item.setForeground(QColor("#D97706"))
        self.table.setItem(row, 7, status_item)
        prot = cls.get("protection", "none")
        self.table.setItem(row, 8, QTableWidgetItem(
            get_protection_label(prot)))

    def _apply_filter(self, text):
        text = text.lower()
        visible = 0
        for row in range(self.table.rowCount()):
            match = not text
            if not match:
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item and text in item.text().lower():
                        match = True
                        break
            self.table.setRowHidden(row, not match)
            if match:
                visible += 1
        self._update_count(visible)

    def _update_count(self, visible=None):
        total = self.table.rowCount()
        if visible is None:
            visible = sum(1 for r in range(total) if not self.table.isRowHidden(r))
        self._count_label.setText(f"{visible}/{total}")

    def _selected_rows(self):
        """Return sorted list of selected visible row indices."""
        return sorted(set(idx.row() for idx in self.table.selectedIndexes()
                         if not self.table.isRowHidden(idx.row())))

    def _on_double_click(self, index):
        row = index.row()
        if row < len(self._class_refs):
            self._edit_single(row)

    def _edit_selected(self):
        rows = self._selected_rows()
        if not rows:
            QMessageBox.information(self, tr("dialogs.edit_classes.title"),
                                   tr("dialogs.edit_classes.select_to_edit"))
            return
        # Edit one at a time
        for row in rows:
            self._edit_single(row)

    def _edit_single(self, row):
        if row >= len(self._class_refs):
            return
        cls = self._class_refs[row]
        if self._edit_callback:
            self._edit_callback(cls)
            # Refresh this row after editing
            self._populate_row(row, cls)
        else:
            # Fallback: open AddClassDialog directly
            dlg = AddClassDialog(self, self.state, edit_cls=cls)
            if dlg.exec() == AddClassDialog.DialogCode.Accepted and dlg.result:
                updated = dlg.result
                for key in updated:
                    cls[key] = updated[key]
                self._edited_classes.append(cls)
                self._populate_row(row, cls)

    def _delete_selected(self):
        rows = self._selected_rows()
        if not rows:
            return
        reply = QMessageBox.question(
            self, tr("dialogs.edit_classes.confirm_delete_title"),
            tr("dialogs.edit_classes.confirm_delete").format(n=len(rows)),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply != QMessageBox.StandardButton.Yes:
            return
        # Collect classes to delete (reverse order for safe removal)
        classes_to_delete = [self._class_refs[r] for r in rows if r < len(self._class_refs)]
        for cls in classes_to_delete:
            if cls in self.state["classes"]:
                self.state["classes"].remove(cls)
                self._deleted_classes.append(cls)
        # Rebuild table
        self._class_refs.clear()
        self._build_table()
        self._search_edit.clear()
        QMessageBox.information(
            self, tr("dialogs.edit_classes.title"),
            tr("dialogs.edit_classes.deleted_count").format(n=len(classes_to_delete)))

    def _export_to_excel(self):
        rows = self._selected_rows()
        if rows:
            classes = [self._class_refs[r] for r in rows if r < len(self._class_refs)]
        else:
            # Export all visible
            classes = [self._class_refs[r] for r in range(self.table.rowCount())
                       if not self.table.isRowHidden(r) and r < len(self._class_refs)]
        if not classes:
            classes = list(self.state.get("classes", []))
        io_rows = [_class_to_io_row(c) for c in classes]
        _export_class_rows(self, io_rows)
