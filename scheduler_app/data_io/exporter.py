"""Export pipeline for the scheduler.

Exports final timetable to Excel (.xlsx), CSV, and optionally PDF.
Does not modify scheduling logic or internal data.
"""

import csv
import html
import warnings
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    XlFont = Font          # the moved ST-ARCH-003 writer spells it this way
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from scheduler_app.logic import (
    get_placed_classes, occupied_slots_of, classroom_of, total_duration,
    target_for_slot_offset,
    get_year_color, lighten_color, build_virtual_classroom_day_layout,
    find_schedule_conflicts, conflict_partner_index,
)
# ST-UI-005. The XLSX and PDF cells paint this text on the SAME
# lighten_color(year, 0.45) background the screen uses, so they carried the
# identical WCAG failure -- in print, where it is worst. One source, three
# surfaces; fixing only the renderer would have created a fresh
# screen-vs-export divergence, which is the shape Phase 4 spent a whole batch
# closing.
from scheduler_app.constants import (
    CELL_FG_CODE, CELL_FG_NAME, CELL_FG_LECTURER, CELL_FG_ROOM,
    CELL_FG_BRANCH,
)
from scheduler_app.core.text_safety import escape_pdf_markup, csv_safe
from scheduler_app.data_io.spreadsheet_safety import neutralize_formula_cells
from scheduler_app.models import (
    get_classroom_export_labels, effective_day, effective_time,
    slot_offset_for_target, cls_key, find_off_grid_placements,
)
from scheduler_app.translations import tr
from scheduler_app.i18n.badge_formatter import get_badge
from scheduler_app.i18n.day_keys import display_day


class FinalSchedule:
    """Wrapper around scheduler state for export purposes."""

    def __init__(self, state: dict):
        self.state = state

    @property
    def days(self):
        return self.state["days"]

    @property
    def slots(self):
        return self.state["slots"]

    @property
    def classrooms(self):
        return self.state["classrooms"]

    @property
    def lecturers(self):
        return self.state.get("lecturers", [])

    @property
    def years(self):
        return self.state.get("years", {})

    def placed_classes(self):
        return get_placed_classes(self.state)

    def build_grid(self):
        """Build a (day, slot) -> list of class info dicts grid."""
        grid = {}
        for cls in self.placed_classes():
            for day, slot in occupied_slots_of(self.state, cls):
                entry = {
                    "class_code": cls.get("class_code", ""),
                    "name": cls["name"],
                    "lecturer": cls["lecturer"],
                    "room": classroom_of(cls),
                    "targets": cls.get("targets", []),
                    "cls": cls,
                }
                grid.setdefault((day, slot), []).append(entry)
        return grid


def plain_cell_text(entry):
    """Plain single-line text for a schedule entry (CSV / clipboard).

    ST-ARCH-009. Lived in `ui/cell_formatter.py`, which this module imported --
    one of the 22 upward layering violations. It needs nothing at all: no
    translation, no logic, just string assembly over a dict. So it moved to its
    only caller rather than into the i18n leaf, where it would have been the
    one member with no language content.

    Expects an entry dict with keys: name, lecturer, room, class_code.
    """
    parts = []
    code = entry.get("class_code", "")
    if code:
        parts.append(code)
    parts.append(entry["name"])
    if entry["lecturer"]:
        parts.append(entry["lecturer"])
    if entry["room"]:
        parts.append(f"[{entry['room']}]")
    return "\n".join(parts)


def _strip_hash(color):
    """Strip '#' from hex color string for openpyxl."""
    return color.lstrip("#")


def _sheet_name_for_export(base, used):
    invalid = set("[]:*?/\\")
    default_sheet = tr("labels.sheet")
    cleaned = "".join(ch for ch in str(base or default_sheet) if ch not in invalid).strip()
    if not cleaned:
        cleaned = default_sheet
    name = cleaned[:31]
    if name not in used:
        used.add(name)
        return name
    counter = 2
    while True:
        suffix = f" ({counter})"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        counter += 1


def _write_unplaceable_sheet(wb, s, drawn, used_sheet_names, days, slots):
    """List placed lessons the everything-matrix had no column for.

    ST-ARCH-003 / ST-FUNC-013. The matrix is indexed by ``state["years"]``, so
    a lesson whose target year has since been deleted -- or which carries no
    targets at all -- matches no column and used to disappear from the
    workbook entirely. Silence is the dangerous outcome here: the file looks
    complete, and a school prints a timetable with a lesson missing.

    Returns the number of rows written (0 when nothing was orphaned, in which
    case no sheet is created).
    """
    from scheduler_app.core.models import effective_day, effective_time

    missing = [c for c in get_placed_classes(s) if cls_key(c) not in drawn]
    if not missing:
        return 0

    # Reuses the PDF appendix's own heading rather than minting a key: this is
    # the workbook's version of that appendix, and the string is already
    # translated in all 22 locales (adding one would move the ST-UI-011 ratchet
    # by 21 pairs and need a translator for a sheet tab).
    ws = wb.create_sheet(
        title=_sheet_name_for_export(tr("export.appendix_offgrid"),
                                     used_sheet_names))
    headers = [
        tr("labels.class_code"), tr("labels.course"), tr("labels.lecturer"),
        tr("labels.day"), tr("labels.time"), tr("labels.classroom"),
        tr("labels.year"), tr("labels.branch"),
    ]
    hdr_fill = PatternFill("solid", fgColor="334155")
    hdr_font = XlFont(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for row, c in enumerate(sorted(missing, key=lambda x: x.get("name", "")),
                            start=2):
        targets = c.get("targets", [])
        ws.cell(row=row, column=1, value=c.get("class_code", ""))
        ws.cell(row=row, column=2, value=c.get("name", ""))
        ws.cell(row=row, column=3, value=c.get("lecturer", ""))
        ws.cell(row=row, column=4, value=display_day(effective_day(c) or ""))
        ws.cell(row=row, column=5, value=effective_time(c) or "")
        ws.cell(row=row, column=6, value=classroom_of(c) or "")
        ws.cell(row=row, column=7,
                value=", ".join(t.get("year", "") for t in targets))
        ws.cell(row=row, column=8,
                value=", ".join(t.get("branch", "") for t in targets))
    for col, width in enumerate((14, 30, 24, 14, 10, 14, 18, 12), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    return len(missing)


def _export_excel(schedule, filepath, mode="everything"):
    """Write the multi-sheet workbook the Excel menu produces.

    ST-ARCH-003. This is the engine the app has always used; it lived
    in ``ui/app.py`` as ``SchedulerApp._write_excel`` while this module
    kept a second, thinner one that nothing called. The duplication was
    not inert: Phase 5's WCAG fix landed on the copy with no users, so
    the workbook schools print kept the failing palette for a phase
    (room text at 1.55:1). The dead copy is gone; this is the only one.

    Touches no Qt and no window state -- measured before the move, it
    read exactly ``self.state_data`` and ``self._sheet_name_for_export``
    and used zero Qt symbols, which is what made the move mechanical.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(tr("errors.openpyxl_required_install"))
    s = schedule.state
    placed = get_placed_classes(s)
    days = s["days"]
    slots = s["slots"]
    # ST-UI-001: one scan for the whole workbook, so every sheet marks the
    # same clashes and the file agrees with the screen and the PDF.
    conflict_partners = conflict_partner_index(find_schedule_conflicts(s))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_sheet_names = set()

    # ── Theme matching the app's blue/slate UI ──
    grid_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=grid_side, right=grid_side,
                         top=grid_side, bottom=grid_side)
    # Day header: dark slate (#334155) with white text
    day_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    day_font = XlFont(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    # Branch sub-header: slate (#475569) with white text
    branch_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    branch_font = XlFont(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    # Corner: gray-blue (#94A3B8) with white text
    corner_fill = PatternFill(start_color="94A3B8", end_color="94A3B8", fill_type="solid")
    corner_font = XlFont(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    # Session number: light gray (#F1F5F9) with dark text
    session_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    session_font = XlFont(name="Segoe UI", size=10, bold=True, color="333333")
    # Time column: slate (#475569) with white text
    time_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
    time_font = XlFont(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    # Empty cell: near-white (#F8FAFC)
    empty_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_nowrap = Alignment(horizontal="center", vertical="center")

    def _append_rich_cell_blocks(blocks, cls, include_room=False, include_targets=False):
        """Append rich text fragments for one class into *blocks*.

        ST-ARCH-003 / ST-UI-005. These colours used to be literals here, and
        because this writer is the one the Excel menu actually reaches --
        ``data_io/exporter.py``'s Excel path has no production caller -- the
        Phase 5 contrast fix never reached the workbook a school prints.
        Measured on the exported file: room 1.55:1, class code 3.15:1,
        lecturer 3.56:1, branch 3.34:1, against a 4.5:1 requirement, while
        the screen and the PDF were correct. They now come from the same
        source those two already use.
        """
        code = cls.get("class_code", "")
        if code:
            blocks.append(TextBlock(
                InlineFont(b=True, sz=9, color=_strip_hash(CELL_FG_CODE)), code + "\n"))
        blocks.append(TextBlock(
            InlineFont(b=True, sz=10, color=_strip_hash(CELL_FG_NAME)), cls["name"] + "\n"))
        if cls.get("lecturer"):
            blocks.append(TextBlock(
                InlineFont(sz=9, color=_strip_hash(CELL_FG_LECTURER)), cls["lecturer"]))
        if include_room:
            room = classroom_of(cls)
            if room:
                blocks.append(TextBlock(
                    InlineFont(sz=9, color=_strip_hash(CELL_FG_ROOM)), "\n" + room))
        if include_targets:
            groups = ", ".join(f"{t['year']}/{t['branch']}" for t in cls.get("targets", []))
            if groups:
                blocks.append(TextBlock(
                    InlineFont(sz=8, color=_strip_hash(CELL_FG_BRANCH)), "\n" + groups))
        # Protection badge -- one source, shared with the screen and the PDF.
        emoji, label, badge_color = get_badge(cls)
        if emoji and label:
            blocks.append(TextBlock(
                InlineFont(b=True, sz=8, color=_strip_hash(badge_color)),
                "\n" + emoji + " " + label))
        return blocks

    def _build_rich_cell(cls, include_room=False, include_targets=False):
        """Build rich text cell with color-coded content matching app view."""
        blocks = []
        _append_rich_cell_blocks(
            blocks, cls, include_room=include_room,
            include_targets=include_targets,
        )
        return CellRichText(*blocks)

    def _build_stacked_rich_cell(classes, include_room=False, include_targets=False):
        """Build a stacked rich-text cell for multiple overlapping classes."""
        blocks = []
        for idx, cls in enumerate(classes):
            if idx > 0:
                blocks.append(TextBlock(
                    InlineFont(sz=8, color="94A3B8"), "\n---\n"))
            _append_rich_cell_blocks(
                blocks, cls, include_room=include_room,
                include_targets=include_targets,
            )
        return CellRichText(*blocks)

    def _apply_page_setup(ws):
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
            fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A3

    def _lesson_fill_for_class(cls):
        year_key = cls["targets"][0]["year"] if cls.get("targets") else None
        if year_key:
            yr_hex = get_year_color(s, year_key).lstrip("#")
            light_hex = lighten_color(f"#{yr_hex}", 0.45).lstrip("#")
        else:
            light_hex = "F3F4F6"
        return PatternFill(
            start_color=light_hex.upper(),
            end_color=light_hex.upper(),
            fill_type="solid",
        )

    def _apply_lesson_cell(cell, cls, rich):
        cell.value = rich
        cell.fill = _lesson_fill_for_class(cls)
        cell.alignment = center_wrap
        cell.border = cell_border

    def _write_filtered_sheet(ws, filter_fn, include_room=False, include_targets=False):
        c = ws.cell(row=1, column=1, value=tr("labels.time"))
        c.fill = corner_fill
        c.font = corner_font
        c.border = cell_border
        c.alignment = center_nowrap

        for d_idx, day in enumerate(days):
            col = d_idx + 2
            dc = ws.cell(row=1, column=col, value=tr(f"weekdays.{day}"))
            dc.fill = day_fill
            dc.font = day_font
            dc.border = cell_border
            dc.alignment = center_nowrap

        filtered_entries = []
        slot_claims = {}
        for order, cls in enumerate(placed):
            if not filter_fn(cls):
                continue
            c_day = effective_day(cls)
            c_start = effective_time(cls)
            if c_day not in days or c_start not in slots:
                continue
            start_si = slots.index(c_start)
            span = min(total_duration(cls), len(slots) - start_si)
            if span <= 0:
                continue
            entry = {
                "cls": cls,
                "day": c_day,
                "start_si": start_si,
                "span": span,
                "order": order,
            }
            filtered_entries.append(entry)
            for off in range(span):
                slot_claims.setdefault((start_si + off, c_day), []).append(entry)

        overlapping_ids = set()
        for entries in slot_claims.values():
            if len(entries) > 1:
                overlapping_ids.update(cls_key(entry["cls"]) for entry in entries)

        occupied_start = {}
        covered = set()
        overlap_cells = {}
        for entry in filtered_entries:
            cls = entry["cls"]
            c_day = entry["day"]
            start_si = entry["start_si"]
            span = entry["span"]
            if cls_key(cls) in overlapping_ids:
                for off in range(span):
                    overlap_cells.setdefault((start_si + off, c_day), []).append(entry)
                continue
            key = (start_si, c_day)
            occupied_start[key] = (cls, span)
            for off in range(span):
                covered.add((start_si + off, c_day))

        for si, slot in enumerate(slots):
            row = si + 2
            t_cell = ws.cell(row=row, column=1, value=slot)
            t_cell.fill = time_fill
            t_cell.font = time_font
            t_cell.border = cell_border
            t_cell.alignment = center_nowrap

            for d_idx, day in enumerate(days):
                col = d_idx + 2
                key = (si, day)
                if key in occupied_start:
                    cls, span = occupied_start[key]
                    rich = _build_rich_cell(cls, include_room=include_room,
                                           include_targets=include_targets)
                    if span > 1:
                        ws.merge_cells(start_row=row, start_column=col,
                                       end_row=row + span - 1, end_column=col)
                    data_cell = ws.cell(row=row, column=col)
                    _apply_lesson_cell(data_cell, cls, rich)
                    for rr in range(row, row + span):
                        ws.cell(row=rr, column=col).border = cell_border
                elif key in overlap_cells:
                    entries = sorted(overlap_cells[key], key=lambda item: item["order"])
                    classes = [item["cls"] for item in entries]
                    rich = _build_stacked_rich_cell(
                        classes,
                        include_room=include_room,
                        include_targets=include_targets,
                    )
                    data_cell = ws.cell(row=row, column=col)
                    _apply_lesson_cell(data_cell, classes[0], rich)
                    # ST-UI-001: stacking already kept both lessons here --
                    # what was missing is saying that they clash. A shared
                    # cell that is NOT a conflict (two online lessons) keeps
                    # its normal year colour.
                    keys = {cls_key(x) for x in classes}
                    if any(conflict_partners.get(cls_key(x), frozenset())
                           & (keys - {cls_key(x)}) for x in classes):
                        data_cell.fill = PatternFill(
                            start_color="FEE2E2", end_color="FEE2E2",
                            fill_type="solid")
                elif key in covered:
                    ws.cell(row=row, column=col).border = cell_border
                else:
                    blank = ws.cell(row=row, column=col, value="")
                    blank.fill = empty_fill
                    blank.border = cell_border
                    blank.alignment = center_wrap

        ws.column_dimensions["A"].width = 12
        for c_idx in range(2, len(days) + 2):
            ws.column_dimensions[get_column_letter(c_idx)].width = 24
        ws.row_dimensions[1].height = 30
        for si in range(len(slots)):
            ws.row_dimensions[si + 2].height = 70
        _apply_page_setup(ws)

    def _write_virtual_classroom_sheet(ws, filter_fn, include_room=False, include_targets=False):
        c = ws.cell(row=1, column=1, value=tr("labels.time"))
        c.fill = corner_fill
        c.font = corner_font
        c.border = cell_border
        c.alignment = center_nowrap

        layout = build_virtual_classroom_day_layout(s, filter_fn)
        day_groups = layout["day_groups"]
        blocks = layout["blocks"]
        occupied_subcolumns = layout["occupied_subcolumns"]
        starts = {
            (block["row"], block["subcolumn"]): block
            for block in blocks
        }
        covered = set(occupied_subcolumns)

        for group in day_groups:
            col_start = 2 + group["subcolumn_start"]
            col_end = col_start + group["lane_count"] - 1
            if group["lane_count"] > 1:
                ws.merge_cells(
                    start_row=1, start_column=col_start,
                    end_row=1, end_column=col_end,
                )
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = day_fill
                cell.font = day_font
                cell.border = cell_border
                cell.alignment = center_nowrap
            ws.cell(row=1, column=col_start, value=tr(f"weekdays.{group['day']}"))

        total_subcolumns = layout["total_subcolumns"]
        for si, slot in enumerate(slots):
            row = si + 2
            t_cell = ws.cell(row=row, column=1, value=slot)
            t_cell.fill = time_fill
            t_cell.font = time_font
            t_cell.border = cell_border
            t_cell.alignment = center_nowrap

            for subcolumn in range(total_subcolumns):
                col = 2 + subcolumn
                key = (si, subcolumn)
                if key in starts:
                    block = starts[key]
                    rich = _build_rich_cell(
                        block["cls"],
                        include_room=include_room,
                        include_targets=include_targets,
                    )
                    span = block["span"]
                    if span > 1:
                        ws.merge_cells(
                            start_row=row, start_column=col,
                            end_row=row + span - 1, end_column=col,
                        )
                    data_cell = ws.cell(row=row, column=col)
                    _apply_lesson_cell(data_cell, block["cls"], rich)
                    # ST-UI-001. This sheet lanes concurrent lessons rather
                    # than stacking them, so it never had an "overlap"
                    # branch to hang the marker on -- and marked no clashes
                    # at all, while the physical sheets in the same
                    # workbook did. The rule is the same non-geometric one
                    # the renderer uses: a lesson is marked if the VALIDATOR
                    # says it cannot coexist with something, not if two
                    # blocks happen to share a cell (two online lessons at
                    # one hour is normal and must stay unmarked).
                    if conflict_partners.get(cls_key(block["cls"])):
                        data_cell.fill = PatternFill(
                            start_color="FEE2E2", end_color="FEE2E2",
                            fill_type="solid")
                    for rr in range(row, row + span):
                        ws.cell(row=rr, column=col).border = cell_border
                elif key in covered:
                    ws.cell(row=row, column=col).border = cell_border
                else:
                    blank = ws.cell(row=row, column=col, value="")
                    blank.fill = empty_fill
                    blank.border = cell_border
                    blank.alignment = center_wrap

        ws.column_dimensions["A"].width = 12
        for c_idx in range(2, total_subcolumns + 2):
            ws.column_dimensions[get_column_letter(c_idx)].width = 24
        ws.row_dimensions[1].height = 30
        for si in range(len(slots)):
            ws.row_dimensions[si + 2].height = 70
        _apply_page_setup(ws)

    if mode == "classroom":
        physical_rooms = set(s.get("classrooms", []))
        for room in get_classroom_export_labels(s.get("classrooms", []), placed):
            ws = wb.create_sheet(title=_sheet_name_for_export(room, used_sheet_names))
            writer = _write_virtual_classroom_sheet if room not in physical_rooms else _write_filtered_sheet
            writer(
                ws,
                filter_fn=lambda c, r=room: classroom_of(c) == r,
                include_room=False,
                include_targets=True,
            )

    elif mode == "group":
        for yr in sorted(s.get("years", {}).keys()):
            for br in s["years"][yr]:
                ws = wb.create_sheet(
                    title=_sheet_name_for_export(f"{yr} {br}", used_sheet_names))
                _write_filtered_sheet(
                    ws,
                    filter_fn=lambda c, y=yr, b=br: any(
                        t["year"] == y and t["branch"] == b
                        for t in c.get("targets", [])
                    ),
                    include_room=True,
                    include_targets=False,
                )

    elif mode == "lecturer":
        lecturers = list(s.get("lecturers", []))
        if not lecturers:
            lecturers = sorted({c.get("lecturer", "") for c in placed if c.get("lecturer", "")})
        for lecturer in lecturers:
            if not lecturer:
                continue
            ws = wb.create_sheet(
                title=_sheet_name_for_export(lecturer, used_sheet_names))
            _write_filtered_sheet(
                ws,
                filter_fn=lambda c, l=lecturer: c.get("lecturer", "") == l,
                include_room=True,
                include_targets=True,
            )

    else:
        # Show Everything matrix export (per year), preserving the current matrix layout.
        # ST-ARCH-003: remember what actually reached a sheet, so anything the
        # matrix has no column for can be reported rather than dropped.
        drawn = set()
        for yr in sorted(s["years"].keys()):
            branches = s["years"][yr]
            if not branches:
                continue
            n_branches = len(branches)
            ws = wb.create_sheet(title=_sheet_name_for_export(yr, used_sheet_names))

            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            c_sn = ws.cell(row=1, column=1, value="")
            c_sn.fill = corner_fill
            c_sn.border = cell_border

            ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
            c_time = ws.cell(row=1, column=2, value=tr("labels.time"))
            c_time.fill = corner_fill
            c_time.font = corner_font
            c_time.border = cell_border
            c_time.alignment = center_nowrap

            for d_idx, day in enumerate(days):
                col_start = 3 + d_idx * n_branches
                col_end = col_start + n_branches - 1
                if n_branches > 1:
                    ws.merge_cells(start_row=1, start_column=col_start,
                                   end_row=1, end_column=col_end)
                dc = ws.cell(row=1, column=col_start, value=tr(f"weekdays.{day}"))
                dc.fill = day_fill
                dc.font = day_font
                dc.border = cell_border
                dc.alignment = center_nowrap
                for mc in range(col_start, col_end + 1):
                    ws.cell(row=1, column=mc).fill = day_fill
                    ws.cell(row=1, column=mc).border = cell_border

            ws.cell(row=2, column=1).fill = corner_fill
            ws.cell(row=2, column=1).border = cell_border
            ws.cell(row=2, column=2).fill = corner_fill
            ws.cell(row=2, column=2).border = cell_border

            for d_idx in range(len(days)):
                for b_idx, br in enumerate(branches):
                    col = 3 + d_idx * n_branches + b_idx
                    bc = ws.cell(row=2, column=col, value=br)
                    bc.fill = branch_fill
                    bc.font = branch_font
                    bc.border = cell_border
                    bc.alignment = center_nowrap

            occupied = {}
            claims = {}
            for c in placed:
                c_day = effective_day(c)
                c_start = effective_time(c)
                if c_day not in days or c_start not in slots:
                    continue
                d_idx = days.index(c_day)
                start_si = slots.index(c_start)
                is_joint = c.get("joint_session", True)
                n_targets = len(c.get("targets", []))
                dur = c["duration"]
                for t in c["targets"]:
                    if t["year"] != yr or t["branch"] not in branches:
                        continue
                    b_idx = branches.index(t["branch"])
                    if not is_joint and n_targets > 1:
                        t_idx = c["targets"].index(t)
                        slot_off = t_idx * dur
                    else:
                        slot_off = 0
                    actual_start = start_si + slot_off
                    for d_off in range(dur):
                        si = actual_start + d_off
                        if si >= len(slots):
                            break
                        okey = (si, d_idx, b_idx)
                        # Identity, not equality, and not a plain append:
                        # this runs inside `for t in c["targets"]`, so a
                        # class carrying two IDENTICAL target dicts (a user
                        # typing "A, B, A" as a year's branches) claimed the
                        # same cell twice and was then found "overlapping"
                        # with ITSELF -- pulled out of occupied_start and
                        # stacked against its own duplicate, losing its
                        # merge and its year colour. `is not` because two
                        # distinct classes can compare equal by value.
                        bucket = claims.setdefault(okey, [])
                        if all(x is not c for x in bucket):
                            bucket.append(c)
                        drawn.add(cls_key(c))
                        if d_off == 0:
                            span = min(dur, len(slots) - actual_start)
                            occupied[okey] = ("start", c, span)
                        else:
                            occupied[okey] = ("span", None, 0)

            # ST-UI-001. The dict above keeps the LAST writer, so a second
            # lesson claiming a cell used to erase the first from this
            # sheet -- while the filtered sheets in the SAME workbook
            # already stacked both. Pull every contested claimant out and
            # stack them, exactly as _write_filtered_sheet does.
            overlapping = {id(c) for cs in claims.values() if len(cs) > 1
                           for c in cs}
            overlap_cells = {}
            for okey, cs in claims.items():
                keep = [c for c in cs if id(c) in overlapping]
                if keep:
                    overlap_cells[okey] = keep
            for okey in overlap_cells:
                occupied.pop(okey, None)

            for si, slot in enumerate(slots):
                erow = si + 3
                sn_cell = ws.cell(row=erow, column=1, value=si + 1)
                sn_cell.fill = session_fill
                sn_cell.font = session_font
                sn_cell.border = cell_border
                sn_cell.alignment = center_nowrap

                t_cell = ws.cell(row=erow, column=2, value=slot)
                t_cell.fill = time_fill
                t_cell.font = time_font
                t_cell.border = cell_border
                t_cell.alignment = center_nowrap

                for d_idx in range(len(days)):
                    for b_idx in range(n_branches):
                        col = 3 + d_idx * n_branches + b_idx
                        okey = (si, d_idx, b_idx)
                        if okey in overlap_cells:
                            claimants = overlap_cells[okey]
                            keys = {cls_key(x) for x in claimants}
                            conflicted = any(
                                conflict_partners.get(cls_key(x), frozenset())
                                & (keys - {cls_key(x)})
                                for x in claimants)
                            rich = _build_stacked_rich_cell(
                                claimants, include_room=True,
                                include_targets=False)
                            dc = ws.cell(row=erow, column=col, value=rich)
                            fill_hex = "FEE2E2" if conflicted else "F1F5F9"
                            dc.fill = PatternFill(
                                start_color=fill_hex, end_color=fill_hex,
                                fill_type="solid")
                            dc.border = cell_border
                            dc.alignment = center_wrap
                        elif okey in occupied:
                            kind, cls, span = occupied[okey]
                            if kind == "span":
                                ws.cell(row=erow, column=col).border = cell_border
                                continue
                            rich = _build_rich_cell(cls, include_room=True,
                                                   include_targets=False)
                            if span > 1:
                                ws.merge_cells(start_row=erow, start_column=col,
                                               end_row=erow + span - 1, end_column=col)
                                for mr in range(erow, erow + span):
                                    ws.cell(row=mr, column=col).border = cell_border
                            yr_hex = get_year_color(s, yr).lstrip("#")
                            light_hex = lighten_color(f"#{yr_hex}", 0.45).lstrip("#")
                            cf = PatternFill(start_color=light_hex.upper(),
                                             end_color=light_hex.upper(), fill_type="solid")
                            data_cell = ws.cell(row=erow, column=col, value=rich)
                            data_cell.fill = cf
                            data_cell.border = cell_border
                            data_cell.alignment = center_wrap
                        else:
                            ec = ws.cell(row=erow, column=col, value="")
                            ec.fill = empty_fill
                            ec.border = cell_border

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 12
            total_cols = 2 + n_branches * len(days)
            for c_idx in range(3, total_cols + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 22
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 22
            for si in range(len(slots)):
                ws.row_dimensions[si + 3].height = 80
            _apply_page_setup(ws)

        # ST-ARCH-003. The matrix draws one column per (year, branch) taken
        # from ``state["years"]``, so a placed lesson has nowhere to go when
        # its target year is no longer in that dict, or when it carries no
        # targets at all. It was silently absent from the workbook -- the
        # printout looked complete. Two failing suite cases proved it once the
        # engines were unified ("the group-less lesson was dropped", and
        # "workbook dropped on-grid lessons: ['Ders02', 'Ders04']" after a year
        # was removed); neither could fire before, because the suite was
        # exercising the *other*, unused writer.
        #
        # ST-FUNC-013's rule applies: an export never vanishes a placement, it
        # reports it. Same reasoning as the PDF appendix Phase 4 added.
        _write_unplaceable_sheet(wb, s, drawn, used_sheet_names, days, slots)

    if not wb.worksheets:
        ws = wb.create_sheet(
            title=_sheet_name_for_export(tr("labels.schedule"), used_sheet_names)
        )
        ws.cell(row=1, column=1, value=tr("warnings.no_schedule_data"))

    # ST-UI-008: the workbook is made to be emailed, so a cell openpyxl typed
    # as a formula must not stay one. Done in memory, on the cell attribute,
    # never by prefixing the string -- this app re-imports its own workbooks.
    neutralize_formula_cells(wb)
    wb.save(filepath)

# CSV export

def _export_csv(schedule: FinalSchedule, filepath: str):
    """Export to a flat CSV file.

    Scope note, measured in Phase 7: ``export_schedule(..., "csv", ...)`` has
    **no production caller**. The CSV a user gets comes from
    ``ui/app.py::export_csv``, a separate writer emitting a different product
    (a class list, not a timetable), wired to the menu at ``ui/app.py:1000``.
    ST-FUNC-006's user-facing pin therefore lives against *that* function; this
    one is fixed to match so the two writers cannot drift again, and so that
    wiring this entry point up later does not reintroduce the defect.
    """
    grid = schedule.build_grid()
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            tr("labels.day"),
            tr("labels.time"),
            tr("labels.class_code"),
            tr("labels.course"),
            tr("labels.lecturer"),
            tr("labels.classroom"),
            tr("labels.year"),
            tr("labels.branch"),
        ])
        for cls in schedule.placed_classes():
            room = classroom_of(cls)
            code = cls.get("class_code", "")
            cells = occupied_slots_of(schedule.state, cls)
            on_grid = bool(cells)
            if not on_grid:
                # Orphaned by a deleted hour. A flat file has room for it even
                # though a grid does not, so write it at its stored position
                # rather than dropping it (ST-DATA-003).
                cells = [(effective_day(cls), effective_time(cls))]
            targets = cls.get("targets", [])
            unreported = set(range(len(targets)))

            def _row(day_text, slot, t):
                writer.writerow([
                    day_text, csv_safe(slot), csv_safe(code),
                    csv_safe(cls["name"]), csv_safe(cls["lecturer"]),
                    csv_safe(room),
                    csv_safe(t["year"]) if t else "",
                    csv_safe(t["branch"]) if t else "",
                ])

            for offset, (day, slot) in enumerate(cells):
                # ST-FUNC-006: the header row is Turkish, so the day column
                # must be too. display_day rather than tr("weekdays.<key>") so
                # a day the grid no longer defines prints its stored value
                # instead of the lookup key.
                day_text = display_day(day)
                # One hour of a non-joint lesson belongs to ONE group, not to
                # all of them: this loop used to emit the cross product, so a
                # two-group lesson claimed each group met in both hours. Half
                # of those rows were false. target_for_slot_offset returns
                # None for a joint session, where every group really does
                # share every hour.
                owner = target_for_slot_offset(cls, offset) if on_grid else None
                if owner is not None:
                    indexed = [(owner, targets[owner])]
                else:
                    indexed = list(enumerate(targets))
                if not indexed:
                    _row(day_text, slot, None)
                    continue
                for idx, t in indexed:
                    unreported.discard(idx)
                    _row(day_text, slot, t)

            # A group whose own hour ran off the end of the grid has no cell to
            # be named in. Same rule as above: reported at the stored start
            # rather than dropped (ST-DATA-003).
            for idx in sorted(unreported):
                _row(display_day(effective_day(cls)), effective_time(cls),
                     targets[idx])


# ── PDF export (optional) ───────────────────────────────────────────────────

# ST-FUNC-004. Helvetica is a base-14 Type1 font limited to WinAnsi, so six of
# the twelve Turkish letters -- ğ Ğ ş Ş ı İ -- have no codepoint in it. (ö ü ç
# Ö Ü Ç do, and always drew correctly; the register's "every Turkish-specific
# letter" was wrong about that.) What reportlab does with the other six is
# worse than the missing-glyph box the register described: it splits the
# paragraph at each unmappable codepoint and switches to ZapfDingbats, whose
# ASCII `n` is a filled block. So "Şükrü Işık Öğretmen" printed as a name with
# solid blobs in it, and -- because the substitution also rewrites the text
# layer -- Ctrl-F for "Öğretmen" found nothing and copy-paste yielded
# "Önretmen". A school archiving its timetables could not search them by
# teacher name.
#
# The fix embeds a TrueType font instead. Nothing is bundled: reportlab ships
# Bitstream Vera inside its own wheel (283 glyphs, missing none of the twelve),
# under a permissive licence it already redistributes, and build_nuitka.bat:118
# already carries --include-package-data=reportlab. Installer delta: 0 bytes.
# Per-document delta: +40 KB for the embedded subset, measured.
_PDF_FALLBACK_FONTS = ("Helvetica", "Helvetica-Bold")
_PDF_UNICODE_FONTS = ("DersisSans", "DersisSans-Bold")

# pdfmetrics keeps a module-global registry, so registration is once per
# process rather than once per canvas -- measured: a second export in the same
# process is byte-identical without re-registering. None means "not yet tried".
_pdf_font_names: "tuple[str, str] | None" = None


def _register_unicode_fonts() -> "tuple[str, str]":
    """Register a Turkish-capable TrueType family; return ``(regular, bold)``.

    Degrades to Helvetica rather than raising. That guard is not theoretical:
    ``requirements-lock.txt`` pins ``reportlab==4.4.10`` while the audit venv
    actually runs 5.0.1, and ``Dersis-mac.spec`` does not collect reportlab's
    package data at all -- so a build where ``fonts/Vera.ttf`` is absent is a
    real possibility, and it must cost the user unreadable letters, not a
    failed export.
    """
    global _pdf_font_names
    if _pdf_font_names is not None:
        return _pdf_font_names

    _pdf_font_names = _PDF_FALLBACK_FONTS
    try:
        import reportlab
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return _pdf_font_names

    font_dir = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
    regular_path = os.path.join(font_dir, "Vera.ttf")
    bold_path = os.path.join(font_dir, "VeraBd.ttf")
    if not (os.path.exists(regular_path) and os.path.exists(bold_path)):
        return _pdf_font_names

    regular, bold = _PDF_UNICODE_FONTS
    try:
        pdfmetrics.registerFont(TTFont(regular, regular_path))
        pdfmetrics.registerFont(TTFont(bold, bold_path))
        # The cell markup uses <b>, which reportlab resolves through the family
        # map -- without this the bold runs fall back to Helvetica-Bold and the
        # six letters break again inside every class name.
        pdfmetrics.registerFontFamily(
            regular, normal=regular, bold=bold,
            italic=regular, boldItalic=bold)
    except Exception:
        return _pdf_font_names

    _pdf_font_names = (regular, bold)
    return _pdf_font_names


# ── Which letters the PDF can actually draw ─────────────────────────────────
#
# ST-FUNC-004 was closed on twelve Turkish letters. The product ships 22
# languages, and the bundled face is Vera: 283 glyphs, Latin-1 and no more.
# Measured against the weekday names of every shipped locale, Vera cannot draw
# a single character of ru (22), ar (18), fa (17), hi (17), zh (9), ja (8),
# ko (8), pl (2) or az (1) -- 9 of the 22, all offered by the first-run
# language gate. reportlab ships only Vera and DarkGarden, so there is no
# bundled face to swap in.
#
# THAT 9-OF-22 IS A FACT ABOUT VERA, NOT ABOUT THE SHIPPED APP. Say which one
# you mean; earlier notes did not, and "9 of 22 locales cannot print their
# weekday names" was carried forward as a product claim. Driving
# _resolve_pdf_fonts per locale on a stock Windows install (measured
# 2026-08-29, weekday names plus the 10 UI keys _pdf_document_text collects)
# recovers 6 of the 9 from host faces with an EMPTY unprintable set:
#
#     az, pl, ru -> arial.ttf      ja -> msgothic.ttc
#     zh         -> msyh.ttc       ko -> malgun.ttf
#
# The live figure is 3 of 22 -- ar, fa and hi -- and the cause for those three
# is shaping, not glyph coverage (see _needs_text_shaping below).
#
# The residual, stated honestly: those 6 depend on a face the HOST happens to
# have. A stripped Windows image, a bare Linux container or a locked-down
# fonts directory drops them back to the note page. That is why the tests in
# tests/test_pdf_locale_coverage.py assert the property ("we never silently
# drop a character we could have drawn") and never the host outcome -- "az
# resolves to arial" is true here and false on the CI runner.
#
# Bundling DejaVu is NOT the answer, and the reason is structural rather than
# a coverage table: it would cover ru/pl/az, which _resolve_pdf_fonts already
# recovers from arial on this host, and it cannot touch ar/fa/hi at all,
# because the shaped-script short-circuit below means no substitute face --
# bundled or host -- is ever tried for them. See the note at _resolve_pdf_fonts.
#
# Two things follow, and this section does both:
#
#   * where the host has a face that covers the document, use it. Measured on
#     Windows: arial.ttf covers ru, pl, az, ar and fa; msgothic.ttc covers ja
#     and zh; malgun.ttf covers ko.
#   * where it does not, SAY SO on the page. A row of empty boxes with no
#     explanation is the same silent-failure shape the appendix above exists
#     to close.
#
# The one thing not done here is quietly drawing a script the layout engine
# cannot lay out -- see _needs_text_shaping.

# Scripts reportlab cannot lay out correctly AS THIS PROJECT INSTALLS IT.
# Measured, not assumed: registering arial.ttf and drawing "العربية" emits
# `(\001\002\003\004\005\006\007) Tj` -- the seven codepoints in LOGICAL order,
# each as its isolated form. For these scripts a covering font buys
# confidently wrong output (a word spelled backwards in disconnected letters)
# in place of an honest box. The note is the truthful answer there.
#
# READ THE NEXT PARAGRAPH BEFORE RE-DERIVING ANYTHING FROM THAT PROBE. An
# earlier version of this comment concluded "reportlab has no bidi and no
# shaping engine". That is false, and the probe above cannot tell the
# difference. reportlab 5.0.1 declares BOTH as optional extras --
# reportlab-5.0.1.dist-info/METADATA:29-32, `Provides-Extra: bidi` -> rlbidi
# and `Provides-Extra: shaping` -> uharfbuzz -- and the code paths are live,
# not vestigial: pdfgen/textobject.py:21-23 binds rlbidi.log2vis,
# pdfbase/ttfonts.py:1344-1360 builds a real HarfBuzz face from the embedded
# TTF, fonts/hb-test.ttf ships as the fixture, and ParagraphStyle carries a
# `shaping` attribute (lib/styles.py:133) that platypus/paragraph.py:2102
# consumes. Neither extra is installed here, and TTFont.shapable is
# `bool(self._shapable and uharfbuzz)`, so with them absent reportlab quietly
# behaves exactly like a library that cannot shape.
#
# So the accurate statement is: THE CAPABILITY EXISTS AND IS GATED ON AN
# OPTIONAL DEPENDENCY THIS PROJECT DOES NOT INSTALL. The decision to leave
# ar/fa/hi on the note page may still be the right one -- on cost, and on the
# fact that requirements-lock.txt pins reportlab 4.4.10 into the shipped
# installer, which is not the version measured above -- but it is not
# "reportlab cannot". Anyone taking this up must verify the extras against
# 4.4.10 first, and must assert the emitted TEXT LAYER (visual order, joined
# forms, and a Ctrl-F round-trip), because ST-FUNC-004's failure mode was
# precisely a page that looked plausible over a falsified text layer.
_SHAPED_SCRIPT_RANGES = (
    (0x0590, 0x05FF),  # Hebrew (bidi)
    (0x0600, 0x06FF),  # Arabic
    (0x0700, 0x074F),  # Syriac
    (0x0750, 0x077F),  # Arabic Supplement
    (0x0780, 0x07BF),  # Thaana
    (0x07C0, 0x08FF),  # NKo .. Arabic Extended-A
    (0x0900, 0x0DFF),  # Devanagari .. Sinhala (reordering matras)
    (0x0E00, 0x0EFF),  # Thai, Lao
    (0x0F00, 0x109F),  # Tibetan, Myanmar
    (0x1780, 0x17FF),  # Khmer
    (0x1800, 0x18AF),  # Mongolian
    (0xFB1D, 0xFDFF),  # Hebrew/Arabic presentation forms
    (0xFE70, 0xFEFF),  # Arabic presentation forms-B
)


def _needs_text_shaping(ch: str) -> bool:
    """True if *ch* belongs to a script reportlab cannot lay out."""
    o = ord(ch)
    return any(lo <= o <= hi for lo, hi in _SHAPED_SCRIPT_RANGES)


def _chars_without_glyphs(font_name: str, text: str) -> "set[str]":
    """The characters of *text* that *font_name* has no glyph for.

    Whitespace is excluded: reportlab never draws it as a glyph.
    """
    wanted = {ch for ch in text if not ch.isspace()}
    if not wanted:
        return set()
    try:
        from reportlab.pdfbase import pdfmetrics
        face = pdfmetrics.getFont(font_name).face
    except Exception:
        return set()
    char_to_glyph = getattr(face, "charToGlyph", None)
    if char_to_glyph is None:
        # A base-14 Type1 face. reportlab encodes those as WinAnsi and answers
        # anything outside it by switching to ZapfDingbats mid-word, which is
        # the pre-ST-FUNC-004 failure: a filled blob on the page and a
        # falsified text layer behind it.
        missing = set()
        for ch in wanted:
            try:
                ch.encode("cp1252")
            except UnicodeEncodeError:
                missing.add(ch)
        return missing
    return {ch for ch in wanted if ord(ch) not in char_to_glyph}


# (regular, bold, subfont index). The bold twin may be the same file: a family
# without one loses <b> weight, which is a far smaller loss than a box.
_PDF_SYSTEM_FONT_CANDIDATES = (
    ("arial.ttf", "arialbd.ttf", 0),        # Windows: Latin, Greek, Cyrillic
    ("segoeui.ttf", "segoeuib.ttf", 0),
    ("tahoma.ttf", "tahomabd.ttf", 0),
    ("DejaVuSans.ttf", "DejaVuSans-Bold.ttf", 0),   # Linux
    ("NotoSans-Regular.ttf", "NotoSans-Bold.ttf", 0),
    ("Arial Unicode.ttf", "Arial Unicode.ttf", 0),  # macOS
    ("Arial.ttf", "Arial Bold.ttf", 0),
    ("msgothic.ttc", "msgothic.ttc", 0),    # Windows: Japanese, Chinese
    ("malgun.ttf", "malgunbd.ttf", 0),      # Windows: Korean
    ("msyh.ttc", "msyh.ttc", 0),            # Windows: Simplified Chinese
)


def _system_font_dirs() -> "list[str]":
    """Directories to look for a covering face in, most likely first."""
    dirs = []
    windir = os.environ.get("WINDIR")
    if windir:
        dirs.append(os.path.join(windir, "Fonts"))
    home = os.path.expanduser("~")
    dirs.extend([
        "/Library/Fonts", "/System/Library/Fonts",
        os.path.join(home, "Library", "Fonts"),
        "/usr/share/fonts/truetype/dejavu",
        "/usr/share/fonts/truetype/noto",
        "/usr/share/fonts/TTF",
        "/usr/share/fonts",
    ])
    return [d for d in dirs if os.path.isdir(d)]


# path -> registered (regular, bold) names, or None when the file cannot be
# read. Registration is a pdfmetrics-global side effect, so it is done once per
# process exactly like _register_unicode_fonts.
_pdf_system_fonts: dict = {}


def _register_covering_font(text: str) -> "tuple[str, str] | None":
    """Register the first host face that can draw *every* character of *text*.

    "Every character" and not "the missing ones": swapping in a face that
    draws Cyrillic but not ğ would trade one locale's boxes for another's.
    Measured on Windows, that is not hypothetical -- malgun.ttf covers Korean
    and has no glyph for ı.
    """
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return None

    dirs = _system_font_dirs()
    for regular_file, bold_file, subfont in _PDF_SYSTEM_FONT_CANDIDATES:
        for d in dirs:
            path = os.path.join(d, regular_file)
            if not os.path.exists(path):
                continue
            cached = _pdf_system_fonts.get(path, "unread")
            if cached is None:
                break  # known unreadable; try the next candidate family
            if cached == "unread":
                names = ("DersisSys-" + regular_file,
                         "DersisSys-" + regular_file + "-Bold")
                bold_path = os.path.join(d, bold_file)
                if not os.path.exists(bold_path):
                    bold_path = path
                try:
                    pdfmetrics.registerFont(
                        TTFont(names[0], path, subfontIndex=subfont))
                    pdfmetrics.registerFont(
                        TTFont(names[1], bold_path, subfontIndex=subfont))
                    pdfmetrics.registerFontFamily(
                        names[0], normal=names[0], bold=names[1],
                        italic=names[0], boldItalic=names[1])
                except Exception:
                    _pdf_system_fonts[path] = None
                    break
                _pdf_system_fonts[path] = names
                cached = names
            if not (_chars_without_glyphs(cached[0], text)
                    or _chars_without_glyphs(cached[1], text)):
                return cached
            break  # this family is present but does not cover the document
    return None


def _resolve_pdf_fonts(text: str) -> "tuple[str, str, set[str]]":
    """Return ``(regular, bold, unprintable)`` for a document containing *text*.

    The shaping guard below is a SHORT-CIRCUIT, and it is the reason adding a
    font never helps Arabic, Persian or Hindi. If any missing character belongs
    to a script the layout engine cannot lay out, the covering-font search is
    skipped entirely -- so no substitute face, bundled or host, is ever tried
    for those documents. Bundling DejaVu, or adding an entry to
    _PDF_SYSTEM_FONT_CANDIDATES, changes nothing for ar/fa/hi BY CONSTRUCTION,
    independently of what the added face covers. Deleting this guard is not a
    font change; it is a decision to emit logical-order isolated forms, which
    is the confidently-wrong output the note page exists to avoid.
    """
    regular, bold = _register_unicode_fonts()
    missing = (_chars_without_glyphs(regular, text)
               | _chars_without_glyphs(bold, text))
    if not missing:
        return regular, bold, set()
    if not any(_needs_text_shaping(ch) for ch in missing):
        covering = _register_covering_font(text)
        if covering is not None:
            return covering[0], covering[1], set()
    return regular, bold, missing


def _pdf_document_text(schedule) -> str:
    """Every string the PDF may draw, concatenated.

    Over-collecting is safe (a character listed but never drawn only widens
    the search for a covering face); under-collecting is not, which is why
    this reads the state rather than the built flowables.
    """
    state = schedule.state
    parts = [tr(k) for k in (
        "labels.time", "labels.type", "labels.class_code", "labels.class_name",
        "labels.lecturer", "labels.day", "export.appendix_title",
        "export.appendix_offgrid", "export.appendix_conflict",
        "warnings.no_schedule_data",
    )]
    parts.extend(display_day(d) for d in state.get("days", []))
    parts.extend(str(s) for s in state.get("slots", []))
    parts.extend(str(r) for r in state.get("classrooms", []))
    parts.extend(str(lect) for lect in state.get("lecturers", []))
    for year, branches in (state.get("years") or {}).items():
        parts.append(str(year))
        parts.extend(str(b) for b in branches or [])
    for cls in state.get("classes", []):
        parts.extend(str(cls.get(k, "")) for k in
                     ("class_code", "name", "lecturer"))
        parts.append(str(classroom_of(cls) or ""))
        for t in cls.get("targets", []) or []:
            parts.append(f'{t.get("year", "")}/{t.get("branch", "")}')
        _emoji, label, _color = get_badge(cls)
        parts.append(str(label or ""))
    return "".join(parts)


def _unprintable_note(font_regular: str, unprintable: "set[str]") -> str:
    """The page text explaining which characters the document could not spell.

    Named by codepoint, not by the character: the characters in question are
    precisely the ones this document cannot draw. The wording falls back to
    English when the chosen face cannot draw the localized string either --
    a note that is itself a row of boxes explains nothing.
    """
    from scheduler_app.translations import TRANSLATIONS

    codes = " ".join(f"U+{ord(ch):04X}" for ch in sorted(unprintable))
    text = tr("export.unprintable_note", count=len(unprintable),
              codepoints=codes)
    if _chars_without_glyphs(font_regular, text):
        english = TRANSLATIONS["en"].get("export.unprintable_note", "")
        try:
            text = english.format(count=len(unprintable), codepoints=codes)
        except (KeyError, IndexError, ValueError):
            text = english
    return text


def _pdf_rich_markup(cls, include_room=False, include_targets=False):
    """The colour-coded markup for one class, without wrapping it in a Paragraph.

    Split out from :func:`_pdf_rich_paragraph` so a contested cell can join
    several classes' markup into ONE Paragraph (ST-UI-001). reportlab's ``SPAN``
    merges rows, not columns, so a PDF cell cannot be split into lanes the way
    the screen is — stacking every claimant into the cell is the shape the XLSX
    writer already uses, and it is what makes the two agree.
    """
    # reportlab parses this as markup, so every interpolated value is USER
    # TEXT and must be escaped: a class named "Fizik & Kimya" or a lecturer
    # called "<Vekil>" otherwise mangles the cell or raises out of the whole
    # export. Only the data is escaped, never the template.
    esc = html.escape
    parts = []
    code = cls.get("class_code", "")
    if code:
        parts.append(
            f'<font color="{CELL_FG_CODE}" size="7"><b>{esc(str(code))}</b></font>')
    parts.append(
        f'<font color="{CELL_FG_NAME}" size="8"><b>{esc(str(cls["name"]))}</b></font>')
    if cls.get("lecturer"):
        parts.append(
            f'<font color="{CELL_FG_LECTURER}" size="7">{esc(str(cls["lecturer"]))}</font>')
    if include_room:
        room = classroom_of(cls)
        if room:
            parts.append(
                f'<font color="{CELL_FG_ROOM}" size="7">{esc(str(room))}</font>')
    if include_targets:
        groups = ", ".join(
            f"{t['year']}/{t['branch']}" for t in cls.get("targets", []))
        if groups:
            parts.append(
                f'<font color="{CELL_FG_BRANCH}" size="6">{esc(groups)}</font>')
    # Protection badge
    emoji, label, b_color = get_badge(cls)
    if emoji:
        parts.append(
            f'<font color="{b_color}" size="6"><b>{esc(str(label))}</b></font>')
    return "<br/>".join(parts)


def _pdf_rich_paragraph(cls, cell_style, include_room=False, include_targets=False):
    """Build a color-coded Paragraph for a class in a PDF cell."""
    from reportlab.platypus import Paragraph

    return Paragraph(
        _pdf_rich_markup(cls, include_room=include_room,
                         include_targets=include_targets),
        cell_style)


def _note_cell_height(tall_rows, para, content_w, first_row, span, min_row_h):
    """Grow *tall_rows* so ``para`` fits the rows it is drawn across.

    ``rowHeights`` is fixed, and reportlab does not grow a fixed row to fit its
    content -- it overprints the neighbours. The contested-cell branch already
    measured itself for that reason; the ordinary occupied cell did not, and
    measurement says it should have: at 7pt in the everything layout's narrow
    columns, "Öğrenci Değerlendirme ve Ölçme Çalıştayı / Şükrü Işık Öğretmen /
    A-101" needs 51pt against MIN_ROW_H's 50 under Helvetica and 60pt under the
    embedded TrueType face (ST-FUNC-004 widens the glyphs). The lecturer layout
    needs 51pt for even a one-word lesson, because it prints the group line
    too. So the printed timetable has been overprinting the hour above and
    below it all along, and the font fix would have made that worse silently.

    A cell merged across *span* rows is drawn in the sum of their heights, so
    the requirement is divided rather than applied to each row.
    """
    span = max(span, 1)
    _w, _h = para.wrap(content_w, 1e6)
    per_row = (_h + 6) / span
    for off in range(span):
        tall_rows[first_row + off] = max(
            tall_rows.get(first_row + off, min_row_h), per_row)


def _pdf_conflict_paragraph(classes, cell_style, conflicted,
                            include_room=False, include_targets=False):
    """One Paragraph holding every class that claims a contested PDF cell.

    ST-UI-001. ``_build_filtered_table`` used to ``continue`` past any class
    whose start cell was already taken — keeping the FIRST claimant, where the
    screen kept the LAST. So a user who checked the timetable on screen and
    then printed it got two different, both-incomplete documents. Every
    claimant is now stacked into the cell, and *conflicted* (a validator
    verdict, not "there are two of them here") adds the ÇAKIŞMA marker.
    """
    from reportlab.platypus import Paragraph

    blocks = [_pdf_rich_markup(c, include_room=include_room,
                               include_targets=include_targets)
              for c in classes]
    markup = '<br/><font color="#DC2626">---</font><br/>'.join(blocks)
    if conflicted:
        markup = (f'<font color="#DC2626" size="7"><b>'
                  f'{html.escape(tr("badges.conflict"))}</b></font><br/>'
                  + markup)
    return Paragraph(markup, cell_style)


def _export_pdf(schedule: FinalSchedule, filepath: str, mode: str = "everything"):
    """Export to PDF using reportlab, matching the timetable layout exactly.

    Supports four modes mirroring the Excel export:
    - "everything": one page per year with branch sub-columns and merged cells
    - "classroom":  one page per classroom
    - "group":      one page per year/branch combination
    - "lecturer":   one page per lecturer
    """
    try:
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak,
        )
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise RuntimeError(tr("errors.reportlab_required"))

    state = schedule.state
    placed = schedule.placed_classes()
    days = schedule.days
    slots = schedule.slots
    # ST-FUNC-013: every table builder records the lessons it actually put on a
    # page here, so the appendix can name the ones no page had room for. Same
    # mechanism the workbook's _write_unplaceable_sheet already uses, and it is
    # a report on the built document rather than a second guess at the filter
    # rules -- so a lesson lost for a reason nobody predicted still surfaces.
    drawn: set = set()
    # ST-UI-001: one scan for the whole document, so every page marks the
    # same clashes and the printout agrees with the screen.
    conflicts = find_schedule_conflicts(state)
    conflict_partners = conflict_partner_index(conflicts)

    page_size = landscape(A3)
    doc = SimpleDocTemplate(
        filepath, pagesize=page_size,
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=8 * mm, bottomMargin=8 * mm,
    )
    avail_w = page_size[0] - 16 * mm

    # ── Paragraph styles ──────────────────────────────────────────────
    # ST-FUNC-004: every style names the registered family, never Helvetica.
    # A style left on Helvetica would still substitute ZapfDingbats for ğ ş ı,
    # and the day-header row is exactly where "Çarşamba" lives.
    #
    # ...and no style can name a face that cannot draw this document's own
    # alphabet. The bundled Vera covers Latin-1, which is 13 of the 22 shipped
    # locales; for the other 9 the page was a row of empty boxes and nothing
    # said so. _resolve_pdf_fonts upgrades to a host face that covers the
    # document where one exists, and reports what is left.
    #
    # What that leaves in practice, measured on a stock Windows install: 6 of
    # those 9 (az/pl/ru/ja/zh/ko) come back with a covering host face and an
    # empty `unprintable_chars`, so 3 locales reach the note page, not 9 --
    # ar, fa and hi, blocked by shaping rather than by coverage. On a host
    # without those faces the number rises again, which is the point of
    # reporting `unprintable_chars` instead of assuming either figure.
    FONT_REGULAR, FONT_BOLD, unprintable_chars = _resolve_pdf_fonts(
        _pdf_document_text(schedule))
    cell_style = ParagraphStyle(
        "CellContent", fontSize=7, leading=9,
        alignment=TA_CENTER, fontName=FONT_REGULAR,
    )
    hdr_style = ParagraphStyle(
        "HdrContent", fontSize=9, leading=11, alignment=TA_CENTER,
        textColor=rl_colors.white, fontName=FONT_BOLD,
    )
    branch_hdr_style = ParagraphStyle(
        "BranchHdr", fontSize=8, leading=10, alignment=TA_CENTER,
        textColor=rl_colors.white, fontName=FONT_BOLD,
    )
    time_style = ParagraphStyle(
        "TimeContent", fontSize=8, leading=10, alignment=TA_CENTER,
        textColor=rl_colors.white, fontName=FONT_BOLD,
    )
    session_style = ParagraphStyle(
        "SessionNum", fontSize=8, leading=10, alignment=TA_CENTER,
        textColor=rl_colors.HexColor("#333333"), fontName=FONT_BOLD,
    )
    title_style = ParagraphStyle(
        "PageTitle", fontSize=11, leading=14, alignment=TA_CENTER,
        fontName=FONT_BOLD, textColor=rl_colors.HexColor("#1E293B"),
        spaceAfter=4 * mm,
    )

    # ── Color constants matching the app UI ───────────────────────────
    COL_DAY_HDR = rl_colors.HexColor("#334155")
    COL_BRANCH_HDR = rl_colors.HexColor("#475569")

    COL_CORNER = rl_colors.HexColor("#94A3B8")
    COL_TIME = rl_colors.HexColor("#475569")
    COL_SESSION = rl_colors.HexColor("#F1F5F9")
    COL_EMPTY = rl_colors.HexColor("#F8FAFC")
    COL_GRID = rl_colors.HexColor("#CBD5E1")

    MIN_ROW_H = 50  # minimum data-row height in points

    # ── Helper: build a filtered-sheet table (classroom / group / lecturer) ──

    def _build_filtered_table(filter_fn, include_room=False,
                              include_targets=False):
        """Build a reportlab Table for a simple Day×Slot grid."""
        n_days = len(days)
        time_w = 18 * mm
        day_w = (avail_w - time_w) / max(n_days, 1)
        col_widths = [time_w] + [day_w] * n_days

        # Header row
        header = [Paragraph(tr("labels.time"), hdr_style)]
        for day in days:
            header.append(Paragraph(tr(f"weekdays.{day}"), hdr_style))

        # Build occupancy map: (slot_index, day) -> (cls, span) or "covered"
        #
        # ST-UI-001. This used to `continue` past any class whose start cell
        # was already claimed, keeping the FIRST claimant — while the screen's
        # dict-overwrite kept the LAST. A user who checked the timetable on
        # screen and then printed it therefore got two different, both-
        # incomplete documents, each missing a lesson the other showed.
        #
        # reportlab's SPAN merges rows, not columns, so a PDF cell cannot be
        # split into lanes the way the grid is. Instead every claimant is
        # stacked into each covered cell and the merge is dropped — which is
        # exactly what ui/app.py::_write_filtered_sheet already does for XLSX,
        # so this makes the three surfaces agree rather than inventing a fourth
        # behaviour.
        entries = []
        claims = {}
        for cls in placed:
            if not filter_fn(cls):
                continue
            c_day = effective_day(cls)
            c_start = effective_time(cls)
            if c_day not in days or c_start not in slots:
                continue
            start_si = slots.index(c_start)
            span = min(total_duration(cls), len(slots) - start_si)
            if span <= 0:
                continue
            entry = {"cls": cls, "day": c_day, "start_si": start_si,
                     "span": span}
            entries.append(entry)
            for off in range(span):
                claims.setdefault((start_si + off, c_day), []).append(entry)

        drawn.update(cls_key(e["cls"]) for e in entries)

        overlapping = {id(e["cls"]) for cells in claims.values()
                       if len(cells) > 1 for e in cells}

        occupied_start = {}
        covered = set()
        overlap_cells = {}
        for entry in entries:
            if id(entry["cls"]) in overlapping:
                for off in range(entry["span"]):
                    overlap_cells.setdefault(
                        (entry["start_si"] + off, entry["day"]), []).append(entry)
                continue
            occupied_start[(entry["start_si"], entry["day"])] = (
                entry["cls"], entry["span"])
            for off in range(entry["span"]):
                covered.add((entry["start_si"] + off, entry["day"]))

        table_data = [header]
        style_cmds = [
            # Header row styling
            ("BACKGROUND", (0, 0), (-1, 0), COL_DAY_HDR),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, COL_GRID),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]
        cell_bg_cmds = []
        row_heights = [24]  # header row
        # Rows holding a stacked contested cell need more than MIN_ROW_H:
        # reportlab does NOT grow a fixed-height row to fit its content, it
        # overprints the neighbours. Measured per cell below.
        tall_rows = {}

        for si, slot in enumerate(slots):
            data_row_idx = si + 1
            row = [Paragraph(escape_pdf_markup(slot), time_style)]
            style_cmds.append(
                ("BACKGROUND", (0, data_row_idx), (0, data_row_idx), COL_TIME))

            for di, day in enumerate(days):
                col_idx = di + 1
                key = (si, day)
                if key in occupied_start:
                    cls, span = occupied_start[key]
                    para = _pdf_rich_paragraph(
                        cls, cell_style,
                        include_room=include_room,
                        include_targets=include_targets)
                    row.append(para)
                    _note_cell_height(tall_rows, para, day_w - 4,
                                      data_row_idx, span, MIN_ROW_H)
                    yr_name = cls["targets"][0]["year"] if cls.get("targets") else ""
                    base = get_year_color(state, yr_name)
                    light = lighten_color(base, 0.45)
                    cell_bg_cmds.append(
                        ("BACKGROUND", (col_idx, data_row_idx),
                         (col_idx, data_row_idx), rl_colors.HexColor(light)))
                    if span > 1:
                        style_cmds.append(
                            ("SPAN", (col_idx, data_row_idx),
                             (col_idx, data_row_idx + span - 1)))
                        for off in range(span):
                            cell_bg_cmds.append(
                                ("BACKGROUND",
                                 (col_idx, data_row_idx + off),
                                 (col_idx, data_row_idx + off),
                                 rl_colors.HexColor(light)))
                elif key in overlap_cells:
                    claimants = [e["cls"] for e in overlap_cells[key]]
                    keys = {cls_key(c) for c in claimants}
                    conflicted = any(
                        conflict_partners.get(cls_key(c), frozenset())
                        & (keys - {cls_key(c)})
                        for c in claimants)
                    stack_para = _pdf_conflict_paragraph(
                        claimants, cell_style, conflicted,
                        include_room=include_room,
                        include_targets=include_targets)
                    row.append(stack_para)
                    # Two or more lessons in one cell is taller than one, and
                    # rowHeights is FIXED -- an unmeasured stack silently
                    # overprints the hours above and below it, which on a
                    # printed timetable is worse than the drop it replaced.
                    # day_w - 4 because LEFTPADDING and RIGHTPADDING are both
                    # 2; + 6 for TOPPADDING + BOTTOMPADDING.
                    _w, _h = stack_para.wrap(day_w - 4, 1e6)
                    tall_rows[data_row_idx] = max(
                        tall_rows.get(data_row_idx, MIN_ROW_H), _h + 6)
                    cell_bg_cmds.append(
                        ("BACKGROUND", (col_idx, data_row_idx),
                         (col_idx, data_row_idx),
                         rl_colors.HexColor(
                             "#FEE2E2" if conflicted else "#F1F5F9")))
                elif key in covered:
                    row.append("")
                else:
                    row.append("")
                    cell_bg_cmds.append(
                        ("BACKGROUND", (col_idx, data_row_idx),
                         (col_idx, data_row_idx), COL_EMPTY))
            table_data.append(row)
            row_heights.append(MIN_ROW_H)

        for _ri, _h in tall_rows.items():
            row_heights[_ri] = _h

        # Default empty bg first, then override with lesson colors
        style_cmds.extend(cell_bg_cmds)

        tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights,
                    repeatRows=1)
        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    def _build_virtual_room_table(room):
        """Build a virtual-classroom table with fixed day subcolumns."""
        layout = build_virtual_classroom_day_layout(
            state, lambda cls, r=room: classroom_of(cls) == r)
        total_subcolumns = layout["total_subcolumns"]
        time_w = 18 * mm
        data_w = (avail_w - time_w) / max(total_subcolumns, 1)
        col_widths = [time_w] + [data_w] * total_subcolumns

        header = [Paragraph(tr("labels.time"), hdr_style)]
        for group in layout["day_groups"]:
            header.append(Paragraph(tr(f"weekdays.{group['day']}"), hdr_style))
            header.extend([""] * (group["lane_count"] - 1))

        starts = {
            (block["row"], block["subcolumn"]): block
            for block in layout["blocks"]
        }
        drawn.update(cls_key(block["cls"]) for block in layout["blocks"])
        covered = set(layout["occupied_subcolumns"])

        table_data = [header]
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), COL_DAY_HDR),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, COL_GRID),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]
        for group in layout["day_groups"]:
            col_start = 1 + group["subcolumn_start"]
            col_end = col_start + group["lane_count"] - 1
            if group["lane_count"] > 1:
                style_cmds.append(("SPAN", (col_start, 0), (col_end, 0)))

        cell_bg_cmds = []
        row_heights = [24]
        tall_rows = {}
        for si, slot in enumerate(slots):
            data_row_idx = si + 1
            row = [Paragraph(escape_pdf_markup(slot), time_style)]
            style_cmds.append(
                ("BACKGROUND", (0, data_row_idx), (0, data_row_idx), COL_TIME))

            for subcolumn in range(total_subcolumns):
                col_idx = subcolumn + 1
                key = (si, subcolumn)
                if key in starts:
                    block = starts[key]
                    cls = block["cls"]
                    span = block["span"]
                    para = _pdf_rich_paragraph(
                        cls, cell_style,
                        include_room=False, include_targets=True)
                    row.append(para)
                    _note_cell_height(tall_rows, para, data_w - 4,
                                      data_row_idx, span, MIN_ROW_H)
                    yr_name = cls["targets"][0]["year"] if cls.get("targets") else ""
                    base = get_year_color(state, yr_name)
                    light = lighten_color(base, 0.45)
                    cell_bg_cmds.append(
                        ("BACKGROUND", (col_idx, data_row_idx),
                         (col_idx, data_row_idx), rl_colors.HexColor(light)))
                    if span > 1:
                        style_cmds.append(
                            ("SPAN", (col_idx, data_row_idx),
                             (col_idx, data_row_idx + span - 1)))
                        for off in range(span):
                            cell_bg_cmds.append(
                                ("BACKGROUND",
                                 (col_idx, data_row_idx + off),
                                 (col_idx, data_row_idx + off),
                                 rl_colors.HexColor(light)))
                elif key in covered:
                    row.append("")
                else:
                    row.append("")
                    cell_bg_cmds.append(
                        ("BACKGROUND", (col_idx, data_row_idx),
                         (col_idx, data_row_idx), COL_EMPTY))
            table_data.append(row)
            row_heights.append(MIN_ROW_H)

        for _ri, _h in tall_rows.items():
            row_heights[_ri] = _h

        style_cmds.extend(cell_bg_cmds)
        tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights,
                    repeatRows=1)
        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    # ── Helper: build "everything" table for one year ─────────────────

    def _build_everything_table(yr):
        """Build the matrix-layout table for a single year (like the UI)."""
        branches = state["years"][yr]
        n_branches = len(branches)
        n_days = len(days)

        # Column widths: session# | time | (branches × days)
        session_w = 8 * mm
        time_w = 16 * mm
        data_cols = n_days * n_branches
        data_col_w = (avail_w - session_w - time_w) / max(data_cols, 1)
        col_widths = [session_w, time_w] + [data_col_w] * data_cols

        # Row 0: corner (merged 2 rows) + day headers (merged across branches)
        row0 = ["", Paragraph(tr("labels.time"), hdr_style)]
        for day in days:
            row0.append(Paragraph(tr(f"weekdays.{day}"), hdr_style))
            row0.extend([""] * (n_branches - 1))

        # Row 1: empty corners + branch sub-headers
        row1 = ["", ""]
        for _d in range(n_days):
            for br in branches:
                row1.append(
                    Paragraph(escape_pdf_markup(br), branch_hdr_style))

        table_data = [row0, row1]

        style_cmds = [
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, COL_GRID),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            # Corner cells: merge rows 0-1 for session# and time columns
            ("SPAN", (0, 0), (0, 1)),
            ("SPAN", (1, 0), (1, 1)),
            ("BACKGROUND", (0, 0), (0, 1), COL_CORNER),
            ("BACKGROUND", (1, 0), (1, 1), COL_CORNER),
        ]

        # Day header merges and background
        for d_idx in range(n_days):
            col_start = 2 + d_idx * n_branches
            col_end = col_start + n_branches - 1
            if n_branches > 1:
                style_cmds.append(
                    ("SPAN", (col_start, 0), (col_end, 0)))
            for c in range(col_start, col_end + 1):
                style_cmds.append(
                    ("BACKGROUND", (c, 0), (c, 0), COL_DAY_HDR))
                style_cmds.append(
                    ("BACKGROUND", (c, 1), (c, 1), COL_BRANCH_HDR))

        # Build occupancy map: (slot_idx, day_idx, branch_idx) -> (kind, cls, span)
        #
        # ST-UI-001. This was a plain dict assignment, so a second claimant of a
        # cell overwrote the first and printed nowhere -- the same defect the
        # filtered PDF tables had, left behind when they were fixed. Every
        # claimant is collected first, and any ENTRY that shares a cell is
        # routed to the stacked branch.
        entries = []
        claims = {}
        for c in placed:
            c_day = effective_day(c)
            c_start = effective_time(c)
            if c_day not in days or c_start not in slots:
                continue
            d_idx = days.index(c_day)
            start_si = slots.index(c_start)
            dur = c["duration"]
            for t in c["targets"]:
                if t["year"] != yr or t["branch"] not in branches:
                    continue
                b_idx = branches.index(t["branch"])
                t_idx = c["targets"].index(t)
                slot_off = slot_offset_for_target(c, t_idx)
                actual_start = start_si + slot_off
                if actual_start >= len(slots):
                    continue
                span = min(dur, len(slots) - actual_start)
                if span <= 0:
                    continue
                entry = {"cls": c, "start_si": actual_start, "span": span,
                         "d_idx": d_idx, "b_idx": b_idx}
                entries.append(entry)
                for off in range(span):
                    okey = (actual_start + off, d_idx, b_idx)
                    bucket = claims.setdefault(okey, [])
                    # Identity on the ENTRY, not on the class: a joint class
                    # contested in Year-1/A must keep rendering normally in the
                    # Year-1/B column, where nothing contests it.
                    if all(x is not entry for x in bucket):
                        bucket.append(entry)

        drawn.update(cls_key(e["cls"]) for e in entries)

        overlapping = {id(e) for cells in claims.values() if len(cells) > 1
                       for e in cells}
        occupied = {}
        overlap_cells = {}
        for entry in entries:
            if id(entry) in overlapping:
                for off in range(entry["span"]):
                    overlap_cells.setdefault(
                        (entry["start_si"] + off, entry["d_idx"],
                         entry["b_idx"]), []).append(entry)
                continue
            base = (entry["start_si"], entry["d_idx"], entry["b_idx"])
            occupied[base] = ("start", entry["cls"], entry["span"])
            for off in range(1, entry["span"]):
                occupied[(entry["start_si"] + off, entry["d_idx"],
                          entry["b_idx"])] = ("span", None, 0)

        cell_bg_cmds = []
        row_heights = [24, 18]  # header rows
        tall_rows = {}

        # The branch sub-header is the one header cell holding free user text:
        # branch names come from a comma-split line the user types
        # (dialogs.py), so "9/A Fen Bilimleri Agirlikli Sube" is as legal as
        # "A". 18pt fits one line, and a fixed row does not grow -- measured,
        # a 2-line label spills 1pt past each border into padding (harmless),
        # but a 4-line one is 60pt in an 18pt row and prints over the day
        # header above and the first lesson below. Measured here for the same
        # reason every other Paragraph in this table is.
        for _hdr_cell in row1:
            if isinstance(_hdr_cell, Paragraph):
                _note_cell_height(tall_rows, _hdr_cell, data_col_w - 4, 1, 1, 18)

        for si, slot in enumerate(slots):
            data_row = si + 2  # after 2 header rows
            row = [
                Paragraph(str(si + 1), session_style),
                Paragraph(escape_pdf_markup(slot), time_style),
            ]
            style_cmds.append(
                ("BACKGROUND", (0, data_row), (0, data_row), COL_SESSION))
            style_cmds.append(
                ("BACKGROUND", (1, data_row), (1, data_row), COL_TIME))

            for d_idx in range(n_days):
                for b_idx in range(n_branches):
                    col = 2 + d_idx * n_branches + b_idx
                    okey = (si, d_idx, b_idx)
                    if okey in overlap_cells:
                        claimants = [e["cls"] for e in overlap_cells[okey]]
                        ckeys = {cls_key(x) for x in claimants}
                        conflicted = any(
                            conflict_partners.get(cls_key(x), frozenset())
                            & (ckeys - {cls_key(x)})
                            for x in claimants)
                        stack_para = _pdf_conflict_paragraph(
                            claimants, cell_style, conflicted,
                            include_room=True, include_targets=False)
                        row.append(stack_para)
                        _w, _h = stack_para.wrap(data_col_w - 4, 1e6)
                        tall_rows[data_row] = max(
                            tall_rows.get(data_row, MIN_ROW_H), _h + 6)
                        cell_bg_cmds.append(
                            ("BACKGROUND", (col, data_row), (col, data_row),
                             rl_colors.HexColor(
                                 "#FEE2E2" if conflicted else "#F1F5F9")))
                    elif okey in occupied:
                        kind, cls, span = occupied[okey]
                        if kind == "span":
                            row.append("")
                            continue
                        para = _pdf_rich_paragraph(
                            cls, cell_style,
                            include_room=True, include_targets=False)
                        row.append(para)
                        _note_cell_height(tall_rows, para, data_col_w - 4,
                                          data_row, span, MIN_ROW_H)
                        yr_hex = get_year_color(state, yr).lstrip("#")
                        light_hex = lighten_color(f"#{yr_hex}", 0.45)
                        bg = rl_colors.HexColor(light_hex)
                        cell_bg_cmds.append(
                            ("BACKGROUND", (col, data_row),
                             (col, data_row), bg))
                        if span > 1:
                            style_cmds.append(
                                ("SPAN", (col, data_row),
                                 (col, data_row + span - 1)))
                            for off in range(1, span):
                                cell_bg_cmds.append(
                                    ("BACKGROUND",
                                     (col, data_row + off),
                                     (col, data_row + off), bg))
                    else:
                        row.append("")
                        cell_bg_cmds.append(
                            ("BACKGROUND", (col, data_row),
                             (col, data_row), COL_EMPTY))

            table_data.append(row)
            row_heights.append(MIN_ROW_H)

        for _ri, _h in tall_rows.items():
            row_heights[_ri] = _h

        style_cmds.extend(cell_bg_cmds)

        tbl = Table(table_data, colWidths=col_widths, rowHeights=row_heights,
                    repeatRows=2)
        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    # ── Assemble pages ────────────────────────────────────────────────
    elements: list = []

    if mode == "everything":
        for yr in sorted(schedule.years.keys()):
            branches = schedule.years[yr]
            if not branches:
                continue
            if elements:
                elements.append(PageBreak())
            elements.append(
                Paragraph(escape_pdf_markup(yr), title_style))
            elements.append(_build_everything_table(yr))

    elif mode == "classroom":
        physical_rooms = set(schedule.classrooms)
        for room in get_classroom_export_labels(schedule.classrooms, placed):
            if elements:
                elements.append(PageBreak())
            elements.append(
                Paragraph(escape_pdf_markup(room), title_style))
            if room not in physical_rooms:
                elements.append(_build_virtual_room_table(room))
            else:
                elements.append(_build_filtered_table(
                    filter_fn=lambda c, r=room: classroom_of(c) == r,
                    include_room=False, include_targets=True))
    elif mode == "group":
        for yr in sorted(schedule.years.keys()):
            for br in schedule.years[yr]:
                if elements:
                    elements.append(PageBreak())
                elements.append(
                    Paragraph(escape_pdf_markup(yr) + " / "
                              + escape_pdf_markup(br), title_style))
                elements.append(_build_filtered_table(
                    filter_fn=lambda c, y=yr, b=br: any(
                        t["year"] == y and t["branch"] == b
                        for t in c.get("targets", [])),
                    include_room=True, include_targets=False))

    elif mode == "lecturer":
        lecturers = schedule.lecturers
        if not lecturers:
            lecturers = sorted({
                c.get("lecturer", "") for c in placed if c.get("lecturer")})
        for lecturer in lecturers:
            if not lecturer:
                continue
            if elements:
                elements.append(PageBreak())
            elements.append(
                Paragraph(escape_pdf_markup(lecturer), title_style))
            elements.append(_build_filtered_table(
                filter_fn=lambda c, l=lecturer: c.get("lecturer", "") == l,
                include_room=True, include_targets=True))

    if not elements:
        elements.append(Paragraph(tr("warnings.no_schedule_data"), title_style))

    # ── Appendix: everything the grid could not say ───────────────────
    #
    # ST-FUNC-013 + ST-UI-001. Three different ways a lesson goes missing from
    # a printed timetable, one page:
    #
    #   * a placement on a day or hour the user has since deleted has no cell
    #     to be drawn in, so every grid-shaped page simply omits it;
    #   * a lesson no page had a column for -- the everything and group layouts
    #     are built by filtering each class's `targets`, so a lesson whose
    #     targets list is empty matches no page. That is not exotic:
    #     new_class() initializes targets to [] (core/models.py:578) and
    #     neither class-editor path requires one (ui/dialogs.py:3694, :3915),
    #     so it is the default state of every lesson placed before its groups
    #     were ticked. classroom and lecturer filter on room and teacher and
    #     kept it, which is why only two of the four layouts lost it;
    #   * a double-booking is now stacked into its cell rather than dropped,
    #     but a stacked cell is easy to miss on a dense page.
    #
    # export_schedule() already raises a Python warning per orphan, which the
    # GUI surfaces — but the *printout* said nothing, and the printout is what
    # gets pinned to a noticeboard. Silence is the dangerous outcome precisely
    # because the paper looks complete.
    #
    # Placed AFTER the `if not elements` check on purpose: appending first
    # would make `elements` non-empty and silently delete the "no schedule
    # data" page that the empty-export test pins.
    appendix_rows = []
    off_grid = list(find_off_grid_placements(state))
    off_grid_keys = {cls_key(cls) for cls, _reason in off_grid}
    for cls, _reason in off_grid:
        appendix_rows.append((
            tr("export.appendix_offgrid"),
            cls.get("class_code", ""),
            cls.get("name", ""),
            cls.get("lecturer", ""),
            f"{display_day(effective_day(cls))} {effective_time(cls) or ''}",
        ))
    # The same heading, deliberately: "not on the timetable" is literally what
    # happened to these too, the string is already translated in all 22
    # locales, and _write_unplaceable_sheet reuses it for the workbook's
    # version of this list for the same reason. Minting a key here would move
    # the ST-UI-011 ratchet by 21 pairs for a phrase a reader cannot
    # distinguish from the one above.
    for cls in placed:
        key = cls_key(cls)
        if key in drawn or key in off_grid_keys:
            continue
        appendix_rows.append((
            tr("export.appendix_offgrid"),
            cls.get("class_code", ""),
            cls.get("name", ""),
            cls.get("lecturer", ""),
            f"{display_day(effective_day(cls))} {effective_time(cls) or ''}",
        ))
    for rec in conflicts:
        a, b = rec["a"], rec["b"]
        appendix_rows.append((
            tr("export.appendix_conflict"),
            f'{a.get("class_code", "")} / {b.get("class_code", "")}',
            f'{a.get("name", "")}  /  {b.get("name", "")}',
            f'{a.get("lecturer", "")}  /  {b.get("lecturer", "")}',
            f'{display_day(rec["day"])} {rec["slot"]}',
        ))

    if appendix_rows:
        elements.append(PageBreak())
        elements.append(Paragraph(tr("export.appendix_title"), title_style))
        head = [Paragraph(h, branch_hdr_style) for h in (
            tr("labels.type"), tr("labels.class_code"), tr("labels.class_name"),
            tr("labels.lecturer"), tr("labels.day"))]
        # Escaped: these rows carry class and lecturer names straight from the
        # user's file into reportlab's markup parser.
        data = [head] + [[Paragraph(html.escape(str(v)), cell_style)
                          for v in row]
                         for row in appendix_rows]
        appendix = Table(
            data,
            colWidths=[avail_w * f for f in (0.20, 0.15, 0.33, 0.20, 0.12)],
            repeatRows=1)
        appendix.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, COL_GRID),
            ("BACKGROUND", (0, 0), (-1, 0), COL_DAY_HDR),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(appendix)

    # ── The characters this document could not spell ──────────────────
    #
    # ST-FUNC-004, the other 9 locales. Everything above draws the boxes; this
    # is the page that admits to them. Same reasoning as the appendix: the
    # printout is what gets pinned to a noticeboard, and a page that looks
    # complete while a teacher's name is five empty rectangles is worse than
    # one that says which characters are missing.
    if unprintable_chars:
        note = _unprintable_note(FONT_REGULAR, unprintable_chars)
        elements.append(PageBreak())
        elements.append(Paragraph(html.escape(note), title_style))
        # And once out of band, the way an orphaned placement already is: the
        # page says it to whoever reads the printout, the warning says it to
        # whoever is reading a log or driving the exporter from code.
        warnings.warn(note, stacklevel=2)

    doc.build(elements)


def _warn_about_off_grid_placements(state):
    """Emit one warning per placement that no longer sits on the grid."""
    from scheduler_app.models import find_off_grid_placements

    for cls, reason in find_off_grid_placements(state):
        name = cls.get("name") or cls.get("class_code") or "?"
        warnings.warn(
            tr("warnings.off_grid_placement").format(
                name=name,
                day=display_day(effective_day(cls)),
                slot=effective_time(cls)),
            stacklevel=2)


# ── Public entry point ──────────────────────────────────────────────────────

def export_schedule(schedule, format: str, filepath: str, mode: str = "everything"):
    """Export a schedule to the specified format.

    Args:
        schedule: Either a FinalSchedule instance or a state dict.
        format: One of 'xlsx', 'csv', 'pdf'.
        filepath: Output file path.
        mode: PDF/Excel layout mode – 'everything', 'classroom',
              'group', or 'lecturer'.
    """
    if isinstance(schedule, dict):
        schedule = FinalSchedule(schedule)

    # ST-DATA-003 / ST-FUNC-013: a placement orphaned by a deleted day or hour
    # has no cell to be drawn in, so every grid-shaped export would simply omit
    # it. Silence is the dangerous outcome — the printout looks complete. Say
    # so once per orphan, before writing anything.
    _warn_about_off_grid_placements(schedule.state)

    fmt = format.lower().strip(".")
    if fmt in ("xlsx", "excel"):
        # ST-ARCH-003: `mode` reaches Excel at last. The docstring above has
        # promised "PDF/Excel layout mode" since the first release while the
        # Excel branch silently discarded it -- all four modes produced the
        # identical sheet set. The writer that honours it is now the only one.
        _export_excel(schedule, filepath, mode=mode)
    elif fmt == "csv":
        _export_csv(schedule, filepath)
    elif fmt == "pdf":
        _export_pdf(schedule, filepath, mode=mode)
    else:
        raise ValueError(
            tr("errors.unsupported_format").format(
                fmt=repr(format)))
