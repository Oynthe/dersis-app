"""Excel template generator for the scheduler import pipeline."""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from scheduler_app.models import (
    LOCATION_FACE_TO_FACE,
    LOCATION_ONLINE,
    LOCATION_LECTURER_OFFICE,
    get_location_label,
)
from scheduler_app.translations import tr
from scheduler_app.data_io.schema import (
    WORKBOOK_SHEETS,
    get_workbook_sheet_title,
    get_workbook_sheet_header_map,
    get_workbook_sheet_description_map,
)


def _sheet_examples():
    lecture_room_type = tr("template.workbook_example.room_type_lecture")
    lab_room_type = tr("template.workbook_example.room_type_lab")
    room_1 = tr("template.workbook_example.room_1_name")
    room_2 = tr("template.workbook_example.room_2_name")
    room_3 = tr("template.workbook_example.room_3_name")
    branch_1 = tr("template.workbook_example.branch_1_name")
    branch_2 = tr("template.workbook_example.branch_2_name")
    branch_3 = tr("template.workbook_example.branch_3_name")

    return {
        "teachers": [
            {
                "teacher_id": "T001",
                "name": tr("template.workbook_example.teacher_1_name"),
                "allowed_days": "",
                "allowed_hours": "",
                "excluded_days": "",
                "excluded_hours": "",
            },
            {
                "teacher_id": "T002",
                "name": tr("template.workbook_example.teacher_2_name"),
                "allowed_days": ", ".join(
                    tr(f"weekdays.{day}")
                    for day in ("monday", "wednesday", "friday")
                ),
                "allowed_hours": "",
                "excluded_days": "",
                "excluded_hours": "12:00",
            },
            {
                "teacher_id": "T003",
                "name": tr("template.workbook_example.teacher_3_name"),
                "allowed_days": "",
                "allowed_hours": "09:00, 10:00, 11:00",
                "excluded_days": tr("weekdays.friday"),
                "excluded_hours": "",
            },
        ],
        "rooms": [
            {
                "room_id": "R001",
                "name": room_1,
                "capacity": 30,
                "room_type": lecture_room_type,
            },
            {
                "room_id": "R002",
                "name": room_2,
                "capacity": 20,
                "room_type": lab_room_type,
            },
            {
                "room_id": "R003",
                "name": room_3,
                "capacity": 200,
                "room_type": lecture_room_type,
            },
        ],
        "branches": [
            {"branch_id": "B001", "name": branch_1},
            {"branch_id": "B002", "name": branch_2},
            {"branch_id": "B003", "name": branch_3},
        ],
        "classes": [
            {
                "class_id": "C001",
                "class_code": "CS101",
                "course_name": tr("template.workbook_example.class_1_name"),
                "teacher_id": "T001",
                "branch_id": "B001",
                "duration": 2,
                "student_count": 25,
                "required_room_type": lab_room_type,
                "allowed_rooms": "",
                "excluded_rooms": "",
                "joint_class_group": "",
                "location_type": get_location_label(LOCATION_FACE_TO_FACE),
            },
            {
                "class_id": "C002",
                "class_code": "MATH201",
                "course_name": tr("template.workbook_example.class_2_name"),
                "teacher_id": "T002",
                "branch_id": "B001",
                "duration": 1,
                "student_count": 30,
                "required_room_type": "",
                "allowed_rooms": f"{room_1}, {room_3}",
                "excluded_rooms": "",
                "joint_class_group": "",
                "location_type": get_location_label(LOCATION_FACE_TO_FACE),
            },
            {
                "class_id": "C003",
                "class_code": "",
                "course_name": tr("template.workbook_example.class_3_name"),
                "teacher_id": "T003",
                "branch_id": "B002",
                "duration": 2,
                "student_count": 20,
                "required_room_type": lab_room_type,
                "allowed_rooms": room_2,
                "excluded_rooms": "",
                "joint_class_group": "",
                "location_type": get_location_label(LOCATION_FACE_TO_FACE),
            },
            {
                "class_id": "C004",
                "class_code": "ENG101",
                "course_name": tr("template.workbook_example.class_4_name"),
                "teacher_id": "T002",
                "branch_id": "B001",
                "duration": 1,
                "student_count": 50,
                "required_room_type": "",
                "allowed_rooms": "",
                "excluded_rooms": "",
                "joint_class_group": "J1",
                "location_type": get_location_label(LOCATION_ONLINE),
            },
            {
                "class_id": "C005",
                "class_code": "ENG101",
                "course_name": tr("template.workbook_example.class_4_name"),
                "teacher_id": "T002",
                "branch_id": "B002",
                "duration": 1,
                "student_count": 50,
                "required_room_type": "",
                "allowed_rooms": "",
                "excluded_rooms": "",
                "joint_class_group": "J1",
                "location_type": get_location_label(LOCATION_LECTURER_OFFICE),
            },
        ],
    }


def generate_excel_template(filepath: str):
    """Create a localized Excel template with sheet labels and guidance rows."""
    if not HAS_OPENPYXL:
        raise RuntimeError(tr("errors.openpyxl_template_required"))

    wb = openpyxl.Workbook()
    first = True
    sheet_examples = _sheet_examples()

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    desc_font = Font(italic=True, size=9, color="6B7280")
    example_font = Font(size=10, color="374151")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for sheet_id in WORKBOOK_SHEETS:
        sheet_name = get_workbook_sheet_title(sheet_id)
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        headers = get_workbook_sheet_header_map(sheet_id)
        descriptions = get_workbook_sheet_description_map(sheet_id)
        ordered_fields = [field for field, _, _ in WORKBOOK_SHEETS[sheet_id]["columns"]]
        examples = sheet_examples[sheet_id]

        for ci, field in enumerate(ordered_fields, start=1):
            cell = ws.cell(row=1, column=ci, value=headers[field])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for ci, field in enumerate(ordered_fields, start=1):
            cell = ws.cell(row=2, column=ci, value=descriptions[field])
            cell.font = desc_font
            cell.border = thin_border

        for ri, example_row in enumerate(examples, start=3):
            for ci, field in enumerate(ordered_fields, start=1):
                cell = ws.cell(row=ri, column=ci, value=example_row.get(field, ""))
                cell.font = example_font
                cell.border = thin_border

        for ci, field in enumerate(ordered_fields, start=1):
            header = headers[field]
            desc = descriptions[field]
            max_len = max(len(str(header)), len(str(desc)))
            for example_row in examples:
                max_len = max(max_len, len(str(example_row.get(field, ""))))
            col_letter = ws.cell(row=1, column=ci).column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A3"

    wb.save(filepath)
