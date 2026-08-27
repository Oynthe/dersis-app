"""Export pipeline: Risks 5, 8, 9.

R5 (High):  PDF export uses only Helvetica/Helvetica-Bold (no embedded
            Unicode font). Turkish letters ğ Ğ ş Ş ı İ are not in WinAnsi
            and cannot render (exporter.py:461-485).
R8 (High):  xlsx export builds sheet titles from lecturer/room/branch names
            (T_/R_/B_ prefixes). Names containing  / \ ? * [ ] :  crash
            openpyxl; truncation to 28 chars collides -> duplicate-title crash.
R9 (Med):   CSV export leaks internal day KEYS (e.g. 'monday') and app.py's
            writer uses the OS locale encoding (no encoding=), which is not
            UTF-8 and crashes on non-Turkish locales.
"""
import os, sys, tempfile, csv, io, re, traceback

_sb = tempfile.mkdtemp(prefix="dersis_probe05_")
os.environ["HOME"] = _sb
os.environ["USERPROFILE"] = _sb
sys.path.insert(0, r"C:\dev\dersis-app")
sys.path.insert(0, r"C:\dev\dersis-app\stress-test\tests")

from scheduler_app.translations import set_language
set_language("tr")

from scheduler_app.data_io import exporter
from scheduler_app.core.models import new_state, new_class, mark_placed

EVID = r"C:\dev\dersis-app\stress-test\evidence"
os.makedirs(EVID, exist_ok=True)

TURKISH = "Coğrafya Şubesi ığİ Öğretim"   # ğ ş İ ı Ö Ç


def placed_state(class_name="Ders", lecturer="Hoca", room="R1",
                 day="monday", time_="09:00"):
    st = new_state()
    st["days"] = ["monday", "tuesday", "wednesday", "thursday", "friday"]
    st["slots"] = [f"{9+i:02d}:00" for i in range(8)]
    st["classrooms"] = [room]
    st["lecturers"] = [lecturer]
    st["years"] = {"Year 1": ["A"]}
    c = new_class()
    c["name"] = class_name
    c["lecturer"] = lecturer
    c["duration"] = 1
    c["targets"] = [{"year": "Year 1", "branch": "A"}]
    mark_placed(c, day, time_, room)
    st["classes"] = [c]
    return st


# ─────────────────────────────────────────────────────────────────────────
print("=" * 75)
print("R5: PDF Turkish glyph support — scan /BaseFont in produced PDF")
print("=" * 75)
st = placed_state(class_name=TURKISH, lecturer="Dr. Şükrü Iğdır")
pdf_path = os.path.join(EVID, "turkish_export.pdf")
try:
    exporter.export_schedule(st, "pdf", pdf_path, mode="everything")
    raw = open(pdf_path, "rb").read()
    fonts = sorted(set(re.findall(rb"/BaseFont\s*/([A-Za-z0-9+\-]+)", raw)))
    print("   PDF written:", os.path.getsize(pdf_path), "bytes")
    print("   /BaseFont entries:", [f.decode() for f in fonts])
    embedded = [f for f in fonts if b"+" in f]  # subset fonts contain 'ABCDEF+'
    print("   embedded subset (Unicode) fonts:", [f.decode() for f in embedded] or "NONE")
    # does the raw Turkish text survive as bytes anywhere? (won't for Helvetica)
    has_fontfile = b"/FontFile" in raw
    print("   PDF contains an embedded /FontFile program:", has_fontfile)
    if not embedded and not has_fontfile:
        print("   *** CONFIRMED: only standard non-Unicode fonts; ğ/ş/İ/ı cannot render ***")
except Exception:
    traceback.print_exc()

# Direct reportlab check: what does Helvetica do with these code points?
print("\n   Direct reportlab stringWidth probe (Helvetica):")
try:
    from reportlab.pdfbase.pdfmetrics import stringWidth
    for ch, nm in [("ş", "U+015F"), ("ğ", "U+011F"), ("İ", "U+0130"),
                   ("ı", "U+0131"), ("a", "U+0061")]:
        try:
            w = stringWidth(ch, "Helvetica", 10)
            print(f"     {nm} {ch!r}: width={w:.2f}")
        except Exception as e:
            print(f"     {nm} {ch!r}: *** {type(e).__name__}: {e} ***")
except Exception:
    traceback.print_exc()

# ─────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 75)
print("R8a: xlsx export — lecturer name with '/' (invalid sheet-title char)")
print("=" * 75)
st = placed_state(lecturer="Dr. Ada/Bob")
st["classes"][0]["lecturer"] = "Dr. Ada/Bob"
xlsx_path = os.path.join(_sb, "slash.xlsx")
try:
    exporter.export_schedule(st, "xlsx", xlsx_path)
    print("   export SUCCEEDED (no crash)")
except Exception as e:
    print(f"   *** CONFIRMED CRASH: {type(e).__name__}: {e} ***")

print("\nR8b: xlsx export — lecturer name with ':' ")
st = placed_state(lecturer="Prof: X")
st["classes"][0]["lecturer"] = "Prof: X"
try:
    exporter.export_schedule(st, "xlsx", os.path.join(_sb, "colon.xlsx"))
    print("   export SUCCEEDED (no crash)")
except Exception as e:
    print(f"   *** CONFIRMED CRASH: {type(e).__name__}: {e} ***")

print("\nR8c: two lecturers colliding after [:28] truncation -> duplicate title")
st = new_state()
st["days"] = ["monday"]; st["slots"] = ["09:00"]; st["classrooms"] = ["R1"]
name1 = "Professor Alexander Hamiltonius The First"
name2 = "Professor Alexander Hamiltonius The Second"
st["lecturers"] = [name1, name2]
st["years"] = {"Y": ["A"]}
for nm in (name1, name2):
    c = new_class(); c["name"] = "X"; c["lecturer"] = nm; c["duration"] = 1
    c["targets"] = [{"year": "Y", "branch": "A"}]
    mark_placed(c, "monday", "09:00", "R1"); st["classes"].append(c)
print(f"   T_{name1[:28]!r} vs T_{name2[:28]!r} (equal prefix: {name1[:28]==name2[:28]})")
try:
    exporter.export_schedule(st, "xlsx", os.path.join(_sb, "dup.xlsx"))
    print("   export SUCCEEDED (no crash)")
except Exception as e:
    print(f"   *** CONFIRMED CRASH: {type(e).__name__}: {e} ***")

# ─────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 75)
print("R9: CSV export leaks day KEYS + locale-fragile encoding")
print("=" * 75)
st = placed_state(class_name=TURKISH, lecturer="Şşğ Hoca", day="monday")
csv_path = os.path.join(EVID, "export_turkish.csv")
exporter.export_schedule(st, "csv", csv_path)
raw = open(csv_path, "rb").read()
text = raw.decode("utf-8")
print("   exporter CSV (data_io) day column value:",
      repr([r.split(",")[0] for r in text.strip().splitlines()[1:]]))
print("   -> leaks raw internal key 'monday' instead of localized 'Pazartesi'"
      if "monday" in text else "   -> localized")
print("   file starts with UTF-8 BOM:", raw[:3] == b"\xef\xbb\xbf",
      "(Excel-TR opens no-BOM UTF-8 as cp1254 => mojibake)")

# Replicate app.py:2306 exact writer (open(...,'w',newline='') with NO encoding)
print("\n   app.py export_csv writer replica (open without encoding=):")
appcsv = os.path.join(_sb, "app_style.csv")
import locale
print("   OS preferred encoding:", locale.getpreferredencoding(False))
with open(appcsv, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Ders", "Öğretim", "day"])
    w.writerow([TURKISH, "Şşğ Hoca", "monday"])
b = open(appcsv, "rb").read()
try:
    b.decode("utf-8")
    print("   app-style file decodes as UTF-8:", True)
except UnicodeDecodeError as e:
    print(f"   *** app-style file is NOT valid UTF-8 (locale bytes): {e} ***")

# Simulate a non-Turkish Windows locale (cp1252) doing the same write:
print("\n   Same write under a cp1252 (e.g. US/DE Windows) locale:")
try:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([TURKISH, "Şşğ Hoca"])
    buf.getvalue().encode("cp1252")
    print("   cp1252 encode: OK")
except UnicodeEncodeError as e:
    print(f"   *** CONFIRMED: UnicodeEncodeError on cp1252 locale: {e} ***")

# ─────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 75)
print("R8d: stale placement (placed_day not in state days) — silent drop?")
print("=" * 75)
st = placed_state(day="funday")   # 'funday' is not in state['days']
try:
    xp = os.path.join(_sb, "stale.xlsx")
    exporter.export_schedule(st, "xlsx", xp)
    import openpyxl
    wb = openpyxl.load_workbook(xp)
    ws = wb[wb.sheetnames[0]]
    vals = [c.value for row in ws.iter_rows() for c in row if c.value]
    appears = any(TURKISH.split()[0] in str(v) for v in vals)
    print(f"   master sheet contains the stale-placed class: {appears} "
          f"(False => silently dropped from export)")
except Exception as e:
    print(f"   EXCEPTION: {type(e).__name__}: {e}")

print("\nSandbox:", _sb, "| Evidence:", EVID)
