"""C3 — an ``excluded_rooms`` name no room answers to is accepted in silence.

The claim under test
--------------------
``data_io/importer.py::_process_classes`` validates ``allowed_rooms`` against
the workbook's own Rooms sheet — an unknown name produces
``errors.unknown_rooms`` and is filtered out of ``required_classrooms`` — and
does nothing of the kind for ``excluded_rooms``, which is copied straight into
``cls["excluded_classrooms"]``::

    if excluded_rooms:
        cls["excluded_classrooms"] = excluded_rooms

The handoff filed this as "an asymmetry, not necessarily a defect", on the
reasoning that a name matching no room forbids nothing and is therefore inert.
Inert is exactly the problem. ``get_physical_room_candidates`` matches
exclusions by the same exact ``in`` (``core/models.py::get_physical_room_candidates``), so the *typo* is
inert while the user's *intention* was not: a teacher who writes ``Lab1`` for
``Lab 1`` has forbidden nothing, is told nothing, and gets the lesson placed in
Lab 1. The identical typo one column to the left is reported.

Nothing in the UI recovers it afterwards either. ``reconcile_placements``
strips the dangling name (core/workflow.py, the ``for field in
("required_classrooms", "excluded_classrooms")`` sweep) but deliberately does
not report an exclusion — only a lost ``required_classrooms`` reaches
``lost_room_requirements`` and the toast — and ``AddClassDialog`` builds its
room checkboxes from the live room list, so a name no room answers to has no
checkbox to notice. The import report is the only place this can be caught,
and it is empty.

What this test does NOT ask for
-------------------------------
It asserts a **warning**, and deliberately not the other half of the
``allowed_rooms`` treatment. Filtering ``excluded_classrooms`` down to known
names would be a regression, not a fix: ``room_names`` is
``dataset.state["classrooms"]``, i.e. the rooms of *this workbook only*, and
the importer explicitly supports a workbook with no Rooms sheet at all ("a
school may keep its rooms in a separate file", ``load_scheduler_data_from_excel``).
Filtering would then delete every exclusion in such a workbook, all of which
name rooms that exist perfectly well in the state being merged into.
``test_a_workbook_with_no_rooms_sheet_keeps_its_exclusions`` below pins that,
and passes today — it is here so a fix cannot buy the warning by breaking it.

For the same reason the warning is pinned as *gated on the workbook having a
Rooms sheet*: without the gate, every classes-only workbook would warn about
every exclusion it carries.
"""

import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

from scheduler_app.core.models import get_physical_room_candidates  # noqa: E402
from scheduler_app.data_io import schema  # noqa: E402
from scheduler_app.data_io.importer import (  # noqa: E402
    load_scheduler_data_from_excel,
)

pytestmark = pytest.mark.excel

SHEET_IDS = ["teachers", "rooms", "branches", "classes"]

TEACHERS = [{"teacher_id": "T001", "name": "Ada Lovelace"}]
# "Lab 1" with a space; the workbooks below mistype it as "Lab1", which is the
# whole point — one character apart, and no room answers to the second.
ROOMS = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30, "room_type": "Derslik"},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20, "room_type": "Laboratuvar"},
]
BRANCHES = [{"branch_id": "B001", "name": "Grup A"}]

BASE_CLASS = {"class_id": "C001", "course_name": "Fizik", "teacher_id": "T001",
              "branch_id": "B001", "year": 1, "duration": 1, "student_count": 10}

REAL_ROOM = "Lab 1"
TYPO = "Lab1"


def build_workbook(path, class_row, sheets=SHEET_IDS):
    """Write a workbook whose titles and headers come from the schema itself.

    Row 2 is a data row, not the template's help text: ``_read_sheet`` drops
    row 2 only when its first ``*_id`` cell is one of the strings the template
    writes, and "C001" is not one.
    """
    rows_by_sheet = {"teachers": TEACHERS, "rooms": ROOMS,
                     "branches": BRANCHES, "classes": [class_row]}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in sheets:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]]
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=headers[field])
        for ri, row in enumerate(rows_by_sheet[sheet_id], start=2):
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=ri, column=col, value=row[field])
    wb.save(str(path))
    return str(path)


def _class_warnings(report):
    return [w for w in report.warnings if TYPO in str(w)]


# ── The premise: the typo really does fail to forbid anything ───────────────

def test_a_mistyped_exclusion_leaves_the_room_a_candidate(tmp_path):
    """Establishes the consequence before the next test asks anyone be told.

    Without this, a fix could satisfy the warning test by warning about
    something harmless. ``Lab 1`` is a candidate for a class whose sheet says
    it must not be, and the correctly spelled twin proves the mechanism is the
    exact-name match and nothing else about the fixture.
    """
    typo_path = build_workbook(tmp_path / "typo.xlsx",
                               dict(BASE_CLASS, excluded_rooms=TYPO))
    typo = load_scheduler_data_from_excel(typo_path)
    assert typo.report.is_valid
    cls = typo.state["classes"][0]
    assert cls["excluded_classrooms"] == [TYPO], (
        "the unvalidated name no longer reaches state — if the importer now "
        "filters it, re-read this module's docstring before deleting the test")
    assert REAL_ROOM in get_physical_room_candidates(typo.state, cls)

    spelled_path = build_workbook(tmp_path / "spelled.xlsx",
                                  dict(BASE_CLASS, excluded_rooms=REAL_ROOM))
    spelled = load_scheduler_data_from_excel(spelled_path)
    spelled_cls = spelled.state["classes"][0]
    assert REAL_ROOM not in get_physical_room_candidates(
        spelled.state, spelled_cls), (
        "the correctly spelled exclusion does not forbid the room either — "
        "this test's premise is gone, and the C3 claim with it")


# ── The defect ──────────────────────────────────────────────────────────────

def test_an_unknown_excluded_room_is_reported_like_its_allowed_twin(tmp_path):
    """FAILS TODAY. The import report says nothing about ``Lab1``.

    The user wrote a room name into the Excluded Rooms column, that name
    matches no room in the Rooms sheet of the same workbook, and the class is
    then free to be placed in the room they meant to forbid. Whatever wording
    the fix chooses, the row number and the offending name have to appear —
    those are the two halves a user can act on, and the sibling
    ``errors.unknown_rooms`` already carries both.

    The warning must NOT reuse ``errors.unknown_rooms`` verbatim: all 22
    locales hard-code the literal column name ``allowed_rooms`` in that string
    (``i18n/translations.py``), so attaching it to this row would name the
    wrong cell to edit — the defect commit 935c84b closed for the room-type
    warning one column over.
    """
    path = build_workbook(tmp_path / "unknown_excluded.xlsx",
                          dict(BASE_CLASS, excluded_rooms=TYPO))
    report = load_scheduler_data_from_excel(path).report

    assert report.is_valid, "an unknown exclusion is a warning, never an error"
    named = _class_warnings(report)
    assert named, (
        "no warning names %r: the Excluded Rooms column accepted a room the "
        "workbook does not define, and the class can still be placed there. "
        "Warnings were: %r" % (TYPO, [str(w) for w in report.warnings]))
    assert any("2" in str(w) for w in named), (
        "the warning does not carry the row number: %r" % [str(w) for w in named])
    assert not any("allowed_rooms" in str(w) for w in named), (
        "the warning reuses errors.unknown_rooms, whose text names the "
        "allowed_rooms column in every locale, so it points at the wrong "
        "cell: %r" % [str(w) for w in named])


def test_the_allowed_twin_still_reports_the_identical_typo(tmp_path):
    """Passes today. The discrimination half: same typo, other column.

    If this ever goes red the asymmetry has closed from the wrong end — the
    allowed path having lost its warning rather than the excluded path having
    gained one — and the test above would then be green for a reason nobody
    wants.
    """
    path = build_workbook(tmp_path / "unknown_allowed.xlsx",
                          dict(BASE_CLASS, allowed_rooms=TYPO))
    report = load_scheduler_data_from_excel(path).report
    assert _class_warnings(report), (
        "allowed_rooms no longer reports an unknown room either: %r"
        % [str(w) for w in report.warnings])


# ── The guard on the fix ────────────────────────────────────────────────────

def test_a_workbook_with_no_rooms_sheet_keeps_its_exclusions(tmp_path):
    """Passes today, and must keep passing. Bounds what the fix may do.

    ``room_names`` is the Rooms sheet of *this* workbook. A classes-only
    workbook has none, so mirroring the ``allowed_rooms`` block onto
    ``excluded_rooms`` would call every exclusion unknown, warn about all of
    them and delete all of them — exclusions that name rooms the existing state
    holds perfectly well.

    The allowed path already does this, and it is a live defect rather than a
    precedent to copy: the assertions below record it as measured, so that a
    fix which "makes the two columns converge" has to notice which behaviour it
    would be converging on.
    """
    path = build_workbook(tmp_path / "classes_only.xlsx",
                          dict(BASE_CLASS, excluded_rooms=REAL_ROOM,
                               allowed_rooms="Oda 1"),
                          sheets=["teachers", "branches", "classes"])
    dataset = load_scheduler_data_from_excel(path)
    cls = dataset.state["classes"][0]

    assert cls["excluded_classrooms"] == [REAL_ROOM], (
        "a workbook with no Rooms sheet lost its exclusion; the fix filtered "
        "against a room list it was never given")
    assert not [w for w in dataset.report.warnings if REAL_ROOM in str(w)], (
        "a workbook with no Rooms sheet was warned about an exclusion naming a "
        "room it never claimed to define")

    # Measured, not endorsed — the allowed_rooms half of the same row.
    assert cls["required_classrooms"] == [], (
        "the allowed_rooms path no longer empties its list on a rooms-less "
        "workbook; if that was fixed, this test's warning about copying it is "
        "stale")
