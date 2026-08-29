"""Excel import correctness at the library level (no UI involved).

Everything here drives ``scheduler_app.data_io.importer.load_scheduler_data_from_excel``
directly against workbooks built in ``tmp_path``, so a failure always points at
the import pipeline and never at Qt, at the user's home directory, or at a
checked-in binary fixture.

Findings guarded here
---------------------
* **ST-FUNC-002** (Critical, fixed in Phase 0) — blank ``joint_class_group``
  cells collapse to the string ``'nan'`` and silently merge unrelated classes.
* **ST-FUNC-003** (High, fixed in Phase 0) — one malformed numeric cell raises
  an uncaught ``ValueError`` that aborts the entire import. Split into three
  tests, because "malformed" covers three cases with three different correct
  outcomes: text in a numeric cell (report it), a blank *optional* cell (use
  the documented default silently), and a blank *required* cell (either, but
  never a crash). A single test demanding a report for all three would stay
  red after the register's own recommended fix.
* **ST-FUNC-011** (fixed in Phase 7) — a workbook with no recognized sheet at
  all reports ``is_valid=True``, which the UI announces as a successful import.
* **ST-FUNC-012** (fixed in Phase 7) — two teacher rows sharing a name fuse
  into one lecturer and the first one's availability is overwritten.
* **ST-FUNC-010** (fixed in Phase 7) — the row-2 help-text heuristic guessed
  from shape, so a class id written ``9 A`` was eaten as help text and the
  Chinese and Japanese templates' help text was imported as real data.
* **ST-FUNC-009** (fixed in Phase 8) — ``required_room_type`` was advertised in
  the template and whitelisted by the import schema, then discarded, so the
  app's own shipped template imported its lab class with no room constraint at
  all and a physics lab could be scheduled into a lecture hall. It is now
  resolved to room names at import time against the ``room_type`` column of the
  *same* workbook. The 22-locale sweep below carries the half no single-locale
  test can: a fix that matches the room's *name* instead of its type is right in
  Turkish by accident (the lab room is called ``Laboratuvar A`` and typed
  ``Laboratuvar``, so the name contains the type) and wrong in ``nl`` and
  ``az``, where the lab room is called ``Lab A`` and typed ``Practicum``.
* **The generated template must re-import in all 22 shipped languages** —
  Spanish could not (two languages disagree about what a sheet titled *Aulas*
  holds) and Chinese and Japanese imported phantom rows.

Two conventions used throughout
-------------------------------
1. *Localized strings are never hard-coded.* Sheet titles, column headers and
   example values all come from ``data_io/schema.py`` and ``translations`` with
   the language pinned to Turkish by the session-wide ``_pinned_language``
   fixture in ``tests/conftest.py``.
2. *Hand-built workbooks omit the template's row-2 description row.*
   ``importer._read_sheet`` drops the first data row only when the first
   ``*_id`` column of that row holds one of the help-text strings the template
   writes, in any shipped language. The workbooks below put real IDs there, so
   nothing is dropped. ``test_builder_*`` below proves this both ways, so a
   failure in any other test can never be blamed on the fixture builder.
"""

import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

from scheduler_app.core.models import (  # noqa: E402
    LOCATION_FACE_TO_FACE,
    LOCATION_LECTURER_OFFICE,
    LOCATION_ONLINE,
    get_location_label,
    get_physical_room_candidates,
    new_class,
    new_state,
)
from scheduler_app.data_io import schema  # noqa: E402
from scheduler_app.data_io.importer import load_scheduler_data_from_excel  # noqa: E402
from scheduler_app.data_io.template import generate_excel_template  # noqa: E402
from scheduler_app.translations import (  # noqa: E402
    TRANSLATIONS,
    get_language,
    set_language,
    tr,
)

pytestmark = pytest.mark.excel

SHEET_IDS = ("teachers", "rooms", "branches", "classes")


# ── Workbook builder ────────────────────────────────────────────────────────

def _fields(sheet_id, omit=()):
    """Ordered canonical field names for a sheet, minus anything omitted."""
    return [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]
            if f not in omit]


def build_workbook(path, *, classes, teachers=None, rooms=None, branches=None,
                   omit_columns=None, sheets=SHEET_IDS, description_row=False,
                   shout_headers=False):
    """Write a workbook whose sheet titles and headers come from the schema.

    A field simply left out of a row dict is written as a *truly empty* cell,
    which is what pandas turns into ``NaN`` — the exact shape that triggers
    ST-FUNC-002 and ST-FUNC-003.

    ``shout_headers`` writes each header as ``label.upper()`` instead of the
    shipped casing. Every other workbook in this module writes the header row
    straight from ``schema.get_workbook_sheet_header_map``, i.e. in exactly the
    case the app itself would have written, so nothing here exercised the
    header fold until this option existed.
    """
    omit_columns = omit_columns or {}
    rows_by_sheet = {
        "teachers": DEFAULT_TEACHERS if teachers is None else teachers,
        "rooms": DEFAULT_ROOMS if rooms is None else rooms,
        "branches": DEFAULT_BRANCHES if branches is None else branches,
        "classes": list(classes),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in sheets:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = _fields(sheet_id, omit_columns.get(sheet_id, ()))
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            label = headers[field]
            ws.cell(row=1, column=col,
                    value=label.upper() if shout_headers else label)

        excel_row = 2
        if description_row:
            descriptions = schema.get_workbook_sheet_description_map(sheet_id)
            for col, field in enumerate(fields, start=1):
                ws.cell(row=excel_row, column=col, value=descriptions[field])
            excel_row += 1

        for row in rows_by_sheet[sheet_id]:
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=excel_row, column=col, value=row[field])
            excel_row += 1

    wb.save(str(path))
    return str(path)


# Room types are LITERALS here, not ``tr(...)`` lookups, and that is
# load-bearing rather than lazy (ST-FUNC-009). ``room_type`` is free text the
# app never translates: the importer matches a class's type against the Rooms
# sheet of the *same workbook*, so both sides of the comparison must be the
# same string, and a school that writes "Atölye" must match too. Pulling these
# from the catalogue also read the language at *import* time, before conftest's
# session-wide ``_pinned_language`` fixture switches to Turkish, so the fixture
# said "Lab" while every test body asking for the same value said
# "Laboratuvar" — two clocks that could never agree. The xfail on
# ``test_required_room_type_constrains_the_class_to_matching_rooms`` hid that:
# the test was red for the fixture's reason, not the importer's.
ROOM_TYPE_LECTURE = "Derslik"
ROOM_TYPE_LAB = "Laboratuvar"

DEFAULT_TEACHERS = [
    {"teacher_id": "T001", "name": "Ada Lovelace"},
    {"teacher_id": "T002", "name": "Bora Yildiz"},
]
# Two rooms share the lab type on purpose (ST-FUNC-009). With only one lab,
# "the type narrows allowed_rooms" and "the type replaces allowed_rooms" produce
# the same list for every input, so both readings pass every assertion and the
# intersection rule is untested. Lab 2 is what separates them.
DEFAULT_ROOMS = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30,
     "room_type": ROOM_TYPE_LECTURE},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    {"room_id": "R003", "name": "Lab 2", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
]
DEFAULT_BRANCHES = [
    {"branch_id": "B001", "name": "Grup A"},
    {"branch_id": "B002", "name": "Grup B"},
]


def klass(class_id, **overrides):
    """A minimal, fully valid Classes row; override any field by keyword.

    ``joint_class_group`` defaults to a value unique per class so that the
    joint-merge machinery stays out of the way of tests that are not about it
    (``_resolve_joint_groups`` skips groups with fewer than two members).
    """
    row = {
        "class_id": class_id,
        "course_name": f"Ders {class_id}",
        "teacher_id": "T001",
        "branch_id": "B001",
        "duration": 1,
        "student_count": 10,
        "joint_class_group": f"UNIQ-{class_id}",
    }
    row.update(overrides)
    return row


def messages(report):
    """All human-readable report lines, errors and warnings together."""
    return list(report.errors) + list(report.warnings)


def row_prefix(sheet_id, excel_row):
    """The localized ``[Sheet] Row N:`` prefix the importer puts on row errors."""
    return tr("status.import_row_prefix").format(
        sheet=schema.get_workbook_sheet_title(sheet_id), row=excel_row)


def course_names(dataset):
    return sorted(c["name"] for c in dataset.state["classes"])


def branch_set(cls):
    return {t["branch"] for t in cls.get("targets", [])}


# ── Fixture-builder self-checks ─────────────────────────────────────────────
# These exist so that a failure anywhere below can be attributed to the
# importer and not to build_workbook().

def test_builder_produces_a_clean_importable_workbook(tmp_path):
    """Guard: the hand-built workbook shape imports cleanly.

    If this fails, every other test in this module is untrustworthy — the
    fixture, not the importer, would be the thing under test.
    """
    path = build_workbook(tmp_path / "clean.xlsx",
                          classes=[klass("C001"), klass("C002"), klass("C003")])
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert ds.report.warnings == []
    assert [e["class_id"] for e in ds.raw_classes] == ["C001", "C002", "C003"]
    assert len(ds.state["classes"]) == 3
    assert ds.state["lecturers"] == ["Ada Lovelace", "Bora Yildiz"]
    assert ds.state["classrooms"] == ["Oda 1", "Lab 1", "Lab 2"]


def test_builder_row_two_description_row_is_skipped(tmp_path):
    """Guard: ``_read_sheet``'s description-row heuristic behaves as documented.

    A user importing the shipped template must not lose the first real data
    row, and must not gain a phantom class made of help text.
    """
    path = build_workbook(tmp_path / "described.xlsx", description_row=True,
                          classes=[klass("C001"), klass("C002"), klass("C003")])
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert [e["class_id"] for e in ds.raw_classes] == ["C001", "C002", "C003"]
    assert course_names(ds) == ["Ders C001", "Ders C002", "Ders C003"]


# ── ST-FUNC-002 — template round-trip ───────────────────────────────────────

def test_template_roundtrip_preserves_reference_data(tmp_path):
    """Guard (ST-FUNC-002 control): teachers/rooms/branches survive a round-trip.

    Pairs with the class-count test below: if this passes and that one fails,
    the loss is provably in joint-group resolution and nowhere else.
    """
    path = tmp_path / "template.xlsx"
    generate_excel_template(str(path))
    ds = load_scheduler_data_from_excel(str(path))

    assert ds.report.is_valid, ds.report.summary()
    assert ds.report.errors == []

    expected_teachers = [tr(f"template.workbook_example.teacher_{i}_name")
                         for i in (1, 2, 3)]
    expected_rooms = [tr(f"template.workbook_example.room_{i}_name")
                      for i in (1, 2, 3)]
    expected_branches = [tr(f"template.workbook_example.branch_{i}_name")
                         for i in (1, 2, 3)]

    assert ds.state["lecturers"] == expected_teachers
    assert ds.state["classrooms"] == expected_rooms
    assert sorted(ds.state["classroom_capacities"].values()) == [20, 30, 200]

    year = tr("status.default_year_name").format(n=1)
    assert list(ds.state["years"]) == [year]
    assert ds.state["years"][year] == expected_branches


def test_template_roundtrip_preserves_every_class(tmp_path):
    """ST-FUNC-002 — the app's own template must survive its own importer.

    ``template.py`` ships five example class rows. C001/C002/C003 leave
    ``joint_class_group`` blank and are three independent courses; C004 and
    C005 both carry the group ``J1`` and are deliberately the *same* English
    course taught jointly to branches B001 and B002, so they are supposed to
    collapse into a single joint session carrying both branch targets. The
    correct round-trip therefore yields **4** classes, not 5 and not 2.

    A failure means a school that downloads the built-in template, fills it in
    and imports it silently loses most of its courses with ``is_valid=True``
    and no warning.
    """
    path = tmp_path / "template.xlsx"
    generate_excel_template(str(path))
    ds = load_scheduler_data_from_excel(str(path))

    assert ds.report.is_valid, ds.report.summary()
    # All five rows were read; only the J1 pair may be collapsed afterwards.
    assert [e["class_id"] for e in ds.raw_classes] == [
        "C001", "C002", "C003", "C004", "C005"]

    solo_names = [tr(f"template.workbook_example.class_{i}_name") for i in (1, 2, 3)]
    joint_name = tr("template.workbook_example.class_4_name")
    branch_1 = tr("template.workbook_example.branch_1_name")
    branch_2 = tr("template.workbook_example.branch_2_name")

    assert len(ds.state["classes"]) == 4, (
        "expected C001/C002/C003 to stay separate and C004+C005 to merge into "
        f"one joint session; got {course_names(ds)}")
    assert course_names(ds) == sorted(solo_names + [joint_name])

    by_name = {c["name"]: c for c in ds.state["classes"]}
    for name in solo_names:
        assert len(by_name[name]["targets"]) == 1, (
            f"{name!r} has a blank joint group and must keep exactly one "
            f"branch target, got {by_name[name]['targets']}")

    joint = by_name[joint_name]
    # NOTE: `joint_session` is NOT evidence that a merge happened — new_class()
    # defaults it to True for every class. It is asserted only so that a future
    # fix which starts clearing the flag for single-target classes cannot also
    # clear it on a genuine joint session. The discriminating signal is targets.
    assert joint["joint_session"] is True
    assert branch_set(joint) == {branch_1, branch_2}
    assert joint["class_code"] == "ENG101"


# ── ST-FUNC-002 — blank joint-group cells ───────────────────────────────────

@pytest.mark.parametrize("variant", ["absent-cell", "empty-string"])
def test_blank_joint_group_never_merges_classes(tmp_path, variant):
    """ST-FUNC-002 — blank ``joint_class_group`` cells must not group anything.

    pandas reads a blank cell as ``NaN`` and ``str(NaN) == 'nan'``, so today
    every blank-group class shares the joint key ``'nan'`` and all but the
    first are deleted from the state. Three unrelated courses must come back
    as three courses, each keeping its own single branch target.

    Note on ``joint_session``: ``models.new_class()`` defaults that flag to
    ``True`` for *every* class, and it is only meaningful when a class has more
    than one target (``models.is_split_session``). The observable signature of
    an unwanted merge is therefore the target list, which is what is asserted.
    """
    # Both spellings a user can produce in Excel are covered. Verified that
    # pandas collapses them to the same float NaN, so this parametrization
    # documents two *user* actions, not two code paths.
    if variant == "absent-cell":
        # Field omitted entirely -> openpyxl writes nothing -> pandas NaN.
        rows = [klass(cid, joint_class_group=None) for cid in ("C001", "C002", "C003")]
        rows = [{k: v for k, v in r.items() if k != "joint_class_group"} for r in rows]
    else:
        rows = [klass(cid, joint_class_group="") for cid in ("C001", "C002", "C003")]

    path = build_workbook(tmp_path / f"joint_{variant}.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert len(ds.state["classes"]) == 3, (
        "blank joint-group cells merged unrelated classes; survivors: "
        f"{course_names(ds)}")
    assert course_names(ds) == ["Ders C001", "Ders C002", "Ders C003"]
    for cls in ds.state["classes"]:
        assert len(cls["targets"]) == 1, (
            f"{cls['name']!r} absorbed another class's target: {cls['targets']}")


def test_whitespace_only_joint_group_never_merges_classes(tmp_path):
    """Guard (ST-FUNC-002 control): a whitespace-only joint cell is no group.

    This path already works because ``str('   ').strip()`` is falsy; keeping it
    green guarantees the ST-FUNC-002 fix does not regress the one blank-cell
    spelling that behaved correctly.
    """
    rows = [klass(cid, joint_class_group="   ") for cid in ("C001", "C002", "C003")]
    path = build_workbook(tmp_path / "joint_ws.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert len(ds.state["classes"]) == 3
    assert all(len(c["targets"]) == 1 for c in ds.state["classes"])


def test_shared_joint_group_still_merges_and_keeps_all_targets(tmp_path):
    """Guard (ST-FUNC-002 control): a real joint group must keep merging.

    The ST-FUNC-002 fix must remove the bogus ``'nan'`` group without breaking
    the feature users actually asked for — one lecture attended by two groups.
    """
    rows = [
        klass("C001", branch_id="B001", joint_class_group="J1", course_name="Ortak"),
        klass("C002", branch_id="B002", joint_class_group="J1", course_name="Ortak"),
    ]
    path = build_workbook(tmp_path / "joint_real.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert len(ds.state["classes"]) == 1
    merged = ds.state["classes"][0]
    # See the note in the template round-trip: joint_session is True on every
    # class, so the merge is proved by the target list, not by this flag.
    assert merged["joint_session"] is True
    assert branch_set(merged) == {"Grup A", "Grup B"}


# ── ST-FUNC-003 — per-row numeric parsing ───────────────────────────────────

def test_garbage_numeric_cells_are_reported_per_row_and_do_not_abort(tmp_path):
    """ST-FUNC-003 — one malformed number must not destroy the whole import.

    ``int(row.get('duration', 1) or 1)`` raises ``ValueError`` on ``'iki'``, and
    nothing between here and the UI catches it. A user with one typo in a
    250-row roster currently gets no classes and no dialog; they must instead
    get the 248 good rows plus a message naming the two bad ones.

    Text where a number belongs is *not* something the importer may fix on the
    user's behalf without saying so, so both bad rows must appear in the report
    (as an error if the row is skipped, as a warning if the value is coerced).
    """
    rows = [
        klass("C001"),
        klass("C002", duration="iki"),        # non-numeric duration  (required col)
        klass("C003", student_count="cok"),   # non-numeric count     (optional col)
        klass("C004"),
    ]
    path = build_workbook(tmp_path / "bad_numbers.xlsx", classes=rows)

    # Must not raise: today this escapes as ValueError and aborts the import.
    ds = load_scheduler_data_from_excel(path)

    # The two intact rows survive.
    assert "Ders C001" in course_names(ds)
    assert "Ders C004" in course_names(ds)

    # Each bad row is identified — by its localized row prefix or by its ID.
    # Excel rows: header is row 1, so C002 is row 3 and C003 is row 4.
    lines = messages(ds.report)
    for class_id, excel_row in (("C002", 3), ("C003", 4)):
        assert any(row_prefix("classes", excel_row) in line or class_id in line
                   for line in lines), (
            f"nothing in the report names the bad row {class_id} "
            f"(Excel row {excel_row}); report was: {lines}")


def test_blank_optional_student_count_falls_back_to_zero(tmp_path):
    """ST-FUNC-003 — a blank *optional* numeric cell is not a malformed cell.

    ``student_count`` is in ``CLASS_OPTIONAL``; the code's own intent is
    ``int(row.get('student_count', 0) or 0)``. But pandas hands back ``NaN``
    rather than the ``0`` default, ``NaN`` is truthy so ``or 0`` never fires,
    and ``int(NaN)`` aborts the whole import.

    A user who simply did not know the head-count for one course must get that
    course with 0 participants — not lose the entire roster, and not be nagged
    about a column they were told was optional.
    """
    blank = {k: v for k, v in klass("C001").items() if k != "student_count"}
    path = build_workbook(tmp_path / "blank_count.xlsx",
                          classes=[blank, klass("C002")])

    # Must not raise: today this is `ValueError: cannot convert float NaN to integer`.
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert course_names(ds) == ["Ders C001", "Ders C002"]
    by_name = {c["name"]: c for c in ds.state["classes"]}
    assert by_name["Ders C001"]["participants"] == 0
    assert by_name["Ders C002"]["participants"] == 10


def test_blank_required_duration_does_not_abort_the_import(tmp_path):
    """ST-FUNC-003 — a blank *required* numeric cell must degrade, not explode.

    Whether a missing ``duration`` should be rejected (row skipped + error) or
    defaulted to one hour is a product decision the register leaves open, so
    this asserts the part that is not negotiable: the import completes, the
    other rows survive, and the affected row is either usable or explained.

    Today neither happens — ``int(NaN)`` raises and the user gets an empty
    timetable with no dialog at all.
    """
    blank = {k: v for k, v in klass("C001").items() if k != "duration"}
    path = build_workbook(tmp_path / "blank_duration.xlsx",
                          classes=[blank, klass("C002")])

    # Must not raise: today this is `ValueError: cannot convert float NaN to integer`.
    ds = load_scheduler_data_from_excel(path)

    # The untouched row always survives.
    assert "Ders C002" in course_names(ds)

    by_name = {c["name"]: c for c in ds.state["classes"]}
    imported_sanely = ("Ders C001" in by_name
                       and isinstance(by_name["Ders C001"]["duration"], int)
                       and by_name["Ders C001"]["duration"] >= 1)
    explained = any(row_prefix("classes", 2) in line or "C001" in line
                    for line in messages(ds.report))
    assert imported_sanely or explained, (
        "the row with a blank duration was neither imported with a usable "
        f"duration nor reported; classes={course_names(ds)}, "
        f"report={messages(ds.report)}")


# ── Guards: schema and reference validation ─────────────────────────────────

def test_missing_required_column_is_rejected_with_a_useful_message(tmp_path):
    """Guard: a Classes sheet without ``duration`` fails loudly, not silently.

    A user who deleted a column must be told which one, not handed an empty
    timetable.
    """
    path = build_workbook(tmp_path / "no_duration.xlsx",
                          classes=[klass("C001")],
                          omit_columns={"classes": ("duration",)})
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid is False
    assert ds.state["classes"] == []
    joined = "\n".join(ds.report.errors)
    assert "duration" in joined
    assert schema.get_workbook_sheet_title("classes") in joined


def test_unknown_teacher_and_branch_ids_are_rejected_per_row(tmp_path):
    """Guard: a dangling ``teacher_id``/``branch_id`` kills only its own row.

    Typing ``T999`` in one row must cost the user that row and produce a
    pointed error, not corrupt the other rows.
    """
    rows = [
        klass("C001", teacher_id="T999"),
        klass("C002", branch_id="B999"),
        klass("C003"),
    ]
    path = build_workbook(tmp_path / "bad_refs.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid is False
    assert course_names(ds) == ["Ders C003"]
    joined = "\n".join(ds.report.errors)
    assert "T999" in joined
    assert "B999" in joined
    assert row_prefix("classes", 2) in joined   # C001 sits on Excel row 2
    assert row_prefix("classes", 3) in joined   # C002 sits on Excel row 3


def test_duplicate_class_ids_are_reported(tmp_path):
    """Guard: two rows claiming the same ``class_id`` must be flagged.

    Silently accepting them would make later edits ambiguous for the user.
    """
    rows = [klass("C001"), klass("C001", course_name="Kopya")]
    path = build_workbook(tmp_path / "dupe_ids.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid is False
    assert any("C001" in line for line in ds.report.errors)


def test_location_type_labels_roundtrip(tmp_path):
    """Guard: the translated ``location_type`` labels map back to stable keys.

    If this breaks, online and lecturer-office lessons silently become
    face-to-face and get assigned physical rooms they never needed.
    """
    rows = [
        klass("C001", location_type=get_location_label(LOCATION_FACE_TO_FACE)),
        klass("C002", location_type=get_location_label(LOCATION_ONLINE)),
        klass("C003", location_type=get_location_label(LOCATION_LECTURER_OFFICE)),
    ]
    path = build_workbook(tmp_path / "locations.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    by_name = {c["name"]: c["location_type"] for c in ds.state["classes"]}
    assert by_name["Ders C001"] == LOCATION_FACE_TO_FACE
    assert by_name["Ders C002"] == LOCATION_ONLINE
    assert by_name["Ders C003"] == LOCATION_LECTURER_OFFICE


def test_unknown_allowed_rooms_warn_and_are_filtered(tmp_path):
    """Guard: ``allowed_rooms`` naming a room that does not exist is dropped.

    Keeping a phantom room in ``required_classrooms`` would make the class
    unplaceable with no explanation the user could act on.
    """
    rows = [klass("C001", allowed_rooms="Oda 1, Hayalet Oda")]
    path = build_workbook(tmp_path / "rooms.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert ds.state["classes"][0]["required_classrooms"] == ["Oda 1"]
    assert any("Hayalet Oda" in line for line in ds.report.warnings)


# ── ST-FUNC-011 — the wrong workbook must not look like a success ───────────

def test_workbook_with_no_recognized_sheets_is_invalid(tmp_path):
    """ST-FUNC-011 — importing the wrong file must not look like a success.

    Pointing the importer at an unrelated spreadsheet produced four warnings,
    an empty state and ``is_valid=True``; ``_import_from_excel`` only warns on
    ``not report.is_valid``, so the user got the success dialog over nothing.
    """
    path = tmp_path / "unrelated.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Butce"
    wb.active["A1"] = "Kalem"
    wb.save(str(path))

    ds = load_scheduler_data_from_excel(str(path))

    assert ds.state["classes"] == []
    assert ds.report.is_valid is False, (
        "an unrecognized workbook was reported as a valid import; "
        f"report was: {ds.report.summary()}")


def test_a_workbook_with_one_recognized_sheet_still_imports(tmp_path):
    """Guard (ST-FUNC-011 discrimination): *some* sheets missing is not fatal.

    A school may keep its teacher roster in its own file, and the importer has
    always merged whatever sheets it finds. The ST-FUNC-011 fix must therefore
    fire on *zero* recognized sheets, not on any missing one — a fix that
    errors whenever a sheet is absent would make partial workbooks
    un-importable and would show up here.
    """
    path = build_workbook(tmp_path / "teachers_only.xlsx", classes=[],
                          sheets=("teachers",))
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.errors == [], ds.report.summary()
    assert ds.report.is_valid is True
    assert ds.state["lecturers"] == ["Ada Lovelace", "Bora Yildiz"]
    # The three absent sheets are still named, as warnings.
    assert len(ds.report.warnings) == 3, ds.report.summary()


# ── Every shipped language must be able to read its own template back ───────

TEMPLATE_LOCALES = sorted(TRANSLATIONS)


@pytest.fixture
def ui_language():
    """Switch the UI language for the duration of one test, then put it back.

    The session-wide ``_pinned_language`` fixture pins Turkish so that every
    other test in this file is deterministic; these tests deliberately leave
    that pin and must restore it however they end.
    """
    original = get_language()

    def _switch(lang):
        set_language(lang)
        assert get_language() == lang, f"{lang!r} is not a shipped locale"

    try:
        yield _switch
    finally:
        set_language(original)


@pytest.mark.parametrize("lang", TEMPLATE_LOCALES)
def test_the_generated_template_reimports_in_every_shipped_locale(
        tmp_path, lang, ui_language):
    """File ▸ Generate Template then File ▸ Import Excel, in all 22 languages.

    The template is written in whatever language the app is running in — sheet
    titles, column headers, the row-2 help text and the example rows are all
    translated — and it is the file the bulk-entry workflow hands users. So the
    workbook the app writes is a workbook the app must be able to read, and
    that had only ever been checked in Turkish (``conftest`` pins the suite to
    ``tr``). Three of the 22 languages could not read their own output:

    * ``es`` — the Spanish room sheet is *Aulas* and the Portuguese class sheet
      is *Aulas* too, and the flat name→sheet-id map let one overwrite the
      other, so a Spanish workbook came back with 0 classrooms, 0 classes and
      "Faltan columnas obligatorias" against its rooms sheet.
    * ``zh`` / ``ja`` — the row-2 help text was recognized by "longer than 20
      characters or contains a space", which no CJK sentence satisfies, so the
      help text was imported as a teacher, a classroom *and* a branch.

    Asserted as an exact round-trip rather than "no errors": the zh failure
    mode was extra data, not missing data, and a count-free assertion would
    have passed straight over three phantom entities.

    ST-FUNC-009 rides along here, and this is the only place the *shipped
    template* is checked. The template's first class asks for the lab *type* and
    lists no rooms, so the importer has to find the lab room through the Rooms
    sheet's ``room_type`` column. Matching the room's *name* instead — or
    matching by substring — is right in Turkish and English purely by accident:
    there the lab room is called "Laboratuvar A"/"Lab A" and typed
    "Laboratuvar"/"Lab", so the name contains the type. In ``nl`` the same room
    is called "Lab A" and typed "Practicum", and ``az`` disagrees the same way.

    Measured, on the substring-of-room-name mutation: this sweep goes red on
    ``nl`` and ``az`` while its own ``tr`` and ``en`` rows stay green. The
    hand-built fixture below also catches that mutation, so this is not the sole
    guard — but it is the only one that would notice the app shipping a template
    its own importer misreads, and the only one that reads more than one locale.
    """
    ui_language(lang)

    path = tmp_path / f"template_{lang}.xlsx"
    generate_excel_template(str(path))
    ds = load_scheduler_data_from_excel(str(path))

    assert ds.report.is_valid, ds.report.summary()
    assert ds.report.errors == []

    expected_teachers = [tr(f"template.workbook_example.teacher_{i}_name")
                         for i in (1, 2, 3)]
    expected_rooms = [tr(f"template.workbook_example.room_{i}_name")
                      for i in (1, 2, 3)]
    expected_branches = [tr(f"template.workbook_example.branch_{i}_name")
                         for i in (1, 2, 3)]

    assert ds.state["lecturers"] == expected_teachers
    assert ds.state["classrooms"] == expected_rooms
    assert sorted(b for names in ds.state.get("years", {}).values()
                  for b in names) == sorted(expected_branches)
    # The template ships five class rows, two of which share joint group J1.
    assert [e["class_id"] for e in ds.raw_classes] == [
        "C001", "C002", "C003", "C004", "C005"]
    assert len(ds.state["classes"]) == 4, (
        f"{lang}: expected the five template rows to merge into four classes, "
        f"got {[c['name'] for c in ds.state['classes']]}")

    # ST-FUNC-009 — the template's class 1 carries required_room_type=lab and a
    # blank allowed_rooms, and room 2 is the lab room in all 22 locales.
    by_name = {c["name"]: c for c in ds.state["classes"]}
    lab_class = by_name[tr("template.workbook_example.class_1_name")]
    assert lab_class["required_classrooms"] == [
        tr("template.workbook_example.room_2_name")], (
        f"{lang}: the template's lab class must be restricted to the room the "
        f"Rooms sheet types as {tr('template.workbook_example.room_type_lab')!r}, "
        f"got {lab_class['required_classrooms']}")

    # ...and that room must be able to seat it. Resolving the type was only
    # half the job: the assertion above passed while the template shipped C001
    # with 25 students and its only lab with 20 seats, so
    # `get_physical_room_candidates` returned [] and the app's own example
    # workbook contained a course that could never be placed by any means --
    # reported by the importer as `is_valid=True`, `warnings: []`. The
    # capacities and the head count are hard-coded integers in template.py, so
    # this was identical in all 22 locales and no single-locale test would have
    # been worth more; it is asserted here because this is the only test that
    # imports the shipped template at all.
    for cls in ds.state["classes"]:
        if not cls.get("required_classrooms"):
            continue
        assert get_physical_room_candidates(ds.state, cls), (
            f"{lang}: the shipped template resolves {cls['name']!r} "
            f"({cls['participants']} students) into "
            f"{cls['required_classrooms']} — capacities "
            f"{ds.state['classroom_capacities']} — leaving no room that can "
            f"seat it, so this example course can never be scheduled")


@pytest.mark.parametrize("workbook_lang", TEMPLATE_LOCALES)
def test_sheet_titles_resolve_per_workbook_not_by_the_current_ui_language(
        workbook_lang, ui_language):
    """A workbook's own four titles must resolve the same whoever opens it.

    *Aulas* means classrooms in Spanish and classes in Portuguese, so no flat
    name→sheet-id table can be right for both. The tempting one-line repair —
    let the active UI language's titles win — fixes Spanish by breaking the
    Portuguese workbook opened on a Spanish desktop, which works today. This
    test resolves each language's own title set under all 22 UI languages, so
    that repair fails here even though the round-trip test above would pass.
    """
    ui_language(workbook_lang)
    titles = [schema.get_workbook_sheet_title(s) for s in SHEET_IDS]
    assert len(set(titles)) == len(SHEET_IDS), (
        f"{workbook_lang} reuses one sheet title for two sheets: {titles}")
    expected = dict(zip(SHEET_IDS, titles))

    for ui_lang in TEMPLATE_LOCALES:
        ui_language(ui_lang)
        assert schema.resolve_workbook_sheet_ids(titles) == expected, (
            f"a {workbook_lang} workbook is misread with the UI in {ui_lang}")


# ── ST-FUNC-012 — two teachers must not fuse into one lecturer ──────────────

def test_duplicate_lecturer_names_are_not_silently_accepted(tmp_path):
    """ST-FUNC-012 — two teacher IDs sharing a name corrupt the lecturer list.

    ``state['lecturers']`` and ``state['lecturer_availability']`` are keyed by
    display name, so a duplicate name silently fuses two real teachers into one
    and overwrites the first one's availability. The defect is that this is
    both silent *and* duplicated, so either fix — reporting it or
    deduplicating — clears the pin.
    """
    teachers = [
        {"teacher_id": "T001", "name": "Ada Lovelace"},
        {"teacher_id": "T002", "name": "Ada Lovelace"},
    ]
    path = build_workbook(tmp_path / "dupe_names.xlsx", teachers=teachers,
                          classes=[klass("C001")])
    ds = load_scheduler_data_from_excel(path)

    lecturers = ds.state["lecturers"]
    reported = any("Ada Lovelace" in line for line in messages(ds.report))
    deduplicated = len(lecturers) == len(set(lecturers))
    assert reported or deduplicated, (
        f"duplicate lecturer name accepted silently; lecturers={lecturers}, "
        f"report={messages(ds.report)}")


def test_lecturer_names_differing_only_in_case_are_not_silently_split(tmp_path):
    """ST-FUNC-012 — the importer must use the app's own rule for "same teacher".

    ``SchedulingWorkflow.register_lecturer`` (ST-UI-020, Phase 6) matches typed
    lecturer names with ``casefold()``, so "ayşe yılmaz" is not a second
    teacher beside "Ayşe Yılmaz". The importer compared nothing at all, so a
    workbook with both spellings produced *two* lecturers for one person: the
    two classes came back carrying different lecturer strings, which the core
    compares with ``==``, so the same teacher could be booked twice in one
    hour — while the class form went on treating the two spellings as one.

    Asserted as a disjunction on purpose: collapsing the pair or reporting it
    are both honest, and the register leaves the choice open. What is not
    allowed is a silent split.
    """
    from scheduler_app.core.workflow import SchedulingWorkflow

    first, second = "Ayşe Yılmaz", "ayşe yılmaz"

    # Control: establish that the app really does read these as one teacher,
    # instead of assuming it. If this ever stops holding, the assertion below
    # is measuring the wrong rule and must be revisited, not relaxed.
    probe = {"lecturers": [first]}
    assert SchedulingWorkflow.register_lecturer(probe, second) == first
    assert probe["lecturers"] == [first]

    teachers = [
        {"teacher_id": "T001", "name": first},
        {"teacher_id": "T002", "name": second},
    ]
    path = build_workbook(tmp_path / "case_names.xlsx", teachers=teachers,
                          classes=[klass("C001"), klass("C002", teacher_id="T002")])
    ds = load_scheduler_data_from_excel(path)

    lecturers = ds.state["lecturers"]
    collapsed = len({n.casefold() for n in lecturers}) == len(lecturers)
    reported = any(first in line or second in line for line in messages(ds.report))
    assert collapsed or reported, (
        f"two spellings of one teacher were accepted silently; "
        f"lecturers={lecturers}, report={messages(ds.report)}")
    # And the user must not be told the import was clean while the roster
    # still holds the same teacher twice.
    assert collapsed or ds.report.is_valid is False, (
        f"a roster holding one teacher twice was reported as valid; "
        f"lecturers={lecturers}, report={messages(ds.report)}")


def test_the_dotted_i_refusal_names_both_spellings_the_rule_and_the_row(tmp_path):
    """The refusal above is the *only* thing the user gets, so it must explain.

    ``_import_from_excel`` throws the whole dataset away on ``not is_valid``
    (ui/app.py shows ``report.summary()`` in a QMessageBox and returns), so a
    Turkish school whose roster holds both *Sıla* and *Sila* loses its teachers,
    rooms, branches and classes on the strength of one sentence.

    That sentence used to be the generic ``errors.duplicate_values`` — the same
    key ``_check_duplicates`` uses for two identical ``teacher_id`` cells. For
    identical cells it explains itself. Here it rendered as

        [Öğretmenler] Öğretmen Adı için yinelenen değerler: Sıla Kaya, Sila Kaya

    with no row number: two names the user can see are *not* the same string,
    asserted to be duplicates, with no statement of the rule that merged them
    and no hint of the remedy. The natural reading is that the app is broken.

    Three things are pinned, all of which the generic key failed:
      * both spellings appear, so the user knows which two rows are meant;
      * the message is the dedicated key, not the generic duplicate one;
      * it carries the colliding row number, like every other error this
        function raises.
    """
    first, second = "Sıla Kaya", "Sila Kaya"
    teachers = [{"teacher_id": "T001", "name": first},
                {"teacher_id": "T002", "name": second}]
    path = build_workbook(tmp_path / "dotted_i.xlsx", teachers=teachers,
                          classes=[klass("C001")])
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid is False
    lines = [ln for ln in messages(ds.report) if second in ln]
    assert len(lines) == 1, (
        f"expected exactly one line about the folded pair, got "
        f"{messages(ds.report)}")
    line = lines[0]

    assert first in line, (
        f"the refusal never names the spelling that was already taken: {line}")

    generic = tr("errors.duplicate_values").format(
        id_col=schema.get_workbook_sheet_header_map("teachers")["name"],
        values=f"{first}, {second}")
    assert generic not in line, (
        "the refusal is still the generic duplicate-values message, which "
        "asserts that two visibly different names are the same string and "
        f"never says why: {line}")

    # Row 3 is the second teacher row: header, T001, T002.
    assert line.startswith(row_prefix("teachers", 3)), (
        f"the refusal carries no row number, so a 200-row roster gives the "
        f"user nowhere to look: {line}")


@pytest.mark.parametrize("first,second", [
    ("Ada Lovelace", "ada lovelace"),
    ("AYŞE YILMAZ", "Ayşe Yılmaz"),
    ("DILEK KAYA", "Dilek Kaya"),
    ("IRIS MURDOCH", "Iris Murdoch"),
    ("İLHAN DEMİR", "Ilhan Demir"),
    ("WILLIAM SMITH", "William Smith"),
])
def test_the_importer_and_the_class_form_agree_on_what_counts_as_one_teacher(
        tmp_path, first, second):
    """ST-FUNC-012 — one rule for "same teacher", not two.

    ``state['lecturers']`` is a list of display names and ``cls['lecturer']``
    holds one of those names, so whatever counts as the same teacher has to
    count the same way wherever a name is typed. The class form's rule lives in
    ``SchedulingWorkflow.register_lecturer``; the importer's lives in
    ``_process_teachers``. This asserts they return the same verdict rather
    than asserting either verdict, so it stays true whichever way the rule is
    later sharpened — and goes red the moment only one side is sharpened.

    That was not hypothetical, and it is now closed. ``casefold()`` misses the
    Turkish dotted/dotless I, so ``AYŞE YILMAZ`` and ``Ayşe Yılmaz`` used to be
    two teachers on both sides. Making just the importer *Turkish*-aware was
    proposed and measured here: a Turkish fold merges that pair, and also
    *splits* ``DILEK KAYA`` from ``Dilek Kaya`` and ``IRIS MURDOCH`` from
    ``Iris Murdoch``, which were already one person on both sides — an ASCII
    ``I`` folds to ``ı``. So that fold was not a strict improvement, it was a
    different set of mistakes, and a locale-dependent fold would make one
    roster merge or split according to a UI setting.

    ``scheduler_app.i18n.text_fold.fold_text`` is the rule this docstring asked
    for. It sends every dotted and dotless I to a plain ASCII ``i`` and changes
    nothing else, so it handles all four original rows plus the two added
    since: ``İLHAN DEMİR``/``Ilhan Demir``, which the old fold split, and
    ``WILLIAM SMITH``/``William Smith``, which the *Turkish* fold would have
    split. Both sides now call that one function — ``core/workflow.py`` and
    ``data_io/importer.py`` — so the divergence this test watches for can no
    longer be introduced one side at a time by accident; see
    ``tests/test_text_fold.py``
    ``test_one_fold_serves_the_day_keys_the_class_form_and_the_importer``.

    This still asserts AGREEMENT rather than either verdict, so it remains an
    honest anti-divergence guard whichever way the rule is later sharpened.
    """
    from scheduler_app.core.workflow import SchedulingWorkflow

    probe = {"lecturers": [first]}
    SchedulingWorkflow.register_lecturer(probe, second)
    class_form_says_one = len(probe["lecturers"]) == 1

    teachers = [{"teacher_id": "T001", "name": first},
                {"teacher_id": "T002", "name": second}]
    path = build_workbook(tmp_path / "agreement.xlsx", teachers=teachers,
                          classes=[klass("C001"), klass("C002", teacher_id="T002")])
    ds = load_scheduler_data_from_excel(path)
    lecturers = ds.state["lecturers"]
    importer_says_one = (
        len({n for n in lecturers}) == 1
        or any(first in line or second in line for line in messages(ds.report))
    )

    assert importer_says_one == class_form_says_one, (
        f"the importer and the class form disagree about {first!r} / "
        f"{second!r}: class form says one teacher = {class_form_says_one}, "
        f"importer says one teacher = {importer_says_one} "
        f"(lecturers={lecturers}, report={messages(ds.report)})")


def test_the_dotted_and_dotless_i_do_not_split_one_teacher():
    """Pins ST-FUNC-012 / ST-UI-020 — one teacher, not two.

    ``register_lecturer``'s own docstring has always promised that typing a
    name in a different casing does not create a second teacher. Measured
    before the fix, that promise was false for any Turkish name containing an
    I: with ``İlhan Demir`` already on the roster, typing ``ilhan demir`` into
    ``AddClassDialog``'s editable lecturer combo appended a SECOND entry.

    What that costs the user, all of it silent: lecturer availability is keyed
    on ``state['lecturer_availability']`` by the FIRST spelling, so the new
    teacher has no unavailable hours and the optimizer will happily schedule
    them at 08:00 on a day they told the app they cannot teach. The class is
    still drawn and still counted, so nothing looks wrong until a timetable
    goes out with a clash in it.

    A hard assertion on the returned spelling, not a disjunction: the existing
    spelling must win, because it is the one the availability record is keyed
    on.
    """
    from scheduler_app.core.workflow import SchedulingWorkflow

    probe = {"lecturers": ["İlhan Demir"]}

    assert SchedulingWorkflow.register_lecturer(probe, "ilhan demir") == \
        "İlhan Demir"
    assert probe["lecturers"] == ["İlhan Demir"], (
        "typing an existing teacher's name in another casing created a second "
        "teacher; their availability record is keyed on the first spelling and "
        "will never apply")

    # The other three casings a user or an Excel UPPER() produces.
    for typed in ("Ilhan Demir", "ILHAN DEMIR", "İLHAN DEMİR"):
        assert SchedulingWorkflow.register_lecturer(probe, typed) == "İlhan Demir"
    assert probe["lecturers"] == ["İlhan Demir"]


@pytest.mark.parametrize("label", [
    # Plain ASCII str.upper() of the shipped labels — what Excel's UPPER() on
    # an English-locale machine produces, and the four that measurably failed.
    "ÖĞRETIM ELEMANI",      # tr  labels.lecturer          'Öğretim Elemanı'
    "ÖĞRETMEN ADI",         # tr  import.columns...        'Öğretmen Adı'
    "ÖĞRETIM ELEMANLARI",   # tr  setup.lecturers          'Öğretim Elemanları'
    "MÜƏLLIM ADI",          # az  import.columns...        'Müəllim Adı'
    # And the Turkish-keyboard uppercase of the same label, dotted İ and all.
    "ÖĞRETİM ELEMANI",
])
@pytest.mark.parametrize("running_language", ["tr", "en"])
def test_a_turkish_roster_header_typed_in_capitals_still_names_the_roster(
        label, running_language, ui_language):
    """Pins ST-FUNC-011's guard against its own false negatives.

    ``is_lecturer_name_header`` exists so that Setup ▸ Lecturers ▸ Import Excel
    can tell a roster from a budget spreadsheet — before it, a sheet of
    ``Kalem``/``Tutar`` line items reported three lecturers imported and put
    them in the staff list. The guard folded headers with ``casefold()``, which
    means it answered False for a genuine Turkish roster whose header row is
    capitalised: measured, four real header labels were rejected. The guard
    added to keep budgets out was keeping real Turkish rosters out too, and the
    user is told only that no lecturers were found.

    Asserted under both a Turkish and an English UI because the function scans
    every shipped catalogue on purpose — a roster exported before a language
    change must still import after it — so its answer must not depend on the
    language the app happens to be running in.
    """
    ui_language(running_language)
    assert schema.is_lecturer_name_header(label), (
        f"{label!r} was not recognised as a staff-name header while the app "
        f"was running in {running_language!r}; a real Turkish roster with a "
        f"capitalised header row is rejected as if it were a budget sheet")


def test_a_turkish_workbook_whose_header_row_is_shouted_still_imports(tmp_path):
    """The other half of the capitalised-header story: File ▸ Import Excel.

    ``is_lecturer_name_header`` above covers Setup ▸ Lecturers ▸ Import Excel.
    The whole-workbook importer takes a different road — ``_read_sheet`` calls
    ``schema.canonicalize_workbook_columns``, which folds each header against
    ``get_workbook_sheet_reverse_header_map`` — and nothing exercised it with
    anything but the shipped casing, because ``build_workbook`` writes the
    header row from the very map the importer reverses.

    Under a bare ``.casefold()`` this is not a cosmetic miss. Turkish ``ADI``
    casefolds to ASCII ``adi`` while the shipped ``Adı`` casefolds to dotless
    ``adı``, so six headers across all four sheets stop resolving and the
    import is refused outright:

        [Öğretmenler] Zorunlu sütunlar eksik: name
        [Derslikler]  Zorunlu sütunlar eksik: name
        [Şubeler]     Zorunlu sütunlar eksik: name
        [Dersler]     Zorunlu sütunlar eksik: class_id

    A school that ran its header row through Excel's ``UPPER()``, or typed it
    on a Turkish keyboard, loses its entire roster, room list, branch list and
    class list from a workbook that looks perfectly correct on screen — and the
    reason (which of the two capital I's the keyboard emitted) is invisible.

    Asserted against the shipped-casing import rather than against literals, so
    the two spellings have to agree rather than merely both being non-empty.
    """
    rows = [klass("C001"), klass("C002")]
    plain = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "plain.xlsx", classes=rows))
    shouted = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "shouted.xlsx", classes=rows,
                       shout_headers=True))

    assert plain.report.is_valid, plain.report.summary()
    assert shouted.report.is_valid, (
        "a Turkish workbook whose header row was typed in capitals was "
        "refused:\n" + shouted.report.summary())

    assert shouted.state["lecturers"] == plain.state["lecturers"]
    assert shouted.state["classrooms"] == plain.state["classrooms"]
    assert course_names(shouted) == course_names(plain)
    assert [e["class_id"] for e in shouted.raw_classes] == \
        [e["class_id"] for e in plain.raw_classes]
    # student_count is one of the six headers that stops resolving, so a class
    # that imported with the default head count instead of its own would pass
    # every assertion above.
    assert [c["participants"] for c in shouted.state["classes"]] == \
        [c["participants"] for c in plain.state["classes"]]


# ── Room type constrains the class (ST-FUNC-009) ────────────────────────────

def test_required_room_type_constrains_the_class_to_matching_rooms(tmp_path):
    """ST-FUNC-009 — ``required_room_type`` must actually restrict rooms.

    The template tells the user to write "Laboratory" here; the value used to be
    read, validated as a known column, and thrown away, so a physics lab could
    be scheduled into a lecture hall. A failure here means the user filled in a
    column the app told them to fill in and the solver ignored it.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / "req_type.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    required = ds.state["classes"][0]["required_classrooms"]
    assert "Lab 1" in required
    assert "Oda 1" not in required


def test_required_room_type_narrows_allowed_rooms_rather_than_replacing_it(tmp_path):
    """ST-FUNC-009 — two room columns are two restrictions on one set.

    ``allowed_rooms`` narrows from all rooms and ``required_room_type`` narrows
    by type; neither claims precedence, so they compose by intersection. A
    failure means the user's hand-typed room list was either widened behind
    their back (a lab class becomes schedulable in a lecture hall again, which
    is ST-FUNC-009 itself) or silently replaced by rooms they never listed.

    The exact equality is what does the work: with Lab 2 in the fixture, union
    gives ``["Oda 1", "Lab 1", "Lab 2"]``, allowed_rooms-wins gives
    ``["Oda 1", "Lab 1"]`` and type-wins gives ``["Lab 1", "Lab 2"]``. Only
    intersection gives ``["Lab 1"]``.
    """
    rows = [klass("C001", allowed_rooms="Oda 1, Lab 1",
                  required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / "narrow.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert ds.state["classes"][0]["required_classrooms"] == ["Lab 1"]


def test_a_room_type_no_room_has_is_reported_and_never_widens_the_class(tmp_path):
    """ST-FUNC-009 — an unmatchable type is a warning, never an empty list.

    Both halves are one invariant, so they are one test. An empty
    ``required_classrooms`` means "any room" to ``get_physical_room_candidates``,
    so writing the empty match would take a class the user *restricted* and
    hand the solver every room in the building — the exact inversion of what
    the column says. A failure in half (a) means the user is never told their
    room type matched nothing; a failure in half (b) means a typo in one column
    silently deleted the rooms they typed in the other.
    """
    rows = [klass("C001", required_room_type="Atölye"),
            klass("C002", required_room_type="Atölye", allowed_rooms="Oda 1")]
    path = build_workbook(tmp_path / "unknown_type.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    # (a) A warning, not an error — an error refuses the whole workbook.
    assert ds.report.is_valid, ds.report.summary()
    assert ds.state["classes"][0]["required_classrooms"] == []
    reported = [m for m in messages(ds.report)
                if "Atölye" in m and row_prefix("classes", 2) in m]
    assert reported, messages(ds.report)

    # (b) The unmatchable type did not empty a list the user typed by hand.
    assert ds.state["classes"][1]["required_classrooms"] == ["Oda 1"]


@pytest.mark.parametrize("head_count,should_warn", [
    (25, True),    # every lab seats 20 — nowhere to go
    (20, False),   # exact fit; `room_fits_class` is cap >= participants
    (18, False),
    (0, False),    # head count not filled in; the importer defaults it to 0
])
def test_a_room_type_that_resolves_only_to_rooms_too_small_is_reported(
        tmp_path, head_count, should_warn):
    """ST-FUNC-009, third contradiction — the resolved list versus the head count.

    ``get_physical_room_candidates`` filters by ``required_classrooms`` first
    and by ``room_fits_class`` second, so a type resolving only to rooms that
    cannot seat the class collapses the candidate set to ``[]`` exactly as an
    unmatchable type or an all-excluded type does — and those two both warn.
    This one did not: both halves are in ``dataset.state`` at that moment
    (``_process_rooms`` writes ``classroom_capacities`` two sheets earlier) and
    neither was compared against the other.

    The app's own shipped template was such a row — 25 students, one 20-seat
    lab — and imported as ``is_valid=True`` with ``warnings: []``.

    Left to the solver this is not silent forever: measured on the template,
    the unplaced list said "İzin verilen hiçbir dersliğin yeterli kapasitesi
    yok", which is accurate and localized. But that sentence only arrives after
    the user has entered days and times and run a solve, mixed in with classes
    that failed for ordinary reasons, whereas this is a contradiction between
    two cells of one row that File ▸ Import Excel is holding at the time and
    promises to report.

    A warning that changes nothing, never a relaxation: widening back to "any
    room" is the inversion the test above exists to catch, and a school may
    genuinely be about to buy chairs.
    """
    rows = [klass("C001", student_count=head_count,
                  required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / f"small_{head_count}.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    # Whatever is reported, the constraint the user typed stands untouched.
    assert cls["required_classrooms"] == ["Lab 1", "Lab 2"]

    reported = [m for m in messages(ds.report)
                if ROOM_TYPE_LAB in m and row_prefix("classes", 2) in m]
    if not should_warn:
        assert not reported, (
            f"{head_count} students fit a 20-seat lab; nothing should be "
            f"reported, got {reported}")
        assert get_physical_room_candidates(ds.state, cls), (
            f"{head_count} students should still have a room to go to")
        return

    assert get_physical_room_candidates(ds.state, cls) == [], (
        "fixture drift: this row is supposed to be the unplaceable one")
    assert len(reported) == 1, messages(ds.report)
    # The user is told which room is the biggest and by how much it falls
    # short, so they can act without opening the Rooms sheet to compare.
    assert "20" in reported[0] and str(head_count) in reported[0], (
        f"the warning does not name both the capacity and the head count: "
        f"{reported[0]}")


def test_required_room_type_and_allowed_rooms_that_disagree_keep_the_named_rooms(
        tmp_path):
    """ST-FUNC-009 — a naive intersection re-creates the finding right here.

    Both columns are valid and their intersection is empty: the user allowed
    only ``Oda 1`` (a lecture room) and required the lab type. Writing the empty
    intersection would mean "any room", so the one class the user constrained
    twice would end up the *least* constrained class in the file. A failure here
    means exactly that.
    """
    rows = [klass("C001", allowed_rooms="Oda 1",
                  required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / "disagree.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert ds.state["classes"][0]["required_classrooms"] == ["Oda 1"]
    assert [m for m in messages(ds.report) if row_prefix("classes", 2) in m], \
        messages(ds.report)


@pytest.mark.parametrize("shouted,expected", [
    # Plain ASCII shouting: 'LABORATUVAR' has no I in it at all.
    (ROOM_TYPE_LAB.upper(), ["Lab 1", "Lab 2"]),
    # The one that matters in the shipped default language. A Turkish keyboard
    # with caps lock produces the dotted capital 'İ', and bare str.casefold()
    # turns it into 'i' + U+0307 COMBINING DOT ABOVE, which does *not* equal the
    # 'i' in 'Derslik'. Only fold_text collapses the pair.
    ("DERSLİK", ["Oda 1"]),
    # Python's own locale-free upper() gives the ASCII I instead, and that must
    # match the same room — the user cannot tell the two capitals apart.
    (ROOM_TYPE_LECTURE.upper(), ["Oda 1"]),
])
def test_required_room_type_matches_across_letter_case(tmp_path, shouted, expected):
    """ST-FUNC-009 — a shouted room type is the same room type.

    A user who types LABORATUVAR or DERSLİK in the Classes sheet and
    Laboratuvar / Derslik in the Rooms sheet means one thing. A failure here
    means their class quietly lost its room constraint over letter case alone —
    and for the Turkish rows, over which of the two capital I's their keyboard
    happens to emit, which no user can be expected to notice.

    Both sides go through ``scheduler_app.i18n.text_fold.fold_text``, the one
    case rule this app compares user-typed text with; ``tests/test_text_fold.py``
    owns the rule itself.
    """
    rows = [klass("C001", required_room_type=shouted)]
    path = build_workbook(tmp_path / "shouted.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert ds.state["classes"][0]["required_classrooms"] == expected


def test_the_room_type_fix_adds_no_new_state_or_class_field(tmp_path):
    """ST-FUNC-009 — the fix must not smuggle a field past the shape tests.

    ``required_room_type`` is resolved into the existing ``required_classrooms``
    precisely so that nothing new has to be taught to save/load, to the Edit
    Class dialog, or to the solver. A later "improvement" that stashes
    ``cls["required_room_type"]`` or ``state["classroom_types"]`` would reach
    production invisibly: ``tests/test_domain_shapes.py`` compares the TypedDicts
    to ``new_class``/``new_state``, so it cannot see a key the *importer* bolts
    on afterwards. For a user, that key is a constraint that survives import and
    then vanishes on the first save, with no message.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / "no_new_field.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert set(ds.state) == set(new_state())
    assert set(ds.state["classes"][0]) == set(new_class())


def test_class_id_containing_a_space_is_not_silently_dropped(tmp_path):
    """ST-FUNC-010 — ``_read_sheet`` mistook ``'C 001'`` for a help-text row.

    ``_read_sheet`` used to discard the first data row whenever its ID
    contained a space, so a school that writes class IDs like ``9 A`` lost its
    first course with no error at all. Only the *first* data row was affected,
    which is why the space is placed there. Closed by matching the row against
    the strings the template actually writes instead of guessing from shape —
    the same change that stopped the Chinese template's help text from being
    imported as data, because these are one defect seen from two sides.
    """
    rows = [klass("C 001"), klass("C002")]
    path = build_workbook(tmp_path / "spaced_id.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert [e["class_id"] for e in ds.raw_classes] == ["C 001", "C002"]
    assert course_names(ds) == ["Ders C 001", "Ders C002"]




def test_a_room_type_whose_every_room_is_excluded_does_not_strand_the_class(tmp_path):
    """ST-FUNC-009 — the type must never leave a class with nowhere to go.

    Phase 8 enforced "the type may only narrow" against ``allowed_rooms`` and
    forgot ``excluded_rooms``, which is written independently and subtracts
    from the same candidate set. Measured on the tree that shipped the fix:
    ``required_room_type='Laboratuvar'`` with ``excluded_rooms='Lab 1, Lab 2'``
    produced ``required_classrooms=['Lab 1','Lab 2']`` against
    ``excluded_classrooms=['Lab 1','Lab 2']``, so
    ``get_physical_room_candidates`` returned ``[]`` where 82f558e -- which
    discarded the column entirely -- returned ``['Oda 1']``.

    A failure means upgrading DERSİS turns a class the school had been
    timetabling for years into one that can never be placed, by the optimizer
    or by hand, with an empty import report and only a generic "no classroom
    matches" in the unplaced sidebar to go on.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  excluded_rooms="Lab 1, Lab 2")]
    path = build_workbook(tmp_path / "type_all_excluded.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    assert cls["required_classrooms"] == [], (
        "the type was applied anyway, leaving %r required against %r excluded"
        % (cls["required_classrooms"], cls["excluded_classrooms"]))

    from scheduler_app.core.models import get_physical_room_candidates
    assert get_physical_room_candidates(ds.state, cls) == ["Oda 1"], (
        "the class has no room left to be placed in")

    # The exact sentence, not just "a line mentioning the type". This row is
    # the one where "every room of type Laboratuvar is also in Excluded Rooms"
    # is *true* — no ``allowed_rooms`` narrowed anything, so the resolved list
    # is the whole type. The three-column wording exists for the case where it
    # is not true, and the two must not drift into each other.
    exact = tr("warnings.room_type_all_excluded").format(
        type=ROOM_TYPE_LAB,
        type_field=schema.get_workbook_sheet_header_map("classes")[
            "required_room_type"],
        rooms_field=schema.get_workbook_sheet_header_map("classes")[
            "excluded_rooms"])
    assert any(exact in m and row_prefix("classes", 2) in m
               for m in messages(ds.report)), (
        "the user was not told the room type was dropped: %r"
        % (messages(ds.report),))


def test_a_room_type_still_narrows_when_only_some_rooms_are_excluded(tmp_path):
    """ST-FUNC-009 guard — the excluded-rooms rescue must not over-fire.

    The fallback above only applies when *every* type-matching room is
    excluded. With one lab excluded and one surviving, the type must still do
    its job. A failure means the rescue swallowed a constraint the user gave.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  excluded_rooms="Lab 1")]
    path = build_workbook(tmp_path / "type_some_excluded.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    assert cls["required_classrooms"] == ["Lab 1", "Lab 2"], (
        "the surviving lab was lost: %r" % (cls["required_classrooms"],))

    from scheduler_app.core.models import get_physical_room_candidates
    assert get_physical_room_candidates(ds.state, cls) == ["Lab 2"]


# ── the room-type report must describe the row it is looking at ─────────────
# Three sentences were added to the import report by the ST-FUNC-009 work, and
# all three named ``required_room_type`` over whatever ``required_classrooms``
# happened to hold. That list is only the type-resolved one when the type
# actually decided it: when no room has the type, or when ``allowed_rooms`` and
# the type do not intersect, the list is still ``allowed_rooms`` and every
# clause about "the room type" is then about the wrong column. The behaviour
# was right in each case; only the sentence was wrong, so these tests assert
# the message AND pin the resolved list, which must not move.

def _classes_header(field):
    """The Classes column header as the user's own workbook spells it."""
    return schema.get_workbook_sheet_header_map("classes")[field]


def test_a_type_that_could_not_be_applied_is_not_described_over_allowed_rooms(
        tmp_path):
    """The capacity sentence must not speak for a type that was not applied.

    ``required_room_type='Derslik'`` with ``allowed_rooms='Lab 1'``: the two
    columns do not intersect, so the row keeps ``['Lab 1']`` and the report
    already says the type was not applied. Measured before the fix, a second
    line then said "No room of type Derslik seats 25 — the largest is Lab 1
    with 20. The room type was kept", of which every clause is false: Oda 1 is
    the Derslik room and seats 30, Lab 1 is not a Derslik room at all, the type
    was not kept, and "change the type" is offered as the remedy for a type
    that is not the problem. The cell the user must edit is Allowed Rooms.
    """
    rows = [klass("C001", student_count=25,
                  required_room_type=ROOM_TYPE_LECTURE,
                  allowed_rooms="Lab 1")]
    path = build_workbook(tmp_path / "type_not_applied.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    # Behaviour is unchanged: the named room still wins, as it did before.
    assert cls["required_classrooms"] == ["Lab 1"]

    capacity_line = tr("warnings.room_type_too_small").format(
        type=ROOM_TYPE_LECTURE,
        type_field=_classes_header("required_room_type"),
        participants=25,
        count_field=_classes_header("student_count"),
        room="Lab 1",
        capacity=20)
    assert not [m for m in messages(ds.report) if capacity_line in m], (
        "the head-count warning described the type over a list the type did "
        "not produce:\n  %s" % "\n  ".join(messages(ds.report)))
    # The one true line — "the type matched none of Allowed Rooms" — stays.
    assert [m for m in messages(ds.report) if row_prefix("classes", 2) in m], \
        "the row now reports nothing at all"


def test_an_unapplied_type_does_not_claim_its_rooms_were_all_excluded(tmp_path):
    """The rescue must not announce a fallback it did not perform.

    ``required_room_type='Derslik'``, ``allowed_rooms='Lab 1'``,
    ``excluded_rooms='Lab 1'``. The type never narrowed anything, so the
    "fallback" recomputed the list it already held — a pure no-op. Measured
    before the fix it still warned "Every room of type Derslik is also listed
    in Excluded Rooms, so the room type was not applied — otherwise this class
    could never be placed anywhere", while Oda 1, the only Derslik room, was
    not excluded, the preceding line had already said the type was not applied,
    and the class was still placeable nowhere afterwards. Both clauses false,
    and both checkable by the user against their own sheet.
    """
    rows = [klass("C001", student_count=25,
                  required_room_type=ROOM_TYPE_LECTURE,
                  allowed_rooms="Lab 1", excluded_rooms="Lab 1")]
    path = build_workbook(tmp_path / "noop_rescue.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    # The no-op stays a no-op: same list the pre-fix tree produced.
    assert cls["required_classrooms"] == ["Lab 1"]
    assert cls["excluded_classrooms"] == ["Lab 1"]

    rescued = tr("warnings.room_type_all_excluded").format(
        type=ROOM_TYPE_LECTURE,
        type_field=_classes_header("required_room_type"),
        rooms_field=_classes_header("excluded_rooms"))
    three_columns = tr("warnings.room_type_allowed_all_excluded").format(
        type=ROOM_TYPE_LECTURE,
        type_field=_classes_header("required_room_type"),
        allowed_field=_classes_header("allowed_rooms"),
        excluded_field=_classes_header("excluded_rooms"))
    reported = messages(ds.report)
    assert not [m for m in reported if rescued in m], (
        "the report claimed every %s room was excluded; Oda 1 is not:\n  %s"
        % (ROOM_TYPE_LECTURE, "\n  ".join(reported)))
    # Neither wording belongs here. Both end "the room type was not applied —
    # otherwise this class could never be placed anywhere", which promises a
    # rescue; nothing was rescued (the list is what it already was) and the
    # class is still placeable nowhere, which the row's other line already
    # said. A rescue sentence for a rescue that did not happen reads as the
    # app contradicting itself.
    assert not [m for m in reported if three_columns in m], (
        "a rescue was announced for a fallback that changed nothing:\n  %s"
        % "\n  ".join(reported))
    assert get_physical_room_candidates(ds.state, cls) == [], (
        "fixture drift: this row is supposed to be the one nothing rescued")


def test_the_all_excluded_sentence_names_allowed_rooms_when_it_did_the_narrowing(
        tmp_path):
    """The rescue's premise has to match the row it fired on.

    ``required_room_type='Laboratuvar'``, ``allowed_rooms='Oda 1, Lab 1'``,
    ``excluded_rooms='Lab 1'``. Here the type genuinely narrowed — to
    ``['Lab 1']`` — and that single survivor is excluded, so the fallback is
    correct and stays. But Lab 2 is a Laboratuvar and is not excluded, so
    "every room of type Laboratuvar is also in Excluded Rooms" is false; the
    empty set is the intersection of *three* columns. The sentence must name
    all three, and the other sentence stays reserved for the case where it is
    true (the test above this block).
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  allowed_rooms="Oda 1, Lab 1", excluded_rooms="Lab 1")]
    path = build_workbook(tmp_path / "three_columns.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    # Behaviour unchanged — the fallback still restores the allowed list.
    assert cls["required_classrooms"] == ["Oda 1", "Lab 1"]

    false_line = tr("warnings.room_type_all_excluded").format(
        type=ROOM_TYPE_LAB,
        type_field=_classes_header("required_room_type"),
        rooms_field=_classes_header("excluded_rooms"))
    true_line = tr("warnings.room_type_allowed_all_excluded").format(
        type=ROOM_TYPE_LAB,
        type_field=_classes_header("required_room_type"),
        allowed_field=_classes_header("allowed_rooms"),
        excluded_field=_classes_header("excluded_rooms"))
    reported = messages(ds.report)
    assert not [m for m in reported if false_line in m], (
        "Lab 2 is a %s and is not excluded, so this sentence is false:\n  %s"
        % (ROOM_TYPE_LAB, "\n  ".join(reported)))
    assert [m for m in reported if true_line in m], (
        "the three-column sentence was not reported:\n  %s"
        % "\n  ".join(reported))


@pytest.mark.parametrize("location_type", [LOCATION_ONLINE,
                                           LOCATION_LECTURER_OFFICE])
def test_a_virtual_class_is_not_warned_about_a_room_it_never_needed(
        tmp_path, location_type):
    """ST-ARCH-004 — no physical room is the right answer, not a problem.

    A school that fills Required Room Type on every row, its online lectures
    included, was told once per remote class that the class "cannot be placed
    until the room is enlarged, the head count lowered, or the type changed".
    Nothing is wrong: ``get_room_candidates`` returns ``[None]``, the virtual
    sentinel, the class schedules normally, and ``normalize_class_data``
    discards ``required_classrooms`` — the importer was warning about a list it
    was about to throw away. ``class_uses_physical_room`` is the predicate
    ``room_fits_class`` and ``get_physical_room_candidates`` short-circuit on,
    and the importer's copy of their arithmetic has to short-circuit with them.
    """
    rows = [klass("C001", student_count=25, required_room_type=ROOM_TYPE_LAB,
                  location_type=get_location_label(location_type))]
    path = build_workbook(tmp_path / f"virtual_{location_type}.xlsx",
                          classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    assert cls["location_type"] == location_type, "fixture drift: not virtual"
    assert not [m for m in messages(ds.report) if row_prefix("classes", 2) in m], (
        "a class needing no room was warned about one:\n  %s"
        % "\n  ".join(messages(ds.report)))

    # The control: the identical row face-to-face still warns, so the silence
    # above is about the location type and not about the check going missing.
    f2f = [klass("C001", student_count=25, required_room_type=ROOM_TYPE_LAB)]
    ds2 = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "virtual_control.xlsx", classes=f2f))
    assert [m for m in messages(ds2.report) if row_prefix("classes", 2) in m], \
        "the control row stopped warning; this test would pass vacuously"


# ── the warning must be true of the row it is attached to ───────────────────

_ROOMS_WITH_A_BIG_LAB = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30,
     "room_type": ROOM_TYPE_LECTURE},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    {"room_id": "R003", "name": "Lab 2", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    # The room that makes the head-count sentence false when the type is not
    # what narrowed the list: big enough for the class, and of the named type.
    {"room_id": "R004", "name": "Lab 3", "capacity": 50,
     "room_type": ROOM_TYPE_LAB},
]


def test_the_head_count_warning_is_not_raised_over_a_list_the_type_did_not_choose(tmp_path):
    """ST-FUNC-009 — a warning must be true of the row it is attached to.

    Found by the adversarial pass against Phase 8's own message fix. The gate
    was "the type touched this list", which is True whenever ``allowed_rooms``
    is a SUBSET of the type's rooms — the intersection narrows to the allowed
    list and the flag is set. The sentence then quantifies over the *type*
    while the seats it counted came from *Allowed Rooms*.

    Measured before this fix, with a fourth room Lab 3 (Laboratuvar, seats 50)
    and a row of ``required_room_type=Laboratuvar, allowed_rooms=Lab 1,
    student_count=25``: "No room of type Laboratuvar seats 25 - the largest is
    Lab 1 with 20." Every clause is false of the user's own Rooms sheet, and
    all three remedies it offers name the wrong cell: the fix is to add Lab 3
    to Allowed Rooms.

    A failure means the import report sends a school to enlarge a room or cut a
    class size when the room it needs already exists.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  allowed_rooms="Lab 1", student_count=25)]
    path = build_workbook(tmp_path / "narrowed_by_allowed.xlsx", classes=rows,
                          rooms=_ROOMS_WITH_A_BIG_LAB)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]
    assert cls["required_classrooms"] == ["Lab 1"], (
        "this fix is message-only; the resolved list must not move")
    assert not any(ROOM_TYPE_LAB in m for m in messages(ds.report)), (
        "the row was warned about the room type over a list Allowed Rooms "
        "alone chose: %r" % (messages(ds.report),))


def test_the_head_count_warning_still_fires_when_the_type_did_choose(tmp_path):
    """ST-FUNC-009 guard — the silence above must not swallow the real case.

    With no ``allowed_rooms`` the resolved list IS the type's rooms, so the
    sentence quantifies over exactly what it counted and is true. A failure
    means the fix for the false warning removed the true one with it.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LECTURE,
                  student_count=40)]
    path = build_workbook(tmp_path / "type_chose.xlsx", classes=rows,
                          rooms=_ROOMS_WITH_A_BIG_LAB)
    ds = load_scheduler_data_from_excel(path)

    assert any(ROOM_TYPE_LECTURE in m and row_prefix("classes", 2) in m
               for m in messages(ds.report)), (
        "a class that genuinely fits no room of its own type was not warned: "
        "%r" % (messages(ds.report),))


def test_a_rescue_that_rescued_nothing_is_not_announced_as_a_rescue(tmp_path):
    """ST-FUNC-009 — do not promise a rescue that did not happen.

    Both excluded-rooms sentences end "...so the room type was not applied —
    otherwise this class could never be placed anywhere". That is only true if
    dropping the type leaves somewhere to go. When every fallback room is also
    excluded the fallback is a no-op and the class is unplaceable either way.

    Measured before this fix on ``type=Laboratuvar, allowed_rooms='Lab 1',
    excluded_rooms='Lab 1'``: the row's ONLY line said the type had been
    dropped so the class could still be placed, while
    ``get_physical_room_candidates`` was ``[]``.

    A failure means the import report tells a school a class was rescued when
    it is silently unschedulable. (The row is now unwarned, which is a real and
    separate gap — the importer has the same blind spot for any class made
    unplaceable by Allowed Rooms alone — but a false sentence is worse than a
    missing one, and the gap is recorded in HANDOFF-PHASE9 §C.)
    """
    from scheduler_app.core.models import get_physical_room_candidates

    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  allowed_rooms="Lab 1", excluded_rooms="Lab 1")]
    path = build_workbook(tmp_path / "rescue_noop.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)
    cls = ds.state["classes"][0]

    assert get_physical_room_candidates(ds.state, cls) == [], (
        "premise of this test no longer holds: the class became placeable")
    assert not any("Excluded Rooms" in m or ROOM_TYPE_LAB in m
                   for m in messages(ds.report)), (
        "a rescue was announced for a class that is still unplaceable: %r"
        % (messages(ds.report),))


def test_a_rescue_that_did_happen_is_still_announced(tmp_path):
    """ST-FUNC-009 guard — the silence above must not swallow a true rescue.

    Here ``Oda 1`` survives the exclusion, so dropping the type really does
    leave the class somewhere to go and the sentence is true. A failure means
    the user stops being told why their room type was ignored.
    """
    rows = [klass("C001", required_room_type=ROOM_TYPE_LAB,
                  allowed_rooms="Oda 1, Lab 1", excluded_rooms="Lab 1")]
    path = build_workbook(tmp_path / "rescue_real.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert any(row_prefix("classes", 2) in m and ROOM_TYPE_LAB in m
               for m in messages(ds.report)), (
        "a genuine rescue was not reported: %r" % (messages(ds.report),))
