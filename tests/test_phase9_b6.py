"""B6 — an import report grows one warning per class row and is shown whole.

The defect
----------
``required_room_type`` (Classes sheet) and ``room_type`` (Rooms sheet) are both
optional. A school that fills the class column and names a type none of its
rooms carries gets **one warning per class row** from
``data_io/importer.py:504-514``, and ``DataValidationReport.summary()``
(``importer.py:78-88``) concatenates every one of them into a single string.
``SchedulerApp._import_from_excel`` (``ui/app.py:5271-5273``) then hands that
whole string to ``QMessageBox.information`` — no scroll area, no cap, no
"and 480 more".

At 500 classes the resulting dialog is tens of thousands of pixels tall with
its OK button at the bottom of it, far below the screen. It is still
dismissible from the keyboard, which is why this is Low and not High, but the
report itself is unreadable past the first few dozen lines and the mouse can
never reach the button.

What these tests pin
--------------------
The *observable* thing: the size of the dialog the app actually puts in front
of the user. Nothing here asserts which method built it, how many warnings the
report holds, or what the text says — a fix that adds a scroll area, one that
collapses the repeats into a single line with a count, and one that moves the
detail into ``setDetailedText`` all satisfy these tests, and a fix that only
reworded the warning satisfies none of them.

Two measurement notes, both load-bearing
----------------------------------------
1. ``sizeHint()`` is *not* what the user gets. ``QMessageBox`` clamps its own
   width — and therefore wraps its text, and therefore grows taller — inside
   its show handler. Measured here (real Windows platform, this machine's
   1536x912 available screen): sizeHint says 973x8106 for the 500-row report,
   while the dialog the user is actually handed is 500x24106. So each dialog
   is run through Qt's real show-time layout with
   ``WA_DontShowOnScreen``, which fires polish/showEvent/layout activation
   without ever creating a native window — no flash on a real desktop, real
   numbers under any platform plugin.
2. The absolute pixel count is platform-dependent (offscreen Qt fonts are
   fixed-pitch, so the offscreen figure is *larger* still), which is exactly
   why the assertion is "it fits on the screen it would open on" rather than a
   hard-coded pixel budget.

What turns these green (measured, both platforms)
-------------------------------------------------
Putting the report in a scrollable view passes everywhere (454x336 offscreen,
503x436 on the real platform, with all 500 lines still in it). Capping the
message instead also passes, but the budget is tighter than it looks: showing
the first 5 warnings and a "... and 495 more" line fits on both platforms,
showing the first 10 fits on the real platform and *not* offscreen (1216 px),
and showing 20 fits on neither. The offscreen platform's fixed-pitch font wraps
these ~170-character warnings about twice as often as a real font, so a fix
that only trims the list must be trimmed hard, or be given somewhere to scroll.

The sibling column
------------------
``allowed_rooms`` (``importer.py:477-481``) warns per row in the same way and
explodes at exactly the same rate — 500 rows, 500 warnings, one dialog. Its
lines are shorter, so it wraps less and the dialog is "only" ~8100 px tall on
the same machine. ``test_the_allowed_rooms_sibling_explodes_the_same_way``
holds that half, so a fix that bounds one dialog and not the other stays red.
"""
import os

import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

pytestmark = [pytest.mark.ui, pytest.mark.excel]


# The size of school in the finding. Big, but the app sells an Institutional
# tier with no class limit at all, so this is a customer, not a stress test.
CLASS_ROWS = 500

# Free text the importer matches against the Rooms sheet of the same workbook.
# No room below carries it, which is the whole point.
ROOM_TYPE_NOBODY_HAS = "Atolye"
ROOM_NOBODY_HAS = "Atolye 1"

_SHEET_IDS = ("teachers", "rooms", "branches", "classes")


# ── Workbook builder ────────────────────────────────────────────────────────
#
# Deliberately self-contained (sheet titles and headers still come from the
# schema, never hard-coded): this file must not go red because another module's
# fixtures moved.

def build_workbook(path, *, class_rows, required_room_type=None,
                   allowed_rooms=None):
    """Write a valid workbook whose Classes sheet has *class_rows* rows.

    Every class row gets the same ``required_room_type`` / ``allowed_rooms``
    cell, so the importer has the same complaint about every one of them.
    """
    from scheduler_app.data_io import schema

    rooms = [
        {"room_id": "R001", "name": "Oda 1", "capacity": 30,
         "room_type": "Derslik"},
        {"room_id": "R002", "name": "Lab 1", "capacity": 20,
         "room_type": "Laboratuvar"},
    ]
    classes = []
    for i in range(1, class_rows + 1):
        row = {
            "class_id": "C%04d" % i,
            "course_name": "Ders %04d" % i,
            "teacher_id": "T001",
            "branch_id": "B001",
            "duration": 1,
            "student_count": 10,
            # Unique per row: two classes sharing a joint group are merged.
            "joint_class_group": "UNIQ-%04d" % i,
        }
        if required_room_type is not None:
            row["required_room_type"] = required_room_type
        if allowed_rooms is not None:
            row["allowed_rooms"] = allowed_rooms
        classes.append(row)

    rows_by_sheet = {
        "teachers": [{"teacher_id": "T001", "name": "Ada Lovelace"}],
        "rooms": rooms,
        "branches": [{"branch_id": "B001", "name": "Grup A"}],
        "classes": classes,
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in _SHEET_IDS:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]]
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=headers[field])
        excel_row = 2
        for row in rows_by_sheet[sheet_id]:
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=excel_row, column=col, value=row[field])
            excel_row += 1

    wb.save(str(path))
    return str(path)


def import_report(path):
    """The report the importer produces for *path*, outside the UI."""
    from scheduler_app.data_io.importer import load_scheduler_data_from_excel

    return load_scheduler_data_from_excel(str(path)).report


# ── Recording what the app showed ───────────────────────────────────────────

class _Modal:
    """One modal the app put in front of the user during the import.

    Either a ``QMessageBox`` static (``widget is None``; the box Qt would have
    built is rebuilt identically for measurement) or a dialog the app
    constructed itself and ``exec()``-ed (``widget`` is that dialog).
    """

    def __init__(self, kind, title="", text="", parent=None, widget=None):
        self.kind = kind
        self.title = title
        self.text = text
        self.parent = parent
        self.widget = widget

    def readable_text(self):
        if self.widget is None:
            return self.text
        from PyQt6.QtWidgets import QLabel, QPlainTextEdit, QTextEdit

        parts = [c.text() for c in self.widget.findChildren(QLabel)]
        parts += [c.toPlainText() for c in self.widget.findChildren(QTextEdit)]
        parts += [c.toPlainText()
                  for c in self.widget.findChildren(QPlainTextEdit)]
        return "\n".join(p for p in parts if p)

    def build(self):
        """Return (widget, owned) — the thing the user was shown."""
        if self.widget is not None:
            return self.widget, False
        from PyQt6.QtWidgets import QMessageBox

        icon = getattr(QMessageBox.Icon, self.kind.capitalize(),
                       QMessageBox.Icon.NoIcon)
        box = QMessageBox(icon, self.title, self.text,
                          QMessageBox.StandardButton.Ok, self.parent)
        return box, True


class _Measurement:
    def __init__(self, modal, size, screen_height, button_bottom):
        self.modal = modal
        self.width = size.width()
        self.height = size.height()
        self.screen_height = screen_height
        self.button_bottom = button_bottom

    @property
    def fits(self):
        return self.height <= self.screen_height

    def __str__(self):
        where = ("bottom edge of its accept button at y=%d"
                 % self.button_bottom) if self.button_bottom is not None else \
                "no accept button found"
        return ("%s %dx%d px on a screen %d px tall (%.1fx too tall), %s"
                % (self.modal.kind, self.width, self.height,
                   self.screen_height,
                   self.height / max(self.screen_height, 1), where))


def measure(modal):
    """Lay *modal* out exactly as Qt does when it is shown, and measure it.

    ``WA_DontShowOnScreen`` runs polish, ``showEvent`` and layout activation —
    everything that decides the final geometry, including the width clamp that
    makes ``QMessageBox`` wrap and grow — without a native window ever
    appearing. See the module docstring for why ``sizeHint()`` will not do.
    """
    from PyQt6.QtCore import Qt
    from PyQt6.QtGui import QGuiApplication
    from PyQt6.QtWidgets import QApplication, QDialogButtonBox

    widget, owned = modal.build()
    widget.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, True)
    widget.show()
    QApplication.instance().processEvents()

    size = widget.size()
    screen = widget.screen() or QGuiApplication.primaryScreen()
    screen_height = screen.availableGeometry().height()

    button_bottom = None
    boxes = widget.findChildren(QDialogButtonBox)
    buttons = boxes[0].buttons() if boxes else []
    if buttons:
        button = buttons[0]
        button_bottom = button.mapTo(widget, button.rect().bottomLeft()).y()

    widget.hide()
    if owned:
        widget.deleteLater()
    return _Measurement(modal, size, screen_height, button_bottom)


def offenders(modals):
    """Every recorded modal that does not fit on the screen it would open on."""
    return [m for m in (measure(modal) for modal in modals) if not m.fits]


def explain(what, path, report, bad):
    """A failure message someone reading CI output can act on."""
    first = report.warnings[0] if report.warnings else "(none)"
    lines = [
        "The import report dialog does not fit on the screen.",
        "  workbook : %s" % os.path.basename(str(path)),
        "  scenario : %s" % what,
        "  report   : %d warnings, %d errors, %d characters of dialog text"
        % (len(report.warnings), len(report.errors), len(report.summary())),
        "  distinct : %d of the %d warnings differ only by row number"
        % (len(set(report.warnings)), len(report.warnings)),
        "  first    : %s" % first[:160],
        "",
        "Dialogs the app showed that do not fit:",
    ]
    lines += ["  - %s" % m for m in bad]
    lines += [
        "",
        "importer.py adds one warning per class row and app.py concatenates",
        "every one of them into a single un-scrollable message box, so the OK",
        "button lands thousands of pixels below the bottom of the screen and",
        "the report is unreadable past its first few dozen lines.",
    ]
    return "\n".join(lines)


# ── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def modals(monkeypatch):
    """Record — and never block on — every modal the import path can raise.

    The four ``QMessageBox`` statics are recorded as arguments (Qt builds the
    box inside C++, so it is rebuilt identically at measurement time), and
    ``exec()`` is intercepted on ``QMessageBox`` and ``QDialog`` as well, so a
    future fix that hands the report to a dialog of its own is measured as the
    real widget rather than slipping past unnoticed.
    """
    from PyQt6.QtWidgets import QDialog, QMessageBox, QWidget

    recorded = []

    def _static(kind, ret):
        def _call(*args, **kwargs):
            strings = [a for a in args if isinstance(a, str)]
            strings += [v for v in kwargs.values() if isinstance(v, str)]
            parent = args[0] if args and isinstance(args[0], QWidget) else None
            recorded.append(_Modal(
                kind,
                title=strings[0] if strings else "",
                text=strings[1] if len(strings) > 1 else "",
                parent=parent,
            ))
            return ret
        return _call

    for kind, ret in (("information", QMessageBox.StandardButton.Ok),
                      ("warning", QMessageBox.StandardButton.Ok),
                      ("critical", QMessageBox.StandardButton.Ok),
                      ("question", QMessageBox.StandardButton.Yes)):
        monkeypatch.setattr(QMessageBox, kind, staticmethod(_static(kind, ret)))

    def _exec(kind, ret):
        def _call(self, *args, **kwargs):
            recorded.append(_Modal(kind, title=self.windowTitle(), widget=self))
            return ret
        return _call

    monkeypatch.setattr(QMessageBox, "exec",
                        _exec("messagebox", QDialog.DialogCode.Accepted.value))
    monkeypatch.setattr(QDialog, "exec",
                        _exec("dialog", QDialog.DialogCode.Rejected.value))
    return recorded


_TIER_REGISTRIES = ("_gated_widgets", "_gated_actions", "_on_tier_changed",
                    "_export_submenu_refreshers")


@pytest.fixture
def window(qapp, dersis_home, monkeypatch):
    """A real, fully constructed ``SchedulerApp`` — never shown.

    Same isolation as ``tests/test_import_ui_flow.py``: the first-run
    controller is disarmed (it arms QTimers that outlive the test), the tier is
    pinned to Institutional by direct assignment so the entity-limit check in
    ``_import_from_excel`` cannot short-circuit a 500-class import, and the
    process-wide ``TierEnforcement`` registries are snapshotted and restored.
    """
    from scheduler_app.i18n.day_keys import DAY_KEYS
    from scheduler_app.plans import TIER_INSTITUTIONAL
    from scheduler_app.ui.first_run import FirstRunController
    from scheduler_app.ui.tier_enforcement import TierEnforcement

    monkeypatch.setattr(FirstRunController, "start", lambda self: None)

    enforcer = TierEnforcement.instance()
    prev_slug, prev_confirmed = enforcer._tier_slug, enforcer._tier_confirmed
    prev_registries = {name: list(getattr(enforcer, name))
                       for name in _TIER_REGISTRIES if hasattr(enforcer, name)}
    enforcer._tier_slug, enforcer._tier_confirmed = TIER_INSTITUTIONAL, True

    from scheduler_app.ui.app import SchedulerApp

    win = SchedulerApp()
    win.state_data["days"] = list(DAY_KEYS[:5])
    win.state_data["slots"] = ["09:00", "10:00", "11:00", "12:00"]
    try:
        yield win
    finally:
        win.close()
        win.deleteLater()
        qapp.processEvents()
        enforcer._tier_slug, enforcer._tier_confirmed = prev_slug, prev_confirmed
        for name, value in prev_registries.items():
            setattr(enforcer, name, value)


@pytest.fixture
def choose_file(monkeypatch):
    """Pin ``QFileDialog.getOpenFileName`` to one path."""
    from PyQt6.QtWidgets import QFileDialog

    def _choose(path):
        monkeypatch.setattr(
            QFileDialog, "getOpenFileName",
            staticmethod(lambda *a, **k: (str(path), "")),
        )

    return _choose


# ── The defect ──────────────────────────────────────────────────────────────

def test_a_500_row_room_type_report_fits_on_the_screen(
        window, modals, choose_file, tmp_path):
    """B6 — 500 classes naming an unknown room type must not build a dialog
    taller than the display it opens on.

    Every warning here is true and worth telling the user about; what is broken
    is that all 500 of them are rendered at once, in one un-scrollable box,
    whose OK button ends up below the bottom of every monitor ever made.
    """
    path = build_workbook(tmp_path / "unknown_room_type_500.xlsx",
                          class_rows=CLASS_ROWS,
                          required_room_type=ROOM_TYPE_NOBODY_HAS)

    # Premise check, not the assertion under test: if the importer ever stops
    # producing a warning per row this test is measuring nothing, and it should
    # say so rather than pass quietly.
    report = import_report(path)
    assert report.is_valid, (
        "the workbook must import successfully — the defect is on the "
        "success path, not the rejection path")
    assert len(report.warnings) >= CLASS_ROWS / 2, (
        "expected roughly one warning per class row, got %d for %d rows; "
        "this test's premise no longer holds"
        % (len(report.warnings), CLASS_ROWS))

    choose_file(path)
    window._import_from_excel()

    assert modals, "the import finished without telling the user anything"
    bad = offenders(modals)
    assert not bad, explain(
        "%d class rows requiring room type %r, which no room in the workbook "
        "has" % (CLASS_ROWS, ROOM_TYPE_NOBODY_HAS), path, report, bad)


def test_the_allowed_rooms_sibling_explodes_the_same_way(
        window, modals, choose_file, tmp_path):
    """The ``allowed_rooms`` column warns per row too, and must be bounded too.

    The per-row form of the room-type warning was chosen for parity with this
    sibling (``errors.unknown_rooms``, same function, ``importer.py:477-481``),
    and the parity is real: 500 rows produce 500 warnings on either column.
    Measured on the same machine, the sibling's dialog is ~8100 px tall against
    the room type's ~24100 (its lines are shorter, so it wraps less) — smaller
    number, same defect. Whatever bounds one dialog must bound both, and this
    test is what makes that true instead of merely advisable.
    """
    path = build_workbook(tmp_path / "unknown_allowed_rooms_500.xlsx",
                          class_rows=CLASS_ROWS,
                          allowed_rooms=ROOM_NOBODY_HAS)

    report = import_report(path)
    assert report.is_valid
    assert len(report.warnings) >= CLASS_ROWS / 2, (
        "expected roughly one warning per class row, got %d for %d rows; "
        "this test's premise no longer holds"
        % (len(report.warnings), CLASS_ROWS))

    choose_file(path)
    window._import_from_excel()

    assert modals, "the import finished without telling the user anything"
    bad = offenders(modals)
    assert not bad, explain(
        "%d class rows whose allowed_rooms names %r, a room the workbook does "
        "not define" % (CLASS_ROWS, ROOM_NOBODY_HAS), path, report, bad)


# ── Discrimination guard ────────────────────────────────────────────────────

def test_a_two_row_report_fits_and_proves_the_measurement_discriminates(
        window, modals, choose_file, tmp_path):
    """The same code path, the same measurement, two warnings instead of 500.

    Without this, a bug in the harness — measuring the wrong widget, or a
    screen height of zero — would make the two tests above red for a reason
    that has nothing to do with the defect. This one is green today and must
    stay green: it is what proves the failures above are caused by the number
    of warnings and by nothing else.
    """
    path = build_workbook(tmp_path / "unknown_room_type_2.xlsx",
                          class_rows=2,
                          required_room_type=ROOM_TYPE_NOBODY_HAS)

    report = import_report(path)
    assert len(report.warnings) == 2, (
        "expected exactly one warning per class row, got %d"
        % len(report.warnings))

    choose_file(path)
    window._import_from_excel()

    assert modals, "the import finished without telling the user anything"
    bad = offenders(modals)
    assert not bad, (
        "a two-warning import report already overflows the screen, so the "
        "measurement in this module is wrong, not the app:\n%s"
        % "\n".join("  - %s" % m for m in bad))
