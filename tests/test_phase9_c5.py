"""C5 / ST-ARCH-004 — room-type warnings fired at classes that need no room.

An online or lecturer-office lesson happens in no classroom.
``get_room_candidates`` answers it with the ``None`` sentinel, it schedules
normally, and ``normalize_class_data`` discards ``required_classrooms`` and
``excluded_classrooms`` outright — so every room-type sentence the importer
attaches to such a row is about a field that is thrown away before anything
reads it.

Phase 8 established both the principle and the mechanism for exactly this. The
head-count sentence (``warnings.room_type_too_small``) was gated on
``class_uses_physical_room`` — the same predicate ``room_fits_class`` and
``get_physical_room_candidates`` short-circuit on — and pinned by
``test_a_virtual_class_is_not_warned_about_a_room_it_never_needed``
(``tests/test_import_roundtrip.py``), which asserts that a virtual row draws no
report line at all. 935c84b then stated the rule in one sentence: *a warning
must be true of the row it is attached to*.

Four sentences in ``_process_classes`` were never brought under that gate, and
the existing test does not reach them because its one fixture row (a matchable
type, no ``allowed_rooms``, no ``excluded_rooms``) is precisely the shape none
of the four fire on. Measured on this tree, each still fires for both virtual
location types:

* ``warnings.unknown_room_type`` — "…so the class was not restricted by room
  type", of a class that was never going to be restricted by one.
* ``warnings.room_type_excludes_allowed_rooms`` — "…so only Allowed Rooms was
  applied". False: Allowed Rooms was not applied either. ``required_classrooms``
  is ``[]`` on the imported class.
* ``warnings.room_type_all_excluded`` and
  ``warnings.room_type_allowed_all_excluded`` — both end "…so the room type was
  not applied — otherwise this class could never be placed anywhere". False:
  the class is placeable either way; ``get_room_candidates`` returns ``[None]``.

The user-visible cost is the one ST-ARCH-004 already named. A school that fills
Required Room Type on every row, its remote lectures included, gets a report
line per online class telling it to go fix a room problem it does not have, and
the lines that matter are buried among them.

The builder below is a deliberate copy of the one in
``tests/test_import_roundtrip.py`` rather than an import of it: this module has
to stand on its own while other agents edit that file in parallel.
"""

import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

from scheduler_app.core.models import (  # noqa: E402
    LOCATION_LECTURER_OFFICE,
    LOCATION_ONLINE,
    get_location_label,
    get_room_candidates,
)
from scheduler_app.data_io import schema  # noqa: E402
from scheduler_app.data_io.importer import load_scheduler_data_from_excel  # noqa: E402
from scheduler_app.translations import tr  # noqa: E402

pytestmark = pytest.mark.excel

SHEET_IDS = ("teachers", "rooms", "branches", "classes")

# Free text the app never translates — the importer matches the Classes sheet
# against the Rooms sheet of the *same* workbook, so both sides are literals.
ROOM_TYPE_LECTURE = "Derslik"
ROOM_TYPE_LAB = "Laboratuvar"
ROOM_TYPE_ABSENT = "Atolye"  # no room in the workbook carries it

DEFAULT_TEACHERS = [{"teacher_id": "T001", "name": "Ada Lovelace"}]
# Two labs and one lecture hall: enough for "the type narrows", "the type and
# Allowed Rooms are disjoint" and "every matching room is excluded" to be three
# different rows.
DEFAULT_ROOMS = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30,
     "room_type": ROOM_TYPE_LECTURE},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    {"room_id": "R003", "name": "Lab 2", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
]
DEFAULT_BRANCHES = [{"branch_id": "B001", "name": "Grup A"}]


def build_workbook(path, *, classes):
    """Write a workbook whose sheet titles and headers come from the schema."""
    rows_by_sheet = {
        "teachers": DEFAULT_TEACHERS,
        "rooms": DEFAULT_ROOMS,
        "branches": DEFAULT_BRANCHES,
        "classes": list(classes),
    }
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in SHEET_IDS:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]]
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=headers[field])
        excel_row = 2
        for row in rows_by_sheet[sheet_id]:
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=excel_row, column=col, value=row[field])
            excel_row += 1
    wb.save(str(path))
    return str(path)


def klass(class_id, **overrides):
    """A minimal, fully valid Classes row; override any field by keyword."""
    row = {
        "class_id": class_id,
        "course_name": f"Ders {class_id}",
        "teacher_id": "T001",
        "branch_id": "B001",
        "duration": 1,
        "student_count": 10,
        "joint_class_group": f"UNIQ-{class_id}",
    }
    row.update(overrides)
    return row


def messages(report):
    """All human-readable report lines, errors and warnings together."""
    return list(report.errors) + list(report.warnings)


def row_prefix(sheet_id, excel_row):
    """The localized ``[Sheet] Row N:`` prefix the importer puts on row lines."""
    return tr("status.import_row_prefix").format(
        sheet=schema.get_workbook_sheet_title(sheet_id), row=excel_row)


# The four room-type sentences that are still ungated, each with the row shape
# that triggers it. ``id`` is the translation key so a failure names the
# sentence the user reads.
UNGATED_SENTENCES = [
    pytest.param({"required_room_type": ROOM_TYPE_ABSENT},
                 id="unknown_room_type"),
    pytest.param({"required_room_type": ROOM_TYPE_LECTURE,
                  "allowed_rooms": "Lab 1"},
                 id="room_type_excludes_allowed_rooms"),
    pytest.param({"required_room_type": ROOM_TYPE_LAB,
                  "excluded_rooms": "Lab 1, Lab 2"},
                 id="room_type_all_excluded"),
    pytest.param({"required_room_type": ROOM_TYPE_LAB,
                  "allowed_rooms": "Oda 1, Lab 1",
                  "excluded_rooms": "Lab 1"},
                 id="room_type_allowed_all_excluded"),
]


@pytest.mark.parametrize("location_type", [LOCATION_ONLINE,
                                           LOCATION_LECTURER_OFFICE])
@pytest.mark.parametrize("room_cells", UNGATED_SENTENCES)
def test_a_class_needing_no_room_is_not_warned_about_a_room_type(
        tmp_path, room_cells, location_type):
    """ST-ARCH-004 — the room-type report has to short-circuit where the room
    logic does.

    ``class_uses_physical_room`` is the single source of truth for room
    branching, and ``warnings.room_type_too_small`` was already brought under
    it in Phase 8. These four sentences were not, so a class that needs no
    classroom is still told what is wrong with the classrooms it will never
    use — and two of them say something outright false of the row: that
    Allowed Rooms "was applied" when the imported class holds
    ``required_classrooms == []``, and that the class "could never be placed
    anywhere" when ``get_room_candidates`` hands it the ``[None]`` sentinel and
    it places normally.
    """
    label = get_location_label(location_type)
    rows = [klass("C001", location_type=label, **room_cells)]
    path = build_workbook(tmp_path / f"virtual_{location_type}.xlsx",
                          classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    cls = ds.state["classes"][0]

    # Fixture guards: this really is a roomless class, it really did keep no
    # room constraint, and it really is placeable. Without these a failure
    # below could be blamed on the workbook rather than on the report.
    assert cls["location_type"] == location_type, "fixture drift: not virtual"
    assert cls["required_classrooms"] == [], (
        "fixture drift: the room list survived normalization: %r"
        % (cls["required_classrooms"],))
    assert get_room_candidates(ds.state, cls) == [None], (
        "fixture drift: this class is supposed to need no physical room")

    assert not [m for m in messages(ds.report) if row_prefix("classes", 2) in m], (
        "a class that needs no classroom was warned about one:\n  %s"
        % "\n  ".join(messages(ds.report)))


@pytest.mark.parametrize("room_cells", UNGATED_SENTENCES)
def test_the_same_room_cells_still_warn_a_face_to_face_class(
        tmp_path, room_cells):
    """The control, so the silence above is about the location type.

    Each of the four row shapes must still produce its line when the class does
    occupy a classroom. Without this, deleting the four warnings outright would
    turn the test above green while making the app worse.
    """
    rows = [klass("C001", **room_cells)]
    path = build_workbook(tmp_path / "control.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    assert [m for m in messages(ds.report) if row_prefix("classes", 2) in m], (
        "the face-to-face control stopped warning; the test above would pass "
        "vacuously:\n  %s" % "\n  ".join(messages(ds.report)))
