"""User text must survive the renderers that parse markup.

ST-UI-007 / ST-UI-008 (Medium) — "Escape HTML/CSV/PDF inputs (injection)".

Every field in this app is free text a school administrator typed: a class name,
a room code, a branch letter, a time-slot label. Three of the surfaces DERSİS
hands them to parse markup, and each parses a different dialect.

What the register got wrong, and what this module therefore pins
---------------------------------------------------------------
ST-UI-007's headline — the warning panel's ``setHtml`` — is **closed**. Phase 4
added ``html.escape`` in ``WarningLogPanel._line``. The live half is elsewhere.

ST-UI-008 blames the **class name** for the reportlab failure. The class name is
the one field Phase 4 already escaped. Measured with a payload reportlab parses
as a tag, across every (field × mode) combination of the PDF export:

    slot label   4 of 4 modes -> ValueError, NO FILE WRITTEN
    branch       2 of 4       -> ValueError (everything, group)
    year         2 of 4       -> ValueError (everything, group)
    room         1 of 4       -> ValueError (classroom)
    lecturer     1 of 4       -> ValueError (lecturer)
    class name   0 of 4       -> already escaped in Phase 4
                 ----
                 10 of 24 produced no file at all

Why this matters to a user: they press "PDF olarak dışa aktar", the app raises,
and **no file appears** — with nothing naming the cause. There is no partial
output and no message pointing at the room or branch that did it.

Two failure modes, and the realistic one is the quieter
-------------------------------------------------------
reportlab fails differently depending on whether it recognises the tag:

    "9-A <B> Subesi"  -> ValueError, whole export dies
    "<Vekil> Dersi"   -> builds, and the text is SILENTLY DROPPED  (' Dersi')
    "<TR> Sinifi"     -> builds, silently dropped                  (' Sinifi')
    "R&D Lab"         -> builds, and renders as "R&D; Lab"

The last one needs no angle bracket at all and is the most likely to occur in a
real school. Phase 4's note that "a bare & is tolerated" is true of the class
name it had just escaped; in an unescaped field a bare ``&`` corrupts.

``html.escape(..., quote=False)`` round-trips all four through
``Paragraph.getPlainText()``. The worry that reportlab needs its own escape
because "the dialect is not HTML" is a false alarm — it decodes ``&amp;``,
``&lt;``, ``&gt;``, ``&quot;`` and ``&#x27;`` correctly.
"""
import os

import pytest

from scheduler_app.core.text_safety import (
    escape_pdf_markup, csv_safe, qt_tooltip, escape_qt_rich,
)

reportlab = pytest.importorskip("reportlab", reason="reportlab not installed")

pytestmark = pytest.mark.pdf

# Payloads chosen for what they do, not for looking like an attack:
#   a known tag  -> ValueError, no file
#   an unknown tag -> silently dropped
#   a bare &     -> corrupted, no angle bracket involved
MARKUP_PAYLOADS = [
    "9-A <B> Subesi",
    "<Vekil> Dersi",
    "<TR> Sinifi",
    "R&D Lab",
    "Salon <A>",
    "Muzik & Sanat",
    "O'Brien Salonu",
]


def _plain_text(markup):
    """Render *markup* through reportlab and read back the text it kept."""
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    return Paragraph(markup, getSampleStyleSheet()["Normal"]).getPlainText()


@pytest.mark.parametrize("payload", MARKUP_PAYLOADS)
def test_escaped_user_text_reaches_the_pdf_unchanged(payload):
    """ST-UI-007 — what the user typed is what the PDF prints.

    A failure means a printed timetable is missing a room or a lecturer, or the
    export died with no file. Both are silent to the app.
    """
    assert _plain_text(escape_pdf_markup(payload)) == payload


def test_the_payloads_really_are_dangerous_unescaped():
    """ST-UI-007 — anti-vacuity: prove the escape is doing the work.

    Without this, every assertion above would still pass if reportlab happened
    to be harmless on these strings, and the module would pin nothing. At least
    one payload must raise and at least one must be silently corrupted, because
    those are the two distinct failure modes the fix addresses.
    """
    raised, corrupted = [], []
    for payload in MARKUP_PAYLOADS:
        try:
            out = _plain_text(payload)
        except Exception:
            raised.append(payload)
            continue
        if out != payload:
            corrupted.append(payload)

    assert raised, (
        "no payload raises unescaped — reportlab's behaviour changed and these "
        "tests no longer pin the ValueError/no-file mode"
    )
    assert corrupted, (
        "no payload is silently corrupted unescaped — these tests no longer "
        "pin the silent-drop mode, which is the likelier one in practice"
    )


@pytest.mark.parametrize("field", ["slot_label", "branch", "year", "room",
                                   "lecturer", "cls_name"])
@pytest.mark.parametrize("mode", ["everything", "classroom", "group",
                                  "lecturer"])
def test_every_pdf_mode_exports_a_file_whatever_the_school_called_things(
        field, mode, tmp_path, dersis_home):
    """ST-UI-007 — the export must produce a file for all 24 combinations.

    Measured before the fix: 10 of these 24 raised ``ValueError`` and wrote
    nothing. The class-name case (which the register blames) was never one of
    them — it is the field Phase 4 had already escaped.
    """
    from scheduler_app.core.models import new_state, new_class, mark_placed
    from scheduler_app.data_io.exporter import export_schedule

    payload = "<A>"          # a tag reportlab recognises
    year, branch, room = "Year-1", "A", "R001"
    lecturer, name, slot = "Lect-01", "Fizik", "09:00"

    state = new_state()
    state["days"] = ["monday"]
    state["slots"] = ["09:00", "10:00"]
    state["classrooms"] = ["R001"]
    state["years"] = {"Year-1": ["A"]}
    state["lecturers"] = ["Lect-01"]

    if field == "slot_label":
        state["slots"] = [payload, "10:00"]
        slot = payload
    elif field == "branch":
        state["years"] = {"Year-1": [payload]}
        branch = payload
    elif field == "year":
        state["years"] = {payload: ["A"]}
        year = payload
    elif field == "room":
        state["classrooms"] = [payload]
        room = payload
    elif field == "lecturer":
        state["lecturers"] = [payload]
        lecturer = payload
    elif field == "cls_name":
        name = payload

    cls = new_class()
    cls["name"] = name
    cls["lecturer"] = lecturer
    cls["class_code"] = "AAA111"
    cls["targets"] = [{"year": year, "branch": branch}]
    state["classes"] = [cls]
    mark_placed(cls, "monday", slot, room)

    out = tmp_path / ("%s_%s.pdf" % (field, mode))
    export_schedule(state, "pdf", str(out), mode=mode)

    assert out.exists(), (
        "PDF export (%s / %s) wrote no file for a %s containing %r"
        % (mode, field, field, payload)
    )
    assert out.stat().st_size > 0


# ── CSV / clipboard: export-only, so a quote prefix is safe here ────────

@pytest.mark.parametrize("value,expected_prefix", [
    ("=1+1", True),
    ("+1", True),
    ("-9A Matematik", True),
    ("@SUM(A1)", True),
    ("\t=cmd", True),
    ("9-A Matematik", False),
    ("Fizik", False),
    ("", False),
])
def test_csv_neutralises_exactly_the_formula_triggers(value, expected_prefix):
    """ST-UI-008 — a spreadsheet must not evaluate a class name.

    The exported .csv gets emailed to colleagues; a cell starting ``=`` is
    executed by their spreadsheet, not read.
    """
    out = csv_safe(value)
    assert out.startswith("'") is expected_prefix, (
        "csv_safe(%r) -> %r" % (value, out))
    assert out.lstrip("'") == value


def test_csv_prefix_is_only_ever_used_where_nothing_reads_it_back():
    """ST-UI-008 — the register's own recommendation is a data-corruption bug.

    "Prefix risky cells with an apostrophe" is right for CSV and **wrong** for
    XLSX, because DERSİS re-imports its own workbooks: measured, 7 of 8
    round-tripped values came back renamed, including an innocent
    ``-9A Matematik`` turning into ``'-9A Matematik``.

    So the guard is architectural: ``csv_safe`` must not be reachable from the
    Excel writers. This pins that it is not.
    """
    import inspect
    from scheduler_app.data_io import exporter

    src = inspect.getsource(exporter)
    for marker in ("def _rich_cell", "def _export_excel", "def _export_csv"):
        assert marker in src, "exporter.py no longer defines %r" % marker

    # The Excel writers only. `_export_csv` sits between `_export_excel` and
    # `_export_pdf`, so bounding this region at `_export_pdf` would sweep the
    # CSV writer in and fail on the one place `csv_safe` is *correct*.
    excel_region = src[src.index("def _rich_cell"):src.index("def _export_csv")]
    assert "csv_safe" not in excel_region, (
        "csv_safe reached an XLSX writer; a literal apostrophe there is "
        "re-imported as part of the name (see the module docstring)"
    )
    # ...and the workbook path must still be neutralised, by the other means.
    assert "neutralize_formula_cells" in src, (
        "the XLSX formula sweep is gone; exported workbooks can carry live "
        "formulas again"
    )


# ── Qt: the format switch, not an escape ───────────────────────────────

def test_qt_tooltip_is_deterministic_regardless_of_what_the_user_typed():
    """ST-UI-007 — a tooltip must not render differently per class name.

    Qt's AutoText guesses per string: ``<b>`` makes it rich text, ``<Vekil>``
    does not. So today the *same* tooltip template renders as markup on one
    lesson and literally on the next, decided by the user's own text.
    """
    for payload in MARKUP_PAYLOADS:
        out = qt_tooltip("Ders: %s\nDerslik: R001" % payload)
        assert out.startswith("<qt>") and out.endswith("</qt>")
        assert "<br>" in out
        # the payload's own markup must not survive as markup
        assert "<B>" not in out and "<A>" not in out


def test_escape_qt_rich_and_escape_pdf_markup_differ_on_quotes():
    """ST-UI-007 — the two escapes are deliberately not the same function.

    ``escape_pdf_markup`` passes ``quote=False`` so a printed lecturer name
    keeps its apostrophe instead of showing ``&#x27;``. Pinning the difference
    stops a future "these look identical, let's merge them" refactor.
    """
    assert escape_qt_rich("O'Brien") != escape_pdf_markup("O'Brien")
    assert escape_pdf_markup("O'Brien") == "O'Brien"


# ═══════════════════════════════════════════════════════════════════════
#  ST-UI-008 — a spreadsheet must not execute a class name
# ═══════════════════════════════════════════════════════════════════════
#
# The .xlsx and .csv exist to be emailed to colleagues, so a cell whose text
# begins with "=" is executed on someone else's machine.
#
# The register's own recommendation -- "prefix risky cells with an apostrophe"
# -- is right for the CSV and a DATA-CORRUPTION BUG for the workbook, because
# DERSİS re-imports its own workbooks through six symmetric export/import pairs.
# Measured across one export/re-import round trip with the value prefixed:
#
#     '=1+1'           -> "'=1+1"           RENAMED
#     '-9A Matematik'  -> "'-9A Matematik"  RENAMED   <- a real class name
#                                           5 of 8 values renamed
#
# Excel stores the same protection as a cell attribute, `quotePrefix`, which
# suppresses evaluation without touching the stored string: 0 of 8 renamed,
# 0 <f> elements in the saved file.
#
# Scope, corrected against the spec that proposed it: the apostrophe is safe in
# the .csv (zero csv.reader / read_csv call sites; the import filters offer only
# *.xlsx) and in the timetable-grid clipboard copy, but NOT in the three
# dialogs.py clipboard writers -- `_copy_rows`, the years copy and the shared
# table copy all round-trip through `_paste_rows` and the rooms/lecturers paste
# handlers, so a prefix there would be read back as part of the name.

# WHICH FIELD IS ACTUALLY AT RISK, measured -- and it is not the one the
# register names. A class NAME is written through `CellRichText`, which openpyxl
# never types as a formula; the TIME-SLOT LABEL is written as a plain string in
# column A of every sheet and does become one. Same shape as the PDF finding
# above: the class name is the field that is already safe.
#
#     name  = "=1+1"   ->  0 formula cells
#     room  = "=1+1"   ->  0 formula cells
#     slot  = "=1+1"   ->  4 formula cells  (Master Schedule!A2, T_L1!A2, ...)
#
# The first version of these tests put the payload in the name and passed with
# the neutralisation deleted.

FORMULA_FIELDS = ["slot", "name", "room", "lecturer"]


def _state_with_payload(field, payload):
    from scheduler_app.core.models import new_state, new_class, mark_placed
    name, slot, room, lecturer = "Fizik", "09:00", "R001", "L1"
    if field == "slot":
        slot = payload
    elif field == "name":
        name = payload
    elif field == "room":
        room = payload
    elif field == "lecturer":
        lecturer = payload

    state = new_state()
    state["days"] = ["monday"]
    state["slots"] = [slot]
    state["classrooms"] = [room]
    state["years"] = {"Year-1": ["A"]}
    state["lecturers"] = [lecturer]
    cls = new_class()
    cls["name"] = name
    cls["lecturer"] = lecturer
    cls["class_code"] = "AAA111"
    cls["targets"] = [{"year": "Year-1", "branch": "A"}]
    state["classes"] = [cls]
    mark_placed(cls, "monday", slot, room)
    return state


@pytest.mark.excel
@pytest.mark.parametrize("field", FORMULA_FIELDS)
def test_the_workbook_carries_no_formula_cells(field, tmp_path, dersis_home):
    """ST-UI-008 — no exported cell may be a formula.

    A failure means the colleague who receives the file opens it and their
    spreadsheet evaluates whatever the school called one of its hours.
    """
    openpyxl = pytest.importorskip("openpyxl")
    from scheduler_app.data_io.exporter import export_schedule

    out = tmp_path / ("sched_%s.xlsx" % field)
    export_schedule(_state_with_payload(field, "=1+1"), "xlsx", str(out))

    book = openpyxl.load_workbook(str(out))
    formulas = ["%s!%s" % (ws.title, cell.coordinate)
                for ws in book.worksheets
                for row in ws.iter_rows() for cell in row
                if cell.data_type == "f"]
    assert not formulas, "formula cells in the export: %r" % (formulas,)


@pytest.mark.excel
def test_the_slot_label_really_is_the_injectable_field(tmp_path, dersis_home):
    """ST-UI-008 — anti-vacuity: prove the sweep has something to sweep.

    If openpyxl stops typing a leading-``=`` string as a formula, or the writer
    stops putting the slot label in a plain cell, the parametrised test above
    becomes trivially true and pins nothing. This asserts the premise directly
    by neutralising a workbook and counting what changed.
    """
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook
    from scheduler_app.data_io.spreadsheet_safety import neutralize_formula_cells

    book = Workbook()
    sheet = book.active
    sheet.cell(row=1, column=1, value="=1+1")
    sheet.cell(row=2, column=1, value="Fizik")
    assert sheet.cell(row=1, column=1).data_type == "f", (
        "openpyxl no longer types a leading '=' as a formula; ST-UI-008's "
        "premise has changed and these tests need re-deriving")
    assert neutralize_formula_cells(book) == 1, (
        "the sweep did not neutralise the one formula cell present")
    assert sheet.cell(row=1, column=1).value == "=1+1", (
        "the sweep altered the stored string; it must set quotePrefix instead")
    assert sheet.cell(row=1, column=1).quotePrefix is True


@pytest.mark.excel
@pytest.mark.parametrize("payload", ["=1+1", "-9A Matematik", "Fizik"])
def test_neutralising_the_workbook_does_not_rename_anything(
        payload, tmp_path, dersis_home):
    """ST-UI-008 — the protection must not corrupt what it protects.

    This is the assertion the register's own recommendation fails. DERSİS
    re-imports its own workbooks, so a literal apostrophe written into the value
    comes back as part of the name — measured 5 of 8 values renamed, including
    the perfectly innocent ``-9A Matematik``. ``quotePrefix`` renames 0 of 8.
    """
    openpyxl = pytest.importorskip("openpyxl")
    from scheduler_app.data_io.exporter import export_schedule

    out = tmp_path / "sched.xlsx"
    export_schedule(_state_with_payload("slot", payload), "xlsx", str(out))

    book = openpyxl.load_workbook(str(out))
    found = [cell.value
             for ws in book.worksheets
             for row in ws.iter_rows() for cell in row
             if isinstance(cell.value, str) and payload in cell.value]
    assert found, "the payload %r does not appear in the export at all" % payload
    assert not any(v.startswith("'") for v in found), (
        "the export prefixed the stored string; re-import would rename it: %r"
        % (found[:3],))


def test_the_csv_neutralises_a_formula_name(tmp_path, dersis_home):
    """ST-UI-008 — the flat export is the one that gets emailed most."""
    import csv as _csv
    from scheduler_app.data_io.exporter import export_schedule

    out = tmp_path / "sched.csv"
    export_schedule(_state_with_payload("slot", "=1+1"), "csv", str(out))

    with open(str(out), encoding="utf-8", newline="") as fh:
        cells = [c for row in _csv.reader(fh) for c in row]
    assert not any(c.startswith("=") for c in cells), (
        "a CSV cell still begins with '=': %r"
        % ([c for c in cells if c.startswith("=")],))
    assert any(c == "'=1+1" for c in cells), (
        "the neutralised name is not in the file: %r" % (cells,))


def test_the_generated_template_carries_no_formula_cells(tmp_path, dersis_home):
    """ST-UI-008 — the blank template travels furthest of all.

    It is the file a user hands to colleagues to fill in, so it is the one most
    likely to be opened on a machine that is not theirs.
    """
    openpyxl = pytest.importorskip("openpyxl")
    from scheduler_app.data_io.template import generate_excel_template

    out = tmp_path / "template.xlsx"
    generate_excel_template(str(out))
    book = openpyxl.load_workbook(str(out))
    formulas = [cell.coordinate
                for ws in book.worksheets
                for row in ws.iter_rows() for cell in row
                if cell.data_type == "f"]
    assert not formulas, "formula cells in the blank template: %r" % (formulas,)
