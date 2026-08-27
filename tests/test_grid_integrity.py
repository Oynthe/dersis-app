"""Grid integrity — the timetable may only ever reference axes that exist.

Three findings from the stress-test audit are, structurally, **one defect seen
from three directions**: nothing in DERSİS ever intersects a class's own
constraint lists with the actual grid axes, and ``logic.slot_index`` resolves a
slot name with a bare ``state["slots"].index(name)``.

======================  ==================================================
ST-SCHED-003 (High)     Ghost-day / ghost-slot placements — ``allowed_days``
                        / ``allowed_times`` are never intersected with
                        ``state["days"]`` / ``state["slots"]``, so a class
                        allowed only on Saturday is *committed* onto
                        Saturday on a Mon–Fri grid.
ST-SCHED-004 (High)     An ``allowed_times`` value absent from the grid
                        reaches ``slot_index`` and raises an uncaught
                        ``ValueError: '20:00' is not in list``, killing the
                        whole reschedule.
ST-DATA-003  (High)     Removing one time slot *after* placement leaves
                        ``placed_time`` pointing at it, and 8 of 9
                        downstream operations then crash on the same bare
                        ``.index()``. PDF is the lone survivor — and it
                        silently drops the orphaned lesson.
======================  ==================================================

What "correct" means here (the decision this module encodes)
------------------------------------------------------------
The register offers two recommendations that pull in opposite directions:
*intersect with the grid* (ST-SCHED-003) and *drop/flag stale constraint values
during normalization* (ST-SCHED-004). Taken literally, "drop the stale value"
is **wrong** for an allow-list, because in ``models.filter_class_days`` an empty
``allowed_days`` means *no restriction*:

    days = cls.get("allowed_days") or list(all_days)

So normalizing ``allowed_days=["saturday"]`` down to ``[]`` on a Mon–Fri grid
would silently convert "only Saturday" into "any day at all" and place the
lesson on Monday — turning a loud bug into a quiet one.

This module therefore asserts the *intersection* semantics throughout:

* a constraint value that is not on the grid contributes **nothing** to the
  search space (it is not a ghost candidate, and it is not a crash);
* if the intersection is empty the class stays **unplaced and is reported** —
  through ``AutoPlaceResult.success is False`` or through
  ``RescheduleResult.unplaced``, never by being placed somewhere the user did
  not allow;
* a placement already stored against an axis the user has since deleted must be
  *survivable* by every downstream operation, and must not vanish without a
  word from the exports.

Conventions
-----------
* Every test here is **fail-now / pass-after**: these findings are all being
  fixed in Phase 1, so nothing is xfailed. Section 4 and the two "design guard"
  tests pass today and must keep passing.
* "Not silently dropped" accepts *either* legitimate fix — keep the data, or
  tell the user it was left out — mirroring ``test_export_smoke.py``. Pinning
  only one of the two would turn red against a correct implementation of the
  other.
* The optimizer is **not** deterministic in the schedule it produces
  (ST-SCHED-013), so nothing below asserts *where* a class lands, only that it
  lands somewhere that exists — or does not land at all.

Cross-module notes for whoever lands the fix
--------------------------------------------
``tests/test_export_smoke.py`` carries two ``xfail(strict=True)`` pins on the
same underlying defect (``test_pdf_does_not_silently_drop_offgrid_placements``
and ``test_offgrid_slot_does_not_crash_csv_and_xlsx``). ``strict=True`` means
they go **red the moment the fix works**. Deleting those two markers is part of
the fix, not a regression.

The same module also holds a *live, currently-green*
``test_csv_still_reports_offgrid_placements``, which asserts the CSV keeps a
placement whose **day** is off-grid. So ``occupied_slots_of`` must NOT start
dropping off-grid-day placements, and the CSV must not start dropping off-grid
*hours* either — see
``test_csv_does_not_silently_drop_a_class_orphaned_by_slot_removal`` below,
which holds the CSV to the same bar this module holds the PDF to.

The crash also *moves* in two places rather than disappearing. Guarding
``slots_fit`` exposes ``workflow.py:476``; guarding ``get_consecutive_slots``
exposes ``logic.py:228``. Both are covered here
(``test_validate_drop_...`` and ``test_axis_removal_does_not_crash_conflict_detection``).
"""
import base64
import csv
import re
import warnings
import zlib

import pytest

from _support.schedule_oracle import check_schedule
from scheduler_app.core import analytics
from scheduler_app.core.candidate_generator import CandidateGenerator
from scheduler_app.core.constraint_validator import ConstraintValidator
from scheduler_app.core.logic import (
    classroom_of, get_placed_classes, occupied_slots_of, slot_index, slots_fit,
    total_duration,
)
from scheduler_app.core.logic import find_conflicts as occupancy_conflicts
from scheduler_app.core.models import (
    effective_day, effective_time, mark_placed, new_class, new_state,
)
from scheduler_app.core.workflow import SchedulingWorkflow
from scheduler_app.data_io import exporter
from scheduler_app.translations import tr

# The grid every "ghost" fixture below lives on. Saturday and 20:00 are the
# two values deliberately kept *off* it.
GRID_DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday"]
GRID_SLOTS = ["09:00", "10:00", "11:00", "12:00"]
OFF_GRID_DAY = "saturday"
OFF_GRID_SLOT = "20:00"

PDF_MODES = ("everything", "classroom", "group", "lecturer")


# ══════════════════════════════════════════════════════════════════════════
#  Fixtures / builders — all hand-built and deterministic
# ══════════════════════════════════════════════════════════════════════════

def _grid_state():
    """A minimal Mon–Fri × 09:00–12:00 grid with two rooms and one lecturer."""
    state = new_state()
    state["days"] = list(GRID_DAYS)
    state["slots"] = list(GRID_SLOTS)
    state["classrooms"] = ["R1", "R2"]
    state["classroom_capacities"] = {"R1": 0, "R2": 0}
    state["lecturers"] = ["L1"]
    state["years"] = {"Year-1": ["A"]}
    assert OFF_GRID_DAY not in state["days"]
    assert OFF_GRID_SLOT not in state["slots"]
    return state


def _class(name="Hayalet", **overrides):
    cls = new_class()
    cls.update({
        "class_code": "G001",
        "name": name,
        "lecturer": "L1",
        "duration": 1,
        "participants": 0,
        "targets": [{"year": "Year-1", "branch": "A"}],
    })
    cls.update(overrides)
    return cls


def _state_with(**class_overrides):
    """``(state, cls)`` — one class on the grid carrying *class_overrides*."""
    state = _grid_state()
    cls = _class(**class_overrides)
    state["classes"].append(cls)
    return state, cls


def _placed_state(make_state, n_classes=8):
    """A small, fully-placed, conflict-free schedule built without the optimizer.

    Every class is duration 1 and gets its own ``(day, slot)`` cell, so the
    oracle sees a perfectly clean schedule — which makes any violation reported
    after an axis is removed unambiguously the removal's fault. The optimizer is
    never invoked, so this is fully deterministic (cf. ST-SCHED-013).
    """
    state = make_state(n_days=5, n_slots=4, n_rooms=2, n_lecturers=8,
                       n_years=2, n_classes=n_classes, density=0.0, seed=11,
                       max_duration=1, online_fraction=0.0)
    n_days = len(state["days"])
    for i, cls in enumerate(state["classes"]):
        cls["duration"] = 1
        cls["location_type"] = "face_to_face"
        cls["participants"] = 0
        cls["name"] = f"Ders{i + 1:02d}"
        mark_placed(cls, state["days"][i % n_days],
                    state["slots"][i // n_days], state["classrooms"][0])
    return state


#: ``kind -> (removal callable, predicate telling whether a class is orphaned)``
#: Each callable returns the removed value.
def _remove_slot(state):
    victim = state["slots"][1]
    state["slots"].remove(victim)
    return victim


def _remove_day(state):
    victim = state["days"][1]
    state["days"].remove(victim)
    return victim


def _remove_room(state):
    victim = state["classrooms"][0]
    state["classrooms"].remove(victim)
    state["classroom_capacities"].pop(victim, None)
    return victim


def _remove_lecturer(state):
    victim = state["classes"][0]["lecturer"]
    state["lecturers"].remove(victim)
    state["lecturer_availability"].pop(victim, None)
    return victim


def _remove_year(state):
    victim = state["classes"][0]["targets"][0]["year"]
    del state["years"][victim]
    return victim


REMOVALS = {
    "slot": (_remove_slot, lambda c, v: c["placed_time"] == v),
    "day": (_remove_day, lambda c, v: c["placed_day"] == v),
    "room": (_remove_room, lambda c, v: c["placed_classroom"] == v),
    "lecturer": (_remove_lecturer, lambda c, v: c["lecturer"] == v),
    "year": (_remove_year,
             lambda c, v: any(t["year"] == v for t in c.get("targets", []))),
}


def _on_grid_placed_names(state):
    """Names of placed lessons that still sit on a real ``(day, slot)`` cell.

    These are the *control* lessons: whatever an operation does about the
    orphans, it must never lose these, so asserting on them is what stops
    "it did not crash" from being satisfied by an operation that produced
    nothing at all.
    """
    return {c["name"] for c in get_placed_classes(state)
            if effective_day(c) in state["days"]
            and effective_time(c) in state["slots"]}


def _two_on_grid(state):
    """``(mover, blocker)`` — two distinct lessons still on real grid cells."""
    on_grid = [c for c in get_placed_classes(state)
               if effective_day(c) in state["days"]
               and effective_time(c) in state["slots"]]
    assert len(on_grid) >= 2, (
        "VACUOUS FIXTURE: fewer than two on-grid lessons left, so the control "
        "assertions below cannot distinguish 'no crash' from 'did nothing'.")
    return on_grid[0], on_grid[1]


def _orphaned_state(make_state, kind):
    """``(state, removed_value, orphans)`` after deleting one axis value.

    Refuses to return a state where the removal orphaned nothing — a vacuous
    fixture would make every "survives the removal" assertion below meaningless.
    """
    state = _placed_state(make_state)
    remove, is_orphan = REMOVALS[kind]
    victim = remove(state)
    orphans = [c["name"] for c in state["classes"]
               if c["placed"] and is_orphan(c, victim)]
    assert orphans, (
        f"VACUOUS FIXTURE: removing {kind}={victim!r} orphaned no placed class, "
        "so nothing downstream is actually being exercised.")
    return state, victim, orphans


# ── PDF introspection ───────────────────────────────────────────────────────
#
# No PDF parser is installed in the audit venv; reportlab writes content streams
# through /Filter [/ASCII85Decode /FlateDecode], so undoing both yields the page
# operators, in which drawn text appears as literal (...) strings.

_STREAM_RE = re.compile(rb"stream\r?\n(.*?)endstream", re.DOTALL)


def _decode_pdf_stream(body):
    data = body.strip()
    if data.endswith(b"~>"):
        try:
            data = base64.a85decode(data, adobe=True)
        except ValueError:
            return body
    try:
        return zlib.decompress(data)
    except zlib.error:
        return data


def _pdf_content_text(raw):
    return b"\n".join(_decode_pdf_stream(b) for b in _STREAM_RE.findall(raw))


def _read_csv_rows(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.reader(fh))


# ══════════════════════════════════════════════════════════════════════════
#  0. Fixture sanity — none of the assertions below may be vacuous
# ══════════════════════════════════════════════════════════════════════════

def test_the_placed_fixture_is_clean_and_fully_populated(make_state):
    """ST-DATA-003 — proves the axis-removal fixture starts from a clean grid.

    A failure means the "removing an axis breaks X" tests are measuring a
    schedule that was already broken, so their verdicts say nothing about the
    removal.
    """
    state = _placed_state(make_state)
    placed = get_placed_classes(state)
    assert len(placed) == len(state["classes"]) == 8, \
        "fixture did not place every class"

    audit = check_schedule(state)
    assert audit["counts"] == {}, \
        f"fixture schedule is not clean before any removal: {audit['counts']}"
    assert audit["n_placed"] == 8


# ══════════════════════════════════════════════════════════════════════════
#  1. ST-SCHED-003 — a day/slot that is not on the grid is not a candidate
# ══════════════════════════════════════════════════════════════════════════

def test_search_space_days_are_a_subset_of_the_grid():
    """ST-SCHED-003 — the day search space must never leave ``state['days']``.

    A failure means the scheduler is free to consider a weekday the school does
    not teach on, which is how lessons end up parked on a day the timetable
    cannot even render.
    """
    state, cls = _state_with(allowed_days=[OFF_GRID_DAY])
    days, times, rooms = CandidateGenerator(state).get_search_space(cls)

    assert set(days) <= set(state["days"]), (
        f"search space offers days outside the grid: "
        f"{sorted(set(days) - set(state['days']))}")
    # The intersection is empty here, which is the whole point: there is no
    # legal day left, so the class must end up unplaced (section 4), not parked.
    assert days == []


def test_candidate_generator_emits_no_off_grid_day():
    """ST-SCHED-003 — ``generate()`` must not return a ghost-day candidate.

    A failure means every consumer of candidate generation (auto-place, batch
    scheduling, the optimizer) is being handed placements onto a day that does
    not exist.
    """
    state, cls = _state_with(allowed_days=[OFF_GRID_DAY])
    candidates = CandidateGenerator(state).generate(cls)

    ghosts = [c for c in candidates if c[0] not in state["days"]]
    assert not ghosts, (
        f"{len(ghosts)} of {len(candidates)} candidates are on days outside "
        f"the grid, e.g. {ghosts[:3]}")


def test_validator_rejects_a_placement_on_a_day_outside_the_grid():
    """ST-SCHED-003 — ``check_placement`` must not bless an off-grid day.

    A failure means the last line of defence before a commit approves a lesson
    on a day the user removed (or never had), so nothing downstream can catch it.
    """
    state, cls = _state_with(allowed_days=[OFF_GRID_DAY])
    validator = ConstraintValidator(state)

    assert validator.check_placement(cls, OFF_GRID_DAY, "09:00", "R1") is False, \
        "check_placement approved a day that is not in state['days']"

    # Control: the same call on a real grid day is still approved, so the
    # assertion above cannot pass by the validator having become a no-op.
    assert validator.check_placement(_class(allowed_days=[]),
                                     "monday", "09:00", "R1") is True


def test_auto_place_never_commits_a_class_onto_a_ghost_day():
    """ST-SCHED-003 — auto-place must refuse rather than invent a Saturday.

    A failure means clicking "yerleştir" on a Saturday-only class silently
    writes ``placed_day='saturday'`` into a Mon–Fri timetable: the lesson
    disappears from every view while still counting as placed.
    """
    state, cls = _state_with(allowed_days=[OFF_GRID_DAY])
    result = SchedulingWorkflow(state, lambda: {}).auto_place(cls)

    if result.success:
        day, slot, _room = result.placed_info
        pytest.fail(
            f"auto_place committed the class to {day}/{slot}; "
            f"grid days are {state['days']}")
    assert cls["placed"] is False
    assert cls["placed_day"] is None

    audit = check_schedule(state)
    assert audit["counts"].get("off_grid_day", 0) == 0, \
        f"oracle still reports off-grid placements: {audit['counts']}"


@pytest.mark.engine
def test_reschedule_never_proposes_or_commits_a_ghost_day():
    """ST-SCHED-003 — a full reschedule must stay inside the grid end to end.

    A failure means "Yeniden planla" proposes a Saturday slot that the commit
    step then throws away without telling anyone, so the user's class quietly
    ends up unplaced with no explanation.
    """
    state, cls = _state_with(allowed_days=[OFF_GRID_DAY])
    # An unconstrained companion, so "no off-grid proposals" cannot be
    # satisfied by an optimizer run that proposed nothing at all.
    healthy = _class(name="Saglikli", class_code="G002")
    state["classes"].append(healthy)

    workflow = SchedulingWorkflow(state, lambda: {})
    result = workflow.reschedule({}, use_cpsat=False)

    proposed = {c.get("name") for c, _d, _s, _r in result.placed}
    assert healthy["name"] in proposed, \
        "the reschedule proposed nothing — the assertions below are vacuous"

    off_grid = [(c.get("name"), d, s) for c, d, s, _r in result.placed
                if d not in state["days"] or s not in state["slots"]]
    assert not off_grid, f"reschedule proposed off-grid placements: {off_grid}"

    rejected = workflow.apply_reschedule(result)
    assert rejected == [], (
        "apply_reschedule had to discard proposals to keep the grid intact: "
        f"{rejected} — the proposal should never have left the grid")
    assert healthy["placed"] is True

    audit = check_schedule(state)
    assert audit["counts"].get("off_grid_day", 0) == 0
    assert audit["counts"].get("off_grid_slot", 0) == 0


# ══════════════════════════════════════════════════════════════════════════
#  2. ST-SCHED-004 — an out-of-grid slot name must not raise
# ══════════════════════════════════════════════════════════════════════════

def test_search_space_survives_an_allowed_time_outside_the_grid():
    """ST-SCHED-004 — a stale ``allowed_times`` value must not raise.

    A failure means ``ValueError: '20:00' is not in list``: one class left
    pointing at a lesson hour the user has since deleted makes the whole
    scheduler unusable.
    """
    state, cls = _state_with(allowed_times=[OFF_GRID_SLOT])
    days, times, rooms = CandidateGenerator(state).get_search_space(cls)

    assert set(times) <= set(state["slots"]), (
        f"search space offers slots outside the grid: "
        f"{sorted(set(times) - set(state['slots']))}")
    assert times == []
    assert CandidateGenerator(state).generate(cls) == []


def test_slots_fit_reports_false_for_a_slot_outside_the_grid():
    """ST-SCHED-004 — ``slots_fit`` must answer, not explode.

    A failure means every caller that asks "does this lesson fit from here?"
    about a stale hour crashes instead of getting a plain "no".
    """
    state = _grid_state()
    assert slots_fit(state, OFF_GRID_SLOT, 1) is False, \
        "an hour that is not on the grid cannot fit anything"
    # Control: a real slot still fits, and a real slot that overflows still does not.
    assert slots_fit(state, "09:00", 1) is True
    assert slots_fit(state, "12:00", 2) is False


def test_check_placement_rejects_a_slot_outside_the_grid():
    """ST-SCHED-004 — the validator must reject a stale hour, not raise.

    A failure means validating a class whose constraint (or placement) names a
    deleted hour aborts the caller with an uncaught ``ValueError``.
    """
    state, cls = _state_with(allowed_times=[OFF_GRID_SLOT])
    validator = ConstraintValidator(state)

    assert validator.check_placement(cls, "monday", OFF_GRID_SLOT, "R1") is False

    valid, reasons = validator.check_placement_explained(
        cls, "monday", OFF_GRID_SLOT, "R1")
    assert valid is False
    assert reasons, "the explained form rejected the slot without saying why"

    # Control: a real hour is still approved, so the assertions above cannot
    # pass by the validator having become a blanket "no".
    unconstrained = _class(name="Kontrol", class_code="G009")
    assert validator.check_placement(
        unconstrained, "monday", "09:00", "R1") is True
    assert validator.check_placement_explained(
        unconstrained, "monday", "09:00", "R1")[0] is True


@pytest.mark.engine
def test_reschedule_completes_with_an_allowed_time_outside_the_grid():
    """ST-SCHED-004 — one stale hour must not kill a whole reschedule.

    A failure means a single class carrying a deleted lesson hour makes
    "Yeniden planla" raise, so the user cannot re-plan *anything* until they
    find and hand-edit the offending class.
    """
    state, cls = _state_with(allowed_times=[OFF_GRID_SLOT])
    # A second, unconstrained class: the reschedule must still do real work.
    healthy = _class(name="Saglikli", class_code="G002")
    state["classes"].append(healthy)

    workflow = SchedulingWorkflow(state, lambda: {})
    result = workflow.reschedule({}, use_cpsat=False)
    workflow.apply_reschedule(result)

    assert healthy["placed"] is True, \
        "the reschedule placed nothing — 'it completed' would be vacuous"
    assert cls["placed"] is False, (
        f"the class was placed at {cls['placed_day']}/{cls['placed_time']} "
        f"even though its only allowed hour ({OFF_GRID_SLOT}) is not on the grid")

    audit = check_schedule(state)
    assert audit["counts"].get("off_grid_slot", 0) == 0
    assert audit["counts"].get("off_grid_day", 0) == 0


def test_auto_place_completes_with_an_allowed_time_outside_the_grid():
    """ST-SCHED-004 — auto-placing a class with a stale hour must not raise.

    A failure means the user clicks auto-place on one stale class and the app
    surfaces a raw ``ValueError`` instead of "bu ders yerleştirilemedi".
    """
    state, cls = _state_with(allowed_times=[OFF_GRID_SLOT])
    result = SchedulingWorkflow(state, lambda: {}).auto_place(cls)

    assert result.success is False
    assert cls["placed"] is False


def test_validate_drop_reports_an_out_of_grid_slot_instead_of_crashing():
    """ST-SCHED-004 — the drag-drop validator must reject a stale hour cleanly.

    A failure means dropping a lesson onto a cell whose hour is no longer on the
    grid raises out of the Qt drop handler. Note this path holds a *second* bare
    ``state['slots'].index(slot)`` (``workflow.py:476``) that only becomes
    reachable once ``slots_fit`` stops raising — fixing ``slot_index`` alone
    moves the crash rather than removing it.
    """
    state, cls = _state_with()
    validation = SchedulingWorkflow.validate_drop(
        state, cls, "monday", OFF_GRID_SLOT)

    assert validation.valid is False
    assert validation.reasons, "rejected the drop without a reason for the user"
    # Control: a legal drop is still accepted.
    assert SchedulingWorkflow.validate_drop(
        state, cls, "monday", "09:00").valid is True


def test_slot_index_never_reports_a_usable_index_for_an_unknown_slot():
    """ST-SCHED-004 (design guard) — a guarded ``slot_index`` must not return -1.

    Passes today (the bare ``.index()`` raises). It exists to stop the tempting
    "return -1 on miss" fix: -1 is a *valid* Python index, so it would silently
    place lessons in the last hour of the day instead of crashing. A failure
    means an unknown slot now resolves to a real grid position.
    """
    state = _grid_state()
    try:
        idx = slot_index(state, OFF_GRID_SLOT)
    except (LookupError, ValueError):
        return  # raising a lookup error is a legitimate design choice
    n = len(state["slots"])
    assert idx is None or not isinstance(idx, int) or not (-n <= idx < n), (
        f"slot_index({OFF_GRID_SLOT!r}) returned {idx!r}, which indexes "
        f"state['slots'][{idx}] = {state['slots'][idx]!r} — a silently wrong "
        "hour is worse than the ValueError it replaced")


# ══════════════════════════════════════════════════════════════════════════
#  3. ST-DATA-003 — removing an axis after placement
# ══════════════════════════════════════════════════════════════════════════

def _analytics_op(state, tmp_path):
    metrics = analytics.compute_all_metrics(state)
    # Non-vacuity: "did not crash" must not be satisfiable by a metrics dict
    # that simply reports nothing. Verified to hold today for every tolerated
    # removal, so tightening this does not turn a green guard red.
    assert metrics["placed_count"] == len(get_placed_classes(state)), (
        f"analytics counted {metrics['placed_count']} of "
        f"{len(get_placed_classes(state))} placed lessons")
    assert metrics["total_classes"] == len(state["classes"])
    for key in ("room_utilization", "lecturer_load", "busiest_slots",
                "lec_gaps_total", "student_gaps_total"):
        assert key in metrics, f"analytics stopped reporting {key!r}"


def _csv_op(state, tmp_path):
    out = tmp_path / "orphan.csv"
    exporter.export_schedule(state, "csv", str(out))
    rows = _read_csv_rows(out)
    assert len(rows) > 1, "CSV contains a header and nothing else"
    # Non-vacuity: an exporter that emitted a plausible skeleton while dropping
    # every lesson would otherwise pass. Every lesson still on a real cell must
    # be written. (What happens to the *orphan* is asserted separately, in
    # test_csv_does_not_silently_drop_a_class_orphaned_by_slot_removal.)
    written = {r[3] for r in rows[1:] if len(r) > 3}
    missing = sorted(_on_grid_placed_names(state) - written)
    assert not missing, f"CSV dropped on-grid lessons: {missing}"


def _xlsx_op(state, tmp_path):
    pytest.importorskip("openpyxl", reason="openpyxl not installed")
    out = tmp_path / "orphan.xlsx"
    exporter.export_schedule(state, "xlsx", str(out))
    assert out.stat().st_size > 0

    import openpyxl
    workbook = openpyxl.load_workbook(out)
    assert workbook.sheetnames, "workbook has no sheets"
    # Same non-vacuity control as the CSV: an empty-but-valid workbook is not
    # a survived export.
    text = " ".join(
        str(value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row if value is not None)
    missing = sorted(n for n in _on_grid_placed_names(state) if n not in text)
    assert not missing, f"workbook dropped on-grid lessons: {missing}"


DOWNSTREAM_OPS = {
    "analytics.compute_all_metrics": _analytics_op,
    "export.csv": _csv_op,
    "export.xlsx": _xlsx_op,
}


@pytest.mark.parametrize("kind", sorted(REMOVALS))
@pytest.mark.parametrize("op_name", ["analytics.compute_all_metrics",
                                     "export.csv",
                                     pytest.param("export.xlsx",
                                                  marks=pytest.mark.excel)])
def test_axis_removal_does_not_crash_downstream(make_state, tmp_path,
                                                kind, op_name):
    """ST-DATA-003 — deleting a grid axis must not break analytics or export.

    A failure means that after the user shortens the teaching day (or removes a
    weekday/room/lecturer/year) in Setup, opening the statistics panel or
    exporting the timetable dies with ``ValueError: '10:00' is not in list`` —
    the schedule is intact on disk but the app can no longer show or print it.

    ``kind`` in {room, lecturer, year, day} is a *regression guard*: those are
    tolerated today and must stay tolerated. ``kind='slot'`` is the finding.
    """
    state, victim, orphans = _orphaned_state(make_state, kind)
    DOWNSTREAM_OPS[op_name](state, tmp_path)


@pytest.mark.parametrize("kind", sorted(REMOVALS))
def test_axis_removal_does_not_crash_occupancy(make_state, kind):
    """ST-DATA-003 — ``occupied_slots_of`` underpins every view and every export.

    A failure means the single helper that answers "which cells does this lesson
    cover?" raises for a lesson orphaned by a Setup edit, which is what takes
    the grid refresh, the CSV writer and the warning panel down together.
    """
    state, victim, orphans = _orphaned_state(make_state, kind)
    seen_on_grid = 0
    for cls in get_placed_classes(state):
        cells = occupied_slots_of(state, cls)
        assert all(slot in state["slots"] for _day, slot in cells), (
            f"occupied_slots_of reported cells outside the grid for "
            f"{cls['name']}: {cells}")
        # Non-vacuity: ``all(... for x in [])`` is True, so a helper that
        # answered "no cells" for EVERY lesson would satisfy the assertion
        # above while destroying every grid view. A lesson still sitting on a
        # real (day, slot) must still report exactly the cells it covers.
        if (effective_day(cls) in state["days"]
                and effective_time(cls) in state["slots"]):
            seen_on_grid += 1
            assert len(cells) == total_duration(cls), (
                f"{cls['name']} is placed on a real cell but "
                f"occupied_slots_of reported {len(cells)} of "
                f"{total_duration(cls)} slots: {cells}")
    assert seen_on_grid >= 2, "no on-grid control lessons were checked"


@pytest.mark.parametrize("kind", sorted(REMOVALS))
def test_axis_removal_does_not_crash_constraint_validator(make_state, kind):
    """ST-DATA-003 — the validator must be *constructible* on an orphaned state.

    A failure means that once one lesson is stranded on a deleted hour, merely
    building the object that every placement check, drag-drop and auto-place
    goes through raises, so the whole application is dead at once rather than
    one feature being degraded.
    """
    state, victim, orphans = _orphaned_state(make_state, kind)
    validator = ConstraintValidator(state)

    # Non-vacuity: a validator whose occupancy map was built empty would also
    # "not crash" — and would bless every cell. It must still see the lessons
    # that are on the grid.
    mover, blocker = _two_on_grid(state)
    assert validator.check_placement(
        mover, effective_day(blocker), effective_time(blocker),
        classroom_of(blocker)) is False, (
        "the validator approved a cell another lesson already occupies in the "
        "same room — its occupancy map was built empty")


@pytest.mark.parametrize("kind", sorted(REMOVALS))
def test_axis_removal_does_not_crash_conflict_detection(make_state, kind):
    """ST-DATA-003 — drag-and-drop conflict detection must survive an orphan.

    A failure means that after one lesson is stranded on a deleted hour the
    user can no longer drag *any* lesson anywhere: choosing a drop target walks
    every placed class, reaches the stranded one and raises.

    This path holds a SECOND bare ``.index()`` — ``logic.py:228``
    (``ex_start_idx = slot_index(state, ex_start)``) — that only becomes
    reachable once ``occupied_slots_of`` stops raising, so guarding
    ``get_consecutive_slots`` alone MOVES this crash rather than removing it.
    """
    state, victim, orphans = _orphaned_state(make_state, kind)
    mover, blocker = _two_on_grid(state)
    day, slot = effective_day(blocker), effective_time(blocker)

    conflicts = occupancy_conflicts(
        state, mover, day, slot, classroom_of(blocker))
    # Non-vacuity: dropping a lesson straight on top of another one in the same
    # room MUST report a conflict, so an empty answer is not a survived call.
    assert conflicts, (
        f"dropping {mover['name']} onto {blocker['name']}'s cell "
        f"({day}/{slot}) reported no conflict at all")

    room, _drop_conflicts = SchedulingWorkflow.find_drop_classroom(
        state, mover, day, slot)
    assert room is None or room in state["classrooms"], (
        f"the drop-target search offered classroom {room!r}, which is not in "
        f"{state['classrooms']}")


@pytest.mark.engine
@pytest.mark.parametrize("kind", sorted(REMOVALS))
def test_axis_removal_does_not_crash_reschedule(make_state, kind):
    """ST-DATA-003 — the user must be able to re-plan after a Setup edit.

    A failure means the one action that would *repair* the orphaned placements
    is itself the action that crashes, leaving the user with no in-app way out.
    """
    state, victim, orphans = _orphaned_state(make_state, kind)
    workflow = SchedulingWorkflow(state, lambda: {})
    result = workflow.reschedule({}, use_cpsat=False)
    # Non-vacuity: a reschedule that proposed nothing leaves the pre-existing
    # placements untouched, so both oracle assertions below would pass for the
    # wrong reason on every kind except 'day'.
    assert result.placed, "the reschedule proposed no placements at all"
    workflow.apply_reschedule(result)

    audit = check_schedule(state)
    assert audit["counts"].get("off_grid_slot", 0) == 0, (
        f"after removing {kind}={victim!r}, reschedule+commit left placements "
        f"on hours that no longer exist: {audit['counts']}")
    assert audit["counts"].get("off_grid_day", 0) == 0
    assert get_placed_classes(state), \
        "the reschedule unplaced everything — 'no violations' would be vacuous"


def test_csv_does_not_silently_drop_a_class_orphaned_by_slot_removal(
        make_state, tmp_path):
    """ST-DATA-003 / ST-FUNC-013 — the CSV must not quietly lose the orphan.

    A failure means the flat export — the format a user reaches for precisely
    to check what the printout left out — is itself missing lessons, with no
    warning anywhere.

    Same three-way bar as the PDF test (keep it, warn about it, or reconcile it
    away), and the same bar ``test_export_smoke.py`` already holds the CSV to
    for an off-grid *day* (``test_csv_still_reports_offgrid_placements``). A fix
    that only makes ``get_consecutive_slots`` return ``[]`` stops the crash but
    turns a loud failure into a silent one here, which is why this is asserted
    separately from the "does not crash" guard above.
    """
    state, victim, orphans = _orphaned_state(make_state, "slot")
    out = tmp_path / "orphan_drop.csv"
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        exporter.export_schedule(state, "csv", str(out))

    rows = _read_csv_rows(out)
    written = {r[3] for r in rows[1:] if len(r) > 3}
    control = _on_grid_placed_names(state)
    assert control and control <= written, (
        f"on-grid lessons {sorted(control - written)} are missing from the CSV "
        "— the test cannot tell a drop apart from an empty export")

    still_placed = {c["name"] for c in state["classes"] if c["placed"]}
    for name in orphans:
        warned = any(name in str(w.message) or str(victim) in str(w.message)
                     for w in caught)
        assert name in written or warned or name not in still_placed, (
            f"{name} is stored as placed at the deleted hour {victim!r} and the "
            "CSV neither wrote it, nor warned about it, nor reconciled it away "
            "— it simply vanished from the export")


@pytest.mark.pdf
@pytest.mark.parametrize("mode", PDF_MODES)
def test_pdf_does_not_silently_drop_a_class_orphaned_by_slot_removal(
        make_state, tmp_path, mode):
    """ST-DATA-003 — PDF survives the removal but loses the lesson without a word.

    A failure means the user shortens the teaching day, prints the timetable,
    and gets a document that is quietly missing lessons — the most dangerous
    outcome of the three, because nothing looks wrong.

    Either legitimate fix passes: draw the orphan, or tell the user it was left
    out. A third acceptable outcome is that the export path reconciles the state
    and the lesson is no longer marked placed — data that was *deliberately*
    removed is not data that vanished.
    """
    pytest.importorskip("reportlab", reason="reportlab not installed")
    state, victim, orphans = _orphaned_state(make_state, "slot")
    kept = [c["name"] for c in state["classes"]
            if c["placed"] and c["placed_time"] != victim]
    assert kept, "no on-grid control lesson left — the test cannot tell a drop apart"

    out = tmp_path / f"orphan_{mode}.pdf"
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        exporter.export_schedule(state, "pdf", str(out), mode=mode)

    raw = out.read_bytes()
    assert raw.startswith(b"%PDF-") and raw.rstrip().endswith(b"%%EOF"), \
        "the PDF is not structurally well formed"
    content = _pdf_content_text(raw)

    # Control: on-grid lessons are drawn, so a miss below is a genuine drop.
    assert any(name.encode() in content for name in kept), \
        f"none of the on-grid lessons {kept} were drawn — the test is broken"

    still_placed = {c["name"] for c in state["classes"] if c["placed"]}
    for name in orphans:
        drawn = name.encode() in content
        warned = any(name in str(w.message) or str(victim) in str(w.message)
                     for w in caught)
        reconciled = name not in still_placed
        assert drawn or warned or reconciled, (
            f"{name} is stored as placed at the deleted hour {victim!r} and the "
            f"mode={mode} PDF neither drew it, nor warned about it, nor "
            "reconciled it away — it simply vanished from the printout")


# ══════════════════════════════════════════════════════════════════════════
#  4. Guards — the constraints that DO make sense must keep working
# ══════════════════════════════════════════════════════════════════════════

def test_a_valid_allowed_days_subset_is_still_honoured():
    """ST-SCHED-003 (guard) — intersecting with the grid must not over-filter.

    A failure means the fix for ghost days threw the baby out: a class legally
    restricted to Wednesday/Thursday can no longer be placed at all, or is
    placed on a day the user forbade.
    """
    allowed = ["wednesday", "thursday"]
    state, cls = _state_with(allowed_days=allowed)

    days, times, rooms = CandidateGenerator(state).get_search_space(cls)
    assert set(days) == set(allowed)

    result = SchedulingWorkflow(state, lambda: {}).auto_place(cls)
    assert result.success is True, "a perfectly placeable class was refused"
    assert result.placed_info[0] in allowed
    assert cls["placed_day"] in allowed
    assert check_schedule(state)["counts"] == {}


def test_a_valid_allowed_times_subset_is_still_honoured():
    """ST-SCHED-004 (guard) — guarding ``slot_index`` must not drop real hours.

    A failure means a class legitimately restricted to 10:00/11:00 stops being
    placeable, so the fix for the stale-hour crash has broken normal use.
    """
    allowed = ["10:00", "11:00"]
    state, cls = _state_with(allowed_times=allowed)

    days, times, rooms = CandidateGenerator(state).get_search_space(cls)
    assert set(times) == set(allowed)

    result = SchedulingWorkflow(state, lambda: {}).auto_place(cls)
    assert result.success is True
    assert result.placed_info[1] in allowed


@pytest.mark.engine
@pytest.mark.parametrize("field,value", [
    ("allowed_days", [OFF_GRID_DAY]),
    ("allowed_times", [OFF_GRID_SLOT]),
])
def test_an_empty_intersection_leaves_the_class_unplaced_and_reported(
        field, value):
    """ST-SCHED-003/004 — "nowhere legal to go" must be said out loud.

    A failure means a class whose only allowed day/hour no longer exists is
    either parked somewhere the user forbade, or dropped from the plan with no
    entry in the unplaced list — in both cases the user is never told.

    This is also why the fix must NOT simply delete the stale value from the
    allow-list: ``models.filter_class_days`` reads an *empty* ``allowed_days``
    as "no restriction", so normalizing ``['saturday']`` to ``[]`` would place
    the lesson on Monday and call it a success.
    """
    state, cls = _state_with(**{field: value})
    healthy = _class(name="Saglikli", class_code="G002")
    state["classes"].append(healthy)

    workflow = SchedulingWorkflow(state, lambda: {})
    result = workflow.reschedule({}, use_cpsat=False)

    proposed = {c.get("name") for c, _d, _s, _r in result.placed}
    assert healthy["name"] in proposed, \
        "the reschedule proposed nothing — the assertions below are vacuous"
    assert cls["name"] not in proposed, (
        f"the class was proposed for placement despite {field}={value} "
        f"having no overlap with the grid")

    reported = {c.get("name"): reason for c, reason in result.unplaced}
    assert cls["name"] in reported, (
        f"the class is neither placed nor listed in RescheduleResult.unplaced — "
        f"unplaced={list(reported)}")
    assert isinstance(reported[cls["name"]], str) and reported[cls["name"]].strip(), \
        "the class was reported as unplaced with an empty reason"


@pytest.mark.parametrize("field,value,reason_key", [
    ("allowed_days", [OFF_GRID_DAY], "negotiation.no_allowed_days_configured"),
    ("allowed_times", [OFF_GRID_SLOT], "negotiation.no_allowed_times_configured"),
])
def test_unplaced_reason_names_the_empty_grid_intersection(
        field, value, reason_key):
    """ST-SCHED-003/004 — the reason shown must be the *right* reason.

    A failure means the user is told something misleading — "all candidate slots
    are full" instead of "no allowed day/hour is left after your constraints" —
    and has no way to work out that a deleted grid axis is to blame.
    """
    state, cls = _state_with(**{field: value})
    reason = CandidateGenerator(state).unplaced_reason(cls)

    assert reason == tr(reason_key), (
        f"unplaced_reason said {reason!r}; expected {tr(reason_key)!r}")
