"""End-to-end regression for the Excel-import UI flow (ST-FUNC-001).

ST-FUNC-001 (Critical): ``SchedulerApp._import_from_excel`` merges the imported
workbook into ``self.state_data`` and *then* calls ``self._on_state_changed()``
and ``self.refresh()`` -- neither of which exists anywhere in the 8-class MRO
(the real methods are ``refresh_grid()`` and ``_update_status()``). So every
successful import raises ``AttributeError`` *after* the state has already been
mutated, leaving the app half-updated: data changed, screen not repainted, no
success dialog.

These tests drive the real ``SchedulerApp`` headlessly (offscreen Qt, sandboxed
HOME) through the whole File -> Import Excel path. They are **fail-now /
pass-after** for the Phase 0 fix -- deliberately no ``xfail``: they must go
green the moment the success handler is repaired and the merge is made
transactional.

Deliberately NOT asserted here: the number of classes the importer produces
from the generated template. That is ST-FUNC-002 (blank joint-group cells
collapse 5 template rows into 2), which is a separate Phase 0 fix. Every
"was it merged?" assertion below compares the window's state against an
*independent* run of ``load_scheduler_data_from_excel`` on the same file, so
these tests stay green whichever way ST-FUNC-002 lands.

Discrimination (verified empirically, see the module report): against a
``_import_from_excel`` gutted to ``pass`` all 7 tests fail; against a "lazy"
fix that only deletes the two bogus lines (crash gone, but no repaint and no
rollback) exactly 3 fail -- ``..._refreshes_the_ui`` and the two rollback
tests; against the full recommended fix all 7 pass.
"""
import copy
import os

import pytest

pytestmark = [pytest.mark.ui, pytest.mark.excel]


# ── Helpers ─────────────────────────────────────────────────────────────────

class _Recorder:
    """Stand-in for a modal QMessageBox static; records instead of blocking."""

    def __init__(self, ret):
        self._ret = ret
        self.calls = []

    def __call__(self, *args, **kwargs):
        self.calls.append((args, kwargs))
        return self._ret

    @property
    def called(self):
        return bool(self.calls)

    def texts(self):
        """Every string argument of every recorded call, flattened."""
        out = []
        for args, kwargs in self.calls:
            out.extend(a for a in args if isinstance(a, str))
            out.extend(v for v in kwargs.values() if isinstance(v, str))
        return out


def _strip_uids(classes):
    """Class dicts carry a random ``class_uid``; drop it before comparing."""
    return [{k: v for k, v in c.items() if k != "class_uid"} for c in classes]


def _make_template(tmp_path):
    """The app's own generated import template -- the canonical happy path."""
    from scheduler_app.data_io.template import generate_excel_template

    path = os.path.join(str(tmp_path), "scheduler_template.xlsx")
    generate_excel_template(path)
    assert os.path.exists(path), "template generator produced no file"
    return path


def _make_rejected_workbook(tmp_path):
    """A workbook the importer must reject (Classes sheet missing 'duration')."""
    import openpyxl

    from scheduler_app.data_io.schema import (
        get_workbook_sheet_header_map,
        get_workbook_sheet_title,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = get_workbook_sheet_title("classes")
    headers = get_workbook_sheet_header_map("classes")
    ws.append([label for field, label in headers.items() if field != "duration"])
    path = os.path.join(str(tmp_path), "missing_duration.xlsx")
    wb.save(path)

    # Guard the fixture itself: this file must actually be rejected, otherwise
    # the "rejected import" tests below would silently assert nothing.
    from scheduler_app.data_io.importer import load_scheduler_data_from_excel

    assert not load_scheduler_data_from_excel(path).report.is_valid
    return path


def _seed_classes(window, names):
    """Put pre-existing classes in the window so a rollback has something to
    protect (an empty list would make the rollback assertion vacuous)."""
    from scheduler_app.core.models import new_class

    for name in names:
        cls = new_class()
        cls["name"] = name
        cls["lecturer"] = "Seed Lecturer"
        window.state_data["classes"].append(cls)


# ── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def message_boxes(monkeypatch):
    """Neutralize every modal the import path can raise.

    An unpatched modal blocks the whole suite under the offscreen platform, so
    this is a hard requirement, not a convenience. ``UpgradeDialog`` is included
    because ``require_entity_limit`` calls ``dialog.exec()`` directly: if the
    tier ever stopped being Institutional, an unpatched ``exec()`` would *hang*
    the run instead of failing it.
    """
    from PyQt6.QtWidgets import QDialog, QMessageBox

    from scheduler_app.ui.tier_enforcement import UpgradeDialog

    recorders = {
        "information": _Recorder(QMessageBox.StandardButton.Ok),
        "warning": _Recorder(QMessageBox.StandardButton.Ok),
        "critical": _Recorder(QMessageBox.StandardButton.Ok),
        "question": _Recorder(QMessageBox.StandardButton.Yes),
    }
    for name, rec in recorders.items():
        monkeypatch.setattr(QMessageBox, name, staticmethod(rec))

    upgrade = _Recorder(QDialog.DialogCode.Rejected.value)
    recorders["upgrade"] = upgrade
    monkeypatch.setattr(UpgradeDialog, "exec", upgrade)
    return recorders


_TIER_REGISTRIES = (
    "_gated_widgets", "_gated_actions", "_on_tier_changed",
    "_export_submenu_refreshers",
)


@pytest.fixture
def window(qapp, dersis_home, monkeypatch):
    """A real, fully constructed SchedulerApp -- never shown.

    Two things are neutralized so the flow is deterministic:
      * the first-run controller (tutorial overlay / setup dialog), which is
        otherwise armed by a QTimer in ``__init__``;
      * the licence tier, pinned to Institutional so the entity-limit checks in
        ``_import_from_excel`` cannot short-circuit the import before the merge.

    Isolation note: every ``SchedulerApp`` registers 5 gated ``QAction``s and a
    tier-change callback into the *process-wide* ``TierEnforcement`` singleton
    and never unregisters them, so without the snapshot/restore below this
    module would leak ~35 dead QActions into whatever runs next in the same
    session. The tier is set by direct assignment rather than ``set_tier()``
    because ``set_tier()`` sweeps every registered gate, including ones
    belonging to windows other tests already destroyed.
    """
    from scheduler_app.plans import TIER_INSTITUTIONAL
    from scheduler_app.i18n.day_keys import DAY_KEYS
    from scheduler_app.ui.first_run import FirstRunController
    from scheduler_app.ui.tier_enforcement import TierEnforcement

    monkeypatch.setattr(FirstRunController, "start", lambda self: None)

    enforcer = TierEnforcement.instance()
    prev_slug, prev_confirmed = enforcer._tier_slug, enforcer._tier_confirmed
    prev_registries = {
        name: list(getattr(enforcer, name))
        for name in _TIER_REGISTRIES if hasattr(enforcer, name)
    }
    enforcer._tier_slug, enforcer._tier_confirmed = TIER_INSTITUTIONAL, True

    from scheduler_app.ui.app import SchedulerApp

    win = SchedulerApp()
    # A brand-new profile has days == slots == [], i.e. a 0x0 grid, which makes
    # refresh_grid() almost a no-op and would let a repaint that crashes on a
    # real timetable slip through. Give the window the grid a real user would
    # have configured (the first-run wizard we disabled above normally does it).
    win.state_data["days"] = list(DAY_KEYS[:5])
    win.state_data["slots"] = ["09:00", "10:00", "11:00", "12:00"]
    try:
        yield win
    finally:
        win.close()
        win.deleteLater()
        qapp.processEvents()
        enforcer._tier_slug, enforcer._tier_confirmed = prev_slug, prev_confirmed
        for name, value in prev_registries.items():
            setattr(enforcer, name, value)


@pytest.fixture
def choose_file(monkeypatch):
    """Return a callable that pins QFileDialog.getOpenFileName to one path."""
    from PyQt6.QtWidgets import QFileDialog

    def _choose(path):
        monkeypatch.setattr(
            QFileDialog, "getOpenFileName",
            staticmethod(lambda *a, **k: (str(path), "")),
        )

    return _choose


# ── 1. The happy path must not crash ────────────────────────────────────────

def test_import_of_generated_template_does_not_raise(
        window, message_boxes, choose_file, tmp_path):
    """ST-FUNC-001 — importing the app's own template must not raise.

    A failure here is the flagship bug: the user picks a valid workbook and the
    app throws instead of importing, every single time.

    The trailing assertions exist so this cannot be satisfied by a
    ``_import_from_excel`` that simply does nothing: "did not crash" is only
    meaningful if the import also actually ran.
    """
    choose_file(_make_template(tmp_path))

    window._import_from_excel()  # must complete cleanly

    assert window.state_data["classes"], "import completed without importing anything"
    assert message_boxes["information"].called, "import completed without telling the user"


# ── 2. The imported data must actually land in the window's state ───────────

def test_import_merges_workbook_into_state(
        window, message_boxes, choose_file, tmp_path):
    """ST-FUNC-001 — the imported lecturers/classrooms/years/classes must end up
    in ``state_data``.

    A failure means the user's roster silently did not arrive, or arrived only
    in part.
    """
    from scheduler_app.data_io.importer import load_scheduler_data_from_excel

    path = _make_template(tmp_path)
    expected = load_scheduler_data_from_excel(path).state
    assert expected["classes"], "test fixture produced no classes to import"

    assert window.state_data["classes"] == []
    choose_file(path)

    window._import_from_excel()

    s = window.state_data
    assert s["lecturers"] == expected["lecturers"]
    assert s["classrooms"] == expected["classrooms"]
    assert s["years"] == expected["years"]
    # The per-entity detail rides along with the name lists and is just as easy
    # to drop in a rewrite; a lecturer whose availability vanished would be
    # rescheduled into hours they told the app they cannot teach.
    #
    # Availability is compared by lecturer *name* only. The REASON changed in
    # Phase 9 and the old one is worth not leaving here, because it described a
    # design that was itself the C1 defect. It used to read: "the app
    # legitimately rewrites the day values (localized labels like 'Pazartesi'
    # from the workbook -> stable keys like 'monday') ... via
    # normalize_state_day_keys". That rewrite happened on the AUTOSAVE path,
    # long after the import returned, and on a schedule whose week had not been
    # laid out yet it did not convert the labels — it deleted every one of them,
    # permanently, and wrote the emptied roster to disk.
    #
    # C1 moved the label -> key conversion into `_process_teachers`, so both
    # sides of this comparison already hold keys and there is no rewrite left on
    # this path. What can still legitimately differ is PRUNING: a day that is
    # not in `state["days"]` is dropped once a week exists. What must not happen
    # is the dict going missing.
    assert set(s["lecturer_availability"]) == set(expected["lecturer_availability"])
    assert s["classroom_capacities"] == expected["classroom_capacities"]
    assert _strip_uids(s["classes"]) == _strip_uids(expected["classes"])


# ── 3. The UI must be repainted, not just mutated ───────────────────────────

def test_import_refreshes_the_ui(window, message_boxes, choose_file,
                                 monkeypatch, tmp_path):
    """ST-FUNC-001 — a successful import must repaint the grid and the status bar.

    A failure means the user's data changed but the screen still shows the old
    timetable, so they cannot tell whether the import worked. Guards against a
    fix that merely deletes the crashing lines.
    """
    from scheduler_app.ui.app import SchedulerApp

    seen = []
    real_refresh = SchedulerApp.refresh_grid
    real_status = SchedulerApp._update_status

    def spy_refresh(self, *a, **k):
        seen.append("refresh_grid")
        return real_refresh(self, *a, **k)

    def spy_status(self, *a, **k):
        seen.append("_update_status")
        return real_status(self, *a, **k)

    monkeypatch.setattr(SchedulerApp, "refresh_grid", spy_refresh)
    monkeypatch.setattr(SchedulerApp, "_update_status", spy_status)

    choose_file(_make_template(tmp_path))

    window._import_from_excel()

    assert "refresh_grid" in seen, "grid was never re-rendered after import"
    assert "_update_status" in seen, "status bar was never updated after import"


# ── 4. The user must be told what happened ──────────────────────────────────

def test_import_reports_success_to_the_user(
        window, message_boxes, choose_file, tmp_path):
    """ST-FUNC-001 — a successful import must end in a success dialog, not an
    error one.

    A failure means the user is left guessing whether their roster was imported.
    """
    from scheduler_app.translations import tr

    choose_file(_make_template(tmp_path))

    window._import_from_excel()

    assert message_boxes["information"].called, "no success dialog was shown"
    assert any(tr("status.import_successful") in t
               for t in message_boxes["information"].texts())
    assert not message_boxes["critical"].called
    assert not message_boxes["warning"].called
    assert not message_boxes["upgrade"].called, "tier limit blocked an unlimited tier"


def test_rejected_import_warns_and_leaves_state_untouched(
        window, message_boxes, choose_file, tmp_path):
    """ST-FUNC-001 — an invalid workbook must warn and change nothing.

    A failure means a workbook the importer already rejected still altered the
    user's schedule, or was reported as a success.
    """
    _seed_classes(window, ["Existing A", "Existing B"])
    before = copy.deepcopy(window.state_data)
    choose_file(_make_rejected_workbook(tmp_path))

    window._import_from_excel()

    assert message_boxes["warning"].called, "invalid workbook was not reported"
    assert not message_boxes["information"].called, "invalid import claimed success"
    assert window.state_data == before


# ── 5. A failure anywhere must roll the merge back ──────────────────────────

def test_failed_import_rolls_back_the_class_merge(
        window, message_boxes, choose_file, monkeypatch, tmp_path):
    """ST-FUNC-001 — if anything in the import fails, no classes may be left
    merged.

    Today the imported classes are appended *before* the crash, so the user ends
    up with a half-applied import they were never told about. Written purely
    against observable state: it does not care how the rollback is implemented,
    only that the repaint (which is where the real bug lives) is inside the
    transaction.
    """
    from scheduler_app.ui.app import SchedulerApp

    _seed_classes(window, ["Existing A", "Existing B"])
    before = copy.deepcopy(window.state_data["classes"])

    detonated = []

    def boom(self, *a, **k):
        detonated.append(True)
        raise RuntimeError("simulated failure during post-import refresh")

    monkeypatch.setattr(SchedulerApp, "refresh_grid", boom)
    choose_file(_make_template(tmp_path))

    # Whether the failure is swallowed and reported, or propagates, is the
    # implementation's choice; the state guarantee is not.
    try:
        window._import_from_excel()
    except Exception:
        pass

    assert window.state_data["classes"] == before
    # Ordered last on purpose: the state assertion above is the finding; this one
    # only stops the test passing vacuously against an import that never runs.
    assert detonated, "import never reached the repaint, so nothing was rolled back"


def test_failed_import_rolls_back_the_whole_merge(
        window, message_boxes, choose_file, monkeypatch, tmp_path):
    """ST-FUNC-001 — a failed import must not leave lecturers/classrooms/years
    half-replaced either.

    The register's fix is "wrap the whole import in try/except that rolls back
    the merge"; the merge replaces the lecturer, classroom and year lists too,
    so a failure that keeps them would still leave the app in a state the user
    never asked for.
    """
    from scheduler_app.ui.app import SchedulerApp

    _seed_classes(window, ["Existing A"])
    window.state_data["lecturers"] = ["Seed Lecturer"]
    window.state_data["classrooms"] = ["Seed Room"]
    before = copy.deepcopy(window.state_data)

    detonated = []

    def boom(self, *a, **k):
        detonated.append(True)
        raise RuntimeError("simulated failure during post-import refresh")

    monkeypatch.setattr(SchedulerApp, "refresh_grid", boom)
    choose_file(_make_template(tmp_path))

    try:
        window._import_from_excel()
    except Exception:
        pass

    assert window.state_data == before
    assert detonated, "import never reached the repaint, so nothing was rolled back"


# ── 4. Setup ▸ Lecturers ▸ Import Excel is a second front door ──────────────
#
# ST-FUNC-011 guarded ``load_scheduler_data_from_excel`` / File ▸ Import Excel.
# The Setup dialog has five import buttons of its own (days, slots, rooms,
# lecturers, years) that never go near that pipeline: each reads sheet 0 with
# ``pd.read_excel(path, sheet_name=0, header=0)`` and takes column 0 as data.
# Only the lecturer door is closed here, because only it is measured; the other
# four are the same shape and want the same guard as their own change.

def _budget_workbook(tmp_path):
    """An unrelated spreadsheet — the file ST-FUNC-011 was written about."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Butce"
    ws.append(["Kalem", "Tutar"])
    for item, amount in (("Kirtasiye", 1200), ("Temizlik", 800), ("Yakit", 4500)):
        ws.append([item, amount])
    path = os.path.join(str(tmp_path), "butce.xlsx")
    wb.save(path)
    return path


def _lecturer_workbook(tmp_path, lang="tr"):
    """What Setup ▸ Lecturers ▸ Export Excel writes — the round-trip partner.

    Built from the same translation keys ``_export_lecturers_to_excel`` uses, so
    a header rename cannot make this fixture and the export drift apart. ``lang``
    picks the catalogue the headers are drawn from without touching the process
    language: a school that exported before switching the app to another
    language must still be able to import the file it already has.
    """
    import openpyxl

    from scheduler_app.translations import TRANSLATIONS

    labels = TRANSLATIONS[lang]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([labels["labels.lecturer"], labels["setup.allowed_days"],
               labels["setup.allowed_hours"], labels["setup.excluded_days"],
               labels["setup.excluded_hours"]])
    ws.append(["Ada Lovelace", "", "", "", ""])
    ws.append(["Bora Yildiz", "", "", "", ""])
    path = os.path.join(str(tmp_path), f"lecturers_{lang}.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def setup_dialog(qapp, dersis_home, message_boxes):
    """A real ``SetupDialog`` over an empty roster, never shown."""
    from scheduler_app.core.models import new_state
    from scheduler_app.ui.dialogs import SetupDialog

    state = new_state()
    state["days"] = ["monday", "tuesday"]
    state["slots"] = ["09:00", "10:00"]
    dlg = SetupDialog(None, state)
    try:
        yield dlg
    finally:
        dlg.deleteLater()
        qapp.processEvents()


def test_setup_lecturer_import_refuses_an_unrelated_spreadsheet(
        setup_dialog, message_boxes, choose_file, tmp_path):
    """A budget sheet must not become three members of staff.

    Measured before the fix: the button's own slot read "Kalem"/"Tutar",
    inserted Kirtasiye, Temizlik and Yakit into the lecturer table and showed
    ``('information', 'İçe aktarma başarılı', '3 Öğretim Elemanları')``. OK then
    wrote those three strings into ``state['lecturers']``, where they are
    indistinguishable from real staff — the lecturer list is keyed by name.
    """
    choose_file(_budget_workbook(tmp_path))

    setup_dialog._import_lecturers_from_excel()

    names = [setup_dialog.lec_table.item(r, 0).text()
             for r in range(setup_dialog.lec_table.rowCount())]
    assert names == [], f"budget line items were imported as lecturers: {names}"
    assert not message_boxes["information"].called, (
        "the user was told an unrelated spreadsheet imported successfully: "
        f"{message_boxes['information'].texts()}")
    assert message_boxes["warning"].called, (
        "the import was refused without telling the user anything")


@pytest.mark.parametrize("lang", ["tr", "en", "zh"])
def test_setup_lecturer_import_still_reads_its_own_export(
        setup_dialog, message_boxes, choose_file, tmp_path, lang):
    """Discrimination: the export/import loop must keep working.

    Without this, "refuse everything" passes the test above. The workbook here
    is exactly what ``_export_lecturers_to_excel`` writes — in three languages,
    because the app is Turkish-first but ships 22 and a roster exported before
    a language change is still the user's roster.
    """
    choose_file(_lecturer_workbook(tmp_path, lang))

    setup_dialog._import_lecturers_from_excel()

    names = [setup_dialog.lec_table.item(r, 0).text()
             for r in range(setup_dialog.lec_table.rowCount())]
    assert names == ["Ada Lovelace", "Bora Yildiz"], (
        f"the dialog cannot read back its own export: {names}")
    assert message_boxes["information"].called
    assert not message_boxes["warning"].called, (
        f"a valid roster was refused: {message_boxes['warning'].texts()}")
