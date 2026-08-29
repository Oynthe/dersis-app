"""B5 / ST-FUNC-009 — a room type declared by a *later* row of a joint group.

``_process_classes`` resolves ``required_room_type`` into ``required_classrooms``
per row, but ``_resolve_joint_groups``
(``scheduler_app/data_io/importer.py``) merges a group by keeping ``classes[0]``
as the primary, copying only ``targets`` off the other members and deleting
them from state. ``required_classrooms`` is never merged, so the room
constraint of every non-first row of a joint group is thrown away — an empty
``required_classrooms`` is exactly what ``get_physical_room_candidates`` reads
as "any room", which is the ST-FUNC-009 inversion the Phase 8 fix exists to
prevent. A joint lab session whose lab row is not first can still be scheduled
into a lecture hall.

The workbook builder here is a deliberate copy of the one in
``tests/test_import_roundtrip.py`` rather than an import of it: this module has
to stand on its own while other agents edit that file in parallel.

The second half of the file guards the *remediation* of that fix. Merging the
room constraints introduced a regression of its own — the new
``excluded_classrooms`` union emptied the candidate set of joint groups that
imported placeable before it, silently — and left the merge's own stated
invariant ("the same two rows swapped must produce the same session") false of
``location_type``. Everything below the marker exists because the first four
tests above were not enough: the whole exclusion union and the whole warning
apparatus could be deleted with the suite still green.
"""

import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

from scheduler_app.core.models import (  # noqa: E402
    LOCATION_FACE_TO_FACE, LOCATION_ONLINE, get_location_label,
    get_physical_room_candidates,
)
from scheduler_app.data_io import schema  # noqa: E402
from scheduler_app.data_io.importer import load_scheduler_data_from_excel  # noqa: E402
from scheduler_app.translations import tr  # noqa: E402

pytestmark = pytest.mark.excel

SHEET_IDS = ("teachers", "rooms", "branches", "classes")

# Free text the app never translates — the importer matches the Classes sheet
# against the Rooms sheet of the *same* workbook, so both sides are literals.
ROOM_TYPE_LECTURE = "Derslik"
ROOM_TYPE_LAB = "Laboratuvar"

DEFAULT_TEACHERS = [
    {"teacher_id": "T001", "name": "Ada Lovelace"},
    {"teacher_id": "T002", "name": "Bora Yildiz"},
]
# Two labs, so "the type narrows the list" and "the type replaces the list"
# cannot produce the same answer; one lecture hall, which is the room a joint
# lab session must never be allowed into.
DEFAULT_ROOMS = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30,
     "room_type": ROOM_TYPE_LECTURE},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    {"room_id": "R003", "name": "Lab 2", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
]
DEFAULT_BRANCHES = [
    {"branch_id": "B001", "name": "Grup A"},
    {"branch_id": "B002", "name": "Grup B"},
    # A third branch so a joint group can have THREE rows. The exclusion union
    # is the only merge rule whose failure needs three of them: no single row
    # can rule out every room in the building, but three rows ruling out one
    # room each can, and that is the shape no two-row test reaches.
    {"branch_id": "B003", "name": "Grup C"},
]


def build_workbook(path, *, classes, teachers=None, rooms=None, branches=None):
    """Write a workbook whose sheet titles and headers come from the schema."""
    rows_by_sheet = {
        "teachers": DEFAULT_TEACHERS if teachers is None else teachers,
        "rooms": DEFAULT_ROOMS if rooms is None else rooms,
        "branches": DEFAULT_BRANCHES if branches is None else branches,
        "classes": list(classes),
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in SHEET_IDS:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]]
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=headers[field])
        excel_row = 2
        for row in rows_by_sheet[sheet_id]:
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=excel_row, column=col, value=row[field])
            excel_row += 1

    wb.save(str(path))
    return str(path)


def klass(class_id, **overrides):
    """A minimal, fully valid Classes row; override any field by keyword."""
    row = {
        "class_id": class_id,
        "course_name": f"Ders {class_id}",
        "teacher_id": "T001",
        "branch_id": "B001",
        "duration": 1,
        "student_count": 10,
        "joint_class_group": f"UNIQ-{class_id}",
    }
    row.update(overrides)
    return row


def joint_rows(*, lab_row_index, joint="J1", course="Ortak Fizik",
               first_type=None, second_type=None):
    """Two rows of one joint group; ``lab_row_index`` says which declares the lab.

    ``lab_row_index`` is 0 or 1. ``first_type``/``second_type`` override it
    outright for the two-different-types case.
    """
    types = [first_type, second_type]
    if first_type is None and second_type is None:
        types = [None, None]
        types[lab_row_index] = ROOM_TYPE_LAB
    rows = []
    for i, (cid, bid) in enumerate((("C001", "B001"), ("C002", "B002"))):
        extra = {}
        if types[i]:
            extra["required_room_type"] = types[i]
        rows.append(klass(cid, branch_id=bid, course_name=course,
                          joint_class_group=joint, **extra))
    return rows


def merged_joint_class(dataset, branches=("Grup A", "Grup B")):
    """The single surviving class of a joint group.

    Asserts the merge itself happened first, so a failure below is never a
    failure of the fixture.
    """
    classes = dataset.state["classes"]
    assert len(classes) == 1, (
        "fixture drift: the rows of joint group J1 did not merge into one "
        f"session; state holds {[c['name'] for c in classes]}")
    merged = classes[0]
    assert merged["joint_session"] is True
    got = {t["branch"] for t in merged.get("targets", [])}
    assert got == set(branches), (
        f"fixture drift: the merged session lost a branch target; got {got}")
    return merged


def joint_warning(dataset, key, **fields):
    """The report line the importer rendered from *key*, or ``None``.

    Builds the expected sentence through ``tr`` and the schema's own header map
    rather than matching English text, so the assertion pins the message the
    user actually reads in whatever locale the suite runs under — and fails if
    the placeholders the importer fills stop matching the ones the catalogue
    declares. Unused keyword arguments are ignored by ``str.format``, so one
    call site serves every joint-group key.
    """
    headers = schema.get_workbook_sheet_header_map("classes")
    text = tr(key).format(
        group_field=headers["joint_class_group"],
        type_field=headers["required_room_type"],
        rooms_field=headers["allowed_rooms"],
        excluded_field=headers["excluded_rooms"],
        location_field=headers["location_type"],
        **fields)
    for line in dataset.report.warnings:
        if line.endswith(text):
            return line
    return None


def joint_klass(class_id, branch_id, **overrides):
    """A row of joint group ``J1``."""
    return klass(class_id, branch_id=branch_id, course_name="Ortak Fizik",
                 joint_class_group="J1", **overrides)


def test_a_joint_group_keeps_a_room_type_declared_by_its_second_row(tmp_path):
    """B5 / ST-FUNC-009 — the lab row is not first, and its lab is discarded.

    Two rows of one joint group; the SECOND one carries
    ``required_room_type = Laboratuvar``. ``_process_classes`` resolves that row
    to ``['Lab 1', 'Lab 2']`` correctly, and then ``_resolve_joint_groups``
    keeps ``classes[0]`` — whose ``required_classrooms`` is empty — copies only
    ``targets`` off the second row and deletes it. The merged session therefore
    imports with no room constraint at all, and an empty
    ``required_classrooms`` means "any room" to
    ``get_physical_room_candidates``: the physics lab is free to land in the
    lecture hall Oda 1, which is the headline case ST-FUNC-009 is about and the
    one the user has been told is fixed.

    Row order in a spreadsheet carries no meaning here — nothing in the
    template tells a user that the row declaring the lab must be typed first —
    so the merged session must carry the same constraint it would have carried
    had the two rows been swapped.
    """
    path = build_workbook(tmp_path / "joint_lab_second.xlsx",
                          classes=joint_rows(lab_row_index=1))
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert sorted(merged["required_classrooms"]) == ["Lab 1", "Lab 2"], (
        "the joint session lost the room type declared by its second row: "
        f"required_classrooms={merged['required_classrooms']!r}, expected the "
        f"two {ROOM_TYPE_LAB} rooms ['Lab 1', 'Lab 2']")

    candidates = get_physical_room_candidates(ds.state, merged)
    assert "Oda 1" not in candidates, (
        f"a joint {ROOM_TYPE_LAB} session can still be scheduled into the "
        f"lecture hall Oda 1; candidate rooms are {candidates}")


def test_a_joint_group_keeps_a_room_type_declared_by_its_first_row(tmp_path):
    """Control for the test above — the same workbook with the rows swapped.

    This is the arrangement the merge happens to handle: the lab row is
    ``classes[0]``, so its ``required_classrooms`` survives by accident of
    position. If this one fails too, the defect is not about row order and the
    room type is being lost somewhere else entirely.
    """
    path = build_workbook(tmp_path / "joint_lab_first.xlsx",
                          classes=joint_rows(lab_row_index=0))
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert sorted(merged["required_classrooms"]) == ["Lab 1", "Lab 2"], (
        "the control case is broken as well: a room type on the FIRST row of a "
        f"joint group is also being lost; got {merged['required_classrooms']!r}")
    assert "Oda 1" not in get_physical_room_candidates(ds.state, merged)


def test_a_joint_merge_never_widens_a_group_whose_rows_declare_room_types(
        tmp_path):
    """Forward guard — two rows, two *different* types, and no way to widen.

    ``Derslik`` resolves to ``['Oda 1']`` and ``Laboratuvar`` to
    ``['Lab 1', 'Lab 2']``; the two do not intersect. A fix for the test above
    that merges by plain intersection would write ``[]`` here, and ``[]`` is
    read by ``get_physical_room_candidates`` as "any room" — so the one session
    in the file that two rows constrained would become the least constrained
    thing in it. That inversion is the ST-FUNC-009 finding itself, and Phase 8
    shipped exactly it once inside ``_process_classes`` before catching it, so
    it is guarded here before the fix is written rather than after.

    This test passes today (nothing merges, so the primary's ``['Oda 1']``
    stands); it exists to stay passing. It deliberately does not dictate
    *which* of the two types wins — only that the answer is never "all of
    them".
    """
    rows = joint_rows(lab_row_index=0, first_type=ROOM_TYPE_LECTURE,
                      second_type=ROOM_TYPE_LAB)
    path = build_workbook(tmp_path / "joint_two_types.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert merged["required_classrooms"], (
        "merging two disagreeing room types emptied required_classrooms, which "
        "means 'any room': the joint session is now allowed into every room in "
        f"the building — {get_physical_room_candidates(ds.state, merged)}")
    assert len(get_physical_room_candidates(ds.state, merged)) < 3, (
        "the merged session is allowed into every room in the building despite "
        "both of its rows naming a room type")


# ── Remediation of the fix above ────────────────────────────────────────────
#
# Everything from here down was written after an adversarial pass measured the
# fix. Four of its seven mutants survived: the entire `excluded_classrooms`
# union could be deleted, its write-back removed, the dropped-room bookkeeping
# deleted and the whole warning suppressed, with the suite still green. The
# three tests above are all about `required_classrooms`, so none of them could
# fire. Each test below names the mutant it kills.


def test_a_joint_group_unions_the_exclusions_of_its_rows(tmp_path):
    """The exclusion half of the merge, pinned. Kills mutants M3 and M5.

    Row 1 asks for the lab type (``['Lab 1', 'Lab 2']``); row 2 excludes
    ``Lab 1``. A joint session is ONE session in ONE room, so a room any member
    row rules out is ruled out for the session — the merge takes the UNION of
    the exclusions, the opposite rule to the intersection it takes for the
    required lists, because an exclusion is a prohibition rather than a
    preference.

    The commit that introduced the union quoted this exact before/after
    measurement (``excluded_classrooms=[]`` with candidates
    ``['Oda 1', 'Lab 1', 'Lab 2']``, now ``['Lab 1']`` and ``['Oda 1', 'Lab 2']``
    — here narrowed further by row 1's lab type) and never turned it into an
    assertion, so both the union and its write-back to the primary could be
    deleted with the suite green.
    """
    rows = [joint_klass("C001", "B001", required_room_type=ROOM_TYPE_LAB),
            joint_klass("C002", "B002", excluded_rooms="Lab 1")]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_union.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert merged["excluded_classrooms"] == ["Lab 1"], (
        "the joint session lost the room its second row excluded: "
        f"excluded_classrooms={merged['excluded_classrooms']!r}")
    assert get_physical_room_candidates(ds.state, merged) == ["Lab 2"], (
        "the merged session can still be scheduled into a room one of its own "
        "rows ruled out; candidates are "
        f"{get_physical_room_candidates(ds.state, merged)}")
    # A union that leaves somewhere to go is not a conflict and must stay
    # silent — the warning below has to mean something when it does appear.
    assert ds.report.warnings == [], (
        f"an unremarkable joint group was warned about: {ds.report.warnings}")


def test_a_joint_merge_never_leaves_a_group_with_nowhere_to_meet(tmp_path):
    """REGRESSION — the exclusion union made placeable groups unplaceable.

    Row 1 requires the lecture type, which resolves to ``['Oda 1']``; row 2
    excludes ``Oda 1``. ``get_physical_room_candidates`` intersects
    ``required_classrooms`` with the live rooms and THEN subtracts
    ``excluded_classrooms``, so the union wrote a pair of lists with no room
    between them. Measured against the pre-merge importer
    (``git show f049964:scheduler_app/data_io/importer.py``):

        before  req=['Oda 1']  exc=[]         candidates=['Oda 1']
        after   req=['Oda 1']  exc=['Oda 1']  candidates=[]  warnings=0

    A joint session the school had been timetabling for years became impossible
    to place, and the import report was empty — the solver's unplaced list was
    the first the user heard of it. ``_process_classes`` already rescues and
    reports the identical contradiction when both cells sit on ONE row.
    """
    rows = [joint_klass("C001", "B001", required_room_type=ROOM_TYPE_LECTURE),
            joint_klass("C002", "B002", excluded_rooms="Oda 1")]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_nowhere.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert get_physical_room_candidates(ds.state, merged) == ["Oda 1"], (
        "the joint session has nowhere left to meet: required_classrooms="
        f"{merged['required_classrooms']!r} and excluded_classrooms="
        f"{merged['excluded_classrooms']!r} leave "
        f"{get_physical_room_candidates(ds.state, merged)}")
    assert joint_warning(ds, "warnings.joint_group_room_unplaceable",
                         group="J1", dropped="Oda 1", kept="Oda 1"), (
        "the merge dropped a row's room restriction to keep the session "
        "placeable and did not say so; the report holds "
        f"{ds.report.warnings}")


def test_three_rows_excluding_a_room_each_do_not_exclude_the_whole_building(
        tmp_path):
    """REGRESSION — the union shape no single row can produce.

    Three rows of one group, each excluding a different one of the three rooms.
    Every row is individually harmless; the union is every room in the
    building. Measured before the merge existed: candidates
    ``['Lab 1', 'Lab 2']``. After it: ``[]``, with an empty report.

    The third row's exclusion is the one that empties the group, so it is the
    one dropped — the same "earlier row wins, and the user is told" rule the
    merge already applies to two disjoint room types. ``Lab 2`` is therefore
    the only room left, and the two exclusions that fit are both honoured.
    """
    rows = [joint_klass("C001", "B001", excluded_rooms="Oda 1"),
            joint_klass("C002", "B002", excluded_rooms="Lab 1"),
            joint_klass("C003", "B003", excluded_rooms="Lab 2")]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_three_excl.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds, branches=("Grup A", "Grup B", "Grup C"))

    assert get_physical_room_candidates(ds.state, merged) == ["Lab 2"], (
        "three rows excluding one room each between them excluded the whole "
        f"building; excluded_classrooms={merged['excluded_classrooms']!r} "
        f"leaves {get_physical_room_candidates(ds.state, merged)}")
    assert merged["excluded_classrooms"] == ["Oda 1", "Lab 1"], (
        "the two exclusions that DO fit were not applied: "
        f"{merged['excluded_classrooms']!r}")
    assert joint_warning(ds, "warnings.joint_group_room_unplaceable",
                         group="J1", dropped="Lab 2", kept="Lab 2"), (
        f"the dropped exclusion was not reported; report: {ds.report.warnings}")


def test_a_room_type_that_every_other_row_excludes_is_reported_either_order(
        tmp_path):
    """REGRESSION — and it is not an artefact of which row comes first.

    The lab type on one row against an exclusion of every lab on the other,
    written both ways round. Both orderings emptied the candidate set silently;
    both must now leave the session somewhere to go and say what they dropped.
    The two answers differ — the rule is "the earlier row wins" — but neither
    is silent and neither is empty, which is the property that matters.
    """
    lab_first = [joint_klass("C001", "B001", required_room_type=ROOM_TYPE_LAB),
                 joint_klass("C002", "B002", excluded_rooms="Lab 1, Lab 2")]
    excl_first = [joint_klass("C001", "B001", excluded_rooms="Lab 1, Lab 2"),
                  joint_klass("C002", "B002", required_room_type=ROOM_TYPE_LAB)]

    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_lab_then_excl.xlsx", classes=lab_first))
    merged = merged_joint_class(ds)
    assert get_physical_room_candidates(ds.state, merged) == ["Lab 1", "Lab 2"], (
        "the lab row came first and the session still has nowhere to meet: "
        f"{get_physical_room_candidates(ds.state, merged)}")
    assert joint_warning(ds, "warnings.joint_group_room_unplaceable",
                         group="J1", dropped="Lab 1, Lab 2",
                         kept="Lab 1, Lab 2"), ds.report.warnings

    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_excl_then_lab.xlsx", classes=excl_first))
    merged = merged_joint_class(ds)
    assert get_physical_room_candidates(ds.state, merged) == ["Oda 1"], (
        "the exclusion row came first and the session still has nowhere to "
        f"meet: {get_physical_room_candidates(ds.state, merged)}")
    assert joint_warning(ds, "warnings.joint_group_room_unplaceable",
                         group="J1", dropped="Lab 1, Lab 2",
                         kept="Oda 1"), ds.report.warnings


def test_a_joint_group_says_which_room_type_it_had_to_ignore(tmp_path):
    """The disjoint-type warning, pinned. Kills mutants M4 and M6.

    ``Derslik`` resolves to ``['Oda 1']`` and ``Laboratuvar`` to
    ``['Lab 1', 'Lab 2']``. The merge cannot honour both, keeps the earlier
    row's list and must name what it ignored: this is the one place the fix
    tells the user it made a choice on their behalf, and both the choice and
    the sentence could be removed with the suite green.

    The sentence is rebuilt from the catalogue rather than quoted, so it also
    fails if the key stops shipping the placeholders the importer fills.
    """
    rows = joint_rows(lab_row_index=0, first_type=ROOM_TYPE_LECTURE,
                      second_type=ROOM_TYPE_LAB)
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_two_types_warned.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged_joint_class(ds)
    assert joint_warning(ds, "warnings.joint_group_room_conflict", group="J1",
                         kept="Oda 1", dropped="Lab 1, Lab 2"), (
        "the merge silently picked one of two disagreeing room types; the "
        f"report holds {ds.report.warnings}")


def test_a_joint_group_whose_rows_disagree_on_where_it_meets_says_so(tmp_path):
    """``location_type`` is not merged, and the silence was the defect.

    The fix above documented "the same two rows swapped must produce the same
    session". True of ``required_classrooms``; false of ``location_type``,
    which is still whatever ``classes[0]`` says.
    ``normalize_class_location_fields`` blanks the room lists of any
    non-physical class, and ``normalize_state_classes`` runs again after the
    merge, so an online row typed FIRST wipes the merged room list outright:

        online, then face-to-face + Laboratuvar -> online, required=[], report empty
        the same two rows swapped              -> face_to_face, ['Lab 1', 'Lab 2']

    A joint physics lab whose online row happens to be typed first imported as
    an online session with no room requirement at all. Merging the column was
    rejected — there is no defensible winner between "online" and
    "face-to-face" for one session, and picking one would flip whole sessions
    in existing workbooks on the strength of a stray cell — so both orderings
    now report the disagreement instead, which is what the merge already does
    for two disjoint room types.
    """
    online = get_location_label(LOCATION_ONLINE)
    in_person = get_location_label(LOCATION_FACE_TO_FACE)

    rows = [joint_klass("C001", "B001", location_type=online),
            joint_klass("C002", "B002", required_room_type=ROOM_TYPE_LAB)]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_online_first.xlsx", classes=rows))
    assert ds.report.is_valid, ds.report.summary()
    merged_joint_class(ds)
    assert joint_warning(ds, "warnings.joint_group_location_conflict",
                         group="J1", kept=online, dropped=in_person), (
        "an online first row swallowed a face-to-face row's lab requirement "
        f"without a word; the report holds {ds.report.warnings}")

    rows = [joint_klass("C001", "B001", required_room_type=ROOM_TYPE_LAB),
            joint_klass("C002", "B002", location_type=online)]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_online_second.xlsx", classes=rows))
    merged = merged_joint_class(ds)
    assert sorted(merged["required_classrooms"]) == ["Lab 1", "Lab 2"], (
        "the swapped ordering lost the lab constraint too: "
        f"{merged['required_classrooms']!r}")
    assert joint_warning(ds, "warnings.joint_group_location_conflict",
                         group="J1", kept=in_person, dropped=online), (
        "the same disagreement went unreported with the rows the other way "
        f"round; the report holds {ds.report.warnings}")


def test_a_joint_group_that_agrees_on_where_it_meets_is_not_warned(tmp_path):
    """Control for the test above — an all-online group must stay silent.

    Both rows say online, so there is no disagreement to report. Without this,
    the previous test is satisfied by a warning that fires on every joint group
    with a ``Location Type`` cell in it.
    """
    online = get_location_label(LOCATION_ONLINE)
    rows = [joint_klass("C001", "B001", location_type=online),
            joint_klass("C002", "B002", location_type=online)]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_all_online.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)
    assert merged["location_type"] == LOCATION_ONLINE
    assert ds.report.warnings == [], (
        f"a joint group whose rows agree was warned about: {ds.report.warnings}")


def test_a_room_named_twice_is_not_read_back_to_the_user_twice(tmp_path):
    """``allowed_rooms='Oda 1, Oda 1'`` reached state, and then the report.

    ``_parse_comma_list`` kept repeats verbatim, which was invisible until the
    merge started rendering ``required_classrooms`` into a sentence: the import
    report read "...the joint session keeps Oda 1, Oda 1 and ignores...". Every
    reader of these lists is a membership test, so de-duplicating changes no
    candidate set — it only stops the repeat reaching state and the user.
    """
    rows = [joint_klass("C001", "B001", allowed_rooms="Oda 1, Oda 1"),
            joint_klass("C002", "B002", required_room_type=ROOM_TYPE_LAB)]
    ds = load_scheduler_data_from_excel(
        build_workbook(tmp_path / "joint_dupe_rooms.xlsx", classes=rows))

    assert ds.report.is_valid, ds.report.summary()
    merged = merged_joint_class(ds)

    assert merged["required_classrooms"] == ["Oda 1"], (
        "a room named twice in one cell reached the merged session twice: "
        f"{merged['required_classrooms']!r}")
    assert get_physical_room_candidates(ds.state, merged) == ["Oda 1"]
    assert joint_warning(ds, "warnings.joint_group_room_conflict", group="J1",
                         kept="Oda 1", dropped="Lab 1, Lab 2"), (
        "the room-conflict sentence still repeats the room name: "
        f"{ds.report.warnings}")
