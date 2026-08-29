"""C6 — ``allowed_rooms`` too small for the head count is imported in silence.

The defect
----------
``_process_classes`` (``scheduler_app/data_io/importer.py``) contains one
head-count-versus-capacity check, and it lives *inside* ``if required_type:``
behind a second gate, ``type_decides_the_list = cls["required_classrooms"] ==
matching``. Both gates are deliberate and both are documented at the check
itself::

    # Scoped to the type-resolved list on purpose. A class whose
    # `allowed_rooms` are all too small reaches the same dead end and
    # is equally silent, but that path behaves exactly as it did at
    # 82f558e and is not this change's to fix.

So a row that names its rooms by hand instead of by type reaches the identical
dead end — ``get_physical_room_candidates`` is ``[]``, the class can never be
placed anywhere — and ``File ▸ Import Excel`` reports zero errors and zero
warnings.

Measured on this module's own three-room sheet (Oda 1/30, Lab 1/20, Lab 2/20):

===============================================  ========  ==========  ========
row                                              warnings  resolved    candidates
===============================================  ========  ==========  ========
``allowed_rooms='Lab 1, Lab 2'``, 25 students    **0**     Lab 1,Lab 2  ``[]``
``required_room_type='Laboratuvar'``, 25          1        Lab 1,Lab 2  ``[]``
``type='Laboratuvar' + allowed='Lab 1'``, 25     **0**     Lab 1        ``[]``
===============================================  ========  ==========  ========

Three rows, one dead end, one warning. The middle row is told at the moment of
import that the class "cannot be placed until the room is enlarged, the head
count lowered, or the type changed". The other two are told nothing until the
user has built a full Setup and run a solve.

What these tests pin
--------------------
Only that the row is *reported at import*, and reported truthfully. Nothing
here asserts a message id, a wording, or which function emits it. In
particular the third row must NOT be described over the room type — the seats
that were counted came from Allowed Rooms, and
``test_the_head_count_warning_is_not_raised_over_a_list_the_type_did_not_choose``
in ``test_import_roundtrip.py`` already holds that half. A fix satisfies both
only with a sentence quantified over Allowed Rooms.

The two guards below are green today and must stay green: a fix that warns on
every ``allowed_rooms`` row, or that warns an online class about rooms
``normalize_class_data`` is about to discard, fails them.
"""
import pytest

pytest.importorskip("pandas", reason="the Excel importer needs pandas")
pytest.importorskip("openpyxl", reason="workbook fixtures need openpyxl")

import openpyxl  # noqa: E402

from scheduler_app.core.models import (  # noqa: E402
    LOCATION_FACE_TO_FACE,
    LOCATION_ONLINE,
    get_location_label,
    get_physical_room_candidates,
)
from scheduler_app.data_io import schema  # noqa: E402
from scheduler_app.data_io.importer import load_scheduler_data_from_excel  # noqa: E402
from scheduler_app.translations import tr  # noqa: E402

pytestmark = pytest.mark.excel

SHEET_IDS = ("teachers", "rooms", "branches", "classes")

# Free text, matched against the Rooms sheet of the same workbook, never
# translated — same reasoning as ``test_import_roundtrip``'s literals.
ROOM_TYPE_LECTURE = "Derslik"
ROOM_TYPE_LAB = "Laboratuvar"

TEACHERS = [{"teacher_id": "T001", "name": "Ada Lovelace"}]
BRANCHES = [{"branch_id": "B001", "name": "Grup A"}]

# Oda 1 is big enough for every head count used here; the two labs are not.
ROOMS = [
    {"room_id": "R001", "name": "Oda 1", "capacity": 30,
     "room_type": ROOM_TYPE_LECTURE},
    {"room_id": "R002", "name": "Lab 1", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
    {"room_id": "R003", "name": "Lab 2", "capacity": 20,
     "room_type": ROOM_TYPE_LAB},
]

# A fourth room that makes "no room of type Laboratuvar seats 25" *false*, so
# the subset case below can only be reported over Allowed Rooms.
ROOMS_WITH_A_BIG_LAB = ROOMS + [
    {"room_id": "R004", "name": "Lab 3", "capacity": 50,
     "room_type": ROOM_TYPE_LAB},
]


def _fields(sheet_id):
    return [f for f, _, _ in schema.WORKBOOK_SHEETS[sheet_id]["columns"]]


def build_workbook(path, *, classes, rooms=None):
    """A workbook whose sheet titles and headers come from the schema itself."""
    rows_by_sheet = {
        "teachers": TEACHERS,
        "rooms": ROOMS if rooms is None else rooms,
        "branches": BRANCHES,
        "classes": list(classes),
    }
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_id in SHEET_IDS:
        ws = wb.create_sheet(schema.get_workbook_sheet_title(sheet_id))
        fields = _fields(sheet_id)
        headers = schema.get_workbook_sheet_header_map(sheet_id)
        for col, field in enumerate(fields, start=1):
            ws.cell(row=1, column=col, value=headers[field])
        excel_row = 2
        for row in rows_by_sheet[sheet_id]:
            for col, field in enumerate(fields, start=1):
                if field in row:
                    ws.cell(row=excel_row, column=col, value=row[field])
            excel_row += 1
    wb.save(str(path))
    return str(path)


def klass(class_id, **overrides):
    """A minimal, fully valid Classes row; override any field by keyword."""
    row = {
        "class_id": class_id,
        "course_name": f"Ders {class_id}",
        "teacher_id": "T001",
        "branch_id": "B001",
        "duration": 1,
        "student_count": 10,
        "joint_class_group": f"UNIQ-{class_id}",
        "location_type": get_location_label(LOCATION_FACE_TO_FACE),
    }
    row.update(overrides)
    return row


def messages(report):
    return list(report.errors) + list(report.warnings)


def row_prefix(sheet_id, excel_row):
    return tr("status.import_row_prefix").format(
        sheet=schema.get_workbook_sheet_title(sheet_id), row=excel_row)


def lines_for_row_2(report):
    """Report lines attached to the first Classes data row."""
    return [m for m in messages(report) if row_prefix("classes", 2) in m]


# ── the defect ──────────────────────────────────────────────────────────────

def test_allowed_rooms_that_are_all_too_small_are_reported_at_import(tmp_path):
    """A hand-typed room list that cannot seat the class must not import mute.

    ``allowed_rooms='Lab 1, Lab 2'`` with 25 students, against a Rooms sheet
    whose labs seat 20 each. ``get_physical_room_candidates`` is ``[]`` before
    the user has drawn a single day on the grid: this class can never be
    placed, in any Setup, by any solver run.

    Measured on 42e1943: ``is_valid=True``, ``errors=[]``, ``warnings=[]``.
    The identical contradiction expressed as ``required_room_type`` warns —
    see the control below — so the school that types room names is the one
    that finds out last.
    """
    rows = [klass("C001", student_count=25, allowed_rooms="Lab 1, Lab 2")]
    path = build_workbook(tmp_path / "allowed_too_small.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    cls = ds.state["classes"][0]
    # The dead end, first — if this ever stops being empty the test is no
    # longer about what it says it is about.
    assert cls["required_classrooms"] == ["Lab 1", "Lab 2"]
    assert get_physical_room_candidates(ds.state, cls) == [], (
        "premise broken: the class is placeable, so there is nothing to warn "
        "about")

    said = lines_for_row_2(ds.report)
    assert said, (
        "the row imported in silence; a class that can never be placed "
        "anywhere was added with no warning at all. Whole report: %r"
        % (messages(ds.report),))
    # Truthful about the row it is attached to: the head count it counted and
    # the largest room the user actually allowed.
    assert any("25" in m and "Lab 1" in m and "20" in m for m in said), (
        "the row was reported, but not with the numbers the user must act on "
        "(25 students, largest allowed room Lab 1 with 20): %r" % (said,))


def test_allowed_rooms_narrowing_under_a_type_is_reported_over_the_right_column(
        tmp_path):
    """The second silent shape: Allowed Rooms narrows *inside* a room type.

    ``required_room_type='Laboratuvar', allowed_rooms='Lab 1'`` with 25
    students, against a Rooms sheet that also carries Lab 3 (Laboratuvar,
    seats 50). ``type_decides_the_list`` is False here — the resolved list is
    ``['Lab 1']``, not the type's ``['Lab 1', 'Lab 2', 'Lab 3']`` — so the one
    capacity check in the importer is skipped and the row says nothing.

    The gate is right about the *wording*: "no room of type Laboratuvar seats
    25" is false, Lab 3 seats 50, and
    ``test_the_head_count_warning_is_not_raised_over_a_list_the_type_did_not_choose``
    exists to keep it false. It is wrong about the *silence*: the row is
    unplaceable and the cell to edit is Allowed Rooms.
    """
    rows = [klass("C001", student_count=25,
                  required_room_type=ROOM_TYPE_LAB, allowed_rooms="Lab 1")]
    path = build_workbook(tmp_path / "narrowed_too_small.xlsx", classes=rows,
                          rooms=ROOMS_WITH_A_BIG_LAB)
    ds = load_scheduler_data_from_excel(path)

    cls = ds.state["classes"][0]
    assert cls["required_classrooms"] == ["Lab 1"], (
        "premise broken: the resolved list moved")
    assert get_physical_room_candidates(ds.state, cls) == []

    said = lines_for_row_2(ds.report)
    assert said, (
        "the row imported in silence: %r" % (messages(ds.report),))
    # The existing contract, restated so a fix cannot satisfy one and break
    # the other: the sentence may not quantify over the room type.
    assert not any(ROOM_TYPE_LAB in m for m in said), (
        "the row was described over the room type while the seats counted "
        "came from Allowed Rooms — Lab 3 is a Laboratuvar and seats 50: %r"
        % (said,))
    assert any("25" in m and "Lab 1" in m and "20" in m for m in said), (
        "the row was reported without the numbers the user must act on: %r"
        % (said,))


# ── the control: the half Phase 8 did close ─────────────────────────────────

def test_the_type_resolved_half_still_reports(tmp_path):
    """Green today. Present so the asymmetry is visible in one file.

    Same head count, same rooms, same empty candidate list — expressed as a
    room type instead of a room list, and the importer speaks.
    """
    rows = [klass("C001", student_count=25,
                  required_room_type=ROOM_TYPE_LAB)]
    path = build_workbook(tmp_path / "type_too_small.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    cls = ds.state["classes"][0]
    assert get_physical_room_candidates(ds.state, cls) == []
    assert lines_for_row_2(ds.report), (
        "the control failed: the type-resolved half stopped reporting, which "
        "means this file is measuring something else now")


# ── guards: a fix must not over-fire ────────────────────────────────────────

def test_allowed_rooms_that_do_fit_are_not_warned(tmp_path):
    """Green today, and a fix that warns on every Allowed Rooms row breaks it."""
    rows = [klass("C001", student_count=25, allowed_rooms="Oda 1, Lab 1")]
    path = build_workbook(tmp_path / "allowed_fits.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    cls = ds.state["classes"][0]
    assert get_physical_room_candidates(ds.state, cls) == ["Oda 1"]
    assert not lines_for_row_2(ds.report), (
        "a placeable row was warned: %r" % (messages(ds.report),))


def test_an_online_class_is_not_warned_about_rooms_it_will_never_use(tmp_path):
    """Green today. ``class_uses_physical_room`` must gate any widening.

    An online class gets the ``[None]`` virtual sentinel from
    ``get_room_candidates`` and places normally; ``normalize_class_data`` then
    throws ``required_classrooms`` away entirely. Warning it that it "cannot
    be placed" would be false of the row — the same mistake the type-resolved
    check already had to be taught not to make.
    """
    rows = [klass("C001", student_count=25, allowed_rooms="Lab 1, Lab 2",
                  location_type=get_location_label(LOCATION_ONLINE))]
    path = build_workbook(tmp_path / "online_too_small.xlsx", classes=rows)
    ds = load_scheduler_data_from_excel(path)

    assert not lines_for_row_2(ds.report), (
        "an online class was warned about classroom capacity: %r"
        % (messages(ds.report),))
